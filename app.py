import os
import io
import csv
import json
import uuid
import math
import contextlib
import smtplib
import logging
import secrets
import calendar
import unicodedata
from collections import deque
from contextlib import asynccontextmanager
from datetime import date, datetime, time, timedelta, timezone
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from pathlib import Path
from threading import RLock
from typing import Any, Optional
from urllib.error import HTTPError, URLError
from urllib.parse import urlparse
from urllib.request import Request as UrlLibRequest, urlopen

from fastapi import Depends, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import Boolean, Date, DateTime, Float, ForeignKey, Integer, String, Text, UniqueConstraint, and_, create_engine, or_, text
from sqlalchemy.dialects.postgresql import UUID
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import DeclarativeBase, Mapped, Session, mapped_column, relationship, sessionmaker
from sqlalchemy.sql import func
from openpyxl import load_workbook
from auth_runtime import (
    PasswordPolicyConfig,
    SessionConfig,
    SessionManager,
    hash_password as runtime_hash_password,
    new_salt as runtime_new_salt,
    password_policy_error as runtime_password_policy_error,
    verify_password as runtime_verify_password,
)
from bootstrap_runtime import ensure_frankstein_tables as runtime_ensure_frankstein_tables
from security_utils import decrypt_pii, encrypt_pii, generate_pii_encryption_key, hash_email, hash_optional, hash_token, last4_digits, mask_document, mask_email, pii_encryption_enabled
from simulacao_engine import SimulacaoInput, engine_calculo_imobiliario
from frankstein_operacional import AnaliseInput as FranksteinAnaliseInput
from frankstein_operacional import RespostaFrankstein as FranksteinRespostaOperacional
from frankstein_operacional import analisar_operacao_frankstein
from workflow_constants import (
    CAIXA_ASSINATURA_APTA_STATUSES,
    CSV_IMPORT_DELIMITERS,
    CSV_IMPORT_ENCODINGS,
    ESTAGIOS_DASH_COMERCIAL,
    ESTAGIOS_REPASSE_COMERCIAL,
    ESTAGIO_COMERCIAL_INDEX,
    ESTAGIO_COMERCIAL_SET,
    ESTAGIO_COMERCIAL_VALUES,
    IMPORT_COLUMN_ALIASES,
    IMPORT_REQUIRED_COLUMNS,
    LEAD_CCA_DECISION_SET,
    LEAD_CCA_DECISION_VALUES,
    LEAD_STAGE_SET,
    LEAD_STAGE_VALUES,
    PLANEJAMENTO_DIA_TODO_META_PREFIX,
    PLANEJAMENTO_ENTREGA_META_PREFIX,
    PLANEJAMENTO_ENTREGA_STATUS_LABELS,
    PLANEJAMENTO_STATUS_LABELS,
    PLANEJAMENTO_TIPO_LABELS,
    PROCESS_AGEHAB_STATUSES,
    PROCESS_CAIXA_STATUSES,
    PROCESS_CCA_FINAL_STATUSES,
    PROCESS_CREDITO_STATUSES,
    PROCESS_FIADOR_STATUSES,
    PROCESS_GERAL_ARQUIVO_IMEDIATO,
    PROCESS_GERAL_FINAL_STATUSES,
    PROCESS_GERAL_STATUSES,
    PROCESS_OVERVIEW_LABELS,
    PROCESS_READY_STATUS_KEYS,
    PROCESS_RECOLHA_FGTS_STATUSES,
    PROCESS_SINAL_STATUSES,
    REPASSE_ETAPAS_SET,
    REPASSE_ETAPAS_VALUES,
    SLA_OWNER_ANALISTA,
    SLA_OWNER_CCA,
    SLA_OWNER_CORRETOR,
    SLA_OWNER_NONE,
    SLA_OWNER_VALUES,
    UNIDADE_STATUS_SET,
    UNIDADE_STATUS_VALUES,
)
import psycopg

logger = logging.getLogger("sistema_credito")

WEB_DIR = Path(__file__).resolve().parent / "web"
REACT_DIST_DIR = Path(__file__).resolve().parent / "frontend-react" / "dist"
DATA_DIR = Path(__file__).resolve().parent / "data"
DATA_DIR.mkdir(exist_ok=True, parents=True)
FRANKSTEIN_RAW_DIR = DATA_DIR / "raw"
FRANKSTEIN_PROCESSED_DIR = DATA_DIR / "processed"
FRANKSTEIN_MODELS_DIR = DATA_DIR / "models"
FRANKSTEIN_CURRENT_MODELS_DIR = FRANKSTEIN_MODELS_DIR / "current"
FRANKSTEIN_CANDIDATES_DIR = FRANKSTEIN_MODELS_DIR / "candidates"
FRANKSTEIN_ARCHIVE_MODELS_DIR = FRANKSTEIN_MODELS_DIR / "archive"
FRANKSTEIN_RAW_DIR.mkdir(exist_ok=True, parents=True)
IA_FEEDBACK_PATH = DATA_DIR / "ia_feedback.json"
FRANKSTEIN_EVENTS_PATH = FRANKSTEIN_RAW_DIR / "frankstein_events.json"
FRANKSTEIN_EVENTS_LEGACY_PATH = DATA_DIR / "yvy_events.json"
FRANKSTEIN_MODEL_PATH = DATA_DIR / "frankstein_model.json"
FRANKSTEIN_MODEL_LEGACY_PATH = DATA_DIR / "yvy_model.json"
FRANKSTEIN_CURRENT_SKLEARN_MODEL_PATH = FRANKSTEIN_CURRENT_MODELS_DIR / "frankstein_model.pkl"
FRANKSTEIN_CURRENT_SKLEARN_MODEL_LEGACY_PATH = FRANKSTEIN_CURRENT_MODELS_DIR / "yvy_model.pkl"
FRANKSTEIN_CURRENT_FEATURE_COLUMNS_PATH = FRANKSTEIN_CURRENT_MODELS_DIR / "feature_columns.json"
FRANKSTEIN_CURRENT_METRICS_PATH = FRANKSTEIN_CURRENT_MODELS_DIR / "metrics.json"
FRANKSTEIN_CURRENT_MODEL_INFO_PATH = FRANKSTEIN_CURRENT_MODELS_DIR / "model_info.json"
FRANKSTEIN_REGISTRY_DIR = Path(__file__).resolve().parent / "registry"
FRANKSTEIN_MODEL_REGISTRY_PATH = FRANKSTEIN_REGISTRY_DIR / "model_registry.json"
FRANKSTEIN_METRICS_HISTORY_PATH = FRANKSTEIN_REGISTRY_DIR / "metrics_history.json"
TABELA_PRECO_PATH = DATA_DIR / "tabela_precos.json"
TABELA_PRECO_CACHE: list[dict[str, Any]] = []
TABELA_PRECO_CACHE_LOCK = RLock()
FRANKSTEIN_DB_URL = os.getenv("FRANKSTEIN_DB_URL") or os.getenv("YVY_DB_URL") or os.getenv("DATABASE_URL")

ANALISE_DB_PATH = DATA_DIR / "analises.db"
analise_engine = create_engine(f"sqlite:///{ANALISE_DB_PATH}", future=True)
AnaliseSessionLocal = sessionmaker(analise_engine, expire_on_commit=False)

SESSION_COOKIE_NAME = "sc_session"
SESSION_TTL_SECONDS = int(os.getenv("SESSION_TTL_SECONDS", "43200"))
SESSION_IDLE_TIMEOUT_SECONDS = int(os.getenv("SESSION_IDLE_TIMEOUT_SECONDS", "1800"))
SESSION_DB_SYNC_INTERVAL_SECONDS = int(os.getenv("SESSION_DB_SYNC_INTERVAL_SECONDS", "60"))
SESSION_IDLE_PASSIVE_PATHS = {
    "/app/api/processos",
}
AUTH_SESSION_CONFIG = SessionConfig(
    session_cookie_name=SESSION_COOKIE_NAME,
    ttl_seconds=SESSION_TTL_SECONDS,
    idle_timeout_seconds=SESSION_IDLE_TIMEOUT_SECONDS,
    db_sync_interval_seconds=SESSION_DB_SYNC_INTERVAL_SECONDS,
    idle_passive_paths=set(SESSION_IDLE_PASSIVE_PATHS),
)
SESSION_COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "false").lower() in {"1", "true", "yes"}
CORRETOR_ROUTE_ENABLED = os.getenv("CORRETOR_ROUTE_ENABLED", "true").lower() in {"1", "true", "yes"}
try:
    GESTOR_META_MENSAL = max(0, int(os.getenv("GESTOR_META_MENSAL", "0")))
except ValueError:
    GESTOR_META_MENSAL = 0
try:
    GESTOR_META_SEMANAL = max(0, int(os.getenv("GESTOR_META_SEMANAL", "0")))
except ValueError:
    GESTOR_META_SEMANAL = 0
META_MENSAL_RUNTIME_KEY = "gestor_meta_mensal"
META_SEMANAL_RUNTIME_KEY = "gestor_meta_semanal"
LAYOUT_BLACKHOLE_RUNTIME_KEY = "layout_blackhole_enabled"
BLACKHOLE_LAYOUT_ALLOWED = os.getenv("BLACKHOLE_LAYOUT_ALLOWED", "false").lower() in {"1", "true", "yes", "on"}
FRANKSTEIN_ALERT_EMAIL_TO_RUNTIME_KEY = "frankstein_alert_email_to"
USERS_SEED_MODE_RUNTIME_KEY = "users_seed_mode"
USERS_SEED_MODE_FULL = "full"
USERS_SEED_MODE_ADMIN_ONLY = "admin_only"
REPASSE_ARQUIVO_PERIODO_RUNTIME_KEY = "repasse_arquivo_periodo"
RESET_ADMIN_USERNAME = "douglasadm"
ROLE_CORRETOR = "corretor"
ROLE_CCA = "cca"
ROLE_ANALISTA = "analista"
ROLE_ADMIN = "admin"
ROLE_GESTOR = "gestor"
ROLE_GESTOR_CREDITO = "gestor_credito"

VALID_ROLES = {ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_ADMIN, ROLE_GESTOR, ROLE_GESTOR_CREDITO}
CHECKLIST_PAGE_ROLES = (ROLE_CORRETOR, ROLE_ANALISTA, ROLE_CCA, ROLE_ADMIN)
PROCESSO_INTAKE_ROLES = (ROLE_CORRETOR, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)
PROCESSO_FULL_ROLES = (ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)
PROCESSO_DOCUMENTOS_UPSERT_ROLES = PROCESSO_FULL_ROLES

APP_CCA_USER = os.getenv("APP_CCA_USER", os.getenv("APP_LOGIN_USER", "cca"))
APP_CCA_PASSWORD = os.getenv("APP_CCA_PASSWORD", os.getenv("APP_LOGIN_PASSWORD", ""))
APP_ANALISTA_USER = os.getenv("APP_ANALISTA_USER", "analista")
APP_ANALISTA_PASSWORD = os.getenv("APP_ANALISTA_PASSWORD", "")
APP_CORRETOR_USER = os.getenv("APP_CORRETOR_USER", "corretor")
APP_CORRETOR_PASSWORD = os.getenv("APP_CORRETOR_PASSWORD", "")
APP_ADMIN_USER = os.getenv("APP_ADMIN_USER", "douglasadm")
APP_ADMIN_PASSWORD = os.getenv("APP_ADMIN_PASSWORD", "")
APP_GESTOR_USER = os.getenv("APP_GESTOR_USER", "gestor")
APP_GESTOR_PASSWORD = os.getenv("APP_GESTOR_PASSWORD", "")
APP_GESTOR_CREDITO_USER = os.getenv("APP_GESTOR_CREDITO_USER", "")
APP_GESTOR_CREDITO_PASSWORD = os.getenv("APP_GESTOR_CREDITO_PASSWORD", "")


class TabelaPrecoItem(BaseModel):
  empreendimento: str
  unidade: str
  garantido_minimo: float
  preco: float
  sobrepreco: float
  is_maximo: float
  prosoluto_minimo: float


class TabelaPrecoUploadResponse(BaseModel):
  linhas: int
  filename: str


class AnaliseCreate(BaseModel):
  empreendimento: str
  unidade: str
  preco_imovel: float
  valor_obtido: float
  prosoluto_calculado: float
  prosoluto_liquido: float
  sinal: float
  sinal_produto: float
  financiamento: float
  subsidio: float
  cheque_moradia: float
  renda_bruta: float
  perc_construcao: float
  is_agora: float
  is_pos_chaves: float
  tabela_referencia: list[dict[str, Any]] = Field(default_factory=list)
  data_referencia: datetime

  model_config = ConfigDict(arbitrary_types_allowed=True)


def normalize_header(value: str) -> str:
  cleaned = unicodedata.normalize("NFKD", (value or "")).encode("ascii", "ignore").decode()
  return "_".join(cleaned.strip().lower().split())


def parse_number(value: Any) -> float:
  try:
    return float(value)
  except (TypeError, ValueError):
    return 0.0


def _tabela_precos_payload(rows: list["TabelaPreco"]) -> list[dict[str, Any]]:
  return [
      {
          "empreendimento": row.empreendimento,
          "unidade": row.unidade,
          "garantido_minimo": row.garantido_minimo,
          "preco": row.preco,
          "sobrepreco": row.sobrepreco,
          "is_maximo": row.is_maximo,
          "prosoluto_minimo": row.prosoluto_minimo,
      }
      for row in rows
  ]


def carregar_tabela_precos() -> list[TabelaPrecoItem]:
  if SessionLocal is not None:
    try:
      with SessionLocal() as db:
        payload = _tabela_precos_payload(
            db.query(TabelaPreco).order_by(TabelaPreco.empreendimento.asc(), TabelaPreco.unidade.asc()).all()
        )
        with TABELA_PRECO_CACHE_LOCK:
          TABELA_PRECO_CACHE[:] = payload
        return payload
    except SQLAlchemyError:
      logger.exception("Falha ao carregar tabela de precos do banco.")
      with TABELA_PRECO_CACHE_LOCK:
        if TABELA_PRECO_CACHE:
          return list(TABELA_PRECO_CACHE)
      raise

  with TABELA_PRECO_CACHE_LOCK:
    if TABELA_PRECO_CACHE:
      return list(TABELA_PRECO_CACHE)
  if not TABELA_PRECO_PATH.exists():
    return []
  try:
    payload = json.loads(TABELA_PRECO_PATH.read_text(encoding="utf-8"))
  except json.JSONDecodeError:
    return []
  with TABELA_PRECO_CACHE_LOCK:
    TABELA_PRECO_CACHE[:] = payload
  return payload


def salvar_tabela_precos(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
  if SessionLocal is not None:
    try:
      with SessionLocal() as db:
        db.query(TabelaPreco).delete()
        db.add_all([TabelaPreco(**row) for row in rows])
        db.commit()
        payload = _tabela_precos_payload(
            db.query(TabelaPreco).order_by(TabelaPreco.empreendimento.asc(), TabelaPreco.unidade.asc()).all()
        )
        with TABELA_PRECO_CACHE_LOCK:
          TABELA_PRECO_CACHE[:] = payload
        return payload
    except SQLAlchemyError:
      logger.exception("Falha ao salvar tabela de precos no banco.")
      raise

  with TABELA_PRECO_CACHE_LOCK:
    TABELA_PRECO_CACHE[:] = rows
  TABELA_PRECO_PATH.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
  return rows


class AnaliseBase(DeclarativeBase):
  pass


class Analise(AnaliseBase):
  __tablename__ = "analises"

  id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
  created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=datetime.utcnow)
  empreendimento: Mapped[str] = mapped_column(String(120))
  unidade: Mapped[str] = mapped_column(String(120))
  preco_imovel: Mapped[float] = mapped_column(Float)
  valor_obtido: Mapped[float] = mapped_column(Float)
  prosoluto_calculado: Mapped[float] = mapped_column(Float)
  prosoluto_liquido: Mapped[float] = mapped_column(Float)
  sinal: Mapped[float] = mapped_column(Float)
  sinal_produto: Mapped[float] = mapped_column(Float)
  financiamento: Mapped[float] = mapped_column(Float)
  subsidio: Mapped[float] = mapped_column(Float)
  cheque_moradia: Mapped[float] = mapped_column(Float)
  renda_bruta: Mapped[float] = mapped_column(Float)
  perc_construcao: Mapped[float] = mapped_column(Float)
  is_agora: Mapped[float] = mapped_column(Float)
  is_pos_chaves: Mapped[float] = mapped_column(Float)
  tabela_referencia: Mapped[str] = mapped_column(Text)  # JSON string de tabela no upload
  data_referencia: Mapped[datetime] = mapped_column(DateTime(timezone=True))


AnaliseBase.metadata.create_all(analise_engine)


async def processar_tabela_upload(upload: UploadFile) -> list[TabelaPrecoItem]:
  content = await upload.read()
  if not content:
    raise HTTPException(status_code=400, detail="Arquivo vazio.")

  filename = (upload.filename or "").lower()
  required = {
    "empreendimento": "empreendimento",
    "unidade": "unidade",
    "garantido_minimo": "garantido_minimo",
    "preco": "preco",
    "sobrepreco": "sobrepreco",
    "is_maximo": "is_maximo",
    "prosoluto_minimo": "prosoluto_minimo",
  }
  rows: list[TabelaPrecoItem] = []

  if filename.endswith(".csv"):
    text = content.decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
      raise HTTPException(status_code=400, detail="Cabecalho nao encontrado no CSV.")
    header_map = {normalize_header(name): name for name in reader.fieldnames}
    missing = [col for col in required if col not in header_map]
    if missing:
      raise HTTPException(status_code=400, detail=f"Colunas ausentes: {', '.join(missing)}")
    for line in reader:
      rows.append(
        {
          "empreendimento": str(line.get(header_map["empreendimento"], "")).strip(),
          "unidade": str(line.get(header_map["unidade"], "")).strip(),
          "garantido_minimo": parse_number(line.get(header_map["garantido_minimo"])),
          "preco": parse_number(line.get(header_map["preco"])),
          "sobrepreco": parse_number(line.get(header_map["sobrepreco"])),
          "is_maximo": parse_number(line.get(header_map["is_maximo"])),
          "prosoluto_minimo": parse_number(line.get(header_map["prosoluto_minimo"])),
        }
      )
    return rows

  # XLSX / XLS
  workbook = load_workbook(io.BytesIO(content), data_only=True)
  sheet = workbook.active
  header_cells = [normalize_header(str(cell.value or "")) for cell in next(sheet.iter_rows(max_row=1))]
  header_map = {name: idx for idx, name in enumerate(header_cells)}
  missing = [col for col in required if col not in header_map]
  if missing:
    raise HTTPException(status_code=400, detail=f"Colunas ausentes: {', '.join(missing)}")

  for row in sheet.iter_rows(min_row=2):
    rows.append(
      {
        "empreendimento": str(row[header_map["empreendimento"]].value or "").strip(),
        "unidade": str(row[header_map["unidade"]].value or "").strip(),
        "garantido_minimo": parse_number(row[header_map["garantido_minimo"]].value),
        "preco": parse_number(row[header_map["preco"]].value),
        "sobrepreco": parse_number(row[header_map["sobrepreco"]].value),
        "is_maximo": parse_number(row[header_map["is_maximo"]].value),
        "prosoluto_minimo": parse_number(row[header_map["prosoluto_minimo"]].value),
      }
    )

  return rows
FORCE_RECOVER_ADMIN_ON_STARTUP = os.getenv("FORCE_RECOVER_ADMIN_ON_STARTUP", "false").lower() in {"1", "true", "yes"}
EMAIL_SMTP_HOST = (os.getenv("EMAIL_SMTP_HOST", "") or "").strip()
try:
    EMAIL_SMTP_PORT = int(os.getenv("EMAIL_SMTP_PORT", "587"))
except ValueError:
    EMAIL_SMTP_PORT = 587
EMAIL_SMTP_USER = (os.getenv("EMAIL_SMTP_USER", "") or "").strip()
EMAIL_SMTP_PASSWORD = os.getenv("EMAIL_SMTP_PASSWORD", "")
EMAIL_SMTP_FROM = (os.getenv("EMAIL_SMTP_FROM", "") or "").strip()
EMAIL_SMTP_STARTTLS = os.getenv("EMAIL_SMTP_STARTTLS", "true").lower() in {"1", "true", "yes"}
EMAIL_FROM = (os.getenv("EMAIL_FROM", "") or EMAIL_SMTP_FROM).strip()
AI_ASSISTANT_DISPLAY_NAME = (os.getenv("AI_ASSISTANT_DISPLAY_NAME", "") or "Foguetinho").strip()
EMAIL_FROM_NAME = (os.getenv("EMAIL_FROM_NAME", "") or f"{AI_ASSISTANT_DISPLAY_NAME} SioCred").strip()
EMAIL_DELIVERY_PROVIDER = (os.getenv("EMAIL_DELIVERY_PROVIDER", "") or os.getenv("EMAIL_PROVIDER", "")).strip().lower()
EMAIL_BREVO_API_KEY = (os.getenv("EMAIL_BREVO_API_KEY", "") or os.getenv("BREVO_API_KEY", "")).strip()
EMAIL_BREVO_API_URL = (os.getenv("EMAIL_BREVO_API_URL", "") or "https://api.brevo.com/v3/smtp/email").strip()
EMAIL_CONFIRM_LINK_BASE_URL = (os.getenv("EMAIL_CONFIRM_LINK_BASE_URL", "") or "").strip()
try:
    EMAIL_CONFIRM_TOKEN_TTL_HOURS = max(1, int(os.getenv("EMAIL_CONFIRM_TOKEN_TTL_HOURS", "72")))
except ValueError:
    EMAIL_CONFIRM_TOKEN_TTL_HOURS = 72

PASSWORD_HASH_ITERATIONS = int(os.getenv("PASSWORD_HASH_ITERATIONS", "200000"))
PASSWORD_MIN_LENGTH = int(os.getenv("PASSWORD_MIN_LENGTH", "10"))
PASSWORD_REQUIRE_UPPER = os.getenv("PASSWORD_REQUIRE_UPPER", "true").lower() in {"1", "true", "yes"}
PASSWORD_REQUIRE_LOWER = os.getenv("PASSWORD_REQUIRE_LOWER", "true").lower() in {"1", "true", "yes"}
PASSWORD_REQUIRE_DIGIT = os.getenv("PASSWORD_REQUIRE_DIGIT", "true").lower() in {"1", "true", "yes"}
PASSWORD_REQUIRE_SYMBOL = os.getenv("PASSWORD_REQUIRE_SYMBOL", "true").lower() in {"1", "true", "yes"}
PASSWORD_POLICY = PasswordPolicyConfig(
    iterations=PASSWORD_HASH_ITERATIONS,
    min_length=PASSWORD_MIN_LENGTH,
    require_upper=PASSWORD_REQUIRE_UPPER,
    require_lower=PASSWORD_REQUIRE_LOWER,
    require_digit=PASSWORD_REQUIRE_DIGIT,
    require_symbol=PASSWORD_REQUIRE_SYMBOL,
)
ALLOW_WEAK_SEED_PASSWORDS = os.getenv("ALLOW_WEAK_SEED_PASSWORDS", "false").lower() in {"1", "true", "yes"}
ENABLE_LEGACY_DEMO_USERS = os.getenv("ENABLE_LEGACY_DEMO_USERS", "false").lower() in {"1", "true", "yes"}
ACTIVE_SESSIONS: dict[str, dict[str, Any]] = {}
ACTIVE_SESSIONS_LOCK = RLock()
AUTH_SESSION_MANAGER: Optional[SessionManager] = None
PROCESS_LIST_CACHE_TTL_SECONDS = int(os.getenv("PROCESS_LIST_CACHE_TTL_SECONDS", "8"))
try:
    FALL_RISK_DAYS = max(1, int(os.getenv("FALL_RISK_DAYS", "15")))
except ValueError:
    FALL_RISK_DAYS = 15
try:
    DASHBOARD_MAX_DIAS_EM_ABERTO = max(30, int(os.getenv("DASHBOARD_MAX_DIAS_EM_ABERTO", "730")))
except ValueError:
    DASHBOARD_MAX_DIAS_EM_ABERTO = 730
try:
    DASHBOARD_IMPORT_DATE_BACKFILL_TOLERANCE_DAYS = max(
        0,
        int(os.getenv("DASHBOARD_IMPORT_DATE_BACKFILL_TOLERANCE_DAYS", "365")),
    )
except ValueError:
    DASHBOARD_IMPORT_DATE_BACKFILL_TOLERANCE_DAYS = 365
RUNTIME_SCHEMA_REVISION = "2026-03-01-cca-analise-financeira-v2"
PENDENCIA_INFO_MIN_LENGTH = 0
PROCESS_LIST_CACHE: dict[str, dict[str, Any]] = {}
PROCESS_LIST_CACHE_LOCK = RLock()
SEED_USERS_READY = False
CREDITO_PLANEJAMENTO_TIPOS = {"tarefa", "subtarefa", "agendamento", "entrega", "urgente", "anotacao"}
CREDITO_PLANEJAMENTO_STATUS = {"pendente", "em_andamento", "concluido", "atrasado"}
SIOCRED_ALERT_EMAIL_TO = (os.getenv("SIOCRED_ALERT_EMAIL_TO", "") or "").strip()
SIOCRED_EMAIL_ALERT_TOKEN = (os.getenv("SIOCRED_EMAIL_ALERT_TOKEN", "") or "").strip()
try:
    SIOCRED_EMAIL_ALERT_WINDOW_MINUTES = max(1, int(os.getenv("SIOCRED_EMAIL_ALERT_WINDOW_MINUTES", "6")))
except ValueError:
    SIOCRED_EMAIL_ALERT_WINDOW_MINUTES = 6
ANALISTA_REUNIAO_FOLLOWUP_STATUS = {"seguir", "finalizar_hoje", "assinado"}
ANALISTA_REUNIAO_COMPROMISSO_STATUS = {"pendente", "nao_entregue", "entregue"}
ANALISTA_REUNIAO_ESTAGIOS = {"EM_PROCESSO", "CREDITO", "SECRETARIA_VENDAS"}
KEEPALIVE_TELA = "health_keepalive"
KEEPALIVE_BRT_TZ = timezone(timedelta(hours=-3))
try:
    KEEPALIVE_LOG_MEMORY_LIMIT = max(50, int(os.getenv("KEEPALIVE_LOG_MEMORY_LIMIT", "500")))
except ValueError:
    KEEPALIVE_LOG_MEMORY_LIMIT = 500
KEEPALIVE_RECENT: deque[dict[str, Any]] = deque(maxlen=KEEPALIVE_LOG_MEMORY_LIMIT)


def _normalize_username(value: Optional[str]) -> str:
    return (value or "").strip().lower()


def _normalize_corretor_nome_curto(value: Optional[str]) -> str:
    cleaned = " ".join(str(value or "").strip().split())
    if not cleaned:
        return ""

    # Remove observacoes comuns no fim (ex.: " - CLT", " | Equipe", " / Time").
    for separator in (" - ", " | ", " / "):
        if separator in cleaned:
            cleaned = cleaned.split(separator, 1)[0].strip()

    # Remove sufixo entre parenteses quando vier no final.
    if cleaned.endswith(")") and "(" in cleaned:
        cleaned = cleaned[: cleaned.rfind("(")].strip()

    partes = [parte for parte in cleaned.split(" ") if parte]
    if not partes:
        return ""
    if len(partes) == 1:
        return partes[0].lower()
    if len(partes) == 2:
        return f"{partes[0]} {partes[1]}".lower()
    return f"{partes[0]} {partes[-1]}".lower()


def _build_app_users() -> dict[str, dict[str, str]]:
    users: dict[str, dict[str, str]] = {}
    configs = [
        (APP_CORRETOR_USER, APP_CORRETOR_PASSWORD, ROLE_CORRETOR),
        (APP_CCA_USER, APP_CCA_PASSWORD, ROLE_CCA),
        (APP_ANALISTA_USER, APP_ANALISTA_PASSWORD, ROLE_ANALISTA),
        (APP_ADMIN_USER, APP_ADMIN_PASSWORD, ROLE_ADMIN),
        (APP_GESTOR_USER, APP_GESTOR_PASSWORD, ROLE_GESTOR),
        (APP_GESTOR_CREDITO_USER, APP_GESTOR_CREDITO_PASSWORD, ROLE_GESTOR_CREDITO),
    ]
    if ENABLE_LEGACY_DEMO_USERS:
        configs.extend(
            [
                ("Douglas", "1234", ROLE_ANALISTA),
                ("Fabio", "1234", ROLE_CORRETOR),
                ("Endy", "1234", ROLE_CCA),
                ("Douglasadm", "12345", ROLE_ADMIN),
            ]
        )

    for username_raw, password, role in configs:
        username = _normalize_username(username_raw)
        if not username or not password:
            continue
        if username in users:
            logger.warning("Usuario duplicado nas credenciais da aplicacao: %s", username)
        users[username] = {"password": password, "role": role}
    return users


APP_USERS = _build_app_users()


def _normalize_database_url(raw_url: str) -> str:
    url = raw_url.strip()
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql+psycopg://", 1)
    elif url.startswith("postgresql://"):
        url = url.replace("postgresql://", "postgresql+psycopg://", 1)

    if "sslmode=" not in url:
        joiner = "&" if "?" in url else "?"
        url = f"{url}{joiner}sslmode=require"
    return url


def _looks_like_placeholder_url(raw_url: str) -> bool:
    text = raw_url.upper()
    markers = (
        "YOUR-PASSWORD",
        "<PASSWORD>",
        "<PROJECT-REF>",
        "<REGION>",
    )
    return any(marker in text for marker in markers)


DATABASE_URL = os.getenv("DATABASE_URL")
DB_URL_HAS_PLACEHOLDERS = bool(DATABASE_URL and _looks_like_placeholder_url(DATABASE_URL))
engine = None
SessionLocal = None

if DATABASE_URL:
    try:
        engine = create_engine(
            _normalize_database_url(DATABASE_URL),
            pool_pre_ping=True,
            pool_use_lifo=True,
            pool_size=int(os.getenv("DB_POOL_SIZE", "3")),
            max_overflow=int(os.getenv("DB_MAX_OVERFLOW", "2")),
            pool_timeout=int(os.getenv("DB_POOL_TIMEOUT", "30")),
            pool_recycle=int(os.getenv("DB_POOL_RECYCLE", "1800")),
            # Supabase pooler (PgBouncer) em transaction mode nao funciona bem com prepared statements.
            # Com psycopg3, prepare_threshold=None desativa preparo automatico e evita DuplicatePreparedStatement.
            connect_args={
                "connect_timeout": int(os.getenv("DB_CONNECT_TIMEOUT", "10")),
                "prepare_threshold": None,
            },
        )
        SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    except Exception:
        engine = None
        SessionLocal = None
        logger.exception("DATABASE_URL invalido: falha ao inicializar engine.")


def _is_supabase_direct_host(raw_url: str) -> bool:
    try:
        parsed = urlparse(_normalize_database_url(raw_url))
        host = (parsed.hostname or "").lower()
        return host.startswith("db.") and host.endswith(".supabase.co")
    except Exception:
        logger.warning("DATABASE_URL com formato invalido; nao foi possivel validar host Supabase.")
        return False


def _db_error_hint(exc: SQLAlchemyError) -> str:
    text = str(getattr(exc, "orig", exc)).lower()

    if "password authentication failed" in text:
        return (
            "Autenticacao no banco falhou. Confirme usuario/senha do DATABASE_URL "
            "(incluindo URL encode da senha)."
        )

    if "could not translate host name" in text or "name or service not known" in text:
        return "Host do banco invalido no DATABASE_URL. Revise host/porta do Supabase pooler."

    if "timed out" in text or "timeout" in text:
        return "Timeout ao conectar no banco. Verifique rede e host do Supabase pooler."

    if "connection refused" in text or "could not connect" in text:
        return "Conexao recusada pelo banco. Verifique host, porta e credenciais no DATABASE_URL."

    if DATABASE_URL and _is_supabase_direct_host(DATABASE_URL):
        return (
            "DATABASE_URL parece usar conexao direta Supabase. Em Render, use a URL do "
            "Supavisor transaction mode (porta 6543)."
        )

    return "Falha de conexao com banco. Revise o DATABASE_URL e a conectividade com o Supabase."


def _open_auth_db_session():
    if SessionLocal is None:
        return None
    return SessionLocal()


def _get_auth_session_manager() -> SessionManager:
    global AUTH_SESSION_MANAGER
    if AUTH_SESSION_MANAGER is None:
        AUTH_SESSION_MANAGER = SessionManager(
            config=AUTH_SESSION_CONFIG,
            open_db_session=_open_auth_db_session,
            app_session_model=AppSession,
            app_user_model=AppUser,
            active_sessions=ACTIVE_SESSIONS,
            active_sessions_lock=ACTIVE_SESSIONS_LOCK,
            logger=logger,
            normalize_role=_normalize_role,
            session_token_hash=_session_token_hash,
            utcnow=_utcnow,
        )
    return AUTH_SESSION_MANAGER


def _warn_default_credentials() -> None:
    defaults = {
        "Troque#Corretor123",
        "Troque#Cca123",
        "Troque#Analista123",
        "Troque#Admin123",
        "Troque#Gestor123",
        "1234",
        "12345",
        "",
    }
    checks = [
        (_normalize_username(APP_CORRETOR_USER), "corretor"),
        (_normalize_username(APP_CCA_USER), "cca"),
        (_normalize_username(APP_ANALISTA_USER), "analista"),
        (_normalize_username(APP_ADMIN_USER), "admin"),
        (_normalize_username(APP_GESTOR_USER), "gestor"),
    ]
    for username, role in checks:
        if not username:
            continue
        configured = APP_USERS.get(username, {}).get("password")
        if configured in defaults:
            logger.warning("Credencial insegura ou ausente para perfil '%s'. Defina senha forte no ambiente.", role)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _new_salt() -> str:
    return runtime_new_salt()


def _hash_password(password: str, salt: str) -> str:
    return runtime_hash_password(password, salt, PASSWORD_POLICY.iterations)


def _verify_password(password: str, password_hash: str, password_salt: str) -> bool:
    return runtime_verify_password(password, password_hash, password_salt, PASSWORD_POLICY.iterations)


def _password_policy_error(password: str) -> Optional[str]:
    return runtime_password_policy_error(password, PASSWORD_POLICY)


def _normalize_role(value: Optional[str]) -> str:
    role = (value or "").strip().lower()
    return role if role in VALID_ROLES else ROLE_CORRETOR


def _home_for_session(session: Optional[dict[str, Any]]) -> str:
    if not session:
        return "/login"
    if bool(session.get("must_change_password")):
        return "/app/trocar-senha"
    return _home_for_role(str(session.get("role", "")))


def _home_for_role(role: Optional[str]) -> str:
    role_key = (role or "").strip().lower()
    if role_key == ROLE_ADMIN:
        return "/app/admin"
    if role_key == ROLE_GESTOR:
        return "/app/gestor"
    if role_key == ROLE_GESTOR_CREDITO:
        return "/app/gestor-credito"
    if role_key == ROLE_ANALISTA:
        return "/app/analista"
    if role_key == ROLE_CCA:
        return "/app/cca"
    if role_key == ROLE_CORRETOR:
        return "/app-react/apresentacao" if CORRETOR_ROUTE_ENABLED else "/login"
    return "/app/analista"


def _new_session(user_id: uuid.UUID, username: str, role: str, must_change_password: bool) -> str:
    return _get_auth_session_manager().new_session(
        user_id=user_id,
        username=username,
        role=role,
        must_change_password=must_change_password,
    )


def _session_token_hash(token: str) -> str:
    return hash_token(token)


def _persist_session_record(token: str, session: dict[str, Any]) -> None:
    _get_auth_session_manager().persist_session_record(token, session)


def _load_session_from_store(token: str) -> Optional[dict[str, Any]]:
    return _get_auth_session_manager().load_session_from_store(token)


def _delete_session_record(token: str) -> None:
    _get_auth_session_manager().delete_session_record(token)


def _persist_session_activity(token: str, session: dict[str, Any]) -> None:
    _get_auth_session_manager().persist_session_activity(token, session)


def _sync_session_from_db(token: str, session: dict[str, Any]) -> Optional[dict[str, Any]]:
    return _get_auth_session_manager().sync_session_from_db(token, session)


def _drop_sessions_for_user(user_id: uuid.UUID) -> None:
    _get_auth_session_manager().drop_sessions_for_user(user_id)


def _process_list_cache_key(
    role: str,
    username: str,
    limit: Optional[int],
    offset: int,
) -> str:
    return f"{role}|{username}|{limit if limit is not None else 'all'}|{offset}"


def _get_cached_process_list(cache_key: str) -> Optional[list["ProcessoOverviewOut"]]:
    if PROCESS_LIST_CACHE_TTL_SECONDS <= 0:
        return None
    with PROCESS_LIST_CACHE_LOCK:
        cached = PROCESS_LIST_CACHE.get(cache_key)
        if not cached:
            return None
        expires_at = cached.get("expires_at")
        if not isinstance(expires_at, datetime) or expires_at <= _utcnow():
            PROCESS_LIST_CACHE.pop(cache_key, None)
            return None
        data = cached.get("data")
        if isinstance(data, list):
            return data
    return None


def _set_cached_process_list(cache_key: str, data: list["ProcessoOverviewOut"]) -> None:
    if PROCESS_LIST_CACHE_TTL_SECONDS <= 0:
        return
    with PROCESS_LIST_CACHE_LOCK:
        PROCESS_LIST_CACHE[cache_key] = {
            "data": data,
            "expires_at": _utcnow() + timedelta(seconds=PROCESS_LIST_CACHE_TTL_SECONDS),
        }


def _invalidate_process_list_cache() -> None:
    with PROCESS_LIST_CACHE_LOCK:
        if PROCESS_LIST_CACHE:
            PROCESS_LIST_CACHE.clear()


def _extract_origin_host(value: Optional[str]) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    parsed = urlparse(raw)
    host = (parsed.netloc or "").strip().lower()
    if host.endswith(":80"):
        host = host[:-3]
    if host.endswith(":443"):
        host = host[:-4]
    return host


def _request_host(request: Request) -> str:
    host = (
        (request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc or "")
        .strip()
        .lower()
    )
    if host.endswith(":80"):
        host = host[:-3]
    if host.endswith(":443"):
        host = host[:-4]
    return host


def _request_client_ip(request: Request) -> str:
    forwarded_for = (request.headers.get("x-forwarded-for") or "").strip()
    if forwarded_for:
        candidate = forwarded_for.split(",")[0].strip()
        if candidate:
            return candidate

    for header_name in ("cf-connecting-ip", "true-client-ip", "x-real-ip"):
        candidate = (request.headers.get(header_name) or "").strip()
        if candidate:
            return candidate

    client = getattr(request, "client", None)
    host = (getattr(client, "host", "") or "").strip()
    return host


def _read_session(request: Request) -> Optional[dict[str, Any]]:
    return _get_auth_session_manager().read_session(request)


def _read_session_user(request: Request) -> Optional[str]:
    return _get_auth_session_manager().read_session_user(request)


def _read_session_role(request: Request) -> Optional[str]:
    return _get_auth_session_manager().read_session_role(request)


def require_app_session(request: Request) -> dict[str, Any]:
    return _get_auth_session_manager().require_app_session(request)


def require_fully_authenticated_session(request: Request) -> dict[str, Any]:
    return _get_auth_session_manager().require_fully_authenticated_session(request)


def require_app_user(request: Request) -> str:
    return _get_auth_session_manager().require_app_user(request)


def require_roles(*roles: str):
    return _get_auth_session_manager().require_roles(*roles)


def _status_token(value: Optional[str]) -> str:
    return (value or "").strip().upper()


def _normalize_text_key(value: Optional[str]) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).replace(" ", "_")


def _process_view_label(area: str, value: Optional[str], *, fallback: str = "-") -> str:
    key = _normalize_text_key(value)
    mapped = PROCESS_OVERVIEW_LABELS.get(area, {}).get(key, "")
    if mapped:
        return mapped
    raw = str(value or "").strip()
    return raw or fallback


def _process_status_is_resolved(area: str, value: Optional[str]) -> bool:
    return _normalize_text_key(value) in PROCESS_READY_STATUS_KEYS.get(area, set())


def _process_pending_items(status_cca: Optional[str], status_agehab: Optional[str], status_sinal: Optional[str], status_fiador: Optional[str]) -> list[str]:
    items: list[str] = []
    if not _process_status_is_resolved("status_cca", status_cca):
        items.append(f"Caixa: {_process_view_label('status_cca', status_cca)}")
    if not _process_status_is_resolved("status_agehab", status_agehab):
        items.append(f"Agehab: {_process_view_label('status_agehab', status_agehab)}")
    if not _process_status_is_resolved("status_sinal", status_sinal):
        items.append(f"Sinal: {_process_view_label('status_sinal', status_sinal)}")
    if not _process_status_is_resolved("status_fiador", status_fiador):
        items.append(f"Fiador: {_process_view_label('status_fiador', status_fiador)}")
    return items


def _process_observacao_resumo(value: Optional[str], *, max_len: int = 140) -> str:
    normalized = " ".join(str(value or "").split()).strip()
    if not normalized:
        return "Sem observacao registrada"
    if len(normalized) <= max_len:
        return normalized
    return f"{normalized[: max_len - 3].rstrip()}..."


def _process_documentos_resumo(*, docs_total: int, docs_recebidos: int, sem_documento_enviado: bool) -> str:
    if sem_documento_enviado or docs_total <= 0:
        return "Documentos: nenhum enviado"
    pendentes = max(0, docs_total - docs_recebidos)
    if pendentes <= 0:
        return "Documentos: OK, todos aprovados"
    return f"Documentos pendentes: {pendentes} de {docs_total}"


def _process_repasse_display_key(etapa_repasse: Optional[str], status_cca: Optional[str], status_agehab: Optional[str]) -> str:
    repasse_key = _normalize_text_key(etapa_repasse)
    caixa_key = _normalize_text_key(status_cca)
    agehab_key = _normalize_text_key(status_agehab)
    if agehab_key == "validado_agehab" or caixa_key == "finalizado":
        return "inicio_garantia"
    if caixa_key == "assinatura_caixa" or repasse_key == "assinatura_autorizada":
        return "assinatura_caixa"
    if repasse_key == "inicio_repasse":
        return "inicio_repasse"
    if repasse_key == "em_repasse":
        return "em_repasse"
    return "sem_repasse"


def _process_estagio_comercial(value: Optional[str], fallback: str = "RESERVA") -> str:
    token = _normalize_text_key(value)
    aliases = {
        "reserva": "RESERVA",
        "processo": "EM_PROCESSO",
        "em_processo": "EM_PROCESSO",
        "credito": "CREDITO",
        "secretaria_vendas": "SECRETARIA_VENDAS",
        "secretaria_de_vendas": "SECRETARIA_VENDAS",
        "assinatura_diretoria": "ASSINATURA_DIRETORIA",
        "aprovacao_diretoria": "AUTORIZACAO_DIRETORIA",
        "aprovacao_da_diretoria": "AUTORIZACAO_DIRETORIA",
        "autorizacao_diretoria": "AUTORIZACAO_DIRETORIA",
        "envio_sienge": "ENVIO_SIENGE",
        "venda_finalizada": "VENDA_FINALIZADA",
    }
    mapped = aliases.get(token, "")
    return mapped if mapped in ESTAGIO_COMERCIAL_SET else fallback


def _process_etapa_repasse(value: Optional[str], fallback: Optional[str] = None) -> Optional[str]:
    token = _normalize_text_key(value)
    aliases = {
        "em_repasse": "EM_REPASSE",
        "inicio_repasse": "INICIO_REPASSE",
        "assinatura_autorizada": "ASSINATURA_AUTORIZADA",
    }
    mapped = aliases.get(token, "")
    if mapped in REPASSE_ETAPAS_SET:
        return mapped
    return fallback


def _lead_stage(value: Optional[str], fallback: str = "LEAD") -> str:
    token = _normalize_text_key(value)
    aliases = {
        "lead": "LEAD",
        "novo": "LEAD",
        "contato_inicial": "LEAD",
        "primeiro_contato": "LEAD",
        "qualificacao": "LEAD",
        "qualificado": "LEAD",
        "agendamento": "AGENDAMENTO",
        "agendado": "AGENDAMENTO",
        "visita": "VISITA",
        "em_atendimento": "VISITA",
        "atendimento": "VISITA",
        "precadastro": "PRECADASTRO",
        "pre_cadastro": "PRECADASTRO",
        "proposta": "PRECADASTRO",
        "proposta_enviada": "PRECADASTRO",
        "negociacao": "PRECADASTRO",
        "negociando": "PRECADASTRO",
        "reserva": "RESERVA",
        "ganho": "RESERVA",
        "fechado": "RESERVA",
        "perdido": "PERDIDO",
    }
    mapped = aliases.get(token, "")
    return mapped if mapped in LEAD_STAGE_SET else fallback


def _lead_cca_decision(value: Optional[str], fallback: str = "EM_ANALISE") -> str:
    token = _normalize_text_key(value)
    aliases = {
        "em_analise": "EM_ANALISE",
        "analise": "EM_ANALISE",
        "pendente": "EM_ANALISE",
        "aprovado": "APROVADO",
        "condicionado": "CONDICIONADO",
        "reprovado": "REPROVADO",
        "bloqueado": "BLOQUEADO",
        "dar_qv": "DAR_QV",
        "dar qv": "DAR_QV",
        "dar-qv": "DAR_QV",
        "darqv": "DAR_QV",
    }
    mapped = aliases.get(token, "")
    return mapped if mapped in LEAD_CCA_DECISION_SET else fallback


def _unidade_status(value: Optional[str], fallback: str = "DISPONIVEL") -> str:
    token = _normalize_text_key(value)
    aliases = {
        "disponivel": "DISPONIVEL",
        "livre": "DISPONIVEL",
        "reservada": "RESERVADA",
        "reservado": "RESERVADA",
        "vendida": "VENDIDA",
        "vendido": "VENDIDA",
        "bloqueada": "BLOQUEADA",
        "bloqueado": "BLOQUEADA",
    }
    mapped = aliases.get(token, "")
    return mapped if mapped in UNIDADE_STATUS_SET else fallback


def _should_be_in_repasse(processo: "Processo") -> bool:
    stage = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
    return stage in ESTAGIOS_REPASSE_COMERCIAL


def _fila_atual_from_processo(processo: "Processo") -> str:
    if _should_be_in_repasse(processo) or _process_etapa_repasse(getattr(processo, "etapa_repasse", None)):
        return "REPASSE"
    return "COMERCIAL"


def _is_caixa_apta_para_assinatura(status_cca: Optional[str]) -> bool:
    return _status_token(status_cca) in CAIXA_ASSINATURA_APTA_STATUSES


def _can_set_assinatura_autorizada(processo: "Processo") -> bool:
    stage = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
    sinal = _process_sinal_status(getattr(processo, "status_sinal", None))
    fiador = _process_fiador_status(getattr(processo, "status_fiador", None))
    agehab = _process_agehab_status(getattr(processo, "status_agehab", None))
    caixa = _process_caixa_status(getattr(processo, "status_cca", None))
    geral = _process_geral_status(getattr(processo, "status_geral", None))
    return (
        stage == "VENDA_FINALIZADA"
        and geral not in {"CANCELADO", "DISTRATO", "REPROVADO"}
        and sinal in {"NAO_TEM", "PAGO"}
        and fiador in {"NAO_TEM", "FINALIZADO"}
        and agehab == "VALIDADO_AGEHAB"
        and _is_caixa_apta_para_assinatura(caixa)
    )


def _sync_estagio_repasse_rules(processo: "Processo", now: Optional[datetime] = None) -> None:
    stage = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
    processo.estagio_comercial = stage
    etapa_repasse = _process_etapa_repasse(getattr(processo, "etapa_repasse", None))
    if stage in ESTAGIOS_REPASSE_COMERCIAL:
        if not etapa_repasse:
            processo.etapa_repasse = "EM_REPASSE"
    else:
        # Se voltou para estagio comercial anterior ao repasse, sai da trilha de repasse.
        # Mantemos etapa apenas quando assinatura de caixa ja foi concluida.
        if etapa_repasse and _status_token(getattr(processo, "status_cca", None)) not in PROCESS_CCA_FINAL_STATUSES:
            processo.etapa_repasse = None

    if stage == "RESERVA":
        processo.status_geral = "NOVO"
        processo.status_credito = "EM_ANALISE"
    elif stage == "VENDA_FINALIZADA":
        # Venda finalizada no comercial nao implica assinatura de caixa.
        # A assinatura depende da trilha de repasse e validacoes (sinal/fiador/agehab/cca).
        if _status_token(processo.status_geral) in {"", "NOVO", "PENDENCIADO"}:
            processo.status_geral = "EM_ANDAMENTO"
        if _status_token(processo.status_credito) not in {"APROVADO", "REPROVADO"}:
            processo.status_credito = "APROVADO"
    else:
        processo.status_geral = "EM_ANDAMENTO"
        if _status_token(processo.status_credito) == "REPROVADO":
            processo.status_credito = "EM_ANALISE"

    # Se o processo voltar para status nao final da Caixa, sai do arquivo mensal.
    if _status_token(getattr(processo, "status_cca", None)) not in PROCESS_CCA_FINAL_STATUSES:
        _clear_processo_archived(processo)


def _normalize_sla_owner(value: Optional[str], fallback: str = SLA_OWNER_NONE) -> str:
    raw = _status_token(value)
    return raw if raw in SLA_OWNER_VALUES else fallback


def _as_utc(value: Optional[datetime]) -> Optional[datetime]:
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _utc_start_of_day(value: Optional[date]) -> Optional[datetime]:
    if not isinstance(value, date):
        return None
    return datetime.combine(value, time.min, tzinfo=timezone.utc)


def _resolve_sla_clock_end(processo: "Processo", now: Optional[datetime] = None) -> datetime:
    now_utc = _as_utc(now) or _utcnow()
    if _is_processo_finalizado(processo):
        finished_at = _as_utc(getattr(processo, "updated_at", None))
        if finished_at is not None:
            return finished_at if finished_at <= now_utc else now_utc
    return now_utc


def _resolve_sla_comercial_start(processo: "Processo") -> Optional[datetime]:
    start_at = _as_utc(getattr(processo, "sla_comercial_inicio_at", None))
    if start_at is not None:
        return start_at

    cliente = getattr(processo, "cliente", None)
    if cliente is not None:
        start_at = _utc_start_of_day(getattr(cliente, "data_cadastro_origem", None))
        if start_at is not None:
            return start_at
        created_cliente = _as_utc(getattr(cliente, "created_at", None))
        if created_cliente is not None:
            return created_cliente

    return _as_utc(getattr(processo, "created_at", None))


def _resolve_sla_credito_start(processo: "Processo") -> Optional[datetime]:
    return _as_utc(getattr(processo, "sla_credito_inicio_at", None))


def _is_sla_comercial_closed(processo: "Processo") -> bool:
    return _process_estagio_comercial(getattr(processo, "estagio_comercial", None)) == "VENDA_FINALIZADA"


def _is_sla_credito_closed(processo: "Processo") -> bool:
    return _status_token(getattr(processo, "status_cca", None)) in PROCESS_CCA_FINAL_STATUSES


def _refresh_sla_fixed_markers(processo: "Processo", now: Optional[datetime] = None) -> None:
    now_utc = _as_utc(now) or _utcnow()

    comercial_closed = _is_sla_comercial_closed(processo)
    comercial_fim = _as_utc(getattr(processo, "sla_comercial_fim_at", None))
    if comercial_closed and comercial_fim is None:
        processo.sla_comercial_fim_at = now_utc
    elif (not comercial_closed) and comercial_fim is not None:
        processo.sla_comercial_fim_at = None

    credito_closed = _is_sla_credito_closed(processo)
    credito_fim = _as_utc(getattr(processo, "sla_credito_fim_at", None))
    if credito_closed and credito_fim is None:
        processo.sla_credito_fim_at = now_utc
    elif (not credito_closed) and credito_fim is not None:
        processo.sla_credito_fim_at = None


def _resolve_sla_fixed_end(processo: "Processo", owner: str, now: Optional[datetime] = None) -> datetime:
    now_utc = _as_utc(now) or _utcnow()
    owner_norm = _normalize_sla_owner(owner)
    if owner_norm == SLA_OWNER_CORRETOR:
        fim = _as_utc(getattr(processo, "sla_comercial_fim_at", None))
        if fim is not None:
            return fim if fim <= now_utc else now_utc
        if _is_sla_comercial_closed(processo):
            closed_at = _as_utc(getattr(processo, "updated_at", None)) or now_utc
            return closed_at if closed_at <= now_utc else now_utc
        return _resolve_sla_clock_end(processo, now_utc)
    elif owner_norm == SLA_OWNER_ANALISTA:
        fim = _as_utc(getattr(processo, "sla_credito_fim_at", None))
        if fim is not None:
            return fim if fim <= now_utc else now_utc
        if _is_sla_credito_closed(processo):
            closed_at = _as_utc(getattr(processo, "updated_at", None)) or now_utc
            return closed_at if closed_at <= now_utc else now_utc
        return _resolve_sla_clock_end(processo, now_utc)
    return _resolve_sla_clock_end(processo, now_utc)


def _elapsed_seconds_between(start_at: Optional[datetime], end_at: Optional[datetime]) -> int:
    if start_at is None or end_at is None or end_at <= start_at:
        return 0
    return int((end_at - start_at).total_seconds())


def _ensure_credito_sla_start(processo: "Processo", actor_role: Optional[str], now: Optional[datetime] = None) -> None:
    if _normalize_role(actor_role) != ROLE_ANALISTA:
        return
    if _as_utc(getattr(processo, "sla_credito_inicio_at", None)) is not None:
        return
    processo.sla_credito_inicio_at = _as_utc(now) or _utcnow()


def _seconds_from_value(value: Any) -> int:
    try:
        parsed = int(value or 0)
    except (TypeError, ValueError):
        return 0
    return max(0, parsed)


def _is_cca_sla_start_condition(processo: "Processo") -> bool:
    return _status_token(processo.status_cca) in {"ANALISE_CCA", "ANALISE_CREDITO"}


def _is_cca_sla_pendencia_condition(processo: "Processo") -> bool:
    return _status_token(processo.status_cca) in {"PENDENTE_CCA", "CONDICIONADO"} or _status_token(processo.status_agehab) == "PENDENTE_AGEHAB"


def _is_cca_sla_end_condition(processo: "Processo") -> bool:
    return _status_token(processo.status_cca) in PROCESS_CCA_FINAL_STATUSES


def _is_processo_finalizado(processo: "Processo") -> bool:
    return _status_token(processo.status_geral) in PROCESS_GERAL_FINAL_STATUSES


def _get_sla_seconds(processo: "Processo", owner: str) -> int:
    owner_norm = _normalize_sla_owner(owner)
    if owner_norm == SLA_OWNER_CORRETOR:
        return _seconds_from_value(getattr(processo, "sla_corretor_seconds", 0))
    if owner_norm == SLA_OWNER_ANALISTA:
        return _seconds_from_value(getattr(processo, "sla_analista_seconds", 0))
    if owner_norm == SLA_OWNER_CCA:
        return _seconds_from_value(getattr(processo, "sla_cca_seconds", 0))
    return 0


def _set_sla_seconds(processo: "Processo", owner: str, seconds: int) -> None:
    owner_norm = _normalize_sla_owner(owner)
    value = max(0, int(seconds))
    if owner_norm == SLA_OWNER_CORRETOR:
        processo.sla_corretor_seconds = value
    elif owner_norm == SLA_OWNER_ANALISTA:
        processo.sla_analista_seconds = value
    elif owner_norm == SLA_OWNER_CCA:
        processo.sla_cca_seconds = value


def _compute_sla_seconds(processo: "Processo", owner: str, now: Optional[datetime] = None) -> int:
    owner_norm = _normalize_sla_owner(owner)
    if owner_norm == SLA_OWNER_CORRETOR:
        start_at = _resolve_sla_comercial_start(processo)
        end_at = _resolve_sla_fixed_end(processo, owner_norm, now)
        return max(0, _elapsed_seconds_between(start_at, end_at))
    if owner_norm == SLA_OWNER_ANALISTA:
        start_at = _resolve_sla_credito_start(processo)
        end_at = _resolve_sla_fixed_end(processo, owner_norm, now)
        return max(0, _elapsed_seconds_between(start_at, end_at))

    total_seconds = _get_sla_seconds(processo, owner_norm)
    active_owner = _normalize_sla_owner(getattr(processo, "sla_owner", SLA_OWNER_NONE))
    if owner_norm != SLA_OWNER_NONE and owner_norm == active_owner:
        started_at = _as_utc(getattr(processo, "sla_active_since", None))
        now_utc = _as_utc(now) or _utcnow()
        if started_at and now_utc > started_at:
            total_seconds += int((now_utc - started_at).total_seconds())
    return max(0, total_seconds)


def _compute_sla_hours(processo: "Processo", owner: str, now: Optional[datetime] = None) -> int:
    total_seconds = _compute_sla_seconds(processo, owner, now)
    return max(0, total_seconds // 3600)


def _refresh_sla_snapshots(processo: "Processo", now: Optional[datetime] = None) -> None:
    now_utc = _as_utc(now) or _utcnow()
    sla_corretor_horas = _compute_sla_hours(processo, SLA_OWNER_CORRETOR, now_utc)
    sla_analista_horas = _compute_sla_hours(processo, SLA_OWNER_ANALISTA, now_utc)
    sla_cca_horas = _compute_sla_hours(processo, SLA_OWNER_CCA, now_utc)
    processo.sla_corretor_dias = sla_corretor_horas // 24
    processo.sla_credito_dias = sla_analista_horas // 24
    processo.sla_cca_dias = sla_cca_horas // 24


def _accrue_current_sla(processo: "Processo", now: Optional[datetime] = None) -> None:
    now_utc = _as_utc(now) or _utcnow()
    owner = _normalize_sla_owner(getattr(processo, "sla_owner", SLA_OWNER_NONE))
    if owner == SLA_OWNER_NONE:
        processo.sla_active_since = None
        return

    started_at = _as_utc(getattr(processo, "sla_active_since", None))
    if started_at is None:
        processo.sla_active_since = now_utc
        return

    if now_utc <= started_at:
        return

    elapsed = int((now_utc - started_at).total_seconds())
    if elapsed <= 0:
        return

    current = _get_sla_seconds(processo, owner)
    _set_sla_seconds(processo, owner, current + elapsed)
    processo.sla_active_since = now_utc


def _switch_sla_owner(processo: "Processo", owner: str, now: Optional[datetime] = None) -> None:
    now_utc = _as_utc(now) or _utcnow()
    target_owner = _normalize_sla_owner(owner)
    current_owner = _normalize_sla_owner(getattr(processo, "sla_owner", SLA_OWNER_NONE))

    if target_owner == current_owner:
        if target_owner == SLA_OWNER_NONE:
            processo.sla_active_since = None
        elif _as_utc(getattr(processo, "sla_active_since", None)) is None:
            processo.sla_active_since = now_utc
        _refresh_sla_snapshots(processo, now_utc)
        return

    _accrue_current_sla(processo, now_utc)
    processo.sla_owner = target_owner
    processo.sla_active_since = None if target_owner == SLA_OWNER_NONE else now_utc
    _refresh_sla_snapshots(processo, now_utc)


def _process_has_enviado_docs(db: Session, processo_id: uuid.UUID) -> bool:
    row = (
        db.query(Documento.id)
        .filter(Documento.processo_id == processo_id, Documento.status_doc.in_(["ENVIADO", "APROVADO", "NAO_APLICA"]))
        .first()
    )
    return row is not None


def _apply_sla_rules(
    processo: "Processo",
    *,
    trigger: Optional[str] = None,
    has_enviado_docs: bool = False,
    now: Optional[datetime] = None,
) -> None:
    now_utc = _as_utc(now) or _utcnow()
    trigger_key = (trigger or "").strip().lower()
    current_owner = _normalize_sla_owner(getattr(processo, "sla_owner", SLA_OWNER_NONE))

    if _is_processo_finalizado(processo) or _is_cca_sla_end_condition(processo):
        _switch_sla_owner(processo, SLA_OWNER_NONE, now_utc)
        return

    if trigger_key == "analista_pendenciou":
        _switch_sla_owner(processo, SLA_OWNER_CORRETOR, now_utc)
        return

    if trigger_key == "cca_pendenciou" or _is_cca_sla_pendencia_condition(processo):
        _switch_sla_owner(processo, SLA_OWNER_ANALISTA, now_utc)
        return

    if _is_cca_sla_start_condition(processo):
        _switch_sla_owner(processo, SLA_OWNER_CCA, now_utc)
        return

    if trigger_key == "corretor_enviou" and has_enviado_docs:
        _switch_sla_owner(processo, SLA_OWNER_ANALISTA, now_utc)
        return

    if current_owner == SLA_OWNER_NONE:
        _switch_sla_owner(processo, SLA_OWNER_ANALISTA, now_utc)
        return

    _switch_sla_owner(processo, current_owner, now_utc)


def _process_credit_status(value: Optional[str], fallback: str = "EM_ANALISE") -> str:
    raw = _status_token(value)
    aliases = {
        "ANALISE": "EM_ANALISE",
        "EM ANALISE": "EM_ANALISE",
        "PENDENTE": "PENDENCIADO",
    }
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_CREDITO_STATUSES else fallback


def _process_geral_status(value: Optional[str], fallback: str = "NOVO") -> str:
    raw = _status_token(value)
    aliases = {
        "EM_ANALISE": "EM_ANDAMENTO",
        "EM ANALISE": "EM_ANDAMENTO",
        "EMANDAMENTO": "EM_ANDAMENTO",
        "PENDENTE": "PENDENCIADO",
    }
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_GERAL_STATUSES else fallback


def _process_caixa_status(value: Optional[str], fallback: str = "ANALISE_CREDITO") -> str:
    raw = _status_token(value)
    aliases = {
        "EM_ANALISE": "ANALISE_CREDITO",
        "EM ANALISE": "ANALISE_CREDITO",
        "PENDENTE": "PENDENTE_CREDITO",
        "CONFORME": "APROVADO",
        "TRATANDO_PRODUTO": "CONDICIONADO",
        "AGUARDANDO_CONFORMIDADE": "CONDICIONADO",
        "AGENDADO": "DAR_QV",
        "DAR QV": "DAR_QV",
        "DAR-QV": "DAR_QV",
        "DARQV": "DAR_QV",
    }
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_CAIXA_STATUSES else fallback


def _process_agehab_status(value: Optional[str], fallback: str = "ANALISE_CREDITO") -> str:
    raw = _status_token(value)
    aliases = {
        "EM_ANALISE": "ANALISE_CREDITO",
        "EM ANALISE": "ANALISE_CREDITO",
        "PENDENTE": "PENDENTE_AGEHAB",
        "APROVADO": "VALIDADO_AGEHAB",
        "REPROVADO": "PENDENTE_AGEHAB",
        "ENVIO_AGEHAG": "ENVIO_AGEHAB",
    }
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_AGEHAB_STATUSES else fallback


def _process_sinal_status(value: Optional[str], fallback: str = "NAO_TEM") -> str:
    raw = _status_token(value)
    aliases = {"NAO TEM": "NAO_TEM"}
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_SINAL_STATUSES else fallback


def _process_fiador_status(value: Optional[str], fallback: str = "NAO_TEM") -> str:
    raw = _status_token(value)
    aliases = {"NAO TEM": "NAO_TEM"}
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_FIADOR_STATUSES else fallback


def _process_recolha_fgts_status(value: Optional[str], fallback: str = "NAO_RECOLHIDO") -> str:
    raw = _status_token(value)
    aliases = {
        "NAO RECOLHIDO": "NAO_RECOLHIDO",
        "VALIDADO BANCO": "VALIDADO_PELO_BANCO",
        "VALIDADO PELO BANCO": "VALIDADO_PELO_BANCO",
    }
    raw = aliases.get(raw, raw)
    return raw if raw in PROCESS_RECOLHA_FGTS_STATUSES else fallback


def _is_pendencia_status(field: str, status_value: Optional[str]) -> bool:
    token = _status_token(status_value)
    if field in {"status_credito", "status_geral"}:
        return token == "PENDENCIADO"
    if field in {"status_cca", "status_agehab"}:
        return token.startswith("PENDENTE_") or token == "CONDICIONADO"
    if field in {"status_sinal", "status_fiador"}:
        return token == "PENDENTE"
    return False


def _validate_status_transition(field: str, current_value: Optional[str], next_value: Optional[str]) -> None:
    current = _status_token(current_value)
    nxt = _status_token(next_value)
    if not current or current == nxt:
        return

    if field == "status_geral" and current in PROCESS_GERAL_FINAL_STATUSES and nxt != current:
        # Permite fechamento administrativo apos aprovacao/reprovacao.
        if current in {"APROVADO", "REPROVADO"} and nxt in {"CANCELADO", "DISTRATO"}:
            return
        raise HTTPException(status_code=422, detail="Processo finalizado nao permite reabertura de status geral.")
    if field == "status_cca" and current in PROCESS_CCA_FINAL_STATUSES and nxt != current:
        raise HTTPException(status_code=422, detail="Status Caixa finalizado nao permite reabertura.")
    if field == "status_geral" and current == "NOVO" and nxt in {"APROVADO", "REPROVADO"}:
        raise HTTPException(status_code=422, detail="Status geral nao pode ir de NOVO direto para resultado final.")


def _validate_estagio_comercial_transition(current_value: Optional[str], next_value: Optional[str]) -> None:
    current = _process_estagio_comercial(current_value, fallback="")
    nxt = _process_estagio_comercial(next_value, fallback="")
    if not current or not nxt or current == nxt:
        return
    if current == "VENDA_FINALIZADA" and nxt != "VENDA_FINALIZADA":
        raise HTTPException(
            status_code=422,
            detail=(
                "Apos VENDA_FINALIZADA o cliente nao pode voltar de estagio comercial. "
                "Se houver desistencia, altere o status geral para DISTRATO."
            ),
        )


def _stringify_audit_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return _as_utc(value).isoformat() if _as_utc(value) else value.isoformat()
    text = str(value).strip()
    return text or None


def _doc_status(value: Optional[str], fallback: str = "PENDENTE") -> str:
    raw = (value or "").strip().upper()
    aliases = {
        "NAO APLICA": "NAO_APLICA",
        "N/A": "NAO_APLICA",
    }
    raw = aliases.get(raw, raw)
    allowed = {"PENDENTE", "ENVIADO", "APROVADO", "NAO_APLICA"}
    return raw if raw in allowed else fallback


def _doc_is_done(status_doc: Optional[str]) -> bool:
    token = _doc_status(status_doc, fallback="PENDENTE")
    return token in {"ENVIADO", "APROVADO", "NAO_APLICA"}


def _credit_status(value: Optional[str], fallback: str = "ANALISE") -> str:
    raw = (value or "").strip().upper()
    aliases = {
        "EM_ANALISE": "ANALISE",
        "EM ANALISE": "ANALISE",
        "AGUARDANDO ENVIO": "AGUARDANDO_ENVIO",
        "NAO APLICA": "NAO_APLICA",
        "N/A": "NAO_APLICA",
    }
    raw = aliases.get(raw, raw)
    allowed = {"AGUARDANDO_ENVIO", "ANALISE", "PENDENCIADO", "APROVADO", "REPROVADO", "NAO_APLICA"}
    return raw if raw in allowed else fallback


def _normalize_pendencia_info(value: Optional[str]) -> Optional[str]:
    text = " ".join(str(value or "").strip().split())
    return text or None


DEFAULT_DOCUMENTOS = [
    {"categoria": "proponente", "nome": "Identidade e CPF"},
    {"categoria": "proponente", "nome": "Comprovante de estado civil"},
    {"categoria": "proponente", "nome": "Comprovante de residencia"},
    {"categoria": "proponente", "nome": "Comprovante de renda"},
    {"categoria": "caixa", "nome": "Ficha de cadastro Caixa"},
    {"categoria": "caixa", "nome": "Ficha de abertura de conta"},
    {"categoria": "caixa", "nome": "MO"},
    {"categoria": "agehab", "nome": "Checklist Agehab"},
    {"categoria": "agehab", "nome": "Ficha Agehab"},
    {"categoria": "agehab", "nome": "Declaracao de endereco"},
]


def _ensure_default_documentos(db: Session, processo_id: uuid.UUID, *, autocommit: bool = True) -> None:
    existing = (
        db.query(Documento.categoria, Documento.nome)
        .filter(Documento.processo_id == processo_id)
        .all()
    )
    existing_keys = {(categoria, nome) for categoria, nome in existing}

    created = False
    for doc in DEFAULT_DOCUMENTOS:
        key = (doc["categoria"], doc["nome"])
        if key in existing_keys:
            continue
        db.add(
            Documento(
                processo_id=processo_id,
                categoria=doc["categoria"],
                nome=doc["nome"],
                status_doc="PENDENTE",
                status_credito="AGUARDANDO_ENVIO",
            )
        )
        created = True

    if created and autocommit:
        db.commit()


class Base(DeclarativeBase):
    pass


class Cliente(Base):
    __tablename__ = "clientes"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    nome: Mapped[str] = mapped_column(Text, nullable=False)
    corretor: Mapped[Optional[str]] = mapped_column(Text)
    obra: Mapped[Optional[str]] = mapped_column(Text)
    imobiliaria: Mapped[Optional[str]] = mapped_column(Text)
    data_reserva_origem: Mapped[Optional[date]] = mapped_column(Date)
    data_cadastro_origem: Mapped[Optional[date]] = mapped_column(Date)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now())

    processos: Mapped[list["Processo"]] = relationship(back_populates="cliente", cascade="all, delete-orphan")


class Processo(Base):
    __tablename__ = "processos"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    cliente_id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), ForeignKey("clientes.id", ondelete="CASCADE"))
    status_credito: Mapped[str] = mapped_column(String, nullable=False, default="EM_ANALISE")
    status_geral: Mapped[str] = mapped_column(String, nullable=False, default="NOVO")
    status_cca: Mapped[str] = mapped_column(String, nullable=False, default="ANALISE_CREDITO")
    status_agehab: Mapped[str] = mapped_column(String, nullable=False, default="ANALISE_CREDITO")
    status_sinal: Mapped[str] = mapped_column(String, nullable=False, default="NAO_TEM")
    valor_sinal: Mapped[Optional[float]] = mapped_column(Float)
    renda_bruta: Mapped[Optional[float]] = mapped_column(Float)
    renda_liquida: Mapped[Optional[float]] = mapped_column(Float)
    valor_parcela: Mapped[Optional[float]] = mapped_column(Float)
    valor_parcela_7lm: Mapped[Optional[float]] = mapped_column(Float)
    renda_complementar_valor: Mapped[Optional[float]] = mapped_column(Float)
    renda_complementar_responsavel: Mapped[Optional[str]] = mapped_column(Text)
    renda_complementar_vinculo: Mapped[Optional[str]] = mapped_column(String(40))
    fgts_futuro_empresa_6m: Mapped[Optional[str]] = mapped_column(String(20))
    unidade_retomada_agehab: Mapped[Optional[str]] = mapped_column(String(20))
    valor_imovel: Mapped[Optional[float]] = mapped_column(Float)
    valor_avaliacao: Mapped[Optional[float]] = mapped_column(Float)
    valor_financiamento: Mapped[Optional[float]] = mapped_column(Float)
    valor_subsidio: Mapped[Optional[float]] = mapped_column(Float)
    valor_cheque_moradia: Mapped[Optional[float]] = mapped_column(Float)
    recolha_fgts: Mapped[str] = mapped_column(String(30), nullable=False, default="NAO_RECOLHIDO")
    status_fiador: Mapped[str] = mapped_column(String, nullable=False, default="NAO_TEM")
    estagio_comercial: Mapped[str] = mapped_column(String(40), nullable=False, default="RESERVA")
    etapa_repasse: Mapped[Optional[str]] = mapped_column(String(40))
    cca_responsavel: Mapped[Optional[str]] = mapped_column(String(120))
    pendente_fiador: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    pendente_sinal: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    nao_contar_mes: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    nao_contar_mes_ref_ano: Mapped[Optional[int]] = mapped_column(Integer)
    nao_contar_mes_ref_mes: Mapped[Optional[int]] = mapped_column(Integer)
    arquivado: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    arquivado_em: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    arquivado_ref_ano: Mapped[Optional[int]] = mapped_column(Integer)
    arquivado_ref_mes: Mapped[Optional[int]] = mapped_column(Integer)
    sla_comercial_inicio_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    sla_credito_inicio_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    sla_comercial_fim_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    sla_credito_fim_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    sla_credito_dias: Mapped[Optional[int]] = mapped_column(Integer)
    sla_corretor_dias: Mapped[Optional[int]] = mapped_column(Integer)
    sla_cca_dias: Mapped[Optional[int]] = mapped_column(Integer)
    sla_owner: Mapped[str] = mapped_column(String(20), nullable=False, default=SLA_OWNER_NONE)
    sla_active_since: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    sla_analista_seconds: Mapped[int] = mapped_column(Integer, nullable=False, default=0)
    sla_corretor_seconds: Mapped[int] = mapped_column(Integer, nullable=False, default=0)
    sla_cca_seconds: Mapped[int] = mapped_column(Integer, nullable=False, default=0)
    frankstein_last_event_id: Mapped[Optional[str]] = mapped_column(String(36))
    frankstein_last_event_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    observacao: Mapped[Optional[str]] = mapped_column(Text)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now())

    cliente: Mapped["Cliente"] = relationship(back_populates="processos")
    documentos: Mapped[list["Documento"]] = relationship(back_populates="processo", cascade="all, delete-orphan")


class Documento(Base):
    __tablename__ = "documentos"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    processo_id: Mapped[uuid.UUID] = mapped_column(
        UUID(as_uuid=True),
        ForeignKey("processos.id", ondelete="CASCADE"),
    )
    categoria: Mapped[str] = mapped_column(String, nullable=False)
    nome: Mapped[str] = mapped_column(Text, nullable=False)
    status_doc: Mapped[str] = mapped_column(String, nullable=False, default="PENDENTE")
    status_credito: Mapped[str] = mapped_column(String, nullable=False, default="AGUARDANDO_ENVIO")
    pendencia_info: Mapped[Optional[str]] = mapped_column(Text)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now())

    processo: Mapped["Processo"] = relationship(back_populates="documentos")


class ProcessoEvento(Base):
    __tablename__ = "processo_eventos"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    processo_id: Mapped[uuid.UUID] = mapped_column(
        UUID(as_uuid=True),
        ForeignKey("processos.id", ondelete="CASCADE"),
        nullable=False,
        index=True,
    )
    actor_username: Mapped[str] = mapped_column(String(120), nullable=False)
    actor_role: Mapped[str] = mapped_column(String(20), nullable=False, default="system")
    event_type: Mapped[str] = mapped_column(String(50), nullable=False)
    field_name: Mapped[Optional[str]] = mapped_column(String(100))
    old_value: Mapped[Optional[str]] = mapped_column(Text)
    new_value: Mapped[Optional[str]] = mapped_column(Text)
    details: Mapped[Optional[str]] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)


class SistemaLog(Base):
    __tablename__ = "sistema_logs"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    actor_username: Mapped[str] = mapped_column(String(120), nullable=False, index=True)
    actor_role: Mapped[str] = mapped_column(String(20), nullable=False, default="system", index=True)
    tela: Mapped[str] = mapped_column(String(60), nullable=False, index=True)
    acao: Mapped[str] = mapped_column(String(80), nullable=False, index=True)
    entidade_tipo: Mapped[Optional[str]] = mapped_column(String(60))
    entidade_id: Mapped[Optional[str]] = mapped_column(String(120))
    details: Mapped[Optional[str]] = mapped_column(Text)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)


class AppUser(Base):
    __tablename__ = "app_users"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    username: Mapped[str] = mapped_column(String(120), unique=True, nullable=False, index=True)
    password_hash: Mapped[str] = mapped_column(String(128), nullable=False)
    password_salt: Mapped[str] = mapped_column(String(64), nullable=False)
    role: Mapped[str] = mapped_column(String(20), nullable=False, default=ROLE_CORRETOR)
    is_active: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)
    must_change_password: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)
    last_login_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now())
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )


class AppSession(Base):
    __tablename__ = "app_sessions"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    session_token_hash: Mapped[str] = mapped_column(String(64), unique=True, nullable=False, index=True)
    user_id: Mapped[uuid.UUID] = mapped_column(
        UUID(as_uuid=True),
        ForeignKey("app_users.id", ondelete="CASCADE"),
        nullable=False,
        index=True,
    )
    username: Mapped[str] = mapped_column(String(120), nullable=False, index=True)
    role: Mapped[str] = mapped_column(String(20), nullable=False, default=ROLE_CORRETOR)
    must_change_password: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False, default=_utcnow)
    last_seen_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False, default=_utcnow)
    db_checked_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False, default=_utcnow)
    expires_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), nullable=False, index=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )


class Empreendimento(Base):
    __tablename__ = "empreendimentos"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    nome: Mapped[str] = mapped_column(String(200), unique=True, nullable=False, index=True)
    is_active: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now())
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )


class EmpreendimentoRegraFinanceira(Base):
    __tablename__ = "empreendimento_regras_financeiras"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    empreendimento_id: Mapped[uuid.UUID] = mapped_column(
        UUID(as_uuid=True),
        ForeignKey("empreendimentos.id", ondelete="CASCADE"),
        nullable=False,
        unique=True,
        index=True,
    )
    valor_cheque_moradia: Mapped[float] = mapped_column(Float, nullable=False, default=0.0)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now())
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )


class TabelaPreco(Base):
    __tablename__ = "tabela_precos"
    __table_args__ = (
        UniqueConstraint("empreendimento", "unidade", name="uq_tabela_preco_empreendimento_unidade"),
    )

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    empreendimento: Mapped[str] = mapped_column(String(200), nullable=False, index=True)
    unidade: Mapped[str] = mapped_column(String(120), nullable=False, index=True)
    garantido_minimo: Mapped[float] = mapped_column(Float, nullable=False, default=0.0)
    preco: Mapped[float] = mapped_column(Float, nullable=False, default=0.0)
    sobrepreco: Mapped[float] = mapped_column(Float, nullable=False, default=0.0)
    is_maximo: Mapped[float] = mapped_column(Float, nullable=False, default=0.0)
    prosoluto_minimo: Mapped[float] = mapped_column(Float, nullable=False, default=0.0)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )


class LeadPreCadastro(Base):
    __tablename__ = "lead_precadastros"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    corretor_username: Mapped[str] = mapped_column(String(120), nullable=False, index=True)
    nome_cliente: Mapped[str] = mapped_column(Text, nullable=False)
    _telefone: Mapped[Optional[str]] = mapped_column("telefone", Text)
    telefone_last4: Mapped[Optional[str]] = mapped_column(String(4), index=True)
    _whatsapp: Mapped[Optional[str]] = mapped_column("whatsapp", Text)
    whatsapp_last4: Mapped[Optional[str]] = mapped_column(String(4), index=True)
    _email: Mapped[Optional[str]] = mapped_column("email", Text)
    email_hash: Mapped[Optional[str]] = mapped_column(String(64), index=True)
    _cpf: Mapped[Optional[str]] = mapped_column("cpf", Text)
    cpf_hash: Mapped[Optional[str]] = mapped_column(String(64), index=True)
    cpf_last4: Mapped[Optional[str]] = mapped_column(String(4), index=True)
    _documento_identificacao: Mapped[Optional[str]] = mapped_column("documento_identificacao", Text)
    documento_hash: Mapped[Optional[str]] = mapped_column(String(64), index=True)
    estado_civil: Mapped[Optional[str]] = mapped_column(String(30))
    _certidao_numero: Mapped[Optional[str]] = mapped_column("certidao_numero", Text)
    _cidade_nascimento: Mapped[Optional[str]] = mapped_column("cidade_nascimento", Text)
    data_nascimento: Mapped[Optional[date]] = mapped_column(Date)
    _endereco: Mapped[Optional[str]] = mapped_column("endereco", Text)
    empreendimento_interesse: Mapped[Optional[str]] = mapped_column(Text)
    localidade_interesse: Mapped[Optional[str]] = mapped_column(Text)
    local_agendamento: Mapped[Optional[str]] = mapped_column(Text)
    tipo_visita: Mapped[Optional[str]] = mapped_column(String(20))
    data_agendamento: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    estagio_lead: Mapped[str] = mapped_column(String(40), nullable=False, default="LEAD")
    decisao_cca: Mapped[str] = mapped_column(String(30), nullable=False, default="EM_ANALISE")
    contrato_assinado: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    contrato_assinado_em: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    assinatura_email_confirmada: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False)
    assinatura_email_confirmada_em: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    assinatura_email_token: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    assinatura_email_token_expires_at: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    assinatura_email_enviado_em: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    processo_id: Mapped[Optional[uuid.UUID]] = mapped_column(UUID(as_uuid=True), ForeignKey("processos.id", ondelete="SET NULL"))
    reservado_em: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    observacoes: Mapped[Optional[str]] = mapped_column(Text)
    ultimo_contato_em: Mapped[Optional[datetime]] = mapped_column(DateTime(timezone=True))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
        index=True,
    )

    @property
    def telefone(self) -> Optional[str]:
        return decrypt_pii(self._telefone)

    @telefone.setter
    def telefone(self, value: Optional[str]) -> None:
        self._telefone = encrypt_pii(value)
        self.telefone_last4 = last4_digits(value)

    @property
    def whatsapp(self) -> Optional[str]:
        return decrypt_pii(self._whatsapp)

    @whatsapp.setter
    def whatsapp(self, value: Optional[str]) -> None:
        self._whatsapp = encrypt_pii(value)
        self.whatsapp_last4 = last4_digits(value)

    @property
    def email(self) -> Optional[str]:
        return decrypt_pii(self._email)

    @email.setter
    def email(self, value: Optional[str]) -> None:
        self._email = encrypt_pii(value)
        self.email_hash = hash_email(value)

    @property
    def cpf(self) -> Optional[str]:
        return decrypt_pii(self._cpf)

    @cpf.setter
    def cpf(self, value: Optional[str]) -> None:
        self._cpf = encrypt_pii(value)
        self.cpf_hash = hash_optional(value)
        self.cpf_last4 = last4_digits(value)

    @property
    def documento_identificacao(self) -> Optional[str]:
        return decrypt_pii(self._documento_identificacao)

    @documento_identificacao.setter
    def documento_identificacao(self, value: Optional[str]) -> None:
        self._documento_identificacao = encrypt_pii(value)
        self.documento_hash = hash_optional(value)

    @property
    def certidao_numero(self) -> Optional[str]:
        return decrypt_pii(self._certidao_numero)

    @certidao_numero.setter
    def certidao_numero(self, value: Optional[str]) -> None:
        self._certidao_numero = encrypt_pii(value)

    @property
    def cidade_nascimento(self) -> Optional[str]:
        return decrypt_pii(self._cidade_nascimento)

    @cidade_nascimento.setter
    def cidade_nascimento(self, value: Optional[str]) -> None:
        self._cidade_nascimento = encrypt_pii(value)

    @property
    def endereco(self) -> Optional[str]:
        return decrypt_pii(self._endereco)

    @endereco.setter
    def endereco(self, value: Optional[str]) -> None:
        self._endereco = encrypt_pii(value)


class IaFeedback(Base):
    __tablename__ = "ia_feedback"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    aceitou: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False, index=True)
    preco_sugerido: Mapped[Optional[float]] = mapped_column(Float)
    contexto_json: Mapped[Optional[str]] = mapped_column(Text)
    origem: Mapped[str] = mapped_column(String(80), nullable=False, default="app", index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)


class AnaliseRegistroDB(Base):
    __tablename__ = "analises"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), default=_utcnow, index=True)
    empreendimento: Mapped[str] = mapped_column(String(120))
    unidade: Mapped[str] = mapped_column(String(120))
    preco_imovel: Mapped[float] = mapped_column(Float)
    valor_obtido: Mapped[float] = mapped_column(Float)
    prosoluto_calculado: Mapped[float] = mapped_column(Float)
    prosoluto_liquido: Mapped[float] = mapped_column(Float)
    sinal: Mapped[float] = mapped_column(Float)
    sinal_produto: Mapped[float] = mapped_column(Float)
    financiamento: Mapped[float] = mapped_column(Float)
    subsidio: Mapped[float] = mapped_column(Float)
    cheque_moradia: Mapped[float] = mapped_column(Float)
    renda_bruta: Mapped[float] = mapped_column(Float)
    perc_construcao: Mapped[float] = mapped_column(Float)
    is_agora: Mapped[float] = mapped_column(Float)
    is_pos_chaves: Mapped[float] = mapped_column(Float)
    tabela_referencia: Mapped[Optional[str]] = mapped_column(Text)
    data_referencia: Mapped[datetime] = mapped_column(DateTime(timezone=True))


class UnidadeDisponivel(Base):
    __tablename__ = "unidades_disponiveis"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    created_by_username: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    empreendimento: Mapped[str] = mapped_column(String(220), nullable=False, index=True)
    unidade: Mapped[str] = mapped_column(String(80), nullable=False, index=True)
    bloco: Mapped[Optional[str]] = mapped_column(String(80))
    tipologia: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    quartos: Mapped[Optional[int]] = mapped_column(Integer)
    banheiros: Mapped[Optional[int]] = mapped_column(Integer)
    vagas: Mapped[Optional[int]] = mapped_column(Integer)
    area_m2: Mapped[Optional[float]] = mapped_column(Float)
    valor: Mapped[Optional[float]] = mapped_column(Float)
    localizacao: Mapped[Optional[str]] = mapped_column(Text)
    diferenciais: Mapped[Optional[str]] = mapped_column(Text)
    url_imagem: Mapped[Optional[str]] = mapped_column(Text)
    visita_disponivel: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)
    status_unidade: Mapped[str] = mapped_column(String(20), nullable=False, default="DISPONIVEL", index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
        index=True,
    )


class CreditoPlanejamentoItem(Base):
    __tablename__ = "credito_planejamento_itens"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    tipo: Mapped[str] = mapped_column(String(20), nullable=False, default="tarefa", index=True)
    titulo: Mapped[str] = mapped_column(String(180), nullable=False)
    descricao: Mapped[Optional[str]] = mapped_column(Text)
    responsavel: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    data_referencia: Mapped[Optional[date]] = mapped_column(Date, index=True)
    hora_inicio: Mapped[Optional[str]] = mapped_column(String(5))
    hora_fim: Mapped[Optional[str]] = mapped_column(String(5))
    status: Mapped[str] = mapped_column(String(20), nullable=False, default="pendente", index=True)
    progresso: Mapped[int] = mapped_column(Integer, nullable=False, default=0)
    urgente: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False, index=True)
    created_by_username: Mapped[Optional[str]] = mapped_column(String(120))
    updated_by_username: Mapped[Optional[str]] = mapped_column(String(120))
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
        index=True,
    )


class AnalistaReuniaoComercial(Base):
    __tablename__ = "analista_reuniao_comercial"

    processo_id: Mapped[uuid.UUID] = mapped_column(
        UUID(as_uuid=True),
        ForeignKey("processos.id", ondelete="CASCADE"),
        primary_key=True,
    )
    conta_no_mes: Mapped[bool] = mapped_column(Boolean, nullable=False, default=True)
    data_prevista_entrega: Mapped[Optional[date]] = mapped_column(Date, index=True)
    probabilidade_queda: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False, index=True)
    solicitar_cancelamento: Mapped[bool] = mapped_column(Boolean, nullable=False, default=False, index=True)
    status_followup: Mapped[str] = mapped_column(String(20), nullable=False, default="seguir", index=True)
    observacao: Mapped[Optional[str]] = mapped_column(Text)
    justificativa_reincidencia: Mapped[Optional[str]] = mapped_column(Text)
    updated_by_username: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
        index=True,
    )


class AnalistaReuniaoCompromisso(Base):
    __tablename__ = "analista_reuniao_comercial_compromissos"

    id: Mapped[uuid.UUID] = mapped_column(UUID(as_uuid=True), primary_key=True, default=uuid.uuid4)
    processo_id: Mapped[uuid.UUID] = mapped_column(
        UUID(as_uuid=True),
        ForeignKey("processos.id", ondelete="CASCADE"),
        nullable=False,
        index=True,
    )
    data_prometida: Mapped[date] = mapped_column(Date, nullable=False, index=True)
    status: Mapped[str] = mapped_column(String(20), nullable=False, default="pendente", index=True)
    justificativa: Mapped[Optional[str]] = mapped_column(Text)
    observacao: Mapped[Optional[str]] = mapped_column(Text)
    created_by_username: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    updated_by_username: Mapped[Optional[str]] = mapped_column(String(120), index=True)
    created_at: Mapped[datetime] = mapped_column(DateTime(timezone=True), server_default=func.now(), index=True)
    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
        index=True,
    )


class ClienteCreate(BaseModel):
    nome: str
    corretor: Optional[str] = None
    obra: Optional[str] = None
    imobiliaria: Optional[str] = None
    data_reserva_origem: Optional[date] = None
    data_cadastro_origem: Optional[date] = None


class ClienteOut(ClienteCreate):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID


class ProcessoCreate(BaseModel):
    cliente_id: uuid.UUID


class ProcessoUpdate(BaseModel):
    status_credito: Optional[str] = None
    status_geral: Optional[str] = None
    status_cca: Optional[str] = None
    status_agehab: Optional[str] = None
    status_sinal: Optional[str] = None
    valor_sinal: Optional[float] = None
    renda_bruta: Optional[float] = None
    renda_liquida: Optional[float] = None
    valor_parcela: Optional[float] = None
    valor_parcela_7lm: Optional[float] = None
    renda_complementar_valor: Optional[float] = None
    renda_complementar_responsavel: Optional[str] = None
    renda_complementar_vinculo: Optional[str] = None
    fgts_futuro_empresa_6m: Optional[str] = None
    unidade_retomada_agehab: Optional[str] = None
    valor_imovel: Optional[float] = None
    valor_avaliacao: Optional[float] = None
    valor_financiamento: Optional[float] = None
    valor_subsidio: Optional[float] = None
    valor_cheque_moradia: Optional[float] = None
    recolha_fgts: Optional[str] = None
    status_fiador: Optional[str] = None
    estagio_comercial: Optional[str] = None
    etapa_repasse: Optional[str] = None
    cca_responsavel: Optional[str] = None
    pendente_fiador: Optional[bool] = None
    pendente_sinal: Optional[bool] = None
    nao_contar_mes: Optional[bool] = None
    sla_credito_dias: Optional[int] = None
    sla_corretor_dias: Optional[int] = None
    sla_cca_dias: Optional[int] = None
    observacao: Optional[str] = None


class ProcessoOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    cliente_id: uuid.UUID
    status_credito: str
    status_geral: str
    status_cca: str
    status_agehab: str
    status_sinal: str
    valor_sinal: Optional[float] = None
    renda_bruta: Optional[float] = None
    renda_liquida: Optional[float] = None
    valor_parcela: Optional[float] = None
    valor_parcela_7lm: Optional[float] = None
    renda_complementar_valor: Optional[float] = None
    renda_complementar_responsavel: Optional[str] = None
    renda_complementar_vinculo: Optional[str] = None
    fgts_futuro_empresa_6m: Optional[str] = None
    unidade_retomada_agehab: Optional[str] = None
    valor_imovel: Optional[float] = None
    valor_avaliacao: Optional[float] = None
    valor_financiamento: Optional[float] = None
    valor_subsidio: Optional[float] = None
    valor_cheque_moradia: Optional[float] = None
    recolha_fgts: str
    status_fiador: str
    estagio_comercial: str
    etapa_repasse: Optional[str] = None
    fila_atual: Optional[str] = None
    cca_responsavel: Optional[str] = None
    pendente_fiador: bool
    pendente_sinal: bool
    nao_contar_mes: bool
    sla_credito_dias: Optional[int] = None
    sla_corretor_dias: Optional[int] = None
    sla_cca_dias: Optional[int] = None
    sla_analista_horas: Optional[int] = None
    sla_corretor_horas: Optional[int] = None
    sla_cca_horas: Optional[int] = None
    sla_analista_seconds: Optional[int] = None
    sla_corretor_seconds: Optional[int] = None
    sla_cca_seconds: Optional[int] = None
    sla_owner: Optional[str] = None
    sla_active_since: Optional[datetime] = None
    observacao: Optional[str] = None
    created_at: Optional[datetime] = None


class DocumentoCreate(BaseModel):
    processo_id: uuid.UUID
    categoria: str
    nome: str


class DocumentoUpdate(BaseModel):
    status_doc: Optional[str] = None
    status_credito: Optional[str] = None
    pendencia_info: Optional[str] = None


class DocumentoOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    processo_id: uuid.UUID
    categoria: str
    nome: str
    status_doc: str
    status_credito: str
    pendencia_info: Optional[str] = None


class ProcessoEventoOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    processo_id: uuid.UUID
    actor_username: str
    actor_role: str
    event_type: str
    field_name: Optional[str] = None
    old_value: Optional[str] = None
    new_value: Optional[str] = None
    details: Optional[str] = None
    created_at: datetime


class SistemaLogOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    actor_username: str
    actor_role: str
    tela: str
    acao: str
    entidade_tipo: Optional[str] = None
    entidade_id: Optional[str] = None
    details: Optional[str] = None
    created_at: datetime


class ProcessoMetricasFunnelOut(BaseModel):
    novo: int
    em_andamento: int
    pendenciado: int
    aprovado: int
    reprovado: int
    distrato: int
    cancelado: int


class ProcessoMetricasSlaOut(BaseModel):
    analista_alerta_24h: int
    analista_critico_48h: int
    corretor_alerta_24h: int
    corretor_critico_48h: int
    cca_alerta_24h: int
    cca_critico_48h: int


class ProcessoMetricasQualidadeOut(BaseModel):
    processos_com_pendencia: int
    first_pass_yield_percent: float
    media_retrabalho_por_processo: float


class ProcessoMetricasOut(BaseModel):
    total_processos: int
    funnel: ProcessoMetricasFunnelOut
    sla: ProcessoMetricasSlaOut
    qualidade: ProcessoMetricasQualidadeOut
    updated_at: datetime


class LoginPayload(BaseModel):
    username: str
    password: str


class ChangePasswordPayload(BaseModel):
    current_password: str
    new_password: str


class AdminUserCreate(BaseModel):
    username: str
    password: str
    role: str


class AdminUserUpdate(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None


class AdminResetPasswordPayload(BaseModel):
    new_password: str
    force_change_password: bool = True


class AdminDeleteRegistroPayload(BaseModel):
    entidade: str
    registro_id: uuid.UUID
    motivo: Optional[str] = None


class AdminRegistroLookupItem(BaseModel):
    id: uuid.UUID
    titulo: str
    detalhe: Optional[str] = None


class AdminRegistroLookupOut(BaseModel):
    entidade: str
    total: int
    itens: list[AdminRegistroLookupItem]


class AdminFranksteinEmailAlertsOut(BaseModel):
    smtp_configured: bool
    smtp_missing: list[str] = Field(default_factory=list)
    smtp_config: dict[str, Any] = Field(default_factory=dict)
    recipients_configured: bool
    destinatarios: str
    destinatarios_mascarados: list[str]
    fonte_destinatarios: str
    window_minutes: int


class AdminFranksteinEmailAlertsPayload(BaseModel):
    destinatarios: str = ""


class AdminFranksteinEmailTestPayload(BaseModel):
    destinatarios: Optional[str] = None


class LayoutPreferencePayload(BaseModel):
    blackhole_enabled: bool = False


class LayoutPreferenceOut(BaseModel):
    blackhole_enabled: bool
    blackhole_allowed: bool = False
    fonte: str = "runtime"


class FranksteinAnalistaAprendizadoPayload(BaseModel):
    processo_id: Optional[uuid.UUID] = None
    contexto: dict[str, Any] = Field(default_factory=dict)
    decisao_analista: dict[str, Any] = Field(default_factory=dict)
    regras_frankstein: list[dict[str, Any]] = Field(default_factory=list)


class FranksteinAnalistaAprendizadoOut(BaseModel):
    ok: bool
    total_interacoes: int
    padroes: list[dict[str, Any]] = Field(default_factory=list)
    fonte: str


class AppUserOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    username: str
    role: str
    is_active: bool
    must_change_password: bool
    last_login_at: Optional[datetime] = None
    created_at: datetime
    updated_at: datetime


class EmpreendimentoCreate(BaseModel):
    nome: str


class EmpreendimentoOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    nome: str
    is_active: bool
    created_at: datetime
    updated_at: datetime


class EmpreendimentoRegraFinanceiraPayload(BaseModel):
    valor_cheque_moradia: float


class EmpreendimentoRegraFinanceiraOut(BaseModel):
    empreendimento_id: uuid.UUID
    empreendimento_nome: str
    valor_cheque_moradia: float
    updated_at: datetime


class LeadPreCadastroCreate(BaseModel):
    nome_cliente: str
    telefone: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    cpf: Optional[str] = None
    documento_identificacao: Optional[str] = None
    estado_civil: Optional[str] = None
    certidao_numero: Optional[str] = None
    cidade_nascimento: Optional[str] = None
    data_nascimento: Optional[date] = None
    endereco: Optional[str] = None
    empreendimento_interesse: Optional[str] = None
    localidade_interesse: Optional[str] = None
    local_agendamento: Optional[str] = None
    tipo_visita: Optional[str] = None
    data_agendamento: Optional[datetime] = None
    estagio_lead: Optional[str] = None
    decisao_cca: Optional[str] = None
    contrato_assinado: Optional[bool] = None
    observacoes: Optional[str] = None
    ultimo_contato_em: Optional[datetime] = None


class LeadPreCadastroUpdate(BaseModel):
    nome_cliente: Optional[str] = None
    telefone: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    cpf: Optional[str] = None
    documento_identificacao: Optional[str] = None
    estado_civil: Optional[str] = None
    certidao_numero: Optional[str] = None
    cidade_nascimento: Optional[str] = None
    data_nascimento: Optional[date] = None
    endereco: Optional[str] = None
    empreendimento_interesse: Optional[str] = None
    localidade_interesse: Optional[str] = None
    local_agendamento: Optional[str] = None
    tipo_visita: Optional[str] = None
    data_agendamento: Optional[datetime] = None
    estagio_lead: Optional[str] = None
    decisao_cca: Optional[str] = None
    contrato_assinado: Optional[bool] = None
    observacoes: Optional[str] = None
    ultimo_contato_em: Optional[datetime] = None


class LeadPreCadastroOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    corretor_username: str
    nome_cliente: str
    telefone: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    cpf: Optional[str] = None
    documento_identificacao: Optional[str] = None
    estado_civil: Optional[str] = None
    certidao_numero: Optional[str] = None
    cidade_nascimento: Optional[str] = None
    data_nascimento: Optional[date] = None
    endereco: Optional[str] = None
    empreendimento_interesse: Optional[str] = None
    localidade_interesse: Optional[str] = None
    local_agendamento: Optional[str] = None
    tipo_visita: Optional[str] = None
    data_agendamento: Optional[datetime] = None
    estagio_lead: str
    decisao_cca: str
    contrato_assinado: bool
    contrato_assinado_em: Optional[datetime] = None
    assinatura_email_confirmada: bool
    assinatura_email_confirmada_em: Optional[datetime] = None
    assinatura_email_enviado_em: Optional[datetime] = None
    assinatura_email_token_expires_at: Optional[datetime] = None
    processo_id: Optional[uuid.UUID] = None
    reservado_em: Optional[datetime] = None
    observacoes: Optional[str] = None
    ultimo_contato_em: Optional[datetime] = None
    created_at: datetime
    updated_at: datetime


class LeadReservaOut(BaseModel):
    lead_id: uuid.UUID
    cliente_id: uuid.UUID
    processo_id: uuid.UUID
    estagio_lead: str
    reservado_em: datetime


class LeadAssinaturaEmailOut(BaseModel):
    lead_id: uuid.UUID
    email: str
    token_expires_at: datetime
    enviado_em: datetime


class UnidadeDisponivelCreate(BaseModel):
    empreendimento: str
    unidade: str
    bloco: Optional[str] = None
    tipologia: Optional[str] = None
    quartos: Optional[int] = None
    banheiros: Optional[int] = None
    vagas: Optional[int] = None
    area_m2: Optional[float] = None
    valor: Optional[float] = None
    localizacao: Optional[str] = None
    diferenciais: Optional[str] = None
    url_imagem: Optional[str] = None
    visita_disponivel: Optional[bool] = True
    status_unidade: Optional[str] = None


class UnidadeDisponivelUpdate(BaseModel):
    empreendimento: Optional[str] = None
    unidade: Optional[str] = None
    bloco: Optional[str] = None
    tipologia: Optional[str] = None
    quartos: Optional[int] = None
    banheiros: Optional[int] = None
    vagas: Optional[int] = None
    area_m2: Optional[float] = None
    valor: Optional[float] = None
    localizacao: Optional[str] = None
    diferenciais: Optional[str] = None
    url_imagem: Optional[str] = None
    visita_disponivel: Optional[bool] = None
    status_unidade: Optional[str] = None


class UnidadeDisponivelOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)
    id: uuid.UUID
    created_by_username: Optional[str] = None
    empreendimento: str
    unidade: str
    bloco: Optional[str] = None
    tipologia: Optional[str] = None
    quartos: Optional[int] = None
    banheiros: Optional[int] = None
    vagas: Optional[int] = None
    area_m2: Optional[float] = None
    valor: Optional[float] = None
    localizacao: Optional[str] = None
    diferenciais: Optional[str] = None
    url_imagem: Optional[str] = None
    visita_disponivel: bool
    status_unidade: str
    created_at: datetime
    updated_at: datetime


class UnidadeDisponivelMetricasOut(BaseModel):
    total: int
    disponiveis: int
    ticket_medio: float
    area_media: float
    empreendimentos: int


class ProcessoIntakeCreate(BaseModel):
    nome: str
    corretor: Optional[str] = None
    obra: Optional[str] = None
    imobiliaria: Optional[str] = None
    data_reserva_origem: Optional[date] = None
    data_cadastro_origem: Optional[date] = None
    estagio_comercial: Optional[str] = None


class ProcessoOverviewOut(BaseModel):
    processo_id: uuid.UUID
    cliente_id: uuid.UUID
    cliente_nome: str
    corretor: Optional[str] = None
    obra: Optional[str] = None
    imobiliaria: Optional[str] = None
    status_credito: str
    status_geral: str
    status_cca: str
    status_agehab: str
    status_sinal: str
    valor_sinal: Optional[float] = None
    renda_bruta: Optional[float] = None
    renda_bruta_duplicada: bool = False
    renda_bruta_duplicada_clientes: list[str] = Field(default_factory=list)
    renda_bruta_duplicada_tooltip: Optional[str] = None
    recolha_fgts: str
    status_fiador: str
    estagio_comercial: str
    etapa_repasse: Optional[str] = None
    fila_atual: Optional[str] = None
    cca_responsavel: Optional[str] = None
    pendente_fiador: bool
    pendente_sinal: bool
    nao_contar_mes: bool
    sla_credito_dias: Optional[int] = None
    sla_corretor_dias: Optional[int] = None
    sla_cca_dias: Optional[int] = None
    sla_analista_horas: Optional[int] = None
    sla_corretor_horas: Optional[int] = None
    sla_cca_horas: Optional[int] = None
    sla_analista_seconds: Optional[int] = None
    sla_corretor_seconds: Optional[int] = None
    sla_cca_seconds: Optional[int] = None
    sla_owner: Optional[str] = None
    sla_active_since: Optional[datetime] = None
    data_reserva_origem: Optional[date] = None
    data_cadastro_origem: Optional[date] = None
    created_at: Optional[datetime] = None
    observacao: Optional[str] = None
    docs_total: int = 0
    docs_recebidos: int = 0
    sem_documento_enviado: bool = True
    aviso_gerar_contrato_agehab: bool = False
    estagio_comercial_key: Optional[str] = None
    estagio_comercial_label: Optional[str] = None
    etapa_repasse_key: Optional[str] = None
    etapa_repasse_label: Optional[str] = None
    repasse_fase_key: Optional[str] = None
    repasse_fase_label: Optional[str] = None
    status_cca_key: Optional[str] = None
    status_cca_label: Optional[str] = None
    status_agehab_key: Optional[str] = None
    status_agehab_label: Optional[str] = None
    status_sinal_key: Optional[str] = None
    status_sinal_label: Optional[str] = None
    status_fiador_key: Optional[str] = None
    status_fiador_label: Optional[str] = None
    docs_pendentes: int = 0
    documentos_resumo: Optional[str] = None
    observacao_resumo: Optional[str] = None
    status_pendencias: list[str] = Field(default_factory=list)
    status_pendencias_resumo: Optional[str] = None
    status_tudo_ok: bool = False


class CcaAnaliseItemOut(BaseModel):
    processo_id: uuid.UUID
    cliente_id: uuid.UUID
    cliente_nome: str
    corretor: Optional[str] = None
    obra: Optional[str] = None
    estagio_comercial: str
    status_cca: str
    status_agehab: str
    renda_bruta: Optional[float] = None
    renda_liquida: Optional[float] = None
    valor_parcela: Optional[float] = None
    valor_imovel: Optional[float] = None
    valor_avaliacao: Optional[float] = None
    valor_financiamento: Optional[float] = None
    valor_subsidio: Optional[float] = None
    valor_cheque_moradia: Optional[float] = None
    recolha_fgts: str
    cpf: Optional[str] = None
    documento_identificacao: Optional[str] = None
    estado_civil: Optional[str] = None
    certidao_numero: Optional[str] = None
    cidade_nascimento: Optional[str] = None
    data_nascimento: Optional[date] = None
    telefone: Optional[str] = None
    whatsapp: Optional[str] = None
    email: Optional[str] = None
    docs_total: int = 0
    docs_recebidos: int = 0
    docs_pendentes: int = 0
    updated_at: Optional[datetime] = None


class CcaAnaliseUpdate(BaseModel):
    status_cca: Optional[str] = None
    renda_bruta: Optional[float] = None
    renda_liquida: Optional[float] = None
    valor_parcela: Optional[float] = None
    valor_imovel: Optional[float] = None
    valor_avaliacao: Optional[float] = None
    valor_financiamento: Optional[float] = None
    valor_subsidio: Optional[float] = None
    recolha_fgts: Optional[str] = None


class ProcessoArquivadoOut(BaseModel):
    processo_id: uuid.UUID
    cliente_id: uuid.UUID
    cliente_nome: str
    corretor: Optional[str] = None
    obra: Optional[str] = None
    imobiliaria: Optional[str] = None
    status_geral: str
    estagio_comercial: str
    etapa_repasse: Optional[str] = None
    status_cca: str
    status_agehab: str
    status_sinal: str
    status_fiador: str
    arquivado_em: Optional[datetime] = None
    arquivado_ref_ano: Optional[int] = None
    arquivado_ref_mes: Optional[int] = None
    data_reserva_origem: Optional[date] = None
    data_cadastro_origem: Optional[date] = None
    created_at: Optional[datetime] = None
    sla_analista_horas: Optional[int] = None
    sla_corretor_horas: Optional[int] = None


class ProcessoArquivadoListOut(BaseModel):
    total_clientes_cadastrados: int
    total_processos_ativos: int
    total_processos_arquivados: int
    itens: list[ProcessoArquivadoOut]


class AdminStorageSummaryOut(BaseModel):
    total_clientes: int
    total_processos: int
    total_processos_ativos: int
    total_processos_arquivados: int
    total_pre_cadastros: int
    total_unidades_disponiveis: int
    total_documentos: int
    total_eventos_processo: int
    total_logs_sistema: int
    total_usuarios: int
    total_empreendimentos: int
    total_registros_monitorados: int


class ImportPlanilhaRowOut(BaseModel):
    linha: int
    nome_cliente: Optional[str] = None
    status: str
    motivo: Optional[str] = None


class ImportPlanilhaOut(BaseModel):
    ok: bool
    total_linhas: int
    importados: int
    ignorados_existentes: int
    invalidados: int
    resultados: list[ImportPlanilhaRowOut]


class GestorMetaPayload(BaseModel):
    meta: int


class GestorMetaOut(BaseModel):
    meta: int
    fonte: str


class GestorMetaPeriodoPayload(BaseModel):
    ano: int
    mes: int
    meta_mensal: int
    meta_semanal: int


class GestorMetaPeriodoOut(BaseModel):
    ano: int
    mes: int
    meta_mensal: int
    meta_semanal: int
    fonte_mensal: str
    fonte_semanal: str


class CreditoPlanejamentoItemCreate(BaseModel):
    tipo: str = "tarefa"
    titulo: str
    descricao: Optional[str] = None
    responsavel: Optional[str] = None
    data_referencia: Optional[date] = None
    hora_inicio: Optional[str] = None
    hora_fim: Optional[str] = None
    status: Optional[str] = "pendente"
    progresso: Optional[int] = 0
    urgente: Optional[bool] = False


class CreditoPlanejamentoItemUpdate(BaseModel):
    tipo: Optional[str] = None
    titulo: Optional[str] = None
    descricao: Optional[str] = None
    responsavel: Optional[str] = None
    data_referencia: Optional[date] = None
    hora_inicio: Optional[str] = None
    hora_fim: Optional[str] = None
    status: Optional[str] = None
    progresso: Optional[int] = None
    urgente: Optional[bool] = None


class CreditoPlanejamentoItemOut(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: uuid.UUID
    tipo: str
    titulo: str
    descricao: Optional[str] = None
    responsavel: Optional[str] = None
    data_referencia: Optional[date] = None
    hora_inicio: Optional[str] = None
    hora_fim: Optional[str] = None
    status: str
    progresso: int
    urgente: bool
    created_by_username: Optional[str] = None
    updated_by_username: Optional[str] = None
    created_at: datetime
    updated_at: datetime
    tipo_label: str = "Item"
    status_label: str = "Pendente"
    display_titulo: Optional[str] = None
    display_descricao: Optional[str] = None
    display_meta: Optional[str] = None
    meta_kind: Optional[str] = None
    meta_cliente: Optional[str] = None
    meta_acao: Optional[str] = None
    meta_observacao: Optional[str] = None
    meta_responsavel: Optional[str] = None
    meta_status_oper: Optional[str] = None
    meta_status_oper_label: Optional[str] = None


class CreditoPlanejamentoEvolucaoOut(BaseModel):
    responsavel: str
    total: int
    concluidas: int
    pendentes: int
    progresso_medio: int
    taxa_conclusao: float


class CreditoPlanejamentoDashboardOut(BaseModel):
    referencia: date
    pendentes_total: int
    tarefas_dia: list[CreditoPlanejamentoItemOut]
    agendamentos_dia: list[CreditoPlanejamentoItemOut]
    entregas_dia: list[CreditoPlanejamentoItemOut]
    urgentes: list[CreditoPlanejamentoItemOut]
    evolucao_time: list[CreditoPlanejamentoEvolucaoOut]
    anotacoes: list[CreditoPlanejamentoItemOut]
    itens: list[CreditoPlanejamentoItemOut]


class AnalistaReuniaoCompromissoCreate(BaseModel):
    data_prometida: date
    observacao: Optional[str] = None


class AnalistaReuniaoCompromissoNaoEntreguePayload(BaseModel):
    justificativa: str
    nova_data_prometida: Optional[date] = None
    nova_observacao: Optional[str] = None


class AnalistaReuniaoComercialUpdate(BaseModel):
    conta_no_mes: Optional[bool] = None
    data_prevista_entrega: Optional[date] = None
    probabilidade_queda: Optional[bool] = None
    solicitar_cancelamento: Optional[bool] = None
    status_followup: Optional[str] = None
    observacao: Optional[str] = None
    justificativa_reincidencia: Optional[str] = None


class AnalistaReuniaoCompromissoOut(BaseModel):
    id: uuid.UUID
    processo_id: uuid.UUID
    data_prometida: date
    status: str
    justificativa: Optional[str] = None
    observacao: Optional[str] = None
    created_by_username: Optional[str] = None
    updated_by_username: Optional[str] = None
    created_at: datetime
    updated_at: datetime


class AnalistaReuniaoClienteOut(BaseModel):
    processo_id: uuid.UUID
    cliente_nome: str
    empreendimento: Optional[str] = None
    corretor: Optional[str] = None
    imobiliaria: Optional[str] = None
    estagio_comercial: str
    status_cca: str
    data_cadastro_origem: Optional[date] = None
    sla_dias: int
    conta_no_mes: bool
    data_prevista_entrega: Optional[date] = None
    entrega_hoje: bool
    probabilidade_queda: bool
    solicitar_cancelamento: bool
    status_followup: str
    observacao: Optional[str] = None
    justificativa_reincidencia: Optional[str] = None
    nao_entregou_count: int
    compromissos: list[AnalistaReuniaoCompromissoOut]


class AnalistaReuniaoComercialDashboardOut(BaseModel):
    referencia: date
    total_clientes: int
    assinados: int
    seguir: int
    finalizar_hoje: int
    entrega_hoje: int
    risco_queda: int
    solicitar_cancelamento: int
    clientes_entrega_hoje: list[str]
    clientes: list[AnalistaReuniaoClienteOut]


class ProcessoFullOut(BaseModel):
    processo: ProcessoOut
    cliente: ClienteOut
    documentos: list[DocumentoOut]


class DocumentoBulkItem(BaseModel):
    categoria: str
    nome: str
    status_doc: Optional[str] = None
    status_credito: Optional[str] = None
    pendencia_info: Optional[str] = None


class DocumentoBulkUpsert(BaseModel):
    documentos: list[DocumentoBulkItem]


def _normalize_credito_planejamento_text(value: Optional[str], *, max_len: int) -> Optional[str]:
    cleaned = " ".join(str(value or "").strip().split())
    if not cleaned:
        return None
    return cleaned[:max_len]


def _credito_planejamento_tipo(value: Optional[str], *, fallback: str = "tarefa") -> str:
    raw = (
        str(value or "")
        .strip()
        .lower()
        .replace("-", "_")
        .replace(" ", "_")
    )
    aliases = {
        "agenda": "agendamento",
        "agendamentos": "agendamento",
        "tarefas": "tarefa",
        "subtarefas": "subtarefa",
        "subtask": "subtarefa",
        "subtasks": "subtarefa",
        "task": "tarefa",
        "tasks": "tarefa",
        "entregas": "entrega",
        "urgencias": "urgente",
        "urgencia": "urgente",
        "nota": "anotacao",
        "notas": "anotacao",
        "anotacoes": "anotacao",
    }
    value_norm = aliases.get(raw, raw)
    if value_norm in CREDITO_PLANEJAMENTO_TIPOS:
        return value_norm
    return fallback if fallback in CREDITO_PLANEJAMENTO_TIPOS else "tarefa"


def _credito_planejamento_status(value: Optional[str], *, fallback: str = "pendente") -> str:
    raw = (
        str(value or "")
        .strip()
        .lower()
        .replace("-", "_")
        .replace(" ", "_")
    )
    aliases = {
        "andamento": "em_andamento",
        "em_andamento": "em_andamento",
        "em_execucao": "em_andamento",
        "feito": "concluido",
        "finalizado": "concluido",
        "done": "concluido",
    }
    value_norm = aliases.get(raw, raw)
    if value_norm in CREDITO_PLANEJAMENTO_STATUS:
        return value_norm
    return fallback if fallback in CREDITO_PLANEJAMENTO_STATUS else "pendente"


def _credito_planejamento_hora(value: Optional[str], *, field_name: str) -> Optional[str]:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        parsed = datetime.strptime(raw, "%H:%M")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=f"{field_name} invalida. Use HH:MM.") from exc
    return parsed.strftime("%H:%M")


def _credito_planejamento_progresso(value: Optional[int]) -> int:
    try:
        parsed = int(value if value is not None else 0)
    except (TypeError, ValueError):
        parsed = 0
    return max(0, min(100, parsed))


def _credito_planejamento_tipo_label(value: Optional[str]) -> str:
    return PLANEJAMENTO_TIPO_LABELS.get(_credito_planejamento_tipo(value, fallback="tarefa"), "Item")


def _credito_planejamento_status_label(value: Optional[str]) -> str:
    return PLANEJAMENTO_STATUS_LABELS.get(_credito_planejamento_status(value, fallback="pendente"), "Pendente")


def _credito_planejamento_time_window(hora_inicio: Optional[str], hora_fim: Optional[str]) -> str:
    start = str(hora_inicio or "").strip()
    end = str(hora_fim or "").strip()
    if start and end:
        return f"{start} - {end}"
    return start or end


def _credito_planejamento_date_label(value: Optional[date]) -> str:
    if not isinstance(value, date):
        return ""
    return value.strftime("%d/%m/%Y")


def _normalize_credito_planejamento_entrega_status(value: Optional[str], *, fallback: str = "pendenciado") -> str:
    token = _normalize_text_key(value)
    if token in {"entregue", "concluido", "finalizado", "done"}:
        return "entregue"
    if token in {"caixa", "analise_caixa"}:
        return "caixa"
    if token in {"agehab", "analise_agehab"}:
        return "agehab"
    return fallback


def _extract_credito_planejamento_meta_payload(raw_description: Optional[str]) -> Optional[dict[str, Any]]:
    raw = str(raw_description or "").strip()
    if not raw:
        return None
    json_source = ""
    if raw.startswith(PLANEJAMENTO_ENTREGA_META_PREFIX):
        json_source = raw[len(PLANEJAMENTO_ENTREGA_META_PREFIX) :]
    elif raw.startswith(PLANEJAMENTO_DIA_TODO_META_PREFIX):
        json_source = raw[len(PLANEJAMENTO_DIA_TODO_META_PREFIX) :]
    else:
        json_start = raw.find("{")
        has_known_prefix = (
            raw.startswith(PLANEJAMENTO_ENTREGA_META_PREFIX)
            or raw.startswith(PLANEJAMENTO_DIA_TODO_META_PREFIX)
            or "entrega_meta" in raw
            or "dia_todo_meta" in raw
        )
        if json_start < 0 or (not has_known_prefix and json_start != 0):
            return None
        json_source = raw[json_start:]
    try:
        payload = json.loads(json_source or "{}")
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    return payload if isinstance(payload, dict) else None


def _credito_planejamento_item_view(item: CreditoPlanejamentoItem) -> dict[str, Optional[str]]:
    tipo = _credito_planejamento_tipo(item.tipo, fallback="tarefa")
    raw_desc = str(item.descricao or "").strip()
    payload = _extract_credito_planejamento_meta_payload(raw_desc)
    time_window = _credito_planejamento_time_window(item.hora_inicio, item.hora_fim)
    date_label = _credito_planejamento_date_label(item.data_referencia)

    display_title = str(item.titulo or "").strip() or "-"
    display_description = raw_desc or None
    meta_kind = "texto"
    meta_cliente = None
    meta_acao = None
    meta_observacao = None
    meta_responsavel = str(item.responsavel or "").strip() or None
    meta_status_oper = None
    meta_status_oper_label = None

    if tipo == "entrega":
        meta_kind = "entrega"
        meta_cliente = str((payload or {}).get("cliente") or item.titulo or "").strip() or None
        meta_acao = str((payload or {}).get("acao") or "").strip() or None
        meta_observacao = str((payload or {}).get("observacao") or ("" if payload else raw_desc)).strip() or None
        meta_responsavel = str((payload or {}).get("responsavel") or item.responsavel or "").strip() or None
        meta_status_oper = _normalize_credito_planejamento_entrega_status((payload or {}).get("statusOper") or item.status)
        meta_status_oper_label = PLANEJAMENTO_ENTREGA_STATUS_LABELS.get(meta_status_oper, "Pendenciado")
        display_title = meta_cliente or display_title
        display_description = meta_acao or meta_observacao or None
    elif raw_desc.startswith(PLANEJAMENTO_DIA_TODO_META_PREFIX):
        meta_kind = "dia_todo"
        meta_acao = str((payload or {}).get("acao") or "").strip() or None
        meta_observacao = str((payload or {}).get("observacao") or "").strip() or None
        display_description = meta_acao or meta_observacao or None

    meta_parts = [meta_responsavel or "", time_window, date_label]
    if meta_status_oper_label:
        meta_parts.append(meta_status_oper_label)

    return {
        "display_titulo": display_title,
        "display_descricao": display_description,
        "display_meta": " - ".join(part for part in meta_parts if part),
        "meta_kind": meta_kind,
        "meta_cliente": meta_cliente,
        "meta_acao": meta_acao,
        "meta_observacao": meta_observacao,
        "meta_responsavel": meta_responsavel,
        "meta_status_oper": meta_status_oper,
        "meta_status_oper_label": meta_status_oper_label,
    }


def _credito_planejamento_item_occurs_on(item: CreditoPlanejamentoItem, target_date: date) -> bool:
    tipo = _credito_planejamento_tipo(item.tipo, fallback="tarefa")
    if tipo == "anotacao":
        return False
    if _credito_planejamento_status(item.status, fallback="pendente") == "concluido":
        return False
    if not item.hora_inicio:
        return False
    if tipo == "subtarefa":
        return item.data_referencia is None or item.data_referencia <= target_date
    return item.data_referencia == target_date


def _credito_planejamento_reminder_at_brt(item: CreditoPlanejamentoItem) -> Optional[datetime]:
    if not item.hora_inicio:
        return None
    try:
        hour_raw, minute_raw = str(item.hora_inicio).strip()[:5].split(":", 1)
        task_time = time(int(hour_raw), int(minute_raw))
    except (TypeError, ValueError):
        return None
    target_date = item.data_referencia or datetime.now(KEEPALIVE_BRT_TZ).date()
    return datetime.combine(target_date, task_time, tzinfo=KEEPALIVE_BRT_TZ) - timedelta(minutes=5)


def _credito_planejamento_alert_key(item: CreditoPlanejamentoItem, reminder_at: datetime) -> str:
    return f"alert_key=planejamento:{item.id}:{reminder_at.date().isoformat()}:{str(item.hora_inicio or '').strip()[:5]}"


def _credito_planejamento_email_body(item: CreditoPlanejamentoItem, view: dict[str, Optional[str]], reminder_at: datetime) -> str:
    titulo = str(view.get("display_titulo") or item.titulo or "Tarefa sem titulo").strip()
    descricao = str(view.get("display_descricao") or "").strip()
    responsavel = str(item.responsavel or view.get("meta_responsavel") or "").strip()
    horario = _credito_planejamento_time_window(item.hora_inicio, item.hora_fim) or str(item.hora_inicio or "").strip()
    data_label = _credito_planejamento_date_label(item.data_referencia or reminder_at.date())
    status_label = _credito_planejamento_status_label(item.status)

    linhas = [
        f"{AI_ASSISTANT_DISPLAY_NAME} alerta supervisionado do SioCred",
        "",
        "Faltam 5 minutos para uma tarefa ou compromisso.",
        "",
        f"Tarefa: {titulo}",
        f"Data: {data_label}",
        f"Horario: {horario or '-'}",
        f"Status atual: {status_label}",
    ]
    if responsavel:
        linhas.append(f"Responsavel: {responsavel}")
    if descricao:
        linhas.extend(["", f"Detalhe: {descricao}"])
    linhas.extend(
        [
            "",
            "Acao sugerida:",
            "Abra a agenda operacional do SioCred, execute a tarefa e marque o check-in/conclusao.",
            "",
            f"Este e um alerta automatico do {AI_ASSISTANT_DISPLAY_NAME}. Ele apenas avisa; a decisao continua supervisionada por voce.",
        ]
    )
    return "\n".join(linhas)


def _processar_alertas_email_planejamento(db: Session, *, now: Optional[datetime] = None) -> dict[str, Any]:
    now_brt = (now or datetime.now(KEEPALIVE_BRT_TZ)).astimezone(KEEPALIVE_BRT_TZ)
    recipients_raw, recipients_source = _get_alert_recipients_source(db)
    recipients = _split_alert_recipients(recipients_raw)
    result: dict[str, Any] = {
        "ok": True,
        "email_configured": _is_email_delivery_configured(),
        "recipients_configured": bool(recipients),
        "recipients_source": recipients_source,
        "now_brt": now_brt.isoformat(),
        "window_minutes": SIOCRED_EMAIL_ALERT_WINDOW_MINUTES,
        "checked": 0,
        "sent": 0,
        "skipped": 0,
        "errors": [],
    }
    if not _is_email_delivery_configured() or not recipients:
        result["ok"] = False
        result["reason"] = "email_not_configured" if not _is_email_delivery_configured() else "recipients_not_configured"
        return result

    today = now_brt.date()
    candidate_rows = (
        db.query(CreditoPlanejamentoItem)
        .filter(func.lower(func.coalesce(CreditoPlanejamentoItem.status, "")) != "concluido")
        .filter(func.lower(func.coalesce(CreditoPlanejamentoItem.tipo, "")) != "anotacao")
        .filter(CreditoPlanejamentoItem.hora_inicio.isnot(None))
        .filter(
            or_(
                CreditoPlanejamentoItem.data_referencia == today,
                and_(
                    func.lower(func.coalesce(CreditoPlanejamentoItem.tipo, "")) == "subtarefa",
                    or_(
                        CreditoPlanejamentoItem.data_referencia.is_(None),
                        CreditoPlanejamentoItem.data_referencia <= today,
                    ),
                ),
            )
        )
        .order_by(CreditoPlanejamentoItem.hora_inicio.asc(), CreditoPlanejamentoItem.updated_at.desc())
        .limit(200)
        .all()
    )

    for item in candidate_rows:
        if not _credito_planejamento_item_occurs_on(item, today):
            continue
        reminder_at = _credito_planejamento_reminder_at_brt(item)
        if not reminder_at:
            continue
        elapsed = now_brt - reminder_at
        if elapsed < timedelta(0) or elapsed > timedelta(minutes=SIOCRED_EMAIL_ALERT_WINDOW_MINUTES):
            continue

        result["checked"] += 1
        alert_key = _credito_planejamento_alert_key(item, reminder_at)
        already_sent = (
            db.query(func.count(SistemaLog.id))
            .filter(SistemaLog.tela == "frankstein_agenda")
            .filter(SistemaLog.acao == "EMAIL_ALERTA_5MIN")
            .filter(SistemaLog.entidade_id == str(item.id))
            .filter(SistemaLog.details == alert_key)
            .scalar()
            or 0
        )
        if already_sent:
            result["skipped"] += 1
            continue

        view = _credito_planejamento_item_view(item)
        subject = f"{AI_ASSISTANT_DISPLAY_NAME}: alerta em 5 min - {view.get('display_titulo') or item.titulo}"
        body = _credito_planejamento_email_body(item, view, reminder_at)
        try:
            for to_email in recipients:
                _send_email_message(to_email=to_email, subject=subject, text_body=body)
            _record_system_log(
                db,
                actor_username="frankstein",
                actor_role="system",
                tela="frankstein_agenda",
                acao="EMAIL_ALERTA_5MIN",
                entidade_tipo="planejamento",
                entidade_id=str(item.id),
                details=alert_key,
            )
            db.commit()
            result["sent"] += 1
        except Exception as exc:  # pragma: no cover - depends on SMTP provider
            db.rollback()
            logger.exception("Falha ao enviar alerta de agenda %s por e-mail.", AI_ASSISTANT_DISPLAY_NAME)
            result["errors"].append({"item_id": str(item.id), "error": exc.__class__.__name__})

    if result["errors"]:
        result["ok"] = False
    return result


def _credito_planejamento_item_out(item: CreditoPlanejamentoItem) -> CreditoPlanejamentoItemOut:
    view = _credito_planejamento_item_view(item)
    return CreditoPlanejamentoItemOut(
        id=item.id,
        tipo=_credito_planejamento_tipo(item.tipo, fallback="tarefa"),
        titulo=item.titulo,
        descricao=item.descricao,
        responsavel=item.responsavel,
        data_referencia=item.data_referencia,
        hora_inicio=item.hora_inicio,
        hora_fim=item.hora_fim,
        status=_credito_planejamento_status(item.status, fallback="pendente"),
        progresso=_credito_planejamento_progresso(item.progresso),
        urgente=bool(item.urgente) or _credito_planejamento_tipo(item.tipo, fallback="tarefa") == "urgente",
        created_by_username=item.created_by_username,
        updated_by_username=item.updated_by_username,
        created_at=item.created_at,
        updated_at=item.updated_at,
        tipo_label=_credito_planejamento_tipo_label(item.tipo),
        status_label=_credito_planejamento_status_label(item.status),
        display_titulo=view["display_titulo"],
        display_descricao=view["display_descricao"],
        display_meta=view["display_meta"],
        meta_kind=view["meta_kind"],
        meta_cliente=view["meta_cliente"],
        meta_acao=view["meta_acao"],
        meta_observacao=view["meta_observacao"],
        meta_responsavel=view["meta_responsavel"],
        meta_status_oper=view["meta_status_oper"],
        meta_status_oper_label=view["meta_status_oper_label"],
    )


def _analista_reuniao_followup_status(value: Optional[str], *, fallback: str = "seguir") -> str:
    raw = (
        str(value or "")
        .strip()
        .lower()
        .replace("-", "_")
        .replace(" ", "_")
    )
    aliases = {
        "assinado": "assinado",
        "seguir_fluxo": "seguir",
        "seguir_no_fluxo": "seguir",
        "finalizar": "finalizar_hoje",
        "finalizar_no_dia": "finalizar_hoje",
    }
    status = aliases.get(raw, raw)
    if status in ANALISTA_REUNIAO_FOLLOWUP_STATUS:
        return status
    return fallback if fallback in ANALISTA_REUNIAO_FOLLOWUP_STATUS else "seguir"


def _analista_reuniao_compromisso_status(value: Optional[str], *, fallback: str = "pendente") -> str:
    raw = (
        str(value or "")
        .strip()
        .lower()
        .replace("-", "_")
        .replace(" ", "_")
    )
    aliases = {
        "naoentregue": "nao_entregue",
        "nao_entregou": "nao_entregue",
        "entregou": "entregue",
        "done": "entregue",
    }
    status = aliases.get(raw, raw)
    if status in ANALISTA_REUNIAO_COMPROMISSO_STATUS:
        return status
    return fallback if fallback in ANALISTA_REUNIAO_COMPROMISSO_STATUS else "pendente"


def _analista_reuniao_compromisso_out(item: AnalistaReuniaoCompromisso) -> AnalistaReuniaoCompromissoOut:
    return AnalistaReuniaoCompromissoOut(
        id=item.id,
        processo_id=item.processo_id,
        data_prometida=item.data_prometida,
        status=_analista_reuniao_compromisso_status(item.status, fallback="pendente"),
        justificativa=item.justificativa,
        observacao=item.observacao,
        created_by_username=item.created_by_username,
        updated_by_username=item.updated_by_username,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _analista_reuniao_data_base(processo: Processo, cliente: Cliente) -> date:
    base = getattr(cliente, "data_cadastro_origem", None) or getattr(cliente, "data_reserva_origem", None)
    if isinstance(base, date):
        return base
    created = _as_utc(getattr(processo, "created_at", None)) or _utcnow()
    return created.date()


def _analista_reuniao_cliente_out(
    processo: Processo,
    cliente: Cliente,
    followup: Optional[AnalistaReuniaoComercial],
    compromissos: list[AnalistaReuniaoCompromisso],
    *,
    referencia: date,
) -> AnalistaReuniaoClienteOut:
    data_base = _analista_reuniao_data_base(processo, cliente)
    sla_dias = max(0, (referencia - data_base).days)
    status_followup = _analista_reuniao_followup_status(getattr(followup, "status_followup", None), fallback="seguir")
    data_prevista_entrega = getattr(followup, "data_prevista_entrega", None)
    entrega_hoje = bool(data_prevista_entrega and data_prevista_entrega == referencia)
    solicitar_cancelamento = bool(getattr(followup, "solicitar_cancelamento", False))
    referencia_dt = datetime.combine(referencia, time.min, tzinfo=timezone.utc)
    fora_contagem_mes = _is_nao_contar_mes_active(processo, referencia_dt)
    conta_no_mes = (not fora_contagem_mes) and not solicitar_cancelamento
    compromissos_sorted = sorted(
        compromissos,
        key=lambda row: (
            row.data_prometida,
            row.created_at or _utcnow(),
        ),
        reverse=True,
    )
    compromissos_out = [_analista_reuniao_compromisso_out(item) for item in compromissos_sorted]
    nao_entregou_count = sum(1 for item in compromissos_out if item.status == "nao_entregue")
    return AnalistaReuniaoClienteOut(
        processo_id=processo.id,
        cliente_nome=cliente.nome,
        empreendimento=cliente.obra,
        corretor=cliente.corretor,
        imobiliaria=getattr(cliente, "imobiliaria", None),
        estagio_comercial=_process_estagio_comercial(processo.estagio_comercial),
        status_cca=_process_caixa_status(processo.status_cca),
        data_cadastro_origem=data_base,
        sla_dias=sla_dias,
        conta_no_mes=conta_no_mes,
        data_prevista_entrega=data_prevista_entrega,
        entrega_hoje=entrega_hoje,
        probabilidade_queda=bool(getattr(followup, "probabilidade_queda", False)),
        solicitar_cancelamento=solicitar_cancelamento,
        status_followup=status_followup,
        observacao=getattr(followup, "observacao", None),
        justificativa_reincidencia=getattr(followup, "justificativa_reincidencia", None),
        nao_entregou_count=nao_entregou_count,
        compromissos=compromissos_out,
    )


def _record_processo_event(
    db: Session,
    *,
    processo_id: uuid.UUID,
    actor_username: str,
    actor_role: str,
    event_type: str,
    field_name: Optional[str] = None,
    old_value: Any = None,
    new_value: Any = None,
    details: Optional[str] = None,
) -> None:
    role_raw = (actor_role or "").strip().lower()
    db.add(
        ProcessoEvento(
            processo_id=processo_id,
            actor_username=_normalize_username(actor_username) or "system",
            actor_role=role_raw if role_raw in VALID_ROLES else "system",
            event_type=(event_type or "").strip().upper() or "EVENT",
            field_name=(field_name or "").strip() or None,
            old_value=_stringify_audit_value(old_value),
            new_value=_stringify_audit_value(new_value),
            details=(details or "").strip() or None,
        )
    )


def _record_system_log(
    db: Session,
    *,
    actor_username: str,
    actor_role: str,
    tela: str,
    acao: str,
    entidade_tipo: Optional[str] = None,
    entidade_id: Optional[str] = None,
    details: Optional[str] = None,
) -> None:
    role_raw = (actor_role or "").strip().lower()
    db.add(
        SistemaLog(
            actor_username=_normalize_username(actor_username) or "system",
            actor_role=role_raw if role_raw in VALID_ROLES else "system",
            tela=(tela or "").strip().lower() or "sistema",
            acao=(acao or "").strip().upper() or "EVENT",
            entidade_tipo=(entidade_tipo or "").strip().lower() or None,
            entidade_id=(entidade_id or "").strip() or None,
            details=(details or "").strip() or None,
        )
    )


def _record_keepalive_ping(request: Request) -> dict[str, Any]:
    now = _utcnow()
    ts_utc = now.isoformat()
    ts_brt = now.astimezone(KEEPALIVE_BRT_TZ).isoformat()
    ip = _request_client_ip(request)
    host = _request_host(request)
    user_agent = (request.headers.get("user-agent") or "").strip()
    forwarded_for = (request.headers.get("x-forwarded-for") or "").strip()
    forward_proto = (request.headers.get("x-forwarded-proto") or "").strip()

    event: dict[str, Any] = {
        "ts_utc": ts_utc,
        "ts_brt": ts_brt,
        "client_ip": ip or "",
        "host": host or "",
        "user_agent": user_agent[:240],
        "x_forwarded_for": forwarded_for[:240],
        "x_forwarded_proto": forward_proto[:40],
        "db_logged": False,
    }
    KEEPALIVE_RECENT.append(event.copy())

    if SessionLocal is None or DB_URL_HAS_PLACEHOLDERS:
        return event

    db = SessionLocal()
    try:
        details = (
            f"ts_utc={ts_utc}; ts_brt={ts_brt}; ip={ip or '-'}; host={host or '-'}; "
            f"x_forwarded_for={forwarded_for or '-'}; x_forwarded_proto={forward_proto or '-'}; "
            f"user_agent={user_agent or '-'}"
        )
        _record_system_log(
            db,
            actor_username="keepalive",
            actor_role="system",
            tela=KEEPALIVE_TELA,
            acao="PING",
            entidade_tipo="health",
            entidade_id=(ip or None),
            details=details[:1800],
        )
        db.commit()
        event["db_logged"] = True
        if KEEPALIVE_RECENT:
            KEEPALIVE_RECENT[-1]["db_logged"] = True
    except Exception as exc:
        db.rollback()
        logger.warning("Falha ao gravar keepalive em sistema_logs: %s", exc.__class__.__name__)
    finally:
        db.close()

    return event


def _frankstein_normalize_pattern_text(value: Any) -> str:
    text_value = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text_value = "".join(ch for ch in text_value if unicodedata.category(ch) != "Mn")
    return " ".join(text_value.split())


def _frankstein_comprometimento_faixa(value: Any) -> str:
    try:
        percent = float(value or 0)
    except (TypeError, ValueError):
        percent = 0.0
    if percent <= 30:
        return "ok"
    if percent <= 40:
        return "atencao"
    if percent <= 45:
        return "critico"
    return "complementar"


def _frankstein_context_signature(contexto: dict[str, Any]) -> str:
    parts = [
        _frankstein_normalize_pattern_text(contexto.get("perfil_renda")),
        _frankstein_normalize_pattern_text(contexto.get("tipo_dependente")),
        _frankstein_normalize_pattern_text(contexto.get("cotista_3_anos")),
        _frankstein_normalize_pattern_text(contexto.get("empresa_atual_6m")),
        _frankstein_comprometimento_faixa(contexto.get("comprometimento_percent")),
        _frankstein_normalize_pattern_text(contexto.get("vencimento_status")),
    ]
    return "|".join(parts)


def _frankstein_learning_suggestions(db: Session, contexto: dict[str, Any]) -> list[dict[str, Any]]:
    rows = (
        db.query(SistemaLog)
        .filter(SistemaLog.tela == "frankstein")
        .filter(SistemaLog.acao == "ANALISTA_APRENDIZADO")
        .order_by(SistemaLog.created_at.desc())
        .limit(700)
        .all()
    )
    signature = _frankstein_context_signature(contexto)
    similar_count = 0
    status_counter: dict[str, int] = {}
    doc_counter: dict[str, dict[str, Any]] = {}
    complementary_count = 0
    total_interacoes = 0

    for row in rows:
        try:
            payload = json.loads(row.details or "{}")
        except json.JSONDecodeError:
            continue
        total_interacoes += 1
        row_context = payload.get("contexto") if isinstance(payload.get("contexto"), dict) else {}
        if _frankstein_context_signature(row_context) != signature:
            continue
        similar_count += 1
        decisao = payload.get("decisao_analista") if isinstance(payload.get("decisao_analista"), dict) else {}
        for field in ("status_caixa", "status_agehab", "status_sinal", "status_fiador"):
            value = _frankstein_normalize_pattern_text(decisao.get(field))
            if value:
                key = f"{field}:{value}"
                status_counter[key] = status_counter.get(key, 0) + 1
        if decisao.get("renda_complementar_preenchida"):
            complementary_count += 1
        docs = decisao.get("documentos")
        if isinstance(docs, list):
            for doc in docs:
                if not isinstance(doc, dict):
                    continue
                categoria = _frankstein_normalize_pattern_text(doc.get("categoria"))
                nome = _frankstein_normalize_pattern_text(doc.get("nome"))
                status = _frankstein_normalize_pattern_text(doc.get("status"))
                motivo = _frankstein_normalize_pattern_text(doc.get("motivo"))
                if not categoria or not nome or status in {"", "aguardando"}:
                    continue
                key = f"{categoria}:{nome}:{status}:{motivo[:80]}"
                entry = doc_counter.setdefault(
                    key,
                    {
                        "count": 0,
                        "categoria": doc.get("categoria"),
                        "nome": doc.get("nome"),
                        "status": doc.get("status"),
                        "motivo": doc.get("motivo"),
                    },
                )
                entry["count"] += 1

    suggestions: list[dict[str, Any]] = []
    if similar_count >= 3:
        for key, count in sorted(status_counter.items(), key=lambda item: item[1], reverse=True)[:3]:
            field, value = key.split(":", 1)
            if count < 3:
                continue
            confidence = round(count / max(similar_count, 1), 2)
            suggestions.append(
                {
                    "tipo": "padrao_status",
                    "status": "dica" if confidence < 0.8 else "atencao",
                    "titulo": "Aprendido com analistas",
                    "mensagem": (
                        f"Em {count} de {similar_count} casos parecidos, o campo {field.replace('_', ' ')} "
                        f"foi salvo como {value}. Confianca {int(confidence * 100)}%."
                    ),
                    "confianca": confidence,
                }
            )
        if complementary_count >= 3:
            confidence = round(complementary_count / max(similar_count, 1), 2)
            suggestions.append(
                {
                    "tipo": "padrao_renda_complementar",
                    "status": "atencao" if confidence >= 0.8 else "dica",
                    "titulo": "Padrao de renda complementar",
                    "mensagem": (
                        f"Em {complementary_count} de {similar_count} casos parecidos, o analista registrou renda complementar. "
                        f"Confianca {int(confidence * 100)}%."
                    ),
                    "confianca": confidence,
                }
            )
        for entry in sorted(doc_counter.values(), key=lambda item: item["count"], reverse=True)[:4]:
            if int(entry["count"]) < 3:
                continue
            confidence = round(int(entry["count"]) / max(similar_count, 1), 2)
            suggestions.append(
                {
                    "tipo": "padrao_documento",
                    "status": "dica" if confidence < 0.8 else "atencao",
                    "titulo": "Padrao documental aprendido",
                    "mensagem": (
                        f"{entry['nome']} costuma ser marcado como {entry['status']} em casos parecidos"
                        + (f" com motivo: {entry['motivo']}." if entry.get("motivo") else ".")
                    ),
                    "confianca": confidence,
                }
            )

    return suggestions


def _fetch_keepalive_db_logs(limit: int) -> list[dict[str, Any]]:
    if SessionLocal is None or DB_URL_HAS_PLACEHOLDERS:
        return []

    db = SessionLocal()
    try:
        rows = (
            db.query(SistemaLog)
            .filter(func.lower(func.trim(SistemaLog.tela)) == KEEPALIVE_TELA)
            .filter(SistemaLog.acao == "PING")
            .order_by(SistemaLog.created_at.desc())
            .limit(limit)
            .all()
        )
        return [
            {
                "id": str(row.id),
                "ts_utc": (_as_utc(row.created_at) or _utcnow()).isoformat(),
                "actor_username": row.actor_username,
                "details": row.details or "",
            }
            for row in rows
        ]
    except Exception as exc:
        logger.warning("Falha ao consultar logs keepalive: %s", exc.__class__.__name__)
        return []
    finally:
        db.close()


def _processo_has_pendencia(processo: "Processo") -> bool:
    status_values = {
        "status_credito": processo.status_credito,
        "status_geral": processo.status_geral,
        "status_cca": processo.status_cca,
        "status_agehab": processo.status_agehab,
        "status_sinal": processo.status_sinal,
        "status_fiador": processo.status_fiador,
    }
    return any(_is_pendencia_status(field, value) for field, value in status_values.items())


def _query_processos_by_scope(db: Session, role: str, username: str, include_archived: bool = False):
    query = db.query(Processo)
    if not include_archived:
        query = query.filter(_processos_ativos_clause())
    if role == ROLE_CORRETOR:
        if not username:
            return query.filter(text("1=0"))
        query = query.join(Cliente, Processo.cliente_id == Cliente.id).filter(
            func.lower(func.trim(func.coalesce(Cliente.corretor, ""))) == username
        )
    elif role == ROLE_CCA:
        if not username:
            return query.filter(text("1=0"))
        query = query.filter(func.lower(func.trim(func.coalesce(Processo.cca_responsavel, ""))) == username)
    return query


def _set_user_password(user: AppUser, password: str, must_change_password: bool = True) -> None:
    salt = _new_salt()
    user.password_salt = salt
    user.password_hash = _hash_password(password, salt)
    user.must_change_password = bool(must_change_password)


def _get_user_by_username(db: Session, username: str) -> Optional[AppUser]:
    username_key = _normalize_username(username)
    if not username_key:
        return None
    return db.query(AppUser).filter(func.lower(AppUser.username) == username_key).first()


def _ensure_seed_users(db: Session, force: bool = False) -> None:
    global SEED_USERS_READY
    if SEED_USERS_READY and not force:
        return

    seed_mode_raw = (_get_runtime_meta(db, USERS_SEED_MODE_RUNTIME_KEY) or USERS_SEED_MODE_FULL).strip().lower()
    seed_mode = seed_mode_raw if seed_mode_raw in {USERS_SEED_MODE_FULL, USERS_SEED_MODE_ADMIN_ONLY} else USERS_SEED_MODE_FULL

    seeds = []
    if seed_mode == USERS_SEED_MODE_ADMIN_ONLY:
        admin_username = _normalize_username(RESET_ADMIN_USERNAME) or _normalize_username(APP_ADMIN_USER)
        admin_password = (APP_ADMIN_PASSWORD or "").strip()
        seeds.append((admin_username, admin_password, ROLE_ADMIN))
        admin_seed_username = admin_username
    else:
        for username, account in APP_USERS.items():
            seeds.append((username, account["password"], _normalize_role(account["role"])))
        admin_seed_username = _normalize_username(APP_ADMIN_USER)

    created = 0
    changed = 0
    for username, password, role in seeds:
        if not username or not (password or "").strip():
            logger.warning("Seed ignorado para usuario '%s': senha ausente no ambiente.", username or "<vazio>")
            continue
        policy_error = _password_policy_error(password)
        if policy_error and not ALLOW_WEAK_SEED_PASSWORDS:
            logger.warning("Seed ignorado para usuario '%s': %s", username, policy_error)
            continue
        if policy_error:
            logger.warning("Senha fraca em usuario seed '%s': %s", username, policy_error)

        existing = _get_user_by_username(db, username)
        if existing:
            is_admin_seed = _normalize_username(existing.username) == admin_seed_username
            if is_admin_seed and FORCE_RECOVER_ADMIN_ON_STARTUP:
                admin_changed = False
                if existing.role != ROLE_ADMIN:
                    existing.role = ROLE_ADMIN
                    admin_changed = True
                if not bool(existing.is_active):
                    existing.is_active = True
                    admin_changed = True
                if not existing.must_change_password:
                    existing.must_change_password = True
                    admin_changed = True
                existing.last_login_at = None
                if not _verify_password(password, existing.password_hash, existing.password_salt):
                    _set_user_password(existing, password, must_change_password=True)
                    admin_changed = True
                if admin_changed:
                    changed += 1
                continue
            if existing.last_login_at is None and not existing.must_change_password:
                existing.must_change_password = True
                changed += 1
            continue
        user = AppUser(
            username=_normalize_username(username),
            role=role,
            is_active=True,
            must_change_password=True,
        )
        _set_user_password(user, password, must_change_password=True)
        db.add(user)
        created += 1

    if created or changed:
        db.commit()
    SEED_USERS_READY = True


def _normalize_empreendimento_nome(value: Optional[str]) -> str:
    return " ".join((value or "").strip().split())


def _normalize_lead_text(value: Optional[str], *, max_len: int = 400) -> Optional[str]:
    text = " ".join(str(value or "").strip().split())
    if not text:
        return None
    return text[:max_len]


def _normalize_lead_email(value: Optional[str]) -> Optional[str]:
    email = _normalize_lead_text(value, max_len=180)
    if not email:
        return None
    return email.lower()


def _normalize_lead_phone(value: Optional[str]) -> Optional[str]:
    phone = _normalize_lead_text(value, max_len=40)
    if not phone:
        return None
    return phone


def _normalize_lead_cpf(value: Optional[str]) -> Optional[str]:
    raw = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not raw:
        return None
    if len(raw) > 11:
        raw = raw[:11]
    return raw


def _normalize_lead_documento(value: Optional[str]) -> Optional[str]:
    doc = _normalize_lead_text(value, max_len=40)
    if not doc:
        return None
    return doc.upper()


def _refresh_lead_security_fields(lead: LeadPreCadastro) -> bool:
    changed = False

    telefone = decrypt_pii(getattr(lead, "_telefone", None))
    whatsapp = decrypt_pii(getattr(lead, "_whatsapp", None))
    email = decrypt_pii(getattr(lead, "_email", None))
    cpf = decrypt_pii(getattr(lead, "_cpf", None))
    documento = decrypt_pii(getattr(lead, "_documento_identificacao", None))
    certidao = decrypt_pii(getattr(lead, "_certidao_numero", None))
    cidade = decrypt_pii(getattr(lead, "_cidade_nascimento", None))
    endereco = decrypt_pii(getattr(lead, "_endereco", None))

    expected_telefone_last4 = last4_digits(telefone)
    expected_whatsapp_last4 = last4_digits(whatsapp)
    expected_email_hash = hash_email(email)
    expected_cpf_hash = hash_optional(cpf)
    expected_cpf_last4 = last4_digits(cpf)
    expected_documento_hash = hash_optional(documento)

    if lead.telefone_last4 != expected_telefone_last4:
        lead.telefone_last4 = expected_telefone_last4
        changed = True
    if lead.whatsapp_last4 != expected_whatsapp_last4:
        lead.whatsapp_last4 = expected_whatsapp_last4
        changed = True
    if lead.email_hash != expected_email_hash:
        lead.email_hash = expected_email_hash
        changed = True
    if lead.cpf_hash != expected_cpf_hash:
        lead.cpf_hash = expected_cpf_hash
        changed = True
    if lead.cpf_last4 != expected_cpf_last4:
        lead.cpf_last4 = expected_cpf_last4
        changed = True
    if lead.documento_hash != expected_documento_hash:
        lead.documento_hash = expected_documento_hash
        changed = True

    if pii_encryption_enabled():
        secure_fields = [
            ("_telefone", telefone),
            ("_whatsapp", whatsapp),
            ("_email", email),
            ("_cpf", cpf),
            ("_documento_identificacao", documento),
            ("_certidao_numero", certidao),
            ("_cidade_nascimento", cidade),
            ("_endereco", endereco),
        ]
        for attr_name, plain in secure_fields:
            current = getattr(lead, attr_name, None)
            expected = encrypt_pii(plain)
            if expected and current != expected:
                setattr(lead, attr_name, expected)
                changed = True

    return changed


def _backfill_lead_security_fields(db: Session) -> int:
    updated = 0
    rows = db.query(LeadPreCadastro).all()
    for lead in rows:
        if _refresh_lead_security_fields(lead):
            updated += 1
    if updated:
        db.commit()
    return updated


def _normalize_estado_civil(value: Optional[str]) -> Optional[str]:
    token = _normalize_text_key(value)
    aliases = {
        "solteiro": "SOLTEIRO",
        "solteira": "SOLTEIRO",
        "casado": "CASADO",
        "casada": "CASADO",
        "divorciado": "DIVORCIADO",
        "divorciada": "DIVORCIADO",
        "viuvo": "VIUVO",
        "viuva": "VIUVO",
        "uniao_estavel": "UNIAO_ESTAVEL",
    }
    mapped = aliases.get(token, "")
    return mapped or None


def _normalize_tipo_visita(value: Optional[str]) -> Optional[str]:
    token = _normalize_text_key(value)
    aliases = {
        "online": "ONLINE",
        "presencial": "PRESENCIAL",
        "hibrida": "HIBRIDA",
        "hibrido": "HIBRIDA",
    }
    mapped = aliases.get(token, "")
    return mapped or None


def _email_delivery_provider() -> str:
    if EMAIL_DELIVERY_PROVIDER in {"brevo", "smtp"}:
        return EMAIL_DELIVERY_PROVIDER
    if EMAIL_BREVO_API_KEY:
        return "brevo"
    return "smtp"


def _is_email_delivery_configured() -> bool:
    if _email_delivery_provider() == "brevo":
        return bool(EMAIL_BREVO_API_KEY and EMAIL_FROM)
    return bool(EMAIL_SMTP_HOST and EMAIL_FROM)


def _send_email_message_brevo(*, to_email: str, subject: str, text_body: str) -> dict[str, Any]:
    payload = {
        "sender": {"name": EMAIL_FROM_NAME, "email": EMAIL_FROM},
        "to": [{"email": to_email}],
        "replyTo": {"email": EMAIL_FROM},
        "subject": subject,
        "textContent": text_body,
        "tags": ["siocred-frankstein"],
    }
    data = json.dumps(payload, ensure_ascii=True).encode("utf-8")
    request = UrlLibRequest(
        EMAIL_BREVO_API_URL,
        data=data,
        method="POST",
        headers={
            "accept": "application/json",
            "api-key": EMAIL_BREVO_API_KEY,
            "content-type": "application/json",
        },
    )
    try:
        with urlopen(request, timeout=20) as response:
            response_body = response.read().decode("utf-8", errors="ignore")
            status_code = int(getattr(response, "status", 0) or 0)
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore").strip()
        raise RuntimeError(f"Brevo API recusou envio ({exc.code}): {detail or exc.reason}") from exc
    except URLError as exc:
        raise RuntimeError(f"Falha de conexao com Brevo API: {exc.reason}") from exc

    try:
        parsed = json.loads(response_body) if response_body else {}
    except json.JSONDecodeError:
        parsed = {}
    message_id = str(parsed.get("messageId") or make_msgid(domain=(EMAIL_FROM.split("@")[-1] if "@" in EMAIL_FROM else None)))
    return {
        "to": to_email,
        "message_id": message_id,
        "subject": subject,
        "provider": "brevo",
        "status_code": status_code,
    }


def _send_email_message(*, to_email: str, subject: str, text_body: str) -> dict[str, Any]:
    if not _is_email_delivery_configured():
        raise RuntimeError("Envio de e-mail nao configurado no ambiente.")

    if _email_delivery_provider() == "brevo":
        return _send_email_message_brevo(to_email=to_email, subject=subject, text_body=text_body)

    message_id = make_msgid(domain=(EMAIL_FROM.split("@")[-1] if "@" in EMAIL_FROM else None))
    msg = EmailMessage()
    msg["From"] = EMAIL_FROM
    msg["To"] = to_email
    msg["Subject"] = subject
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = message_id
    msg["Reply-To"] = EMAIL_FROM
    msg["X-SioCred-Source"] = AI_ASSISTANT_DISPLAY_NAME
    msg.set_content(text_body)

    with smtplib.SMTP(EMAIL_SMTP_HOST, EMAIL_SMTP_PORT, timeout=20) as smtp:
        if EMAIL_SMTP_STARTTLS:
            smtp.starttls()
        if EMAIL_SMTP_USER:
            smtp.login(EMAIL_SMTP_USER, EMAIL_SMTP_PASSWORD)
        refused = smtp.send_message(msg)

    if refused:
        raise smtplib.SMTPRecipientsRefused(refused)

    return {"to": to_email, "message_id": message_id, "subject": subject, "provider": "smtp"}


def _split_alert_recipients(raw: Optional[str] = None) -> list[str]:
    source = str(raw if raw is not None else SIOCRED_ALERT_EMAIL_TO or "").strip()
    if not source:
        return []
    for separator in (";", "\n", "\t"):
        source = source.replace(separator, ",")
    recipients = []
    seen = set()
    for part in source.split(","):
        email = part.strip()
        if not email or email.lower() in seen:
            continue
        seen.add(email.lower())
        recipients.append(email)
    return recipients


def _validate_alert_recipients(raw: Optional[str]) -> list[str]:
    recipients = _split_alert_recipients(raw)
    invalid = [email for email in recipients if "@" not in email or "." not in email.rsplit("@", 1)[-1]]
    if invalid:
        raise HTTPException(status_code=422, detail=f"E-mail invalido: {invalid[0]}")
    return recipients


def _get_alert_recipients_source(db: Optional[Session] = None) -> tuple[str, str]:
    if db is not None:
        runtime_value = _get_runtime_meta(db, FRANKSTEIN_ALERT_EMAIL_TO_RUNTIME_KEY)
        if runtime_value is not None:
            return runtime_value.strip(), "admin"
    return SIOCRED_ALERT_EMAIL_TO, "ambiente"


def _build_frankstein_email_alerts_status(db: Session) -> AdminFranksteinEmailAlertsOut:
    raw, source = _get_alert_recipients_source(db)
    recipients = _split_alert_recipients(raw)
    provider = _email_delivery_provider()
    smtp_missing = []
    if provider == "brevo":
        if not EMAIL_BREVO_API_KEY:
            smtp_missing.append("EMAIL_BREVO_API_KEY")
        if not EMAIL_FROM:
            smtp_missing.append("EMAIL_FROM ou EMAIL_SMTP_FROM")
    else:
        if not EMAIL_SMTP_HOST:
            smtp_missing.append("EMAIL_SMTP_HOST")
        if not EMAIL_FROM:
            smtp_missing.append("EMAIL_FROM ou EMAIL_SMTP_FROM")
    return AdminFranksteinEmailAlertsOut(
        smtp_configured=_is_email_delivery_configured(),
        smtp_missing=smtp_missing,
        smtp_config={
            "provider": provider,
            "api_key_present": bool(EMAIL_BREVO_API_KEY),
            "api_url_present": bool(EMAIL_BREVO_API_URL),
            "host_present": bool(EMAIL_SMTP_HOST),
            "port": EMAIL_SMTP_PORT,
            "user_present": bool(EMAIL_SMTP_USER),
            "password_present": bool(EMAIL_SMTP_PASSWORD),
            "from_present": bool(EMAIL_FROM),
            "starttls": EMAIL_SMTP_STARTTLS,
        },
        recipients_configured=bool(recipients),
        destinatarios=raw,
        destinatarios_mascarados=[mask_email(email) or email for email in recipients],
        fonte_destinatarios=source,
        window_minutes=SIOCRED_EMAIL_ALERT_WINDOW_MINUTES,
    )


def _build_email_confirmation_link(request: Request, token: str) -> str:
    base = EMAIL_CONFIRM_LINK_BASE_URL.rstrip("/")
    if not base:
        base = str(request.base_url).rstrip("/")
    return f"{base}/app/assinatura/confirmar?token={token}"


def _lead_reserva_block_reason(lead: "LeadPreCadastro") -> Optional[str]:
    decisao = _lead_cca_decision(getattr(lead, "decisao_cca", None))
    if decisao not in {"APROVADO", "CONDICIONADO"}:
        return "Reserva bloqueada: CCA precisa aprovar ou condicionar o cliente."
    if not bool(getattr(lead, "contrato_assinado", False)):
        return "Reserva bloqueada: contrato ainda nao foi marcado como assinado."
    if not bool(getattr(lead, "assinatura_email_confirmada", False)):
        return "Reserva bloqueada: assinatura do contrato ainda nao foi confirmada por e-mail."
    return None


def _normalize_currency_value(value: Any) -> Optional[float]:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace("R$", "").replace(" ", "")
    # Aceita formatos "1.234,56" e "1234.56"
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        parsed = float(raw)
    except ValueError:
        return None
    if parsed < 0:
        return None
    return round(parsed, 2)


def _renda_bruta_duplicate_key(value: Any) -> Optional[int]:
    parsed = _normalize_currency_value(value)
    if parsed is None or parsed <= 0:
        return None
    return int(round(parsed * 100))


def _format_currency_brl_from_cents(cents: int) -> str:
    value = abs(int(cents))
    reais = value // 100
    centavos = value % 100
    reais_txt = f"{reais:,}".replace(",", ".")
    sign = "-" if cents < 0 else ""
    return f"R$ {sign}{reais_txt},{centavos:02d}"


def _cliente_nome_sobrenome(value: Optional[str]) -> str:
    partes = [part.strip() for part in str(value or "").split() if part.strip()]
    if not partes:
        return "-"
    if len(partes) == 1:
        return partes[0]
    return f"{partes[0]} {partes[-1]}"


def _build_renda_bruta_duplicate_lookup(rows: list[tuple[Any, Any, Any]]) -> dict[str, dict[str, Any]]:
    grouped: dict[int, list[tuple[str, str]]] = {}
    for processo_id, renda_bruta, cliente_nome in rows:
        key = _renda_bruta_duplicate_key(renda_bruta)
        if key is None:
            continue
        grouped.setdefault(key, []).append((str(processo_id), _cliente_nome_sobrenome(cliente_nome)))

    lookup: dict[str, dict[str, Any]] = {}
    for cents, entries in grouped.items():
        seen: set[str] = set()
        nomes = []
        for _, nome in entries:
            nome_key = nome.lower()
            if nome_key not in seen:
                seen.add(nome_key)
                nomes.append(nome)
        if len(entries) < 2:
            continue
        tooltip = f"RD - Renda bruta igual ({_format_currency_brl_from_cents(cents)}): {', '.join(nomes)}"
        for processo_id, _ in entries:
            lookup[processo_id] = {
                "clientes": nomes,
                "tooltip": tooltip,
            }
    return lookup


def _resolve_cheque_moradia_valor(db: Session, obra_nome: Optional[str]) -> float:
    obra = _normalize_empreendimento_nome(obra_nome)
    if not obra:
        return 0.0
    empreendimento = (
        db.query(Empreendimento)
        .filter(func.lower(Empreendimento.nome) == obra.lower())
        .first()
    )
    if not empreendimento:
        return 0.0
    regra = (
        db.query(EmpreendimentoRegraFinanceira)
        .filter(EmpreendimentoRegraFinanceira.empreendimento_id == empreendimento.id)
        .first()
    )
    if not regra:
        return 0.0
    valor = float(regra.valor_cheque_moradia or 0.0)
    return round(valor if valor > 0 else 0.0, 2)


def _coerce_optional_currency(value: Any, field_label: str) -> Optional[float]:
    if value is None:
        return None
    parsed = _normalize_currency_value(value)
    if parsed is None:
        raise HTTPException(status_code=422, detail=f"{field_label} invalido.")
    return parsed


def _build_cca_analise_item(
    db: Session,
    processo: "Processo",
    cliente: "Cliente",
    *,
    lead: Optional["LeadPreCadastro"] = None,
    docs_total: int = 0,
    docs_recebidos: int = 0,
) -> CcaAnaliseItemOut:
    valor_cheque_moradia = getattr(processo, "valor_cheque_moradia", None)
    if valor_cheque_moradia is None:
        valor_cheque_moradia = _resolve_cheque_moradia_valor(db, getattr(cliente, "obra", None))

    return CcaAnaliseItemOut(
        processo_id=processo.id,
        cliente_id=cliente.id,
        cliente_nome=cliente.nome,
        corretor=cliente.corretor,
        obra=cliente.obra,
        estagio_comercial=_process_estagio_comercial(processo.estagio_comercial),
        status_cca=_process_caixa_status(processo.status_cca),
        status_agehab=_process_agehab_status(processo.status_agehab),
        renda_bruta=float(processo.renda_bruta) if getattr(processo, "renda_bruta", None) is not None else None,
        renda_liquida=float(processo.renda_liquida) if getattr(processo, "renda_liquida", None) is not None else None,
        valor_parcela=float(processo.valor_parcela) if getattr(processo, "valor_parcela", None) is not None else None,
        valor_imovel=float(processo.valor_imovel) if getattr(processo, "valor_imovel", None) is not None else None,
        valor_avaliacao=float(processo.valor_avaliacao) if getattr(processo, "valor_avaliacao", None) is not None else None,
        valor_financiamento=float(processo.valor_financiamento) if getattr(processo, "valor_financiamento", None) is not None else None,
        valor_subsidio=float(processo.valor_subsidio) if getattr(processo, "valor_subsidio", None) is not None else None,
        valor_cheque_moradia=float(valor_cheque_moradia) if valor_cheque_moradia is not None else None,
        recolha_fgts=_process_recolha_fgts_status(getattr(processo, "recolha_fgts", None)),
        cpf=getattr(lead, "cpf", None),
        documento_identificacao=getattr(lead, "documento_identificacao", None),
        estado_civil=getattr(lead, "estado_civil", None),
        certidao_numero=getattr(lead, "certidao_numero", None),
        cidade_nascimento=getattr(lead, "cidade_nascimento", None),
        data_nascimento=getattr(lead, "data_nascimento", None),
        telefone=getattr(lead, "telefone", None),
        whatsapp=getattr(lead, "whatsapp", None),
        email=getattr(lead, "email", None),
        docs_total=max(0, int(docs_total or 0)),
        docs_recebidos=max(0, int(docs_recebidos or 0)),
        docs_pendentes=max(0, int(docs_total or 0) - int(docs_recebidos or 0)),
        updated_at=getattr(processo, "updated_at", None),
    )


def _normalize_unidade_text(value: Optional[str], *, max_len: int = 220) -> Optional[str]:
    text = " ".join(str(value or "").strip().split())
    if not text:
        return None
    return text[:max_len]


def _normalize_unidade_positive_int(value: Optional[int], *, min_value: int = 0, max_value: int = 999) -> Optional[int]:
    if value is None:
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None
    if parsed < min_value:
        return None
    return min(parsed, max_value)


def _normalize_unidade_positive_float(value: Optional[float], *, min_value: float = 0.0, max_value: float = 1_000_000_000.0) -> Optional[float]:
    if value is None:
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if parsed < min_value:
        return None
    return min(parsed, max_value)


def _resolve_empreendimento_nome(db: Session, value: Optional[str]) -> Optional[str]:
    nome = _normalize_empreendimento_nome(value)
    if not nome:
        return None

    empreendimento = (
        db.query(Empreendimento)
        .filter(func.lower(Empreendimento.nome) == nome.lower(), Empreendimento.is_active.is_(True))
        .first()
    )
    if empreendimento:
        return empreendimento.nome
    return None


def _normalize_cliente_key(value: Optional[str]) -> str:
    raw = " ".join((value or "").strip().lower().split())
    if not raw:
        return ""
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_text = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return " ".join(ascii_text.split())


def _parse_import_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        serial = int(value)
        if 1 <= serial <= 60000:
            # Excel serial date (compatible with 1900 date system).
            return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
        return None

    raw = str(value).strip()
    if not raw:
        return None
    try:
        numeric = float(raw.replace(",", "."))
        serial = int(numeric)
        if 1 <= serial <= 60000:
            return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_import_header(value: Any) -> str:
    token = _normalize_text_key(str(value or ""))
    return IMPORT_COLUMN_ALIASES.get(token, token)


def _canonicalize_import_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw in raw_rows:
        item: dict[str, Any] = {}
        for key, value in (raw or {}).items():
            canonical = _normalize_import_header(key)
            if canonical in IMPORT_REQUIRED_COLUMNS:
                item[canonical] = value
        rows.append(item)
    return rows


def _decode_csv_import_content(content: bytes) -> str:
    for encoding in CSV_IMPORT_ENCODINGS:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _detect_csv_delimiter(text_value: str) -> str:
    lines = [line for line in text_value.splitlines() if line.strip()]
    if not lines:
        return ","
    sample = "\n".join(lines[: min(10, len(lines))])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(CSV_IMPORT_DELIMITERS))
        if dialect and dialect.delimiter in CSV_IMPORT_DELIMITERS:
            return dialect.delimiter
    except csv.Error:
        pass
    header = lines[0]
    counts = {delimiter: header.count(delimiter) for delimiter in CSV_IMPORT_DELIMITERS}
    selected = max(counts, key=counts.get)
    return selected if counts[selected] > 0 else ","


def _parse_csv_import(content: bytes) -> list[dict[str, Any]]:
    text_value = _decode_csv_import_content(content)
    delimiter = _detect_csv_delimiter(text_value)
    reader = csv.DictReader(io.StringIO(text_value), delimiter=delimiter)
    rows: list[dict[str, Any]] = []
    for row in reader:
        if row is None:
            continue
        if all(not str(value or "").strip() for value in row.values()):
            continue
        rows.append(dict(row))
    return rows


def _parse_xlsx_import(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise HTTPException(
            status_code=422,
            detail="Importacao XLSX indisponivel no servidor. Instale a dependencia openpyxl.",
        ) from exc

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [str(cell or "").strip() for cell in header_row]
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        item: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            item[header] = values[idx] if idx < len(values) else None
        rows.append(item)
    return rows


def _load_import_rows(filename: Optional[str], content: bytes) -> list[dict[str, Any]]:
    name = (filename or "").lower().strip()
    if name.endswith(".csv") or name.endswith(".cvs"):
        return _canonicalize_import_rows(_parse_csv_import(content))
    if name.endswith(".xlsx"):
        return _canonicalize_import_rows(_parse_xlsx_import(content))
    raise HTTPException(status_code=422, detail="Formato nao suportado. Envie arquivo .csv/.cvs ou .xlsx.")


def _set_runtime_meta(db: Session, key: str, value: str) -> None:
    table_exists = db.execute(text("SELECT to_regclass('public.app_runtime_meta')")).scalar()
    if not table_exists:
        return
    db.execute(
        text(
            """
            INSERT INTO app_runtime_meta (meta_key, meta_value, updated_at)
            VALUES (:key, :value, NOW())
            ON CONFLICT (meta_key) DO UPDATE
            SET meta_value = EXCLUDED.meta_value,
                updated_at = NOW()
            """
        ),
        {"key": key, "value": value},
    )


def _get_runtime_meta(db: Session, key: str) -> Optional[str]:
    table_exists = db.execute(text("SELECT to_regclass('public.app_runtime_meta')")).scalar()
    if not table_exists:
        return None
    return db.execute(
        text("SELECT meta_value FROM app_runtime_meta WHERE meta_key = :key"),
        {"key": key},
    ).scalar()


def _runtime_meta_non_negative_int(raw_value: Optional[str]) -> Optional[int]:
    if raw_value is None:
        return None
    try:
        return max(0, int(str(raw_value).strip() or "0"))
    except ValueError:
        return None


def _runtime_meta_bool(raw_value: Optional[str], default: bool = False) -> bool:
    if raw_value is None:
        return default
    token = str(raw_value).strip().lower()
    if token in {"1", "true", "yes", "on", "blackhole"}:
        return True
    if token in {"0", "false", "no", "off", "corporativo", "corporate"}:
        return False
    return default


def _is_blackhole_layout_enabled(db: Session) -> bool:
    if not BLACKHOLE_LAYOUT_ALLOWED:
        return False
    return _runtime_meta_bool(_get_runtime_meta(db, LAYOUT_BLACKHOLE_RUNTIME_KEY), default=False)


def _current_meta_period() -> tuple[int, int]:
    hoje = _utcnow().date()
    return hoje.year, hoje.month


def _set_nao_contar_mes_period(processo: "Processo", marcado: bool, now: Optional[datetime] = None) -> None:
    if not marcado:
        processo.nao_contar_mes = False
        processo.nao_contar_mes_ref_ano = None
        processo.nao_contar_mes_ref_mes = None
        return
    ref_now = _as_utc(now) or _utcnow()
    processo.nao_contar_mes = True
    processo.nao_contar_mes_ref_ano = ref_now.year
    processo.nao_contar_mes_ref_mes = ref_now.month


def _set_processo_archived(
    processo: "Processo",
    *,
    archived_at: Optional[datetime] = None,
    reference_date: Optional[date] = None,
) -> None:
    ref_archived_at = _as_utc(archived_at) or _utcnow()
    ref_period = reference_date or ref_archived_at.date()
    processo.arquivado = True
    processo.arquivado_em = ref_archived_at
    processo.arquivado_ref_ano = ref_period.year
    processo.arquivado_ref_mes = ref_period.month


def _clear_processo_archived(processo: "Processo") -> None:
    processo.arquivado = False
    processo.arquivado_em = None
    processo.arquivado_ref_ano = None
    processo.arquivado_ref_mes = None


def _is_nao_contar_mes_active(processo: "Processo", now: Optional[datetime] = None) -> bool:
    if not bool(getattr(processo, "nao_contar_mes", False)):
        return False
    ref_ano = getattr(processo, "nao_contar_mes_ref_ano", None)
    ref_mes = getattr(processo, "nao_contar_mes_ref_mes", None)
    if ref_ano is None or ref_mes is None:
        return True
    try:
        ref_ano_int = int(ref_ano)
        ref_mes_int = int(ref_mes)
    except (TypeError, ValueError):
        return True
    if ref_mes_int < 1 or ref_mes_int > 12:
        return True
    ref_now = _as_utc(now) or _utcnow()
    return ref_ano_int == ref_now.year and ref_mes_int == ref_now.month


def _processos_ativos_clause():
    return or_(Processo.arquivado.is_(False), Processo.arquivado.is_(None))


def _sync_reuniao_conta_no_mes_with_processo(
    db: Session,
    processo: "Processo",
    now: Optional[datetime] = None,
) -> None:
    processo_id = getattr(processo, "id", None)
    if not processo_id:
        return
    item = db.get(AnalistaReuniaoComercial, processo_id)
    if not item:
        return
    conta_no_mes = not _is_nao_contar_mes_active(processo, now)
    if bool(getattr(item, "solicitar_cancelamento", False)):
        conta_no_mes = False
    item.conta_no_mes = conta_no_mes


def _ensure_terminal_status_archiving(db: Session, now: Optional[datetime] = None) -> int:
    ref_now = _as_utc(now) or _utcnow()
    processos_ativos = (
        db.query(Processo)
        .filter(
            _processos_ativos_clause(),
            func.upper(func.coalesce(Processo.status_geral, "")).in_(tuple(PROCESS_GERAL_ARQUIVO_IMEDIATO)),
        )
        .all()
    )
    processo_ids = [processo.id for processo in processos_ativos]
    status_final_evento_por_processo: dict[uuid.UUID, datetime] = {}
    if processo_ids:
        status_rows = (
            db.query(ProcessoEvento.processo_id, func.max(ProcessoEvento.created_at))
            .filter(
                ProcessoEvento.processo_id.in_(processo_ids),
                func.lower(func.coalesce(ProcessoEvento.field_name, "")) == "status_geral",
                func.upper(func.coalesce(ProcessoEvento.new_value, "")).in_(tuple(PROCESS_GERAL_ARQUIVO_IMEDIATO)),
            )
            .group_by(ProcessoEvento.processo_id)
            .all()
        )
        for processo_id, created_at in status_rows:
            status_final_evento_por_processo[processo_id] = created_at

    arquivados = 0
    for processo in processos_ativos:
        referencia_dt = (
            _as_utc(status_final_evento_por_processo.get(processo.id))
            or _as_utc(getattr(processo, "updated_at", None))
            or _as_utc(getattr(processo, "created_at", None))
            or ref_now
        )
        _set_nao_contar_mes_period(processo, True, ref_now)
        _set_processo_archived(
            processo,
            archived_at=ref_now,
            reference_date=referencia_dt.date(),
        )
        _sync_reuniao_conta_no_mes_with_processo(db, processo, ref_now)
        arquivados += 1

    if arquivados > 0:
        _invalidate_process_list_cache()
        db.commit()
    return arquivados


def _ensure_monthly_repasse_archiving(db: Session, now: Optional[datetime] = None) -> int:
    ref_now = _as_utc(now) or _utcnow()
    periodo_atual = f"{ref_now.year:04d}-{ref_now.month:02d}"
    periodo_registrado = str(_get_runtime_meta(db, REPASSE_ARQUIVO_PERIODO_RUNTIME_KEY) or "").strip()
    if periodo_registrado == periodo_atual:
        return 0

    processos_ativos = db.query(Processo).filter(_processos_ativos_clause()).all()
    candidatos = [p for p in processos_ativos if _status_token(getattr(p, "status_cca", None)) in PROCESS_CCA_FINAL_STATUSES]
    candidatos_ids = [p.id for p in candidatos]
    assinatura_evento_por_processo: dict[uuid.UUID, datetime] = {}
    if candidatos_ids:
        assinatura_rows = (
            db.query(ProcessoEvento.processo_id, func.max(ProcessoEvento.created_at))
            .filter(
                ProcessoEvento.processo_id.in_(candidatos_ids),
                func.lower(func.coalesce(ProcessoEvento.field_name, "")) == "status_cca",
                func.upper(func.coalesce(ProcessoEvento.new_value, "")).in_(tuple(PROCESS_CCA_FINAL_STATUSES)),
            )
            .group_by(ProcessoEvento.processo_id)
            .all()
        )
        for processo_id, created_at in assinatura_rows:
            assinatura_evento_por_processo[processo_id] = created_at

    arquivados = 0
    for processo in candidatos:
        assinatura_evento = _as_utc(assinatura_evento_por_processo.get(processo.id))
        credito_fim = _as_utc(getattr(processo, "sla_credito_fim_at", None))
        atualizado_em = _as_utc(getattr(processo, "updated_at", None))
        criado_em = _as_utc(getattr(processo, "created_at", None))
        referencia = assinatura_evento or credito_fim or atualizado_em or criado_em
        if referencia is None:
            continue
        if (referencia.year, referencia.month) < (ref_now.year, ref_now.month):
            _set_processo_archived(
                processo,
                archived_at=ref_now,
                reference_date=referencia.date(),
            )
            arquivados += 1

    _set_runtime_meta(db, REPASSE_ARQUIVO_PERIODO_RUNTIME_KEY, periodo_atual)
    if arquivados > 0:
        _invalidate_process_list_cache()
    db.commit()
    return arquivados


def _ensure_process_archiving(db: Session, now: Optional[datetime] = None) -> int:
    ref_now = _as_utc(now) or _utcnow()
    arquivados = _ensure_terminal_status_archiving(db, ref_now)
    arquivados += _ensure_monthly_repasse_archiving(db, ref_now)
    return arquivados


def _resolve_dashboard_reference_date(
    cliente: "Cliente",
    processo_created_date: Optional[date],
    now_date: date,
) -> Optional[date]:
    floor_by_created: Optional[date] = None
    if processo_created_date is not None:
        floor_by_created = processo_created_date - timedelta(days=DASHBOARD_IMPORT_DATE_BACKFILL_TOLERANCE_DAYS)

    raw_candidates: list[Optional[date]] = [
        getattr(cliente, "data_reserva_origem", None),
        getattr(cliente, "data_cadastro_origem", None),
        processo_created_date,
    ]

    for raw in raw_candidates:
        if raw is None:
            continue
        if raw > now_date:
            continue
        if floor_by_created is not None and raw < floor_by_created:
            continue

        dias = (now_date - raw).days
        if dias > DASHBOARD_MAX_DIAS_EM_ABERTO and processo_created_date is not None and processo_created_date <= now_date:
            return processo_created_date
        return raw

    if processo_created_date is not None and processo_created_date <= now_date:
        return processo_created_date
    return None


def _normalize_meta_period(ano: Optional[int], mes: Optional[int]) -> tuple[int, int]:
    current_ano, current_mes = _current_meta_period()
    ano_val = int(ano if ano is not None else current_ano)
    mes_val = int(mes if mes is not None else current_mes)
    if mes_val < 1 or mes_val > 12:
        raise HTTPException(status_code=422, detail="Mes invalido. Use valores de 1 a 12.")
    if ano_val < 2000 or ano_val > 2100:
        raise HTTPException(status_code=422, detail="Ano invalido. Use valores entre 2000 e 2100.")
    return ano_val, mes_val


def _build_meta_period_runtime_key(base_key: str, ano: int, mes: int) -> str:
    return f"{base_key}:{ano:04d}-{mes:02d}"


def _resolve_gestor_meta_periodo(db: Session, ano: Optional[int], mes: Optional[int]) -> tuple[int, int, str, str]:
    ano_val, mes_val = _normalize_meta_period(ano, mes)
    mensal_period_key = _build_meta_period_runtime_key(META_MENSAL_RUNTIME_KEY, ano_val, mes_val)
    semanal_period_key = _build_meta_period_runtime_key(META_SEMANAL_RUNTIME_KEY, ano_val, mes_val)

    meta_mensal = _runtime_meta_non_negative_int(_get_runtime_meta(db, mensal_period_key))
    if meta_mensal is not None:
        fonte_mensal = "runtime_periodo"
    else:
        meta_mensal = _runtime_meta_non_negative_int(_get_runtime_meta(db, META_MENSAL_RUNTIME_KEY))
        if meta_mensal is not None:
            fonte_mensal = "runtime_global"
        else:
            meta_mensal = GESTOR_META_MENSAL
            fonte_mensal = "env"

    meta_semanal = _runtime_meta_non_negative_int(_get_runtime_meta(db, semanal_period_key))
    if meta_semanal is not None:
        fonte_semanal = "runtime_periodo"
    else:
        meta_semanal = _runtime_meta_non_negative_int(_get_runtime_meta(db, META_SEMANAL_RUNTIME_KEY))
        if meta_semanal is not None:
            fonte_semanal = "runtime_global"
        else:
            meta_semanal = GESTOR_META_SEMANAL
            fonte_semanal = "env"

    return meta_mensal, meta_semanal, fonte_mensal, fonte_semanal


def _resolve_gestor_meta_mensal(db: Session) -> tuple[int, str]:
    ano, mes = _current_meta_period()
    meta_mensal, _, fonte_mensal, _ = _resolve_gestor_meta_periodo(db, ano, mes)
    return meta_mensal, fonte_mensal


def _reconcile_active_sla_timers(db: Session, now: Optional[datetime] = None) -> int:
    now_utc = _as_utc(now) or _utcnow()
    processos = (
        db.query(Processo)
        .filter(Processo.sla_owner != SLA_OWNER_NONE, Processo.sla_active_since.isnot(None))
        .all()
    )
    reconciled = 0
    for processo in processos:
        _accrue_current_sla(processo, now_utc)
        _refresh_sla_snapshots(processo, now_utc)
        reconciled += 1

    if reconciled:
        db.commit()
        _invalidate_process_list_cache()
    return reconciled


def _ensure_runtime_schema(db: Session) -> None:
    processos_table = db.execute(text("SELECT to_regclass('public.processos')")).scalar()
    if not processos_table:
        return

    clientes_table = db.execute(text("SELECT to_regclass('public.clientes')")).scalar()
    documentos_table = db.execute(text("SELECT to_regclass('public.documentos')")).scalar()

    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS app_runtime_meta (
                meta_key TEXT PRIMARY KEY,
                meta_value TEXT NOT NULL,
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS processo_eventos (
                id UUID PRIMARY KEY,
                processo_id UUID NOT NULL REFERENCES processos(id) ON DELETE CASCADE,
                actor_username TEXT NOT NULL,
                actor_role VARCHAR(20) NOT NULL DEFAULT 'system',
                event_type VARCHAR(50) NOT NULL,
                field_name VARCHAR(100),
                old_value TEXT,
                new_value TEXT,
                details TEXT,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS sistema_logs (
                id UUID PRIMARY KEY,
                actor_username TEXT NOT NULL,
                actor_role VARCHAR(20) NOT NULL DEFAULT 'system',
                tela VARCHAR(60) NOT NULL,
                acao VARCHAR(80) NOT NULL,
                entidade_tipo VARCHAR(60),
                entidade_id VARCHAR(120),
                details TEXT,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS lead_precadastros (
                id UUID PRIMARY KEY,
                corretor_username TEXT NOT NULL,
                nome_cliente TEXT NOT NULL,
                telefone VARCHAR(40),
                whatsapp VARCHAR(40),
                email VARCHAR(180),
                cpf VARCHAR(20),
                documento_identificacao VARCHAR(40),
                estado_civil VARCHAR(30),
                certidao_numero VARCHAR(60),
                cidade_nascimento VARCHAR(120),
                data_nascimento DATE,
                endereco TEXT,
                empreendimento_interesse TEXT,
                localidade_interesse TEXT,
                local_agendamento TEXT,
                tipo_visita VARCHAR(20),
                data_agendamento TIMESTAMPTZ,
                estagio_lead VARCHAR(40) NOT NULL DEFAULT 'LEAD',
                decisao_cca VARCHAR(30) NOT NULL DEFAULT 'EM_ANALISE',
                contrato_assinado BOOLEAN NOT NULL DEFAULT FALSE,
                contrato_assinado_em TIMESTAMPTZ,
                assinatura_email_confirmada BOOLEAN NOT NULL DEFAULT FALSE,
                assinatura_email_confirmada_em TIMESTAMPTZ,
                assinatura_email_token VARCHAR(120),
                assinatura_email_token_expires_at TIMESTAMPTZ,
                assinatura_email_enviado_em TIMESTAMPTZ,
                processo_id UUID REFERENCES processos(id) ON DELETE SET NULL,
                reservado_em TIMESTAMPTZ,
                observacoes TEXT,
                ultimo_contato_em TIMESTAMPTZ,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS empreendimento_regras_financeiras (
                id UUID PRIMARY KEY,
                empreendimento_id UUID NOT NULL UNIQUE REFERENCES empreendimentos(id) ON DELETE CASCADE,
                valor_cheque_moradia DOUBLE PRECISION NOT NULL DEFAULT 0,
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS unidades_disponiveis (
                id UUID PRIMARY KEY,
                created_by_username TEXT,
                empreendimento VARCHAR(220) NOT NULL,
                unidade VARCHAR(80) NOT NULL,
                bloco VARCHAR(80),
                tipologia VARCHAR(120),
                quartos INTEGER,
                banheiros INTEGER,
                vagas INTEGER,
                area_m2 DOUBLE PRECISION,
                valor DOUBLE PRECISION,
                localizacao TEXT,
                diferenciais TEXT,
                url_imagem TEXT,
                visita_disponivel BOOLEAN NOT NULL DEFAULT TRUE,
                status_unidade VARCHAR(20) NOT NULL DEFAULT 'DISPONIVEL',
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS credito_planejamento_itens (
                id UUID PRIMARY KEY,
                tipo VARCHAR(20) NOT NULL DEFAULT 'tarefa',
                titulo VARCHAR(180) NOT NULL,
                descricao TEXT,
                responsavel VARCHAR(120),
                data_referencia DATE,
                hora_inicio VARCHAR(5),
                hora_fim VARCHAR(5),
                status VARCHAR(20) NOT NULL DEFAULT 'pendente',
                progresso INTEGER NOT NULL DEFAULT 0,
                urgente BOOLEAN NOT NULL DEFAULT FALSE,
                created_by_username VARCHAR(120),
                updated_by_username VARCHAR(120),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS analista_reuniao_comercial (
                processo_id UUID PRIMARY KEY REFERENCES processos(id) ON DELETE CASCADE,
                conta_no_mes BOOLEAN NOT NULL DEFAULT TRUE,
                data_prevista_entrega DATE,
                probabilidade_queda BOOLEAN NOT NULL DEFAULT FALSE,
                solicitar_cancelamento BOOLEAN NOT NULL DEFAULT FALSE,
                status_followup VARCHAR(20) NOT NULL DEFAULT 'seguir',
                observacao TEXT,
                justificativa_reincidencia TEXT,
                updated_by_username VARCHAR(120),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS analista_reuniao_comercial_compromissos (
                id UUID PRIMARY KEY,
                processo_id UUID NOT NULL REFERENCES processos(id) ON DELETE CASCADE,
                data_prometida DATE NOT NULL,
                status VARCHAR(20) NOT NULL DEFAULT 'pendente',
                justificativa TEXT,
                observacao TEXT,
                created_by_username VARCHAR(120),
                updated_by_username VARCHAR(120),
                created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
            )
            """
        )
    )

    statements = [
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS status_credito VARCHAR(30) DEFAULT 'EM_ANALISE'",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS status_sinal VARCHAR(30) DEFAULT 'NAO_TEM'",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_sinal DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS renda_bruta DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS renda_liquida DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_parcela DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_parcela_7lm DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS renda_complementar_valor DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS renda_complementar_responsavel TEXT",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS renda_complementar_vinculo VARCHAR(40)",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS fgts_futuro_empresa_6m VARCHAR(20)",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS unidade_retomada_agehab VARCHAR(20)",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_imovel DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_avaliacao DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_financiamento DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_subsidio DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS valor_cheque_moradia DOUBLE PRECISION",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS recolha_fgts VARCHAR(30) DEFAULT 'NAO_RECOLHIDO'",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS status_fiador VARCHAR(30) DEFAULT 'NAO_TEM'",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS estagio_comercial VARCHAR(40) DEFAULT 'RESERVA'",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS etapa_repasse VARCHAR(40)",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS cca_responsavel VARCHAR(120)",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS nao_contar_mes BOOLEAN DEFAULT FALSE",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS nao_contar_mes_ref_ano INTEGER",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS nao_contar_mes_ref_mes INTEGER",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS arquivado BOOLEAN DEFAULT FALSE",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS arquivado_em TIMESTAMPTZ",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS arquivado_ref_ano INTEGER",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS arquivado_ref_mes INTEGER",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_comercial_inicio_at TIMESTAMPTZ",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_credito_inicio_at TIMESTAMPTZ",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_comercial_fim_at TIMESTAMPTZ",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_credito_fim_at TIMESTAMPTZ",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_cca_dias INTEGER",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_owner VARCHAR(20) DEFAULT 'NONE'",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_active_since TIMESTAMPTZ",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_analista_seconds INTEGER DEFAULT 0",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_corretor_seconds INTEGER DEFAULT 0",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS sla_cca_seconds INTEGER DEFAULT 0",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS frankstein_last_event_id VARCHAR(36)",
        "ALTER TABLE processos ADD COLUMN IF NOT EXISTS frankstein_last_event_at TIMESTAMPTZ",
    ]
    for stmt in statements:
        db.execute(text(stmt))
    db.execute(text("UPDATE processos SET nao_contar_mes = FALSE WHERE nao_contar_mes IS NULL"))
    db.execute(text("UPDATE processos SET arquivado = FALSE WHERE arquivado IS NULL"))
    db.execute(text("UPDATE processos SET recolha_fgts = 'NAO_RECOLHIDO' WHERE COALESCE(TRIM(recolha_fgts), '') = ''"))
    ano_atual, mes_atual = _current_meta_period()
    db.execute(
        text(
            """
            UPDATE processos
               SET nao_contar_mes_ref_ano = :ano_atual,
                   nao_contar_mes_ref_mes = :mes_atual
             WHERE nao_contar_mes = TRUE
               AND (nao_contar_mes_ref_ano IS NULL OR nao_contar_mes_ref_mes IS NULL)
            """
        ),
        {"ano_atual": ano_atual, "mes_atual": mes_atual},
    )

    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_created_at ON processos (created_at DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_cliente_id ON processos (cliente_id)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_status_geral ON processos (status_geral)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_status_credito ON processos (status_credito)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_status_cca ON processos (status_cca)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_status_agehab ON processos (status_agehab)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_estagio_comercial ON processos (estagio_comercial)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_etapa_repasse ON processos (etapa_repasse)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_arquivado ON processos (arquivado)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processos_frankstein_last_event_id ON processos (frankstein_last_event_id)"))
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_processos_cca_responsavel_norm
            ON processos ((LOWER(TRIM(COALESCE(cca_responsavel, '')))))
            """
        )
    )

    if clientes_table:
        db.execute(text("ALTER TABLE clientes DROP COLUMN IF EXISTS reserva"))
        db.execute(text("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS imobiliaria TEXT"))
        db.execute(text("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS data_reserva_origem DATE"))
        db.execute(text("ALTER TABLE clientes ADD COLUMN IF NOT EXISTS data_cadastro_origem DATE"))
        db.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS ix_clientes_corretor_norm
                ON clientes ((LOWER(TRIM(COALESCE(corretor, '')))))
                """
            )
        )

    if documentos_table:
        db.execute(text("ALTER TABLE documentos ADD COLUMN IF NOT EXISTS pendencia_info TEXT"))
        db.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS ix_documentos_processo_categoria_nome
                ON documentos (processo_id, categoria, nome)
                """
            )
        )
        db.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS ix_documentos_processo_status_doc
                ON documentos (processo_id, status_doc)
                """
            )
        )
        db.execute(
            text(
                """
                CREATE INDEX IF NOT EXISTS ix_documentos_processo_status_credito
                ON documentos (processo_id, status_credito)
                """
            )
        )

    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_processo_eventos_processo_created_at
            ON processo_eventos (processo_id, created_at DESC)
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_processo_eventos_created_at ON processo_eventos (created_at DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_sistema_logs_created_at ON sistema_logs (created_at DESC)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_sistema_logs_actor_username ON sistema_logs (LOWER(TRIM(actor_username)))"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_sistema_logs_tela ON sistema_logs (LOWER(TRIM(tela)))"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS corretor_username TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS nome_cliente TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS telefone VARCHAR(40)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS whatsapp VARCHAR(40)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS email VARCHAR(180)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS cpf VARCHAR(20)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS documento_identificacao VARCHAR(40)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS estado_civil VARCHAR(30)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS certidao_numero VARCHAR(60)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS cidade_nascimento VARCHAR(120)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS telefone_last4 VARCHAR(4)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS whatsapp_last4 VARCHAR(4)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS email_hash VARCHAR(64)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS cpf_hash VARCHAR(64)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS cpf_last4 VARCHAR(4)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS documento_hash VARCHAR(64)"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN telefone TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN whatsapp TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN email TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN cpf TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN documento_identificacao TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN certidao_numero TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN cidade_nascimento TYPE TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS data_nascimento DATE"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS endereco TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS empreendimento_interesse TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS localidade_interesse TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS local_agendamento TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS tipo_visita VARCHAR(20)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS data_agendamento TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS estagio_lead VARCHAR(40) DEFAULT 'LEAD'"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN estagio_lead SET DEFAULT 'LEAD'"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS decisao_cca VARCHAR(30) DEFAULT 'EM_ANALISE'"))
    db.execute(text("ALTER TABLE lead_precadastros ALTER COLUMN decisao_cca SET DEFAULT 'EM_ANALISE'"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS contrato_assinado BOOLEAN DEFAULT FALSE"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS contrato_assinado_em TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS assinatura_email_confirmada BOOLEAN DEFAULT FALSE"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS assinatura_email_confirmada_em TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS assinatura_email_token VARCHAR(120)"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS assinatura_email_token_expires_at TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS assinatura_email_enviado_em TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS processo_id UUID REFERENCES processos(id) ON DELETE SET NULL"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS reservado_em TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS observacoes TEXT"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS ultimo_contato_em TIMESTAMPTZ"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("ALTER TABLE lead_precadastros ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("UPDATE lead_precadastros SET estagio_lead = 'LEAD' WHERE COALESCE(TRIM(estagio_lead), '') = ''"))
    db.execute(text("UPDATE lead_precadastros SET decisao_cca = 'EM_ANALISE' WHERE COALESCE(TRIM(decisao_cca), '') = ''"))
    db.execute(text("UPDATE lead_precadastros SET contrato_assinado = FALSE WHERE contrato_assinado IS NULL"))
    db.execute(text("UPDATE lead_precadastros SET assinatura_email_confirmada = FALSE WHERE assinatura_email_confirmada IS NULL"))
    db.execute(
        text(
            """
            UPDATE lead_precadastros
               SET estagio_lead = CASE
                 WHEN UPPER(COALESCE(estagio_lead, '')) IN ('NOVO', 'CONTATO_INICIAL', 'QUALIFICACAO') THEN 'LEAD'
                 WHEN UPPER(COALESCE(estagio_lead, '')) IN ('AGENDADO') THEN 'AGENDAMENTO'
                 WHEN UPPER(COALESCE(estagio_lead, '')) IN ('EM_ATENDIMENTO') THEN 'VISITA'
                 WHEN UPPER(COALESCE(estagio_lead, '')) IN ('PROPOSTA', 'NEGOCIACAO') THEN 'PRECADASTRO'
                 WHEN UPPER(COALESCE(estagio_lead, '')) IN ('GANHO') THEN 'RESERVA'
                 WHEN UPPER(COALESCE(estagio_lead, '')) IN ('PERDIDO') THEN 'PERDIDO'
                 ELSE estagio_lead
               END
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE lead_precadastros
               SET decisao_cca = CASE
                 WHEN UPPER(COALESCE(decisao_cca, '')) IN ('EM_ANALISE', 'APROVADO', 'CONDICIONADO', 'REPROVADO', 'BLOQUEADO', 'DAR_QV') THEN UPPER(decisao_cca)
                 WHEN UPPER(COALESCE(decisao_cca, '')) IN ('DAR QV', 'DAR-QV', 'DARQV') THEN 'DAR_QV'
                 WHEN UPPER(COALESCE(decisao_cca, '')) IN ('PENDENTE', 'ANALISE') THEN 'EM_ANALISE'
                 ELSE 'EM_ANALISE'
               END
            """
        )
    )
    db.execute(text("UPDATE lead_precadastros SET assinatura_email_confirmada = FALSE WHERE assinatura_email_confirmada IS TRUE AND contrato_assinado IS FALSE"))
    db.execute(text("UPDATE lead_precadastros SET contrato_assinado_em = NULL WHERE contrato_assinado IS FALSE"))
    db.execute(text("UPDATE lead_precadastros SET assinatura_email_confirmada_em = NULL WHERE assinatura_email_confirmada IS FALSE"))
    db.execute(text("UPDATE lead_precadastros SET assinatura_email_token = NULL, assinatura_email_token_expires_at = NULL WHERE assinatura_email_confirmada IS TRUE"))
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_lead_precadastros_corretor_norm
            ON lead_precadastros ((LOWER(TRIM(COALESCE(corretor_username, '')))))
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_cpf ON lead_precadastros (cpf)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_email_hash ON lead_precadastros (email_hash)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_cpf_hash ON lead_precadastros (cpf_hash)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_cpf_last4 ON lead_precadastros (cpf_last4)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_telefone_last4 ON lead_precadastros (telefone_last4)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_whatsapp_last4 ON lead_precadastros (whatsapp_last4)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_documento_hash ON lead_precadastros (documento_hash)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_processo_id ON lead_precadastros (processo_id)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_estagio ON lead_precadastros (estagio_lead)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_decisao_cca ON lead_precadastros (decisao_cca)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_assinatura_email_token ON lead_precadastros (assinatura_email_token)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_lead_precadastros_updated_at ON lead_precadastros (updated_at DESC)"))
    db.execute(text("ALTER TABLE empreendimento_regras_financeiras ADD COLUMN IF NOT EXISTS empreendimento_id UUID REFERENCES empreendimentos(id) ON DELETE CASCADE"))
    db.execute(text("ALTER TABLE empreendimento_regras_financeiras ADD COLUMN IF NOT EXISTS valor_cheque_moradia DOUBLE PRECISION DEFAULT 0"))
    db.execute(text("ALTER TABLE empreendimento_regras_financeiras ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("ALTER TABLE empreendimento_regras_financeiras ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("UPDATE empreendimento_regras_financeiras SET valor_cheque_moradia = 0 WHERE valor_cheque_moradia IS NULL"))
    db.execute(
        text(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_empreendimento_regras_financeiras_empreendimento_id
            ON empreendimento_regras_financeiras (empreendimento_id)
            """
        )
    )
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS created_by_username TEXT"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS empreendimento VARCHAR(220)"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS unidade VARCHAR(80)"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS bloco VARCHAR(80)"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS tipologia VARCHAR(120)"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS quartos INTEGER"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS banheiros INTEGER"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS vagas INTEGER"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS area_m2 DOUBLE PRECISION"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS valor DOUBLE PRECISION"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS localizacao TEXT"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS diferenciais TEXT"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS url_imagem TEXT"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS visita_disponivel BOOLEAN DEFAULT TRUE"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS status_unidade VARCHAR(20) DEFAULT 'DISPONIVEL'"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("ALTER TABLE unidades_disponiveis ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("UPDATE unidades_disponiveis SET status_unidade = 'DISPONIVEL' WHERE COALESCE(TRIM(status_unidade), '') = ''"))
    db.execute(text("UPDATE unidades_disponiveis SET visita_disponivel = TRUE WHERE visita_disponivel IS NULL"))
    db.execute(
        text(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_unidades_disponiveis_empreendimento_unidade
            ON unidades_disponiveis ((LOWER(TRIM(COALESCE(empreendimento, '')))), (LOWER(TRIM(COALESCE(unidade, '')))))
            """
        )
    )
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_unidades_disponiveis_status ON unidades_disponiveis (status_unidade)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_unidades_disponiveis_empreendimento ON unidades_disponiveis (empreendimento)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_unidades_disponiveis_tipologia ON unidades_disponiveis (tipologia)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_unidades_disponiveis_updated_at ON unidades_disponiveis (updated_at DESC)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS tipo VARCHAR(20) DEFAULT 'tarefa'"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS titulo VARCHAR(180)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS descricao TEXT"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS responsavel VARCHAR(120)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS data_referencia DATE"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS hora_inicio VARCHAR(5)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS hora_fim VARCHAR(5)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'pendente'"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS progresso INTEGER DEFAULT 0"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS urgente BOOLEAN DEFAULT FALSE"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS created_by_username VARCHAR(120)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS updated_by_username VARCHAR(120)"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("ALTER TABLE credito_planejamento_itens ADD COLUMN IF NOT EXISTS updated_at TIMESTAMPTZ DEFAULT NOW()"))
    db.execute(text("UPDATE credito_planejamento_itens SET tipo = 'tarefa' WHERE COALESCE(TRIM(tipo), '') = ''"))
    db.execute(text("UPDATE credito_planejamento_itens SET status = 'pendente' WHERE COALESCE(TRIM(status), '') = ''"))
    db.execute(text("UPDATE credito_planejamento_itens SET progresso = 0 WHERE progresso IS NULL"))
    db.execute(text("UPDATE credito_planejamento_itens SET urgente = FALSE WHERE urgente IS NULL"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_credito_planejamento_itens_tipo ON credito_planejamento_itens (tipo)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_credito_planejamento_itens_status ON credito_planejamento_itens (status)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_credito_planejamento_itens_data ON credito_planejamento_itens (data_referencia)"))
    db.execute(text("CREATE INDEX IF NOT EXISTS ix_credito_planejamento_itens_urgente ON credito_planejamento_itens (urgente)"))
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_credito_planejamento_itens_responsavel_norm
            ON credito_planejamento_itens ((LOWER(TRIM(COALESCE(responsavel, '')))))
            """
        )
    )
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_credito_planejamento_itens_updated_at
            ON credito_planejamento_itens (updated_at DESC)
            """
        )
    )
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_analista_reuniao_comercial_status_followup
            ON analista_reuniao_comercial (status_followup)
            """
        )
    )
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_analista_reuniao_comercial_data_prevista
            ON analista_reuniao_comercial (data_prevista_entrega)
            """
        )
    )
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_analista_reuniao_compromissos_processo_data
            ON analista_reuniao_comercial_compromissos (processo_id, data_prometida DESC)
            """
        )
    )
    db.execute(
        text(
            """
            CREATE INDEX IF NOT EXISTS ix_analista_reuniao_compromissos_status
            ON analista_reuniao_comercial_compromissos (status)
            """
        )
    )

    current_revision = db.execute(
        text("SELECT meta_value FROM app_runtime_meta WHERE meta_key = 'runtime_schema_revision'")
    ).scalar()
    if current_revision == RUNTIME_SCHEMA_REVISION:
        db.commit()
        return

    if documentos_table:
        db.execute(
            text(
                """
                UPDATE documentos
                SET status_credito = CASE
                    WHEN UPPER(COALESCE(status_credito, '')) IN ('PENDENCIADO', 'APROVADO', 'REPROVADO', 'NAO_APLICA') THEN UPPER(status_credito)
                    WHEN UPPER(COALESCE(status_credito, '')) IN ('AGUARDANDO_ENVIO', 'AGUARDANDO ENVIO') THEN 'AGUARDANDO_ENVIO'
                    WHEN UPPER(COALESCE(status_doc, '')) = 'NAO_APLICA' THEN 'NAO_APLICA'
                    WHEN UPPER(COALESCE(status_doc, '')) = 'PENDENTE' THEN 'AGUARDANDO_ENVIO'
                    ELSE 'ANALISE'
                END
                """
            )
        )

    db.execute(
        text(
            """
            UPDATE processos
            SET status_credito = CASE
                WHEN UPPER(COALESCE(status_credito, '')) IN ('EM_ANALISE', 'PENDENCIADO', 'APROVADO', 'REPROVADO') THEN UPPER(status_credito)
                WHEN UPPER(COALESCE(status_credito, '')) IN ('ANALISE', 'EM ANALISE', 'EMANALISE') THEN 'EM_ANALISE'
                WHEN UPPER(COALESCE(status_credito, '')) IN ('PENDENTE') THEN 'PENDENCIADO'
                ELSE 'EM_ANALISE'
            END
            """
        )
    )

    db.execute(
        text(
            """
            UPDATE processos
            SET status_geral = CASE
                WHEN UPPER(COALESCE(status_geral, '')) IN ('NOVO', 'EM_ANDAMENTO', 'PENDENCIADO', 'APROVADO', 'REPROVADO', 'DISTRATO', 'CANCELADO') THEN UPPER(status_geral)
                WHEN UPPER(COALESCE(status_geral, '')) IN ('EM_ANALISE', 'EM ANALISE', 'EMANALISE') THEN 'EM_ANDAMENTO'
                WHEN UPPER(COALESCE(status_geral, '')) IN ('PENDENTE') THEN 'PENDENCIADO'
                ELSE 'NOVO'
            END
            """
        )
    )

    db.execute(
        text(
            """
            UPDATE processos
            SET status_cca = CASE
                WHEN UPPER(COALESCE(status_cca, '')) IN (
                    'ANALISE_CREDITO', 'PENDENTE_CREDITO', 'ANALISE_CCA', 'PENDENTE_CCA',
                    'APROVADO', 'REPROVADO', 'CONDICIONADO', 'BLOQUEADO', 'DAR_QV',
                    'ASSINATURA_CAIXA', 'FINALIZADO'
                ) THEN UPPER(status_cca)
                WHEN UPPER(COALESCE(status_cca, '')) IN ('DAR QV', 'DAR-QV', 'DARQV') THEN 'DAR_QV'
                WHEN UPPER(COALESCE(status_cca, '')) IN ('CONFORME') THEN 'APROVADO'
                WHEN UPPER(COALESCE(status_cca, '')) IN ('TRATANDO_PRODUTO', 'AGUARDANDO_CONFORMIDADE') THEN 'CONDICIONADO'
                WHEN UPPER(COALESCE(status_cca, '')) IN ('AGENDADO') THEN 'DAR_QV'
                WHEN UPPER(COALESCE(status_cca, '')) IN ('EM_ANALISE', 'EM ANALISE', 'EMANALISE', 'ANALISE') THEN 'ANALISE_CREDITO'
                WHEN UPPER(COALESCE(status_cca, '')) IN ('PENDENTE', 'PENDENCIADO') THEN 'PENDENTE_CREDITO'
                ELSE 'ANALISE_CREDITO'
            END
            """
        )
    )

    db.execute(
        text(
            """
            UPDATE processos
            SET recolha_fgts = CASE
                WHEN UPPER(COALESCE(recolha_fgts, '')) IN ('OK', 'NAO_RECOLHIDO', 'VALIDADO_PELO_BANCO', 'RECOLHENDO') THEN UPPER(recolha_fgts)
                WHEN UPPER(COALESCE(recolha_fgts, '')) IN ('NAO RECOLHIDO') THEN 'NAO_RECOLHIDO'
                WHEN UPPER(COALESCE(recolha_fgts, '')) IN ('VALIDADO BANCO', 'VALIDADO PELO BANCO') THEN 'VALIDADO_PELO_BANCO'
                ELSE 'NAO_RECOLHIDO'
            END
            """
        )
    )

    db.execute(
        text(
            """
            UPDATE processos
            SET status_agehab = CASE
                WHEN UPPER(COALESCE(status_agehab, '')) IN ('ANALISE_CREDITO', 'PENDENTE_CREDITO', 'ENVIO_AGEHAB', 'PENDENTE_AGEHAB', 'VALIDADO_AGEHAB') THEN UPPER(status_agehab)
                WHEN UPPER(COALESCE(status_agehab, '')) IN ('EM_ANALISE', 'EM ANALISE', 'EMANALISE', 'ANALISE') THEN 'ANALISE_CREDITO'
                WHEN UPPER(COALESCE(status_agehab, '')) IN ('PENDENTE', 'PENDENCIADO') THEN 'PENDENTE_AGEHAB'
                WHEN UPPER(COALESCE(status_agehab, '')) IN ('APROVADO') THEN 'VALIDADO_AGEHAB'
                WHEN UPPER(COALESCE(status_agehab, '')) IN ('REPROVADO') THEN 'PENDENTE_AGEHAB'
                WHEN UPPER(COALESCE(status_agehab, '')) IN ('ENVIO_AGEHAG') THEN 'ENVIO_AGEHAB'
                ELSE 'ANALISE_CREDITO'
            END
            """
        )
    )

    db.execute(
        text(
            """
            UPDATE processos
            SET status_sinal = CASE
                WHEN UPPER(COALESCE(status_sinal, '')) IN ('NAO_TEM', 'PENDENTE', 'PAGO') THEN UPPER(status_sinal)
                WHEN pendente_sinal IS TRUE THEN 'PENDENTE'
                ELSE 'NAO_TEM'
            END
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET status_fiador = CASE
                WHEN UPPER(COALESCE(status_fiador, '')) IN ('NAO_TEM', 'PENDENTE', 'FINALIZADO') THEN UPPER(status_fiador)
                WHEN pendente_fiador IS TRUE THEN 'PENDENTE'
                ELSE 'NAO_TEM'
            END
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET estagio_comercial = CASE
                WHEN UPPER(COALESCE(estagio_comercial, '')) IN (
                    'RESERVA', 'EM_PROCESSO', 'CREDITO', 'SECRETARIA_VENDAS',
                    'ASSINATURA_DIRETORIA', 'AUTORIZACAO_DIRETORIA', 'ENVIO_SIENGE', 'VENDA_FINALIZADA'
                ) THEN UPPER(estagio_comercial)
                WHEN UPPER(COALESCE(status_geral, '')) IN ('NOVO') THEN 'RESERVA'
                WHEN UPPER(COALESCE(status_geral, '')) IN ('APROVADO') THEN 'VENDA_FINALIZADA'
                WHEN UPPER(COALESCE(status_cca, '')) IN ('ASSINATURA_CAIXA', 'FINALIZADO') THEN 'VENDA_FINALIZADA'
                ELSE 'EM_PROCESSO'
            END
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET etapa_repasse = CASE
                WHEN UPPER(COALESCE(etapa_repasse, '')) IN ('EM_REPASSE', 'INICIO_REPASSE', 'ASSINATURA_AUTORIZADA')
                    THEN UPPER(etapa_repasse)
                WHEN UPPER(COALESCE(estagio_comercial, '')) IN ('ASSINATURA_DIRETORIA', 'AUTORIZACAO_DIRETORIA', 'ENVIO_SIENGE', 'VENDA_FINALIZADA')
                    THEN 'EM_REPASSE'
                ELSE NULL
            END
            """
        )
    )
    if clientes_table:
        db.execute(
            text(
                """
                UPDATE processos p
                SET sla_comercial_inicio_at = COALESCE(
                    p.sla_comercial_inicio_at,
                    (c.data_cadastro_origem::timestamp AT TIME ZONE 'UTC'),
                    c.created_at,
                    p.created_at
                )
                FROM clientes c
                WHERE c.id = p.cliente_id
                """
            )
        )
    else:
        db.execute(
            text(
                """
                UPDATE processos
                SET sla_comercial_inicio_at = COALESCE(sla_comercial_inicio_at, created_at)
                """
            )
        )
    db.execute(
        text(
            """
            UPDATE processos p
            SET sla_credito_inicio_at = ev.first_analise_at
            FROM (
                SELECT processo_id, MIN(created_at) AS first_analise_at
                FROM processo_eventos
                WHERE UPPER(COALESCE(actor_role, '')) = 'ANALISTA'
                  AND UPPER(COALESCE(event_type, '')) NOT IN ('PROCESSO_CRIADO', 'PROCESSO_IMPORTADO_PLANILHA', 'SLA_OWNER_CHANGE')
                GROUP BY processo_id
            ) ev
            WHERE p.id = ev.processo_id
              AND p.sla_credito_inicio_at IS NULL
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos p
            SET sla_comercial_fim_at = ev.first_venda_finalizada_at
            FROM (
                SELECT processo_id, MIN(created_at) AS first_venda_finalizada_at
                FROM processo_eventos
                WHERE UPPER(COALESCE(field_name, '')) = 'ESTAGIO_COMERCIAL'
                  AND UPPER(COALESCE(new_value, '')) = 'VENDA_FINALIZADA'
                GROUP BY processo_id
            ) ev
            WHERE p.id = ev.processo_id
              AND p.sla_comercial_fim_at IS NULL
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_comercial_fim_at = COALESCE(sla_comercial_fim_at, updated_at)
            WHERE UPPER(COALESCE(estagio_comercial, '')) = 'VENDA_FINALIZADA'
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_comercial_fim_at = NULL
            WHERE UPPER(COALESCE(estagio_comercial, '')) <> 'VENDA_FINALIZADA'
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos p
            SET sla_credito_fim_at = ev.first_assinatura_at
            FROM (
                SELECT processo_id, MIN(created_at) AS first_assinatura_at
                FROM processo_eventos
                WHERE UPPER(COALESCE(field_name, '')) = 'STATUS_CCA'
                  AND UPPER(COALESCE(new_value, '')) IN ('ASSINATURA_CAIXA', 'FINALIZADO')
                GROUP BY processo_id
            ) ev
            WHERE p.id = ev.processo_id
              AND p.sla_credito_fim_at IS NULL
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_credito_fim_at = COALESCE(sla_credito_fim_at, updated_at)
            WHERE UPPER(COALESCE(status_cca, '')) IN ('ASSINATURA_CAIXA', 'FINALIZADO')
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_credito_fim_at = NULL
            WHERE UPPER(COALESCE(status_cca, '')) NOT IN ('ASSINATURA_CAIXA', 'FINALIZADO')
            """
        )
    )

    db.execute(
        text(
            """
            UPDATE processos
            SET sla_owner = CASE
                WHEN UPPER(COALESCE(sla_owner, '')) IN ('NONE', 'CORRETOR', 'ANALISTA', 'CCA') THEN UPPER(sla_owner)
                ELSE 'NONE'
            END
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_corretor_seconds = GREATEST(COALESCE(sla_corretor_seconds, 0), COALESCE(sla_corretor_dias, 0) * 86400)
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_analista_seconds = GREATEST(COALESCE(sla_analista_seconds, 0), COALESCE(sla_credito_dias, 0) * 86400)
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_cca_seconds = COALESCE(sla_cca_seconds, 0)
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_cca_dias = COALESCE(sla_cca_dias, FLOOR(COALESCE(sla_cca_seconds, 0) / 86400.0)::INTEGER)
            """
        )
    )
    db.execute(
        text(
            """
            UPDATE processos
            SET sla_owner = 'ANALISTA',
                sla_active_since = COALESCE(sla_active_since, NOW())
            WHERE UPPER(COALESCE(sla_owner, 'NONE')) = 'NONE'
              AND UPPER(COALESCE(status_geral, '')) NOT IN ('APROVADO', 'REPROVADO', 'DISTRATO', 'CANCELADO')
              AND UPPER(COALESCE(status_cca, '')) NOT IN ('ASSINATURA_CAIXA', 'FINALIZADO')
            """
        )
    )
    db.execute(
        text(
            """
            INSERT INTO app_runtime_meta (meta_key, meta_value, updated_at)
            VALUES ('runtime_schema_revision', :revision, NOW())
            ON CONFLICT (meta_key) DO UPDATE
            SET meta_value = EXCLUDED.meta_value,
                updated_at = NOW()
            """
        ),
        {"revision": RUNTIME_SCHEMA_REVISION},
    )

    db.commit()


def get_db():
    if DB_URL_HAS_PLACEHOLDERS:
        raise HTTPException(
            status_code=503,
            detail="DATABASE_URL ainda contem placeholders. Configure a URL real do Supabase no Render.",
        )

    if SessionLocal is None:
        if DATABASE_URL:
            detail = "DATABASE_URL invalido. Revise usuario/senha/host e URL do pooler Supabase."
        else:
            detail = "DATABASE_URL nao configurado. Defina a conexao do Supabase no Render."
        raise HTTPException(
            status_code=503,
            detail=detail,
        )
    db = SessionLocal()
    try:
        yield db
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


@asynccontextmanager
async def lifespan(_: FastAPI):
    if DB_URL_HAS_PLACEHOLDERS:
        logger.warning("DATABASE_URL parece conter placeholders (<PASSWORD>, YOUR-PASSWORD, etc.).")

    _warn_default_credentials()

    if not pii_encryption_enabled():
        logger.warning(
            "PII_ENCRYPTION_KEY nao configurada. Tokens sensiveis seguem protegidos por hash, "
            "mas a criptografia de PII em repouso nao sera aplicada ate definir a chave."
        )

    if not SESSION_COOKIE_SECURE:
        logger.warning("SESSION_COOKIE_SECURE=false. Use true em producao com HTTPS.")

    if DATABASE_URL and _is_supabase_direct_host(DATABASE_URL):
        logger.warning(
            "DATABASE_URL aponta para conexao direta Supabase (db.<project-ref>.supabase.co), "
            "que normalmente exige IPv6. Em Render, use a URL do Supavisor transaction mode "
            "(aws-0-<region>.pooler.supabase.com:6543)."
        )

    if engine is not None and os.getenv("AUTO_CREATE_TABLES", "false").lower() in {"1", "true", "yes"}:
        try:
            Base.metadata.create_all(bind=engine)
        except SQLAlchemyError:
            logger.exception("Falha ao executar create_all no startup.")
            if os.getenv("STARTUP_DB_STRICT", "false").lower() in {"1", "true", "yes"}:
                raise

    if engine is not None:
        try:
            Base.metadata.create_all(
                bind=engine,
                tables=[
                    AppUser.__table__,
                    AppSession.__table__,
                    Empreendimento.__table__,
                    TabelaPreco.__table__,
                    IaFeedback.__table__,
                    AnaliseRegistroDB.__table__,
                ],
            )
            if SessionLocal is not None:
                db = SessionLocal()
                try:
                    _ensure_seed_users(db)
                    _ensure_runtime_schema(db)
                    backfilled_leads = _backfill_lead_security_fields(db)
                    _set_runtime_meta(db, "lead_security_backfill_count", str(backfilled_leads))
                    _ensure_process_archiving(db, _utcnow())
                    now_utc = _utcnow()
                    _set_runtime_meta(db, "sla_runtime_started_at", now_utc.isoformat())
                    reconciled = _reconcile_active_sla_timers(db, now_utc)
                    _set_runtime_meta(db, "sla_runtime_last_reconcile_count", str(reconciled))
                    db.commit()
                finally:
                    db.close()
        except SQLAlchemyError:
            logger.exception("Falha ao preparar tabela de usuarios da aplicacao.")
            if os.getenv("STARTUP_DB_STRICT", "false").lower() in {"1", "true", "yes"}:
                raise
    _ensure_frankstein_tables()
    try:
        yield
    finally:
        if SessionLocal is not None:
            db = SessionLocal()
            try:
                _set_runtime_meta(db, "sla_runtime_stopped_at", _utcnow().isoformat())
                db.commit()
            except SQLAlchemyError:
                db.rollback()
                logger.exception("Falha ao registrar checkpoint de parada do runtime.")
            finally:
                db.close()


app = FastAPI(title="Sistema Credito API", lifespan=lifespan)
app.mount("/assets", StaticFiles(directory=str(WEB_DIR)), name="web_assets")


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    errors = []
    for err in exc.errors():
        loc = ".".join(str(x) for x in err.get("loc", []))
        msg = err.get("msg", "")
        errors.append(f"{loc}: {msg}")
    detail = "; ".join(errors) or "Dados invÃ¡lidos."
    return JSONResponse(status_code=422, content={"detail": detail})

def _react_app_unavailable_response() -> HTMLResponse:
    return HTMLResponse(
        """
        <!doctype html>
        <html lang="pt-BR">
          <head><meta charset="utf-8" /><title>React indisponivel</title></head>
          <body style="font-family:Segoe UI, sans-serif; padding:24px;">
            <h2>Frontend React ainda nao foi buildado no servidor.</h2>
            <p>Execute <code>npm install && npm run build</code> em <code>frontend-react</code>.</p>
          </body>
        </html>
        """
    )


def _serve_react_app(path: str = ""):
    if not REACT_DIST_DIR.exists():
        return _react_app_unavailable_response()

    normalized_path = path.strip("/")
    index_file = REACT_DIST_DIR / "index.html"
    index_headers = {"Cache-Control": "no-store, max-age=0, must-revalidate"}
    if not path:
        return FileResponse(index_file, headers=index_headers)

    dist_root = REACT_DIST_DIR.resolve()
    candidate = (REACT_DIST_DIR / path).resolve()
    try:
        candidate.relative_to(dist_root)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="Not Found") from exc

    if candidate.is_file():
        return FileResponse(candidate)

    if Path(path).suffix:
        raise HTTPException(status_code=404, detail="Not Found")

    return FileResponse(index_file, headers=index_headers)


def _corretor_presentation_url(request: Optional[Request] = None) -> str:
    base_url = "/app-react/apresentacao"
    if request is None:
        return base_url
    query = (request.url.query or "").strip()
    return f"{base_url}?{query}" if query else base_url


def _ensure_corretor_presentation_access(request: Request):
    session = _read_session(request)
    if not session:
        return None, RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return None, RedirectResponse(url="/app/trocar-senha", status_code=302)

    role = _normalize_role(str(session.get("role", "")))
    if role != ROLE_CORRETOR:
        return None, RedirectResponse(url=_home_for_role(role), status_code=302)

    if not CORRETOR_ROUTE_ENABLED:
        token = request.cookies.get(SESSION_COOKIE_NAME)
        if token:
            _delete_session_record(token)
        response = RedirectResponse(url="/login", status_code=302)
        response.delete_cookie(key=SESSION_COOKIE_NAME)
        return None, response

    return session, None


@app.get("/app-react")
def react_app_root():
    return _serve_react_app("")


@app.get("/app-react/apresentacao")
def react_corretor_presentation_entry(request: Request):
    _, denied = _ensure_corretor_presentation_access(request)
    if denied:
        return denied
    return _serve_react_app("apresentacao")


@app.get("/app-react/{path:path}")
def react_app_entry(path: str):
    return _serve_react_app(path)


@app.middleware("http")
async def enforce_same_origin_for_write_requests(request: Request, call_next):
    method = (request.method or "").upper()
    path = (request.url.path or "").rstrip("/") or "/"
    if method in {"POST", "PUT", "PATCH", "DELETE"} and (path.startswith("/app/api/") or path.startswith("/auth/")):
        origin_host = _extract_origin_host(request.headers.get("origin"))
        if not origin_host:
            origin_host = _extract_origin_host(request.headers.get("referer"))
        if origin_host:
            req_host = _request_host(request)
            if req_host and origin_host != req_host:
                return JSONResponse(status_code=403, content={"detail": "Origem invalida para operacao autenticada."})
    return await call_next(request)


@app.get("/")
def root(request: Request):
    accept = (request.headers.get("accept") or "").lower()
    if "text/html" in accept:
        return RedirectResponse(url="/app", status_code=302)
    return {"service": "sistema-credito-api", "status": "ok"}


def _html_page(filename: str) -> FileResponse:
    response = FileResponse(WEB_DIR / filename)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.get("/login")
def login_page(request: Request):
    session = _read_session(request)
    if session:
        return RedirectResponse(url=_home_for_session(session), status_code=302)
    return _html_page("login.html")


@app.get("/app")
def app_root(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    return RedirectResponse(url=_home_for_session(session), status_code=302)


@app.get("/app/cca")
def app_cca_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_CCA, ROLE_ANALISTA}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("cca.html")


@app.get("/app/cca/analise")
def app_cca_analise_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("cca_analise.html")


@app.get("/app/checklist")
def app_checklist_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in CHECKLIST_PAGE_ROLES:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("checklist.html")


@app.get("/app/corretor")
def app_corretor_page(request: Request):
    _, denied = _ensure_corretor_presentation_access(request)
    if denied:
        return denied
    return RedirectResponse(url=_corretor_presentation_url(request), status_code=302)


@app.get("/app/corretor/precadastro")
def app_corretor_precadastro_page(request: Request):
    _, denied = _ensure_corretor_presentation_access(request)
    if denied:
        return denied
    return RedirectResponse(url=_corretor_presentation_url(request), status_code=302)


@app.get("/app/corretor/apresentacao")
def app_corretor_apresentacao_page(request: Request):
    _, denied = _ensure_corretor_presentation_access(request)
    if denied:
        return denied
    return RedirectResponse(url=_corretor_presentation_url(request), status_code=302)


def _ensure_analista_page_access(request: Request):
    session = _read_session(request)
    if not session:
        return None, RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return None, RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return None, RedirectResponse(url=_home_for_role(role), status_code=302)
    return session, None


@app.get("/app/analista/legado")
def app_analista_legacy_page(request: Request):
    _, denied = _ensure_analista_page_access(request)
    if denied:
        return denied

    processo_id = (request.query_params.get("processo_id") or "").strip()
    if processo_id:
        target = f"/app/analise?processo_id={processo_id}"
        return RedirectResponse(url=target, status_code=302)

    return _html_page("analista_painel.html")


@app.get("/app/analista")
def app_analista_page(request: Request):
    _, denied = _ensure_analista_page_access(request)
    if denied:
        return denied

    processo_id = (request.query_params.get("processo_id") or "").strip()
    if processo_id:
        target = f"/app/analise?processo_id={processo_id}"
        return RedirectResponse(url=target, status_code=302)

    if REACT_DIST_DIR.exists():
        return RedirectResponse(url="/app-react/analista", status_code=302)

    return _html_page("analista_painel.html")


@app.get("/app/analista/acompanhamento")
def app_analista_acompanhamento_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_acompanhamento.html")


@app.get("/app/analista/acompanhamento-operacional")
def app_analista_acompanhamento_operacional_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_acompanhamento_operacional.html")


@app.get("/app/analista/reuniao-comercial")
def app_analista_reuniao_comercial_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_reuniao_comercial.html")


@app.get("/app/analista/repasse")
def app_analista_repasse_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_repasse.html")


@app.get("/app/analista/arquivados")
def app_analista_arquivados_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_arquivados.html")


@app.get("/app/analise")
def app_analise_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista.html")


@app.get("/app/analista/importacao")
def app_analista_importacao_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_importacao.html")


@app.get("/app/analista/crm")
def app_analista_crm_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("analista_crm.html")


@app.get("/app/gestor-credito")
def app_gestor_credito_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    if REACT_DIST_DIR.exists():
        return RedirectResponse(url="/app-react/gestor", status_code=302)
    return _html_page("gestor_credito.html")


@app.get("/app/admin")
def app_admin_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role != ROLE_ADMIN:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    return _html_page("admin.html")


@app.get("/app/gestor")
def app_gestor_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    if bool(session.get("must_change_password")):
        return RedirectResponse(url="/app/trocar-senha", status_code=302)
    role = _normalize_role(str(session.get("role", "")))
    if role not in {ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN}:
        return RedirectResponse(url=_home_for_role(role), status_code=302)
    if REACT_DIST_DIR.exists():
        return RedirectResponse(url="/app-react/gestor", status_code=302)
    return _html_page("gestor.html")


@app.get("/admin")
def admin_shortcut(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    return RedirectResponse(url=_home_for_session(session), status_code=302)


@app.get("/manutencao")
def maintenance_shortcut(request: Request):
    return admin_shortcut(request)


@app.get("/app/trocar-senha")
def app_change_password_page(request: Request):
    session = _read_session(request)
    if not session:
        return RedirectResponse(url="/login", status_code=302)
    return _html_page("change_password.html")


@app.post("/auth/login")
def auth_login(payload: LoginPayload, db: Session = Depends(get_db)):
    _ensure_seed_users(db)

    username = _normalize_username(payload.username)
    if not username:
        raise HTTPException(status_code=401, detail="Credenciais invalidas")

    user = _get_user_by_username(db, username)
    if not user:
        # Auto-recupera seed full caso o runtime tenha ficado em admin_only apos manutencao.
        seed_mode_raw = (_get_runtime_meta(db, USERS_SEED_MODE_RUNTIME_KEY) or USERS_SEED_MODE_FULL).strip().lower()
        if seed_mode_raw == USERS_SEED_MODE_ADMIN_ONLY:
            try:
                _set_runtime_meta(db, USERS_SEED_MODE_RUNTIME_KEY, USERS_SEED_MODE_FULL)
                db.commit()
                global SEED_USERS_READY
                SEED_USERS_READY = False
                _ensure_seed_users(db, force=True)
                user = _get_user_by_username(db, username)
            except Exception:
                db.rollback()

    if not user:
        raise HTTPException(status_code=401, detail="Credenciais invalidas")
    if not user.is_active:
        raise HTTPException(status_code=403, detail="Acesso bloqueado. Procure o administrador.")
    if not _verify_password(payload.password, user.password_hash, user.password_salt):
        raise HTTPException(status_code=401, detail="Credenciais invalidas")
    normalized_role = _normalize_role(user.role)
    if normalized_role == ROLE_CORRETOR and not CORRETOR_ROUTE_ENABLED:
        raise HTTPException(status_code=403, detail="Acesso do perfil corretor esta inativo no momento.")

    user.last_login_at = _utcnow()
    db.commit()
    db.refresh(user)

    token = _new_session(
        user_id=user.id,
        username=user.username,
        role=normalized_role,
        must_change_password=bool(user.must_change_password),
    )
    home = "/app/trocar-senha" if user.must_change_password else _home_for_role(user.role)
    response = JSONResponse(
        {
            "ok": True,
            "username": user.username,
            "role": normalized_role,
            "must_change_password": bool(user.must_change_password),
            "home": home,
        }
    )
    response.set_cookie(
        key=SESSION_COOKIE_NAME,
        value=token,
        httponly=True,
        secure=SESSION_COOKIE_SECURE,
        samesite="lax",
        max_age=SESSION_TTL_SECONDS,
    )
    return response


@app.post("/auth/logout")
def auth_logout(request: Request):
    token = request.cookies.get(SESSION_COOKIE_NAME)
    if token:
        _delete_session_record(token)

    response = JSONResponse({"ok": True})
    response.delete_cookie(key=SESSION_COOKIE_NAME)
    return response


@app.get("/auth/me")
def auth_me(request: Request):
    session = _read_session(request)
    if not session:
        raise HTTPException(status_code=401, detail="Nao autenticado")
    username = str(session.get("username", ""))
    role = _normalize_role(str(session.get("role", "")))
    must_change_password = bool(session.get("must_change_password"))
    home = "/app/trocar-senha" if must_change_password else _home_for_role(role)
    return {
        "ok": True,
        "username": username,
        "role": role,
        "must_change_password": must_change_password,
        "home": home,
    }


@app.post("/auth/change-password")
def auth_change_password(
    payload: ChangePasswordPayload,
    request: Request,
    db: Session = Depends(get_db),
):
    session = require_app_session(request)
    user_id_raw = str(session.get("user_id", "")).strip()
    if not user_id_raw:
        raise HTTPException(status_code=401, detail="Sessao invalida")

    try:
        user_id = uuid.UUID(user_id_raw)
    except ValueError as exc:
        raise HTTPException(status_code=401, detail="Sessao invalida") from exc

    user = db.get(AppUser, user_id)
    if not user or not user.is_active:
        raise HTTPException(status_code=403, detail="Acesso bloqueado")

    current_password = payload.current_password or ""
    new_password = (payload.new_password or "").strip()
    policy_error = _password_policy_error(new_password)
    if policy_error:
        raise HTTPException(status_code=422, detail=policy_error)

    if not _verify_password(current_password, user.password_hash, user.password_salt):
        raise HTTPException(status_code=401, detail="Senha atual invalida")

    _set_user_password(user, new_password, must_change_password=False)
    db.commit()
    db.refresh(user)

    token = request.cookies.get(SESSION_COOKIE_NAME)
    session = None
    if token:
        with ACTIVE_SESSIONS_LOCK:
            session = ACTIVE_SESSIONS.get(token)
            if session:
                session["must_change_password"] = False
                session["role"] = _normalize_role(user.role)
                session["username"] = user.username
                session["db_checked_at"] = _utcnow()
        if session:
            _persist_session_activity(token, session)

    return {
        "ok": True,
        "username": user.username,
        "role": _normalize_role(user.role),
        "must_change_password": False,
        "home": _home_for_role(user.role),
    }


@app.get("/app/api/admin/users", response_model=list[AppUserOut])
def admin_list_users(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    _ensure_seed_users(db)
    users = db.query(AppUser).order_by(func.lower(AppUser.username).asc()).all()
    return users


@app.post("/app/api/admin/users", response_model=AppUserOut)
def admin_create_user(
    payload: AdminUserCreate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    username = _normalize_username(payload.username)
    if not username:
        raise HTTPException(status_code=422, detail="Usuario obrigatorio")
    policy_error = _password_policy_error(payload.password or "")
    if policy_error:
        raise HTTPException(status_code=422, detail=policy_error)
    role = (payload.role or "").strip().lower()
    if role not in VALID_ROLES:
        raise HTTPException(status_code=422, detail="Perfil invalido")
    if _get_user_by_username(db, username):
        raise HTTPException(status_code=409, detail="Usuario ja existe")

    user = AppUser(
        username=username,
        role=role,
        is_active=True,
        must_change_password=True,
    )
    _set_user_password(user, payload.password, must_change_password=True)
    db.add(user)
    db.flush()
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="USUARIO_CRIADO",
        entidade_tipo="usuario",
        entidade_id=str(user.id),
        details=f"username={user.username}; role={user.role}; force_change_password=true",
    )
    db.commit()
    db.refresh(user)
    return user


@app.patch("/app/api/admin/users/{user_id}", response_model=AppUserOut)
def admin_update_user(
    user_id: uuid.UUID,
    payload: AdminUserUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    user = db.get(AppUser, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    current_user_id = str(session.get("user_id", ""))
    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    change_notes: list[str] = []

    if payload.role is not None:
        role = (payload.role or "").strip().lower()
        if role not in VALID_ROLES:
            raise HTTPException(status_code=422, detail="Perfil invalido")
        if user.role != role:
            change_notes.append(f"role:{user.role}->{role}")
        user.role = role

    if payload.is_active is not None:
        if current_user_id and current_user_id == str(user.id) and not bool(payload.is_active):
            raise HTTPException(status_code=422, detail="Admin nao pode bloquear a si mesmo")
        if bool(user.is_active) != bool(payload.is_active):
            change_notes.append(f"is_active:{bool(user.is_active)}->{bool(payload.is_active)}")
        user.is_active = bool(payload.is_active)
        if not user.is_active:
            _drop_sessions_for_user(user.id)

    if change_notes:
        _record_system_log(
            db,
            actor_username=actor_username,
            actor_role=actor_role,
            tela="admin",
            acao="USUARIO_ATUALIZADO",
            entidade_tipo="usuario",
            entidade_id=str(user.id),
            details=f"username={user.username}; " + "; ".join(change_notes),
        )

    db.commit()
    db.refresh(user)

    return user


@app.post("/app/api/admin/users/{user_id}/reset-password", response_model=AppUserOut)
def admin_reset_password(
    user_id: uuid.UUID,
    payload: AdminResetPasswordPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    user = db.get(AppUser, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Usuario nao encontrado")
    policy_error = _password_policy_error(payload.new_password or "")
    if policy_error:
        raise HTTPException(status_code=422, detail=policy_error)

    _set_user_password(user, payload.new_password, must_change_password=payload.force_change_password)
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="USUARIO_RESET_SENHA",
        entidade_tipo="usuario",
        entidade_id=str(user.id),
        details=f"username={user.username}; force_change_password={bool(payload.force_change_password)}",
    )
    db.commit()
    db.refresh(user)
    _drop_sessions_for_user(user.id)
    return user


@app.get("/app/api/ccas", response_model=list[str])
def app_list_ccas(
    _: dict[str, Any] = Depends(require_roles(ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(AppUser.username)
        .filter(AppUser.is_active.is_(True), func.lower(AppUser.role) == ROLE_CCA)
        .order_by(func.lower(AppUser.username).asc())
        .all()
    )
    return [row[0] for row in rows]


@app.get("/app/api/empreendimentos", response_model=list[EmpreendimentoOut])
def app_list_empreendimentos(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(Empreendimento)
        .filter(Empreendimento.is_active.is_(True))
        .order_by(func.lower(Empreendimento.nome).asc())
        .all()
    )
    return rows


@app.get("/app/api/corretor/pre-cadastros/stages")
def corretor_precadastro_stages(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
):
    return {"stages": LEAD_STAGE_VALUES}


@app.get("/app/api/corretor/pre-cadastros", response_model=list[LeadPreCadastroOut])
def corretor_list_pre_cadastros(
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    q: Optional[str] = Query(default=None),
    estagio: Optional[str] = Query(default=None),
    empreendimento: Optional[str] = Query(default=None),
    localidade: Optional[str] = Query(default=None),
    corretor: Optional[str] = Query(default=None),
    limit: int = Query(default=300, ge=1, le=1000),
):
    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))

    query = db.query(LeadPreCadastro)
    if role == ROLE_CORRETOR:
        query = query.filter(func.lower(func.trim(LeadPreCadastro.corretor_username)) == username)
    elif corretor:
        query = query.filter(func.lower(func.trim(LeadPreCadastro.corretor_username)) == _normalize_username(corretor))

    term = (q or "").strip().lower()
    estagio_norm = _lead_stage(estagio, fallback="") if estagio else ""
    empreendimento_term = (empreendimento or "").strip().lower()
    localidade_term = (localidade or "").strip().lower()

    if term:
        like = f"%{term}%"
        pii_filters: list[Any] = []
        if "@" in term:
            pii_filters.append(LeadPreCadastro.email_hash == hash_email(term))
        digits_term = "".join(ch for ch in term if ch.isdigit())
        if len(digits_term) >= 4:
            last4 = digits_term[-4:]
            pii_filters.extend(
                [
                    LeadPreCadastro.telefone_last4 == last4,
                    LeadPreCadastro.whatsapp_last4 == last4,
                    LeadPreCadastro.cpf_last4 == last4,
                ]
            )
        doc_hash = hash_optional(term)
        if doc_hash:
            pii_filters.extend(
                [
                    LeadPreCadastro.cpf_hash == doc_hash,
                    LeadPreCadastro.documento_hash == doc_hash,
                ]
            )
        query = query.filter(
            or_(
                func.lower(func.coalesce(LeadPreCadastro.nome_cliente, "")).like(like),
                func.lower(func.coalesce(LeadPreCadastro.decisao_cca, "")).like(like),
                func.lower(func.coalesce(LeadPreCadastro.empreendimento_interesse, "")).like(like),
                func.lower(func.coalesce(LeadPreCadastro.localidade_interesse, "")).like(like),
                func.lower(func.coalesce(LeadPreCadastro.local_agendamento, "")).like(like),
                *pii_filters,
            )
        )
    if estagio_norm:
        query = query.filter(func.upper(func.coalesce(LeadPreCadastro.estagio_lead, "")) == estagio_norm)
    if empreendimento_term:
        query = query.filter(
            func.lower(func.coalesce(LeadPreCadastro.empreendimento_interesse, "")).like(f"%{empreendimento_term}%")
        )
    if localidade_term:
        query = query.filter(func.lower(func.coalesce(LeadPreCadastro.localidade_interesse, "")).like(f"%{localidade_term}%"))

    return query.order_by(LeadPreCadastro.updated_at.desc(), LeadPreCadastro.created_at.desc()).limit(limit).all()


@app.post("/app/api/corretor/pre-cadastros", response_model=LeadPreCadastroOut)
def corretor_create_pre_cadastro(
    payload: LeadPreCadastroCreate,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    role = _normalize_role(str(session.get("role", "")))
    actor_username = _normalize_username(str(session.get("username", "")))

    nome_cliente = _normalize_lead_text(payload.nome_cliente, max_len=220)
    if not nome_cliente:
        raise HTTPException(status_code=422, detail="Nome do cliente obrigatorio")

    empreendimento = _resolve_empreendimento_nome(db, payload.empreendimento_interesse) or _normalize_lead_text(
        payload.empreendimento_interesse,
        max_len=220,
    )
    estado_civil = _normalize_estado_civil(payload.estado_civil)
    certidao_numero = _normalize_lead_text(payload.certidao_numero, max_len=60)
    if estado_civil != "CASADO":
        certidao_numero = None

    lead = LeadPreCadastro(
        corretor_username=actor_username,
        nome_cliente=nome_cliente,
        telefone=_normalize_lead_phone(payload.telefone),
        whatsapp=_normalize_lead_phone(payload.whatsapp),
        email=_normalize_lead_email(payload.email),
        cpf=_normalize_lead_cpf(payload.cpf),
        documento_identificacao=_normalize_lead_documento(payload.documento_identificacao),
        estado_civil=estado_civil,
        certidao_numero=certidao_numero,
        cidade_nascimento=_normalize_lead_text(payload.cidade_nascimento, max_len=120),
        data_nascimento=payload.data_nascimento,
        endereco=_normalize_lead_text(payload.endereco, max_len=400),
        empreendimento_interesse=empreendimento,
        localidade_interesse=_normalize_lead_text(payload.localidade_interesse, max_len=220),
        local_agendamento=_normalize_lead_text(payload.local_agendamento, max_len=220),
        tipo_visita=_normalize_tipo_visita(payload.tipo_visita),
        data_agendamento=_as_utc(payload.data_agendamento),
        estagio_lead=_lead_stage(payload.estagio_lead, fallback="LEAD"),
        decisao_cca=_lead_cca_decision(payload.decisao_cca, fallback="EM_ANALISE"),
        contrato_assinado=bool(payload.contrato_assinado) if payload.contrato_assinado is not None else False,
        contrato_assinado_em=_utcnow() if bool(payload.contrato_assinado) else None,
        assinatura_email_confirmada=False,
        assinatura_email_confirmada_em=None,
        assinatura_email_token=None,
        assinatura_email_token_expires_at=None,
        assinatura_email_enviado_em=None,
        observacoes=_normalize_lead_text(payload.observacoes, max_len=2000),
        ultimo_contato_em=_as_utc(payload.ultimo_contato_em),
    )

    db.add(lead)
    db.flush()
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=role,
        tela="corretor_precadastro",
        acao="LEAD_PRECADASTRO_CRIADO",
        entidade_tipo="lead_precadastro",
        entidade_id=str(lead.id),
        details=(
            f"cliente={lead.nome_cliente}; estagio={lead.estagio_lead}; "
            f"empreendimento={lead.empreendimento_interesse or '-'}"
        ),
    )
    db.commit()
    db.refresh(lead)
    return lead


@app.patch("/app/api/corretor/pre-cadastros/{lead_id}", response_model=LeadPreCadastroOut)
def corretor_update_pre_cadastro(
    lead_id: uuid.UUID,
    payload: LeadPreCadastroUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    role = _normalize_role(str(session.get("role", "")))
    actor_username = _normalize_username(str(session.get("username", "")))
    lead = db.get(LeadPreCadastro, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead nao encontrado")
    if role == ROLE_CORRETOR and _normalize_username(lead.corretor_username) != actor_username:
        raise HTTPException(status_code=403, detail="Sem permissao para alterar este lead")

    changes = payload.model_dump(exclude_unset=True)
    if not changes:
        return lead

    if "nome_cliente" in changes:
        nome_cliente = _normalize_lead_text(changes.get("nome_cliente"), max_len=220)
        if not nome_cliente:
            raise HTTPException(status_code=422, detail="Nome do cliente obrigatorio")
        lead.nome_cliente = nome_cliente
    if "telefone" in changes:
        lead.telefone = _normalize_lead_phone(changes.get("telefone"))
    if "whatsapp" in changes:
        lead.whatsapp = _normalize_lead_phone(changes.get("whatsapp"))
    if "email" in changes:
        next_email = _normalize_lead_email(changes.get("email"))
        if next_email != lead.email:
            lead.assinatura_email_confirmada = False
            lead.assinatura_email_confirmada_em = None
            lead.assinatura_email_token = None
            lead.assinatura_email_token_expires_at = None
            lead.assinatura_email_enviado_em = None
        lead.email = next_email
    if "cpf" in changes:
        lead.cpf = _normalize_lead_cpf(changes.get("cpf"))
    if "documento_identificacao" in changes:
        lead.documento_identificacao = _normalize_lead_documento(changes.get("documento_identificacao"))
    if "estado_civil" in changes:
        lead.estado_civil = _normalize_estado_civil(changes.get("estado_civil"))
        if lead.estado_civil != "CASADO":
            lead.certidao_numero = None
    if "certidao_numero" in changes:
        lead.certidao_numero = _normalize_lead_text(changes.get("certidao_numero"), max_len=60) if lead.estado_civil == "CASADO" else None
    if "cidade_nascimento" in changes:
        lead.cidade_nascimento = _normalize_lead_text(changes.get("cidade_nascimento"), max_len=120)
    if "data_nascimento" in changes:
        lead.data_nascimento = changes.get("data_nascimento")
    if "endereco" in changes:
        lead.endereco = _normalize_lead_text(changes.get("endereco"), max_len=400)
    if "empreendimento_interesse" in changes:
        lead.empreendimento_interesse = _resolve_empreendimento_nome(db, changes.get("empreendimento_interesse")) or _normalize_lead_text(
            changes.get("empreendimento_interesse"),
            max_len=220,
        )
    if "localidade_interesse" in changes:
        lead.localidade_interesse = _normalize_lead_text(changes.get("localidade_interesse"), max_len=220)
    if "local_agendamento" in changes:
        lead.local_agendamento = _normalize_lead_text(changes.get("local_agendamento"), max_len=220)
    if "tipo_visita" in changes:
        lead.tipo_visita = _normalize_tipo_visita(changes.get("tipo_visita"))
    if "data_agendamento" in changes:
        lead.data_agendamento = _as_utc(changes.get("data_agendamento"))
    if "estagio_lead" in changes:
        lead.estagio_lead = _lead_stage(changes.get("estagio_lead"), fallback=lead.estagio_lead or "LEAD")
    if "decisao_cca" in changes:
        lead.decisao_cca = _lead_cca_decision(changes.get("decisao_cca"), fallback=lead.decisao_cca or "EM_ANALISE")
    if "contrato_assinado" in changes:
        contrato_assinado = bool(changes.get("contrato_assinado"))
        if contrato_assinado and not bool(lead.contrato_assinado):
            lead.contrato_assinado = True
            lead.contrato_assinado_em = _utcnow()
        elif not contrato_assinado and bool(lead.contrato_assinado):
            lead.contrato_assinado = False
            lead.contrato_assinado_em = None
            lead.assinatura_email_confirmada = False
            lead.assinatura_email_confirmada_em = None
            lead.assinatura_email_token = None
            lead.assinatura_email_token_expires_at = None
            lead.assinatura_email_enviado_em = None
    if "observacoes" in changes:
        lead.observacoes = _normalize_lead_text(changes.get("observacoes"), max_len=2000)
    if "ultimo_contato_em" in changes:
        lead.ultimo_contato_em = _as_utc(changes.get("ultimo_contato_em"))

    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=role,
        tela="corretor_precadastro",
        acao="LEAD_PRECADASTRO_ATUALIZADO",
        entidade_tipo="lead_precadastro",
        entidade_id=str(lead.id),
        details=f"cliente={lead.nome_cliente}; estagio={lead.estagio_lead}",
    )
    db.commit()
    db.refresh(lead)
    return lead


@app.post("/app/api/corretor/pre-cadastros/{lead_id}/assinatura/enviar-confirmacao", response_model=LeadAssinaturaEmailOut)
def corretor_enviar_confirmacao_assinatura_email(
    lead_id: uuid.UUID,
    request: Request,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    role = _normalize_role(str(session.get("role", "")))
    actor_username = _normalize_username(str(session.get("username", "")))
    lead = db.get(LeadPreCadastro, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead nao encontrado")
    if role == ROLE_CORRETOR and _normalize_username(lead.corretor_username) != actor_username:
        raise HTTPException(status_code=403, detail="Sem permissao para este lead")
    if not _is_email_delivery_configured():
        raise HTTPException(
            status_code=503,
            detail=(
                "Envio de e-mail nao configurado. No Render Free, use EMAIL_DELIVERY_PROVIDER=brevo, "
                "EMAIL_BREVO_API_KEY e EMAIL_FROM/EMAIL_SMTP_FROM."
            ),
        )
    if not lead.email:
        raise HTTPException(status_code=422, detail="Lead sem e-mail cadastrado para confirmacao de assinatura.")
    if not bool(lead.contrato_assinado):
        raise HTTPException(status_code=422, detail="Marque contrato assinado antes de enviar confirmacao por e-mail.")

    now = _utcnow()
    expires_at = now + timedelta(hours=EMAIL_CONFIRM_TOKEN_TTL_HOURS)
    raw_token = secrets.token_urlsafe(32)
    stored_token_hash = hash_token(raw_token)
    lead.assinatura_email_confirmada = False
    lead.assinatura_email_confirmada_em = None
    lead.assinatura_email_token = stored_token_hash
    lead.assinatura_email_token_expires_at = expires_at
    lead.assinatura_email_enviado_em = now

    link = _build_email_confirmation_link(request, raw_token)
    subject = "Confirmacao de assinatura de contrato - SioCred"
    text_body = (
        f"Ola, {lead.nome_cliente}.\n\n"
        "Recebemos a assinatura do contrato e precisamos da sua confirmacao por e-mail.\n"
        "Clique no link abaixo para confirmar:\n\n"
        f"{link}\n\n"
        f"Este link expira em {EMAIL_CONFIRM_TOKEN_TTL_HOURS} hora(s).\n"
        "Se voce nao reconhece este envio, desconsidere."
    )
    try:
        _send_email_message(to_email=lead.email, subject=subject, text_body=text_body)
    except Exception as exc:
        db.rollback()
        logger.exception("Falha ao enviar e-mail de confirmacao de assinatura.")
        raise HTTPException(status_code=502, detail=f"Falha ao enviar e-mail de confirmacao: {exc}") from exc

    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=role,
        tela="corretor_precadastro",
        acao="LEAD_ASSINATURA_EMAIL_ENVIADO",
        entidade_tipo="lead_precadastro",
        entidade_id=str(lead.id),
        details=f"cliente={lead.nome_cliente}; email={mask_email(lead.email)}; expira={expires_at.isoformat()}",
    )
    db.commit()
    db.refresh(lead)
    return LeadAssinaturaEmailOut(
        lead_id=lead.id,
        email=mask_email(lead.email) or "",
        token_expires_at=lead.assinatura_email_token_expires_at or expires_at,
        enviado_em=lead.assinatura_email_enviado_em or now,
    )


@app.get("/app/assinatura/confirmar", response_class=HTMLResponse)
def app_confirmar_assinatura_email(token: str = Query(default=""), db: Session = Depends(get_db)):
    token_value = (token or "").strip()
    if not token_value:
        return HTMLResponse(
            "<h2>Link invalido</h2><p>Token de confirmacao nao informado.</p>",
            status_code=400,
        )
    now = _utcnow()
    lead = (
        db.query(LeadPreCadastro)
        .filter(
            or_(
                LeadPreCadastro.assinatura_email_token == token_value,
                LeadPreCadastro.assinatura_email_token == hash_token(token_value),
            )
        )
        .first()
    )
    if not lead:
        return HTMLResponse(
            "<h2>Link invalido</h2><p>O link de confirmacao nao foi encontrado ou ja foi utilizado.</p>",
            status_code=404,
        )
    expires_at = _as_utc(getattr(lead, "assinatura_email_token_expires_at", None))
    if expires_at is None or expires_at < now:
        lead.assinatura_email_token = None
        lead.assinatura_email_token_expires_at = None
        db.commit()
        return HTMLResponse(
            "<h2>Link expirado</h2><p>Solicite um novo e-mail de confirmacao com o corretor.</p>",
            status_code=410,
        )
    lead.assinatura_email_confirmada = True
    lead.assinatura_email_confirmada_em = now
    lead.assinatura_email_token = None
    lead.assinatura_email_token_expires_at = None
    db.commit()
    return HTMLResponse(
        "<h2>Assinatura confirmada</h2><p>Obrigada. Sua confirmacao foi registrada com sucesso.</p>",
        status_code=200,
    )


@app.post("/app/api/corretor/pre-cadastros/{lead_id}/reservar", response_model=LeadReservaOut)
def corretor_reservar_pre_cadastro(
    lead_id: uuid.UUID,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    role = _normalize_role(str(session.get("role", "")))
    actor_username = _normalize_username(str(session.get("username", "")))
    lead = db.get(LeadPreCadastro, lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead nao encontrado")
    if role == ROLE_CORRETOR and _normalize_username(lead.corretor_username) != actor_username:
        raise HTTPException(status_code=403, detail="Sem permissao para reservar este lead")

    processo_existente = db.get(Processo, lead.processo_id) if getattr(lead, "processo_id", None) else None
    if processo_existente:
        now_existing = _utcnow()
        processo_existente.estagio_comercial = "EM_PROCESSO"
        lead.estagio_lead = "RESERVA"
        lead.reservado_em = lead.reservado_em or now_existing
        db.commit()
        _invalidate_process_list_cache()
        return LeadReservaOut(
            lead_id=lead.id,
            cliente_id=processo_existente.cliente_id,
            processo_id=processo_existente.id,
            estagio_lead=lead.estagio_lead,
            reservado_em=_as_utc(lead.reservado_em) or now_existing,
        )

    block_reason = _lead_reserva_block_reason(lead)
    if block_reason:
        raise HTTPException(status_code=422, detail=block_reason)

    now = _utcnow()
    decisao_cca = _lead_cca_decision(getattr(lead, "decisao_cca", None), fallback="EM_ANALISE")
    empreendimento_nome = _resolve_empreendimento_nome(db, lead.empreendimento_interesse) or _normalize_lead_text(
        lead.empreendimento_interesse,
        max_len=220,
    )
    corretor_nome = _normalize_corretor_nome_curto(lead.corretor_username) or (lead.corretor_username or "").strip() or None

    cliente = Cliente(
        nome=lead.nome_cliente,
        corretor=corretor_nome,
        obra=empreendimento_nome,
        imobiliaria=None,
        data_reserva_origem=now.date(),
        data_cadastro_origem=(lead.created_at.date() if getattr(lead, "created_at", None) else now.date()),
    )
    db.add(cliente)
    db.flush()

    processo = Processo(
        cliente_id=cliente.id,
        estagio_comercial="EM_PROCESSO",
        status_cca=decisao_cca,
        status_credito=("APROVADO" if decisao_cca == "APROVADO" else ("PENDENCIADO" if decisao_cca == "CONDICIONADO" else "EM_ANALISE")),
        status_geral=("EM_ANDAMENTO" if decisao_cca == "APROVADO" else ("PENDENCIADO" if decisao_cca == "CONDICIONADO" else "NOVO")),
        valor_cheque_moradia=_resolve_cheque_moradia_valor(db, empreendimento_nome),
        sla_comercial_inicio_at=_utc_start_of_day(cliente.data_cadastro_origem) or now,
    )
    _sync_estagio_repasse_rules(processo, now)
    _refresh_sla_fixed_markers(processo, now)
    _switch_sla_owner(processo, SLA_OWNER_ANALISTA, now)
    db.add(processo)
    db.flush()
    _ensure_default_documentos(db, processo.id, autocommit=False)
    _record_processo_event(
        db,
        processo_id=processo.id,
        actor_username=actor_username,
        actor_role=role,
        event_type="PROCESSO_CRIADO_LEAD_RESERVA",
        details=f"lead_id={lead.id}; cliente={cliente.nome}",
    )

    lead.estagio_lead = "RESERVA"
    lead.processo_id = processo.id
    lead.reservado_em = now

    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=role,
        tela="corretor_precadastro",
        acao="LEAD_CONVERTIDO_RESERVA",
        entidade_tipo="lead_precadastro",
        entidade_id=str(lead.id),
        details=(
            f"cliente={cliente.nome}; processo_id={processo.id}; "
            f"empreendimento={cliente.obra or '-'}; cheque_moradia={processo.valor_cheque_moradia or 0}"
        ),
    )

    db.commit()
    _sync_frankstein_events_for_processo(db, processo, lead=lead)
    _invalidate_process_list_cache()
    return LeadReservaOut(
        lead_id=lead.id,
        cliente_id=cliente.id,
        processo_id=processo.id,
        estagio_lead=lead.estagio_lead,
        reservado_em=lead.reservado_em or now,
    )


@app.get("/app/api/corretor/unidades/status")
def corretor_unidades_status(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
):
    return {"status": UNIDADE_STATUS_VALUES}


@app.get("/app/api/corretor/unidades", response_model=list[UnidadeDisponivelOut])
def corretor_list_unidades(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    q: Optional[str] = Query(default=None),
    empreendimento: Optional[str] = Query(default=None),
    tipologia: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    quartos_min: Optional[int] = Query(default=None, ge=0, le=20),
    preco_max: Optional[float] = Query(default=None, ge=0),
    somente_disponiveis: bool = Query(default=True),
    limit: int = Query(default=300, ge=1, le=1000),
):
    query = db.query(UnidadeDisponivel)

    term = (q or "").strip().lower()
    empreendimento_term = (empreendimento or "").strip().lower()
    tipologia_term = (tipologia or "").strip().lower()
    status_norm = _unidade_status(status, fallback="") if status else ""

    if term:
        like = f"%{term}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(UnidadeDisponivel.empreendimento, "")).like(like),
                func.lower(func.coalesce(UnidadeDisponivel.unidade, "")).like(like),
                func.lower(func.coalesce(UnidadeDisponivel.bloco, "")).like(like),
                func.lower(func.coalesce(UnidadeDisponivel.tipologia, "")).like(like),
                func.lower(func.coalesce(UnidadeDisponivel.localizacao, "")).like(like),
                func.lower(func.coalesce(UnidadeDisponivel.diferenciais, "")).like(like),
            )
        )
    if empreendimento_term:
        query = query.filter(func.lower(func.coalesce(UnidadeDisponivel.empreendimento, "")).like(f"%{empreendimento_term}%"))
    if tipologia_term:
        query = query.filter(func.lower(func.coalesce(UnidadeDisponivel.tipologia, "")).like(f"%{tipologia_term}%"))
    if status_norm:
        query = query.filter(func.upper(func.coalesce(UnidadeDisponivel.status_unidade, "")) == status_norm)
    elif somente_disponiveis:
        query = query.filter(func.upper(func.coalesce(UnidadeDisponivel.status_unidade, "")) == "DISPONIVEL")

    if quartos_min is not None:
        query = query.filter(func.coalesce(UnidadeDisponivel.quartos, 0) >= int(quartos_min))
    if preco_max is not None:
        query = query.filter(func.coalesce(UnidadeDisponivel.valor, 0) <= float(preco_max))

    status_order = (
        "CASE UPPER(COALESCE(status_unidade, '')) "
        "WHEN 'DISPONIVEL' THEN 0 "
        "WHEN 'RESERVADA' THEN 1 "
        "WHEN 'BLOQUEADA' THEN 2 "
        "WHEN 'VENDIDA' THEN 3 "
        "ELSE 9 END"
    )
    return (
        query.order_by(
            text(status_order),
            UnidadeDisponivel.valor.asc().nullslast(),
            UnidadeDisponivel.updated_at.desc(),
        )
        .limit(limit)
        .all()
    )


@app.get("/app/api/corretor/unidades/metricas", response_model=UnidadeDisponivelMetricasOut)
def corretor_unidades_metricas(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    total = int(db.query(func.count(UnidadeDisponivel.id)).scalar() or 0)
    disponiveis = int(
        db.query(func.count(UnidadeDisponivel.id))
        .filter(func.upper(func.coalesce(UnidadeDisponivel.status_unidade, "")) == "DISPONIVEL")
        .scalar()
        or 0
    )
    ticket_medio = float(
        db.query(func.avg(UnidadeDisponivel.valor))
        .filter(
            UnidadeDisponivel.valor.isnot(None),
            func.upper(func.coalesce(UnidadeDisponivel.status_unidade, "")) == "DISPONIVEL",
        )
        .scalar()
        or 0.0
    )
    area_media = float(
        db.query(func.avg(UnidadeDisponivel.area_m2))
        .filter(
            UnidadeDisponivel.area_m2.isnot(None),
            func.upper(func.coalesce(UnidadeDisponivel.status_unidade, "")) == "DISPONIVEL",
        )
        .scalar()
        or 0.0
    )
    empreendimentos = int(
        db.query(func.count(func.distinct(func.lower(func.coalesce(UnidadeDisponivel.empreendimento, "")))))
        .filter(UnidadeDisponivel.empreendimento.isnot(None))
        .scalar()
        or 0
    )
    return UnidadeDisponivelMetricasOut(
        total=total,
        disponiveis=disponiveis,
        ticket_medio=ticket_medio,
        area_media=area_media,
        empreendimentos=empreendimentos,
    )


@app.post("/app/api/corretor/unidades", response_model=UnidadeDisponivelOut)
def corretor_create_unidade(
    payload: UnidadeDisponivelCreate,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))

    empreendimento = _resolve_empreendimento_nome(db, payload.empreendimento) or _normalize_unidade_text(
        payload.empreendimento,
        max_len=220,
    )
    unidade = _normalize_unidade_text(payload.unidade, max_len=80)
    if not empreendimento:
        raise HTTPException(status_code=422, detail="Empreendimento obrigatorio")
    if not unidade:
        raise HTTPException(status_code=422, detail="Identificacao da unidade obrigatoria")

    exists = (
        db.query(UnidadeDisponivel.id)
        .filter(
            func.lower(func.trim(func.coalesce(UnidadeDisponivel.empreendimento, ""))) == empreendimento.strip().lower(),
            func.lower(func.trim(func.coalesce(UnidadeDisponivel.unidade, ""))) == unidade.strip().lower(),
        )
        .first()
    )
    if exists:
        raise HTTPException(status_code=409, detail="Unidade ja cadastrada para esse empreendimento")

    unidade_db = UnidadeDisponivel(
        created_by_username=actor_username,
        empreendimento=empreendimento,
        unidade=unidade,
        bloco=_normalize_unidade_text(payload.bloco, max_len=80),
        tipologia=_normalize_unidade_text(payload.tipologia, max_len=120),
        quartos=_normalize_unidade_positive_int(payload.quartos, min_value=0, max_value=20),
        banheiros=_normalize_unidade_positive_int(payload.banheiros, min_value=0, max_value=20),
        vagas=_normalize_unidade_positive_int(payload.vagas, min_value=0, max_value=20),
        area_m2=_normalize_unidade_positive_float(payload.area_m2, min_value=0.0, max_value=5000.0),
        valor=_normalize_unidade_positive_float(payload.valor, min_value=0.0),
        localizacao=_normalize_unidade_text(payload.localizacao, max_len=220),
        diferenciais=_normalize_unidade_text(payload.diferenciais, max_len=1500),
        url_imagem=_normalize_unidade_text(payload.url_imagem, max_len=700),
        visita_disponivel=bool(payload.visita_disponivel if payload.visita_disponivel is not None else True),
        status_unidade=_unidade_status(payload.status_unidade, fallback="DISPONIVEL"),
    )
    db.add(unidade_db)
    db.flush()
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="corretor_apresentacao",
        acao="UNIDADE_CADASTRADA",
        entidade_tipo="unidade",
        entidade_id=str(unidade_db.id),
        details=f"empreendimento={unidade_db.empreendimento}; unidade={unidade_db.unidade}; status={unidade_db.status_unidade}",
    )
    db.commit()
    db.refresh(unidade_db)
    return unidade_db


@app.patch("/app/api/corretor/unidades/{unidade_id}", response_model=UnidadeDisponivelOut)
def corretor_update_unidade(
    unidade_id: uuid.UUID,
    payload: UnidadeDisponivelUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    unidade = db.get(UnidadeDisponivel, unidade_id)
    if not unidade:
        raise HTTPException(status_code=404, detail="Unidade nao encontrada")

    changes = payload.model_dump(exclude_unset=True)
    if not changes:
        return unidade

    empreendimento_normalizado = unidade.empreendimento
    unidade_normalizada = unidade.unidade
    if "empreendimento" in changes:
        empreendimento_normalizado = _resolve_empreendimento_nome(db, changes.get("empreendimento")) or _normalize_unidade_text(
            changes.get("empreendimento"),
            max_len=220,
        )
        if not empreendimento_normalizado:
            raise HTTPException(status_code=422, detail="Empreendimento obrigatorio")
    if "unidade" in changes:
        unidade_normalizada = _normalize_unidade_text(changes.get("unidade"), max_len=80)
        if not unidade_normalizada:
            raise HTTPException(status_code=422, detail="Identificacao da unidade obrigatoria")

    if empreendimento_normalizado != unidade.empreendimento or unidade_normalizada != unidade.unidade:
        exists = (
            db.query(UnidadeDisponivel.id)
            .filter(
                UnidadeDisponivel.id != unidade.id,
                func.lower(func.trim(func.coalesce(UnidadeDisponivel.empreendimento, ""))) == empreendimento_normalizado.strip().lower(),
                func.lower(func.trim(func.coalesce(UnidadeDisponivel.unidade, ""))) == unidade_normalizada.strip().lower(),
            )
            .first()
        )
        if exists:
            raise HTTPException(status_code=409, detail="Ja existe outra unidade com esse identificador")

    unidade.empreendimento = empreendimento_normalizado
    unidade.unidade = unidade_normalizada
    if "bloco" in changes:
        unidade.bloco = _normalize_unidade_text(changes.get("bloco"), max_len=80)
    if "tipologia" in changes:
        unidade.tipologia = _normalize_unidade_text(changes.get("tipologia"), max_len=120)
    if "quartos" in changes:
        unidade.quartos = _normalize_unidade_positive_int(changes.get("quartos"), min_value=0, max_value=20)
    if "banheiros" in changes:
        unidade.banheiros = _normalize_unidade_positive_int(changes.get("banheiros"), min_value=0, max_value=20)
    if "vagas" in changes:
        unidade.vagas = _normalize_unidade_positive_int(changes.get("vagas"), min_value=0, max_value=20)
    if "area_m2" in changes:
        unidade.area_m2 = _normalize_unidade_positive_float(changes.get("area_m2"), min_value=0.0, max_value=5000.0)
    if "valor" in changes:
        unidade.valor = _normalize_unidade_positive_float(changes.get("valor"), min_value=0.0)
    if "localizacao" in changes:
        unidade.localizacao = _normalize_unidade_text(changes.get("localizacao"), max_len=220)
    if "diferenciais" in changes:
        unidade.diferenciais = _normalize_unidade_text(changes.get("diferenciais"), max_len=1500)
    if "url_imagem" in changes:
        unidade.url_imagem = _normalize_unidade_text(changes.get("url_imagem"), max_len=700)
    if "visita_disponivel" in changes:
        unidade.visita_disponivel = bool(changes.get("visita_disponivel"))
    if "status_unidade" in changes:
        unidade.status_unidade = _unidade_status(changes.get("status_unidade"), fallback=unidade.status_unidade or "DISPONIVEL")

    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="corretor_apresentacao",
        acao="UNIDADE_ATUALIZADA",
        entidade_tipo="unidade",
        entidade_id=str(unidade.id),
        details=f"empreendimento={unidade.empreendimento}; unidade={unidade.unidade}; status={unidade.status_unidade}",
    )
    db.commit()
    db.refresh(unidade)
    return unidade


@app.get("/app/api/admin/empreendimentos", response_model=list[EmpreendimentoOut])
def admin_list_empreendimentos(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    rows = db.query(Empreendimento).order_by(func.lower(Empreendimento.nome).asc()).all()
    return rows


@app.post("/app/api/admin/empreendimentos", response_model=EmpreendimentoOut)
def admin_create_empreendimento(
    payload: EmpreendimentoCreate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    nome = _normalize_empreendimento_nome(payload.nome)
    if not nome:
        raise HTTPException(status_code=422, detail="Nome do empreendimento obrigatorio")

    existing = db.query(Empreendimento).filter(func.lower(Empreendimento.nome) == nome.lower()).first()
    if existing:
        if not existing.is_active:
            existing.is_active = True
            _record_system_log(
                db,
                actor_username=_normalize_username(str(session.get("username", ""))),
                actor_role=_normalize_role(str(session.get("role", ""))),
                tela="admin",
                acao="EMPREENDIMENTO_REATIVADO",
                entidade_tipo="empreendimento",
                entidade_id=str(existing.id),
                details=f"nome={existing.nome}",
            )
            db.commit()
            db.refresh(existing)
        return existing

    empreendimento = Empreendimento(nome=nome, is_active=True)
    db.add(empreendimento)
    db.flush()
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="EMPREENDIMENTO_CRIADO",
        entidade_tipo="empreendimento",
        entidade_id=str(empreendimento.id),
        details=f"nome={empreendimento.nome}",
    )
    db.commit()
    db.refresh(empreendimento)
    return empreendimento


@app.delete("/app/api/admin/empreendimentos/{empreendimento_id}")
def admin_delete_empreendimento(
    empreendimento_id: uuid.UUID,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    empreendimento = db.get(Empreendimento, empreendimento_id)
    if not empreendimento:
        raise HTTPException(status_code=404, detail="Empreendimento nao encontrado")
    empreendimento.is_active = False
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="EMPREENDIMENTO_DESATIVADO",
        entidade_tipo="empreendimento",
        entidade_id=str(empreendimento.id),
        details=f"nome={empreendimento.nome}",
    )
    db.commit()
    return {"ok": True}


@app.get("/app/api/admin/empreendimentos/regras-financeiras", response_model=list[EmpreendimentoRegraFinanceiraOut])
def admin_list_empreendimento_regras_financeiras(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    empreendimentos = db.query(Empreendimento).order_by(func.lower(Empreendimento.nome).asc()).all()
    regras = db.query(EmpreendimentoRegraFinanceira).all()
    regras_map: dict[uuid.UUID, EmpreendimentoRegraFinanceira] = {
        regra.empreendimento_id: regra for regra in regras if getattr(regra, "empreendimento_id", None)
    }
    now = _utcnow()
    output: list[EmpreendimentoRegraFinanceiraOut] = []
    for emp in empreendimentos:
        regra = regras_map.get(emp.id)
        output.append(
            EmpreendimentoRegraFinanceiraOut(
                empreendimento_id=emp.id,
                empreendimento_nome=emp.nome,
                valor_cheque_moradia=float(getattr(regra, "valor_cheque_moradia", 0.0) or 0.0),
                updated_at=getattr(regra, "updated_at", None) or getattr(emp, "updated_at", None) or now,
            )
        )
    return output


@app.put("/app/api/admin/empreendimentos/{empreendimento_id}/regra-financeira", response_model=EmpreendimentoRegraFinanceiraOut)
def admin_upsert_empreendimento_regra_financeira(
    empreendimento_id: uuid.UUID,
    payload: EmpreendimentoRegraFinanceiraPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    empreendimento = db.get(Empreendimento, empreendimento_id)
    if not empreendimento:
        raise HTTPException(status_code=404, detail="Empreendimento nao encontrado")

    valor = _coerce_optional_currency(payload.valor_cheque_moradia, "Valor do cheque moradia")
    if valor is None:
        valor = 0.0

    regra = (
        db.query(EmpreendimentoRegraFinanceira)
        .filter(EmpreendimentoRegraFinanceira.empreendimento_id == empreendimento_id)
        .first()
    )
    if not regra:
        regra = EmpreendimentoRegraFinanceira(
            empreendimento_id=empreendimento_id,
            valor_cheque_moradia=valor,
        )
        db.add(regra)
    else:
        regra.valor_cheque_moradia = valor

    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="EMPREENDIMENTO_REGRA_FINANCEIRA_ATUALIZADA",
        entidade_tipo="empreendimento",
        entidade_id=str(empreendimento.id),
        details=f"nome={empreendimento.nome}; valor_cheque_moradia={valor}",
    )
    db.commit()
    db.refresh(regra)
    return EmpreendimentoRegraFinanceiraOut(
        empreendimento_id=empreendimento.id,
        empreendimento_nome=empreendimento.nome,
        valor_cheque_moradia=float(regra.valor_cheque_moradia or 0.0),
        updated_at=regra.updated_at or _utcnow(),
    )


@app.get("/app/api/admin/logs", response_model=list[SistemaLogOut])
def admin_list_system_logs(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
    limit: int = Query(default=300, ge=1, le=2000),
    tela: Optional[str] = Query(default=None),
    username: Optional[str] = Query(default=None),
):
    query = db.query(SistemaLog)
    if tela:
        query = query.filter(func.lower(func.trim(SistemaLog.tela)) == tela.strip().lower())
    if username:
        normalized_username = _normalize_username(username)
        if normalized_username:
            query = query.filter(func.lower(func.trim(SistemaLog.actor_username)) == normalized_username)
    return query.order_by(SistemaLog.created_at.desc()).limit(limit).all()


@app.get("/app/api/admin/storage-summary", response_model=AdminStorageSummaryOut)
def admin_storage_summary(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    _ensure_process_archiving(db, _utcnow())

    total_clientes = int(db.query(func.count(Cliente.id)).scalar() or 0)
    total_processos = int(db.query(func.count(Processo.id)).scalar() or 0)
    total_processos_arquivados = int(
        db.query(func.count(Processo.id)).filter(Processo.arquivado.is_(True)).scalar() or 0
    )
    total_pre_cadastros = int(db.query(func.count(LeadPreCadastro.id)).scalar() or 0)
    total_unidades_disponiveis = int(db.query(func.count(UnidadeDisponivel.id)).scalar() or 0)
    total_processos_ativos = int(
        db.query(func.count(Processo.id)).filter(_processos_ativos_clause()).scalar() or 0
    )
    total_documentos = int(db.query(func.count(Documento.id)).scalar() or 0)
    total_eventos_processo = int(db.query(func.count(ProcessoEvento.id)).scalar() or 0)
    total_logs_sistema = int(db.query(func.count(SistemaLog.id)).scalar() or 0)
    total_usuarios = int(db.query(func.count(AppUser.id)).scalar() or 0)
    total_empreendimentos = int(db.query(func.count(Empreendimento.id)).scalar() or 0)
    total_registros_monitorados = (
        total_clientes
        + total_processos
        + total_pre_cadastros
        + total_unidades_disponiveis
        + total_documentos
        + total_eventos_processo
        + total_logs_sistema
        + total_usuarios
        + total_empreendimentos
    )

    return AdminStorageSummaryOut(
        total_clientes=total_clientes,
        total_processos=total_processos,
        total_processos_ativos=total_processos_ativos,
        total_processos_arquivados=total_processos_arquivados,
        total_pre_cadastros=total_pre_cadastros,
        total_unidades_disponiveis=total_unidades_disponiveis,
        total_documentos=total_documentos,
        total_eventos_processo=total_eventos_processo,
        total_logs_sistema=total_logs_sistema,
        total_usuarios=total_usuarios,
        total_empreendimentos=total_empreendimentos,
        total_registros_monitorados=total_registros_monitorados,
    )


@app.get("/app/api/layout-preference", response_model=LayoutPreferenceOut)
def app_get_layout_preference(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    return LayoutPreferenceOut(
        blackhole_enabled=_is_blackhole_layout_enabled(db),
        blackhole_allowed=BLACKHOLE_LAYOUT_ALLOWED,
        fonte="runtime",
    )


@app.get("/app/api/admin/layout-preference", response_model=LayoutPreferenceOut)
def admin_get_layout_preference(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    return LayoutPreferenceOut(
        blackhole_enabled=_is_blackhole_layout_enabled(db),
        blackhole_allowed=BLACKHOLE_LAYOUT_ALLOWED,
        fonte="runtime",
    )


@app.put("/app/api/admin/layout-preference", response_model=LayoutPreferenceOut)
def admin_set_layout_preference(
    payload: LayoutPreferencePayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    enabled = bool(payload.blackhole_enabled) and BLACKHOLE_LAYOUT_ALLOWED
    _set_runtime_meta(db, LAYOUT_BLACKHOLE_RUNTIME_KEY, "1" if enabled else "0")
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="LAYOUT_PREFERENCIA_ATUALIZADA",
        entidade_tipo="configuracao",
        entidade_id=LAYOUT_BLACKHOLE_RUNTIME_KEY,
        details=f"blackhole_enabled={enabled}",
    )
    db.commit()
    return LayoutPreferenceOut(blackhole_enabled=enabled, blackhole_allowed=BLACKHOLE_LAYOUT_ALLOWED, fonte="runtime")


@app.get("/app/api/admin/frankstein-email-alerts", response_model=AdminFranksteinEmailAlertsOut)
def admin_get_frankstein_email_alerts(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    return _build_frankstein_email_alerts_status(db)


@app.put("/app/api/admin/frankstein-email-alerts", response_model=AdminFranksteinEmailAlertsOut)
def admin_set_frankstein_email_alerts(
    payload: AdminFranksteinEmailAlertsPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    raw = (payload.destinatarios or "").strip()
    recipients = _validate_alert_recipients(raw)
    _set_runtime_meta(db, FRANKSTEIN_ALERT_EMAIL_TO_RUNTIME_KEY, raw)
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="FRANKSTEIN_EMAIL_ALERTA_CONFIGURADO",
        entidade_tipo="configuracao",
        entidade_id=FRANKSTEIN_ALERT_EMAIL_TO_RUNTIME_KEY,
        details=f"destinatarios={','.join(mask_email(email) or email for email in recipients) or '-'}",
    )
    db.commit()
    return _build_frankstein_email_alerts_status(db)


@app.post("/app/api/admin/frankstein-email-alerts/test", response_model=dict[str, Any])
def admin_test_frankstein_email_alerts(
    payload: AdminFranksteinEmailTestPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    if not _is_email_delivery_configured():
        raise HTTPException(
            status_code=422,
            detail=(
                "Envio de e-mail nao configurado. No Render Free, use EMAIL_DELIVERY_PROVIDER=brevo, "
                "EMAIL_BREVO_API_KEY e EMAIL_FROM/EMAIL_SMTP_FROM."
            ),
        )
    raw = payload.destinatarios if payload.destinatarios is not None else _get_alert_recipients_source(db)[0]
    recipients = _validate_alert_recipients(raw)
    if not recipients:
        raise HTTPException(status_code=422, detail="Informe ao menos um e-mail destinatario.")

    now_brt = datetime.now(KEEPALIVE_BRT_TZ)
    subject = f"{AI_ASSISTANT_DISPLAY_NAME}: teste de alerta por e-mail - {now_brt:%d/%m/%Y %H:%M}"
    body = "\n".join(
        [
            f"{AI_ASSISTANT_DISPLAY_NAME} alerta supervisionado do SioCred",
            "",
            "Este e um teste de envio do painel admin.",
            f"Quando houver tarefa ou compromisso com horario, o {AI_ASSISTANT_DISPLAY_NAME} avisara 5 minutos antes.",
            f"Data/hora do teste: {now_brt:%d/%m/%Y %H:%M:%S} BRT",
            f"Canal configurado: {_email_delivery_provider()}",
            f"Remetente configurado: {EMAIL_FROM}",
            "",
            "Se voce recebeu este e-mail, a comunicacao por e-mail esta funcionando.",
        ]
    )
    sent_messages = []
    try:
        for to_email in recipients:
            sent_messages.append(_send_email_message(to_email=to_email, subject=subject, text_body=body))
    except smtplib.SMTPAuthenticationError as exc:
        detail = exc.smtp_error.decode(errors="ignore") if isinstance(exc.smtp_error, bytes) else str(exc.smtp_error)
        raise HTTPException(status_code=422, detail=f"SMTP recusou login. Verifique usuario e senha de app. {detail}") from exc
    except smtplib.SMTPRecipientsRefused as exc:
        refused = ", ".join(exc.recipients.keys())
        raise HTTPException(status_code=422, detail=f"SMTP recusou destinatario(s): {refused}") from exc
    except smtplib.SMTPException as exc:
        raise HTTPException(status_code=422, detail=f"SMTP recusou o envio: {exc}") from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    except OSError as exc:
        raise HTTPException(status_code=422, detail=f"Falha de conexao SMTP: {exc}") from exc

    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="admin",
        acao="FRANKSTEIN_EMAIL_ALERTA_TESTE",
        entidade_tipo="configuracao",
        entidade_id=FRANKSTEIN_ALERT_EMAIL_TO_RUNTIME_KEY,
        details=(
            f"destinatarios={','.join(mask_email(email) or email for email in recipients)}; "
            f"assunto={subject}; "
            f"message_ids={','.join(str(item.get('message_id', '')) for item in sent_messages)}"
        ),
    )
    db.commit()
    return {
        "ok": True,
        "sent": len(sent_messages),
        "accepted_at_brt": now_brt.isoformat(),
        "subject": subject,
        "provider": sent_messages[0].get("provider", _email_delivery_provider()) if sent_messages else _email_delivery_provider(),
        "message_ids": [str(item.get("message_id", "")) for item in sent_messages],
        "destinatarios_mascarados": [mask_email(email) or email for email in recipients],
    }


def _normalize_maintenance_entity(value: Optional[str]) -> str:
    raw = (value or "").strip().lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "cliente": "cliente",
        "clientes": "cliente",
        "processo": "processo",
        "processos": "processo",
        "documento": "documento",
        "documentos": "documento",
        "usuario": "usuario",
        "usuarios": "usuario",
        "app_user": "usuario",
        "app_users": "usuario",
        "empreendimento": "empreendimento",
        "empreendimentos": "empreendimento",
        "unidade": "unidade",
        "unidades": "unidade",
        "unidade_disponivel": "unidade",
        "unidades_disponiveis": "unidade",
        "processo_evento": "processo_evento",
        "processo_eventos": "processo_evento",
        "evento": "processo_evento",
        "eventos": "processo_evento",
        "sistema_log": "sistema_log",
        "sistema_logs": "sistema_log",
        "log": "sistema_log",
        "logs": "sistema_log",
    }
    return aliases.get(raw, "")


@app.get("/app/api/admin/maintenance/search-registros", response_model=AdminRegistroLookupOut)
def admin_search_registros(
    entidade: str = Query(...),
    q: str = Query(..., min_length=1),
    limit: int = Query(default=20, ge=1, le=100),
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    entidade_norm = _normalize_maintenance_entity(entidade)
    if not entidade_norm:
        raise HTTPException(
            status_code=422,
            detail=(
                "Entidade invalida. Use: cliente, processo, documento, usuario, "
                "empreendimento, unidade, processo_evento ou sistema_log."
            ),
        )

    termo = (q or "").strip()
    if len(termo) < 2:
        raise HTTPException(status_code=422, detail="Digite ao menos 2 caracteres para buscar.")

    termo_like = f"%{termo.lower()}%"
    uuid_term: Optional[uuid.UUID] = None
    try:
        uuid_term = uuid.UUID(termo)
    except ValueError:
        uuid_term = None

    itens: list[AdminRegistroLookupItem] = []

    if entidade_norm == "cliente":
        query = db.query(Cliente)
        filtros = or_(
            func.lower(func.coalesce(Cliente.nome, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.corretor, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.obra, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.imobiliaria, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, Cliente.id == uuid_term)
        rows = query.filter(filtros).order_by(Cliente.created_at.desc()).limit(limit).all()
        for item in rows:
            detalhe = f"Corretor: {item.corretor or '-'} | Obra: {item.obra or '-'} | Imobiliaria: {item.imobiliaria or '-'}"
            itens.append(AdminRegistroLookupItem(id=item.id, titulo=item.nome or "-", detalhe=detalhe))

    elif entidade_norm == "processo":
        query = db.query(Processo, Cliente).join(Cliente, Processo.cliente_id == Cliente.id)
        filtros = or_(
            func.lower(func.coalesce(Cliente.nome, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.corretor, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.obra, "")).like(termo_like),
            func.lower(func.coalesce(Processo.estagio_comercial, "")).like(termo_like),
            func.lower(func.coalesce(Processo.status_cca, "")).like(termo_like),
            func.lower(func.coalesce(Processo.status_agehab, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, Processo.id == uuid_term, Processo.cliente_id == uuid_term)
        rows = query.filter(filtros).order_by(Processo.updated_at.desc()).limit(limit).all()
        for processo, cliente in rows:
            detalhe = (
                f"Cliente: {cliente.nome or '-'} | Obra: {cliente.obra or '-'} | "
                f"Comercial: {_process_estagio_comercial(processo.estagio_comercial)} | Caixa: {processo.status_cca}"
            )
            itens.append(AdminRegistroLookupItem(id=processo.id, titulo=f"Processo de {cliente.nome or '-'}", detalhe=detalhe))

    elif entidade_norm == "documento":
        query = (
            db.query(Documento, Processo, Cliente)
            .join(Processo, Documento.processo_id == Processo.id)
            .join(Cliente, Processo.cliente_id == Cliente.id)
        )
        filtros = or_(
            func.lower(func.coalesce(Documento.nome, "")).like(termo_like),
            func.lower(func.coalesce(Documento.categoria, "")).like(termo_like),
            func.lower(func.coalesce(Documento.status_credito, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.nome, "")).like(termo_like),
            func.lower(func.coalesce(Cliente.obra, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, Documento.id == uuid_term, Documento.processo_id == uuid_term, Processo.cliente_id == uuid_term)
        rows = query.filter(filtros).order_by(Documento.updated_at.desc()).limit(limit).all()
        for doc, processo, cliente in rows:
            detalhe = (
                f"Cliente: {cliente.nome or '-'} | Categoria: {doc.categoria or '-'} | "
                f"Status: {doc.status_credito or '-'} | Processo: {str(processo.id)[:8]}"
            )
            itens.append(AdminRegistroLookupItem(id=doc.id, titulo=doc.nome or "-", detalhe=detalhe))

    elif entidade_norm == "usuario":
        query = db.query(AppUser)
        filtros = or_(
            func.lower(func.coalesce(AppUser.username, "")).like(termo_like),
            func.lower(func.coalesce(AppUser.role, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, AppUser.id == uuid_term)
        rows = query.filter(filtros).order_by(func.lower(AppUser.username).asc()).limit(limit).all()
        for user in rows:
            detalhe = f"Perfil: {user.role} | Ativo: {'Sim' if user.is_active else 'Nao'} | Troca senha: {'Sim' if user.must_change_password else 'Nao'}"
            itens.append(AdminRegistroLookupItem(id=user.id, titulo=user.username or "-", detalhe=detalhe))

    elif entidade_norm == "empreendimento":
        query = db.query(Empreendimento)
        filtros = func.lower(func.coalesce(Empreendimento.nome, "")).like(termo_like)
        if uuid_term:
            filtros = or_(filtros, Empreendimento.id == uuid_term)
        rows = query.filter(filtros).order_by(func.lower(Empreendimento.nome).asc()).limit(limit).all()
        for emp in rows:
            detalhe = f"Ativo: {'Sim' if emp.is_active else 'Nao'}"
            itens.append(AdminRegistroLookupItem(id=emp.id, titulo=emp.nome or "-", detalhe=detalhe))

    elif entidade_norm == "unidade":
        query = db.query(UnidadeDisponivel)
        filtros = or_(
            func.lower(func.coalesce(UnidadeDisponivel.empreendimento, "")).like(termo_like),
            func.lower(func.coalesce(UnidadeDisponivel.unidade, "")).like(termo_like),
            func.lower(func.coalesce(UnidadeDisponivel.tipologia, "")).like(termo_like),
            func.lower(func.coalesce(UnidadeDisponivel.status_unidade, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, UnidadeDisponivel.id == uuid_term)
        rows = query.filter(filtros).order_by(UnidadeDisponivel.updated_at.desc()).limit(limit).all()
        for unit in rows:
            detalhe = (
                f"Empreendimento: {unit.empreendimento or '-'} | "
                f"Status: {unit.status_unidade or '-'} | Valor: {unit.valor or '-'}"
            )
            itens.append(AdminRegistroLookupItem(id=unit.id, titulo=unit.unidade or "-", detalhe=detalhe))

    elif entidade_norm == "processo_evento":
        query = db.query(ProcessoEvento)
        filtros = or_(
            func.lower(func.coalesce(ProcessoEvento.actor_username, "")).like(termo_like),
            func.lower(func.coalesce(ProcessoEvento.event_type, "")).like(termo_like),
            func.lower(func.coalesce(ProcessoEvento.field_name, "")).like(termo_like),
            func.lower(func.coalesce(ProcessoEvento.details, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, ProcessoEvento.id == uuid_term, ProcessoEvento.processo_id == uuid_term)
        rows = query.filter(filtros).order_by(ProcessoEvento.created_at.desc()).limit(limit).all()
        for ev in rows:
            detalhe = (
                f"Actor: {ev.actor_username or '-'} | Campo: {ev.field_name or '-'} | "
                f"Novo: {ev.new_value or '-'} | Processo: {str(ev.processo_id)[:8]}"
            )
            itens.append(AdminRegistroLookupItem(id=ev.id, titulo=ev.event_type or "EVENT", detalhe=detalhe))

    elif entidade_norm == "sistema_log":
        query = db.query(SistemaLog)
        filtros = or_(
            func.lower(func.coalesce(SistemaLog.actor_username, "")).like(termo_like),
            func.lower(func.coalesce(SistemaLog.tela, "")).like(termo_like),
            func.lower(func.coalesce(SistemaLog.acao, "")).like(termo_like),
            func.lower(func.coalesce(SistemaLog.details, "")).like(termo_like),
        )
        if uuid_term:
            filtros = or_(filtros, SistemaLog.id == uuid_term)
        rows = query.filter(filtros).order_by(SistemaLog.created_at.desc()).limit(limit).all()
        for log in rows:
            detalhe = f"Tela: {log.tela or '-'} | Usuario: {log.actor_username or '-'} | Acao: {log.acao or '-'}"
            itens.append(AdminRegistroLookupItem(id=log.id, titulo=log.details or "Log do sistema", detalhe=detalhe))

    return AdminRegistroLookupOut(
        entidade=entidade_norm,
        total=len(itens),
        itens=itens,
    )


@app.post("/app/api/admin/maintenance/delete-registro")
def admin_delete_registro(
    payload: AdminDeleteRegistroPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    entidade = _normalize_maintenance_entity(payload.entidade)
    if not entidade:
        raise HTTPException(
            status_code=422,
            detail=(
                "Entidade invalida. Use: cliente, processo, documento, usuario, "
                "empreendimento, unidade, processo_evento ou sistema_log."
            ),
        )

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    entity_map: dict[str, Any] = {
        "cliente": Cliente,
        "processo": Processo,
        "documento": Documento,
        "usuario": AppUser,
        "empreendimento": Empreendimento,
        "unidade": UnidadeDisponivel,
        "processo_evento": ProcessoEvento,
        "sistema_log": SistemaLog,
    }
    model = entity_map[entidade]
    registro = db.get(model, payload.registro_id)
    if not registro:
        raise HTTPException(status_code=404, detail="Registro nao encontrado")

    if entidade == "usuario":
        actor_user_id = str(session.get("user_id", ""))
        if actor_user_id and actor_user_id == str(registro.id):
            raise HTTPException(status_code=422, detail="Admin nao pode excluir a propria conta")
        registro_username = _normalize_username(getattr(registro, "username", ""))
        if registro_username in {_normalize_username(RESET_ADMIN_USERNAME), _normalize_username(APP_ADMIN_USER)}:
            raise HTTPException(status_code=422, detail="Usuario administrador principal nao pode ser excluido")

    identificador = str(payload.registro_id)
    resumo = None
    if hasattr(registro, "username"):
        resumo = f"username={getattr(registro, 'username', '')}"
    elif hasattr(registro, "nome"):
        resumo = f"nome={getattr(registro, 'nome', '')}"
    elif hasattr(registro, "event_type"):
        resumo = f"event_type={getattr(registro, 'event_type', '')}"
    elif hasattr(registro, "acao"):
        resumo = f"acao={getattr(registro, 'acao', '')}"

    if entidade != "sistema_log":
        details = f"entidade={entidade}; registro_id={identificador}"
        if resumo:
            details += f"; {resumo}"
        if payload.motivo:
            details += f"; motivo={payload.motivo.strip()[:400]}"
        _record_system_log(
            db,
            actor_username=actor_username,
            actor_role=actor_role,
            tela="admin",
            acao="REGISTRO_EXCLUIDO_DEFINITIVO",
            entidade_tipo=entidade,
            entidade_id=identificador,
            details=details,
        )

    if entidade == "usuario":
        _drop_sessions_for_user(registro.id)
    db.delete(registro)
    db.commit()
    _invalidate_process_list_cache()

    return {
        "ok": True,
        "entidade": entidade,
        "registro_id": identificador,
    }


@app.post("/app/api/admin/maintenance/purge-historicos")
def admin_purge_historicos(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    eventos_total = int(db.query(func.count(ProcessoEvento.id)).scalar() or 0)
    logs_total = int(db.query(func.count(SistemaLog.id)).scalar() or 0)

    db.query(ProcessoEvento).delete(synchronize_session=False)
    db.query(SistemaLog).delete(synchronize_session=False)
    db.commit()
    _invalidate_process_list_cache()

    return {
        "ok": True,
        "processo_eventos_removidos": eventos_total,
        "sistema_logs_removidos": logs_total,
    }


@app.post("/app/api/admin/maintenance/reset-sistema")
def admin_reset_sistema(
    session: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    global SEED_USERS_READY

    session_user_id = str(session.get("user_id", "")).strip()
    session_username = _normalize_username(str(session.get("username", "")))
    admin_username = session_username or _normalize_username(RESET_ADMIN_USERNAME) or _normalize_username(APP_ADMIN_USER)

    clientes_total = int(db.query(func.count(Cliente.id)).scalar() or 0)
    processos_total = int(db.query(func.count(Processo.id)).scalar() or 0)
    documentos_total = int(db.query(func.count(Documento.id)).scalar() or 0)
    pre_cadastros_total = int(db.query(func.count(LeadPreCadastro.id)).scalar() or 0)
    planejamento_total = int(db.query(func.count(CreditoPlanejamentoItem.id)).scalar() or 0)
    analises_total = int(db.query(func.count(AnaliseRegistroDB.id)).scalar() or 0)
    feedback_total = int(db.query(func.count(IaFeedback.id)).scalar() or 0)
    empreendimentos_total = int(db.query(func.count(Empreendimento.id)).scalar() or 0)
    unidades_total = int(db.query(func.count(UnidadeDisponivel.id)).scalar() or 0)
    eventos_total = int(db.query(func.count(ProcessoEvento.id)).scalar() or 0)
    logs_total = int(db.query(func.count(SistemaLog.id)).scalar() or 0)
    usuarios_total = int(db.query(func.count(AppUser.id)).scalar() or 0)
    analises_legacy_total = 0
    runtime_meta_total = 0
    runtime_meta_exists = bool(db.execute(text("SELECT to_regclass('public.app_runtime_meta')")).scalar())
    if runtime_meta_exists:
        runtime_meta_total = int(db.execute(text("SELECT COUNT(*) FROM app_runtime_meta")).scalar() or 0)
    if AnaliseSessionLocal is not None:
        analise_db = AnaliseSessionLocal()
        try:
            analises_legacy_total = int(analise_db.query(func.count(Analise.id)).scalar() or 0)
        finally:
            analise_db.close()

    # Remove entidades dependentes de processo explicitamente para nao depender
    # apenas de cascades do banco e garantir limpeza completa da carteira.
    db.query(AnalistaReuniaoCompromisso).delete(synchronize_session=False)
    db.query(AnalistaReuniaoComercial).delete(synchronize_session=False)
    db.query(Documento).delete(synchronize_session=False)
    db.query(ProcessoEvento).delete(synchronize_session=False)
    db.query(LeadPreCadastro).delete(synchronize_session=False)
    db.query(Processo).delete(synchronize_session=False)
    db.query(Cliente).delete(synchronize_session=False)
    db.query(CreditoPlanejamentoItem).delete(synchronize_session=False)
    db.query(AnaliseRegistroDB).delete(synchronize_session=False)
    db.query(IaFeedback).delete(synchronize_session=False)
    db.query(Empreendimento).delete(synchronize_session=False)
    db.query(UnidadeDisponivel).delete(synchronize_session=False)
    db.query(SistemaLog).delete(synchronize_session=False)
    if runtime_meta_exists:
        db.execute(text("DELETE FROM app_runtime_meta"))

    admin_user = None
    if session_user_id:
        try:
            admin_user = db.get(AppUser, uuid.UUID(session_user_id))
        except ValueError:
            admin_user = None
    if admin_user is None and admin_username:
        admin_user = db.query(AppUser).filter(func.lower(AppUser.username) == admin_username).first()
    admin_criado = False

    if not admin_user:
        raise HTTPException(status_code=401, detail="Sessao administrativa invalida. Faca login novamente.")
    else:
        admin_user.role = ROLE_ADMIN
        admin_user.is_active = True
        admin_user.username = admin_username or admin_user.username

    # Nao remove usuarios cadastrados no reset geral.
    usuarios_removidos = 0

    _set_runtime_meta(db, META_MENSAL_RUNTIME_KEY, "0")
    _set_runtime_meta(db, LAYOUT_BLACKHOLE_RUNTIME_KEY, "0")
    # Volta o seed para full para manter comportamento padrao sem restringir acesso.
    _set_runtime_meta(db, USERS_SEED_MODE_RUNTIME_KEY, USERS_SEED_MODE_FULL)

    db.commit()
    _invalidate_process_list_cache()
    with ACTIVE_SESSIONS_LOCK:
        ACTIVE_SESSIONS.clear()
    db.query(AppSession).delete()
    db.commit()
    SEED_USERS_READY = False

    analises_legacy_remanescentes = 0
    if AnaliseSessionLocal is not None:
        analise_db = AnaliseSessionLocal()
        try:
            analise_db.query(Analise).delete(synchronize_session=False)
            analise_db.commit()
            analises_legacy_remanescentes = int(analise_db.query(func.count(Analise.id)).scalar() or 0)
        except Exception:
            analise_db.rollback()
            analises_legacy_remanescentes = analises_legacy_total
            logger.exception("Falha ao limpar base legada de analises no reset administrativo.")
        finally:
            analise_db.close()

    remanescentes = {
        "clientes": int(db.query(func.count(Cliente.id)).scalar() or 0),
        "processos": int(db.query(func.count(Processo.id)).scalar() or 0),
        "documentos": int(db.query(func.count(Documento.id)).scalar() or 0),
        "pre_cadastros": int(db.query(func.count(LeadPreCadastro.id)).scalar() or 0),
        "planejamento_itens": int(db.query(func.count(CreditoPlanejamentoItem.id)).scalar() or 0),
        "analises": int(db.query(func.count(AnaliseRegistroDB.id)).scalar() or 0),
        "ia_feedback": int(db.query(func.count(IaFeedback.id)).scalar() or 0),
        "empreendimentos": int(db.query(func.count(Empreendimento.id)).scalar() or 0),
        "unidades_disponiveis": int(db.query(func.count(UnidadeDisponivel.id)).scalar() or 0),
        "processo_eventos": int(db.query(func.count(ProcessoEvento.id)).scalar() or 0),
        "sistema_logs": int(db.query(func.count(SistemaLog.id)).scalar() or 0),
        "analises_legacy": analises_legacy_remanescentes,
    }
    remanescentes = {key: value for key, value in remanescentes.items() if value > 0}

    return {
        "ok": True,
        "cliente_registros_removidos": clientes_total,
        "processos_removidos": processos_total,
        "documentos_removidos": documentos_total,
        "pre_cadastros_removidos": pre_cadastros_total,
        "planejamento_itens_removidos": planejamento_total,
        "analises_removidas": analises_total,
        "ia_feedback_removidos": feedback_total,
        "analises_legacy_removidas": analises_legacy_total,
        "empreendimentos_removidos": empreendimentos_total,
        "unidades_disponiveis_removidas": unidades_total,
        "processo_eventos_removidos": eventos_total,
        "sistema_logs_removidos": logs_total,
        "runtime_meta_removidos": runtime_meta_total,
        "usuarios_removidos": usuarios_removidos,
        "usuarios_anteriores": usuarios_total,
        "usuario_preservado": admin_username,
        "usuario_preservado_criado": admin_criado,
        "relogin_obrigatorio": True,
        "registros_remanescentes": remanescentes,
    }


@app.get("/health")
def health():
    return {
        "ok": True,
        "db_configured": bool(DATABASE_URL),
        "db_url_valid": SessionLocal is not None,
        "db_url_has_placeholders": DB_URL_HAS_PLACEHOLDERS,
        "db_uses_direct_supabase_host": bool(DATABASE_URL and _is_supabase_direct_host(DATABASE_URL)),
    }


@app.get("/health/keepalive")
def health_keepalive(request: Request):
    event = _record_keepalive_ping(request)
    return JSONResponse(
        {
            "ok": True,
            "service": "sistema-credito-api",
            "purpose": "keepalive",
            "ts_utc": event["ts_utc"],
            "ts_brt": event["ts_brt"],
            "client_ip": event["client_ip"],
            "logged_memory": True,
            "logged_db": bool(event.get("db_logged")),
        },
        headers={"Cache-Control": "no-store, max-age=0"},
    )


@app.get("/health/keepalive/logs")
def health_keepalive_logs(
    limit: int = Query(default=30, ge=1, le=200),
    include_db: bool = Query(default=True),
):
    recent_memory = list(KEEPALIVE_RECENT)[-limit:]
    recent_memory.reverse()
    db_logs = _fetch_keepalive_db_logs(limit) if include_db else []
    return {
        "ok": True,
        "memory_limit": KEEPALIVE_LOG_MEMORY_LIMIT,
        "memory_count": len(KEEPALIVE_RECENT),
        "recent_memory": recent_memory,
        "db_logs": db_logs,
        "db_count": len(db_logs),
    }


@app.get("/health/db")
def health_db(db: Session = Depends(get_db)):
    try:
        db.execute(text("SELECT 1"))
        return {"ok": True}
    except SQLAlchemyError as exc:
        logger.warning("Falha no health check de banco: %s", exc.__class__.__name__)
        raise HTTPException(
            status_code=503,
            detail={
                "message": _db_error_hint(exc),
                "error_type": exc.__class__.__name__,
            },
        ) from exc


@app.get("/api/health")
def api_health():
    return {"ok": True}


@app.post("/api/clientes", response_model=ClienteOut)
def create_cliente(
    payload: ClienteCreate,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    cliente = Cliente(**payload.model_dump())
    db.add(cliente)
    db.commit()
    db.refresh(cliente)
    return cliente


@app.get("/api/clientes", response_model=list[ClienteOut])
def list_clientes(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    return db.query(Cliente).order_by(Cliente.created_at.desc()).all()


@app.post("/api/processos", response_model=ProcessoOut)
def create_processo(
    payload: ProcessoCreate,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    cliente = db.get(Cliente, payload.cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente nao encontrado")

    processo = Processo(
        cliente_id=payload.cliente_id,
        sla_comercial_inicio_at=_utc_start_of_day(cliente.data_cadastro_origem) or _utcnow(),
    )
    _refresh_sla_fixed_markers(processo, _utcnow())
    _switch_sla_owner(processo, SLA_OWNER_ANALISTA, _utcnow())
    db.add(processo)
    db.commit()
    db.refresh(processo)
    _sync_frankstein_events_for_processo(db, processo)
    _ensure_default_documentos(db, processo.id)
    _invalidate_process_list_cache()
    return processo


@app.get("/api/processos/{processo_id}", response_model=ProcessoOut)
def get_processo(
    processo_id: uuid.UUID,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")
    return processo


@app.patch("/api/processos/{processo_id}", response_model=ProcessoOut)
def patch_processo(
    processo_id: uuid.UUID,
    payload: ProcessoUpdate,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    changes = payload.model_dump(exclude_unset=True)
    estagio_changed = False
    for field, value in changes.items():
        if field == "status_credito":
            status_credito = _process_credit_status(value)
            _validate_status_transition(field, processo.status_credito, status_credito)
            processo.status_credito = status_credito
        elif field == "status_geral":
            status_geral = _process_geral_status(value)
            _validate_status_transition(field, processo.status_geral, status_geral)
            processo.status_geral = status_geral
        elif field == "status_cca":
            status_cca = _process_caixa_status(value)
            _validate_status_transition(field, processo.status_cca, status_cca)
            processo.status_cca = status_cca
            if status_cca in {"APROVADO", "DAR_QV"}:
                processo.status_credito = "APROVADO"
                if processo.status_geral in {"NOVO", "PENDENCIADO"}:
                    processo.status_geral = "EM_ANDAMENTO"
            elif status_cca == "CONDICIONADO":
                processo.status_credito = "PENDENCIADO"
                processo.status_geral = "PENDENCIADO"
            elif status_cca in {"REPROVADO", "BLOQUEADO"}:
                processo.status_credito = "REPROVADO"
                processo.status_geral = "REPROVADO"
        elif field == "status_agehab":
            status_agehab = _process_agehab_status(value)
            _validate_status_transition(field, processo.status_agehab, status_agehab)
            processo.status_agehab = status_agehab
        elif field == "status_sinal":
            processo.status_sinal = _process_sinal_status(value)
        elif field == "status_fiador":
            processo.status_fiador = _process_fiador_status(value)
        elif field == "recolha_fgts":
            processo.recolha_fgts = _process_recolha_fgts_status(value, fallback=processo.recolha_fgts or "NAO_RECOLHIDO")
        elif field == "estagio_comercial":
            next_stage = _process_estagio_comercial(value, fallback="")
            if not next_stage:
                raise HTTPException(status_code=422, detail="Estagio comercial invalido.")
            old_stage = _process_estagio_comercial(processo.estagio_comercial)
            _validate_estagio_comercial_transition(old_stage, next_stage)
            processo.estagio_comercial = next_stage
            estagio_changed = old_stage != next_stage
        elif field == "etapa_repasse":
            next_step = _process_etapa_repasse(value, fallback=None)
            if value is not None and str(value).strip() and not next_step:
                raise HTTPException(status_code=422, detail="Etapa de repasse invalida.")
            processo.etapa_repasse = next_step
        elif field == "nao_contar_mes":
            _set_nao_contar_mes_period(processo, bool(value), _utcnow())
        elif field in {"sla_credito_dias", "sla_corretor_dias", "sla_cca_dias"}:
            continue
        else:
            setattr(processo, field, value)

    if estagio_changed:
        _sync_estagio_repasse_rules(processo, _utcnow())
    if changes:
        _refresh_sla_fixed_markers(processo, _utcnow())
        _apply_sla_rules(
            processo,
            has_enviado_docs=_process_has_enviado_docs(db, processo.id),
        )
    if "nao_contar_mes" in changes:
        _sync_reuniao_conta_no_mes_with_processo(db, processo, _utcnow())

    db.commit()
    db.refresh(processo)
    _sync_frankstein_events_for_processo(db, processo)
    _invalidate_process_list_cache()
    return processo


@app.post("/api/documentos", response_model=DocumentoOut)
def create_documento(
    payload: DocumentoCreate,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, payload.processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    documento = Documento(**payload.model_dump())
    db.add(documento)
    db.commit()
    db.refresh(documento)
    _invalidate_process_list_cache()
    return documento


@app.get("/api/processos/{processo_id}/documentos", response_model=list[DocumentoOut])
def list_documentos(
    processo_id: uuid.UUID,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    return db.query(Documento).filter(Documento.processo_id == processo_id).all()


@app.patch("/api/documentos/{documento_id}", response_model=DocumentoOut)
def patch_documento(
    documento_id: uuid.UUID,
    payload: DocumentoUpdate,
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    documento = db.get(Documento, documento_id)
    if not documento:
        raise HTTPException(status_code=404, detail="Documento nao encontrado")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(documento, field, value)

    db.commit()
    db.refresh(documento)
    _invalidate_process_list_cache()
    return documento


@app.get("/app/api/processos", response_model=list[ProcessoOverviewOut])
def app_list_processos(
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    limit: Optional[int] = Query(default=120, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
):
    _ensure_process_archiving(db, _utcnow())
    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))
    cache_key = _process_list_cache_key(role, username, limit, offset)

    cached = _get_cached_process_list(cache_key)
    if cached is not None:
        return cached

    query = db.query(Processo, Cliente).join(Cliente, Processo.cliente_id == Cliente.id).filter(_processos_ativos_clause())
    if role == ROLE_CORRETOR:
        if not username:
            return []
        query = query.filter(func.lower(func.trim(func.coalesce(Cliente.corretor, ""))) == username)
    elif role == ROLE_CCA:
        if not username:
            return []
        query = query.filter(func.lower(func.trim(func.coalesce(Processo.cca_responsavel, ""))) == username)

    renda_duplicate_lookup = _build_renda_bruta_duplicate_lookup(
        query.with_entities(Processo.id, Processo.renda_bruta, Cliente.nome).all()
    )

    query = query.order_by(Processo.created_at.desc()).offset(offset)
    if limit is not None:
        query = query.limit(limit)

    rows = query.all()
    now = _utcnow()
    docs_stats: dict[uuid.UUID, dict[str, int]] = {}
    processo_ids = [processo.id for processo, _ in rows]
    if processo_ids:
        docs_rows = (
            db.query(Documento.processo_id, Documento.status_doc)
            .filter(Documento.processo_id.in_(processo_ids))
            .all()
        )
        for processo_id, status_doc in docs_rows:
            stats = docs_stats.setdefault(processo_id, {"docs_total": 0, "docs_recebidos": 0})
            stats["docs_total"] += 1
            if _doc_is_done(status_doc):
                stats["docs_recebidos"] += 1

    output: list[ProcessoOverviewOut] = []
    for processo, cliente in rows:
        stats = docs_stats.get(processo.id, {"docs_total": 0, "docs_recebidos": 0})
        docs_total = int(stats.get("docs_total", 0))
        docs_recebidos = int(stats.get("docs_recebidos", 0))
        sem_documento_enviado = docs_recebidos <= 0
        status_cca_norm = _process_caixa_status(processo.status_cca)
        status_agehab_norm = _process_agehab_status(processo.status_agehab)
        etapa_repasse_norm = _process_etapa_repasse(getattr(processo, "etapa_repasse", None))
        estagio_comercial_norm = _process_estagio_comercial(processo.estagio_comercial)
        status_sinal_norm = _process_sinal_status(getattr(processo, "status_sinal", None))
        status_fiador_norm = _process_fiador_status(getattr(processo, "status_fiador", None))
        status_cca_key = "nao_iniciado" if sem_documento_enviado else _normalize_text_key(status_cca_norm or "ANALISE_CREDITO")
        status_agehab_key = "nao_iniciado" if sem_documento_enviado else _normalize_text_key(status_agehab_norm or "ANALISE_CREDITO")
        status_sinal_key = _normalize_text_key(status_sinal_norm or "NAO_TEM")
        status_fiador_key = _normalize_text_key(status_fiador_norm or "NAO_TEM")
        estagio_comercial_key = _normalize_text_key(estagio_comercial_norm)
        etapa_repasse_key = _normalize_text_key(etapa_repasse_norm or "SEM_REPASSE")
        repasse_fase_key = _process_repasse_display_key(etapa_repasse_norm, status_cca_key, status_agehab_key)
        status_pendencias = _process_pending_items(status_cca_key, status_agehab_key, status_sinal_key, status_fiador_key)
        status_tudo_ok = len(status_pendencias) == 0
        renda_duplicate = renda_duplicate_lookup.get(str(processo.id), {})
        docs_pendentes = max(0, docs_total - docs_recebidos)
        documentos_resumo = _process_documentos_resumo(
            docs_total=docs_total,
            docs_recebidos=docs_recebidos,
            sem_documento_enviado=sem_documento_enviado,
        )
        observacao_resumo = _process_observacao_resumo(getattr(processo, "observacao", None))
        espelho_validado = status_cca_norm in {"APROVADO", "DAR_QV"}
        agehab_validado = status_agehab_norm == "VALIDADO_AGEHAB"
        contrato_ja_acionado = etapa_repasse_norm == "ASSINATURA_AUTORIZADA" or status_cca_norm in {"ASSINATURA_CAIXA", "FINALIZADO"}
        aviso_gerar_contrato_agehab = (
            not sem_documento_enviado
            and espelho_validado
            and agehab_validado
            and _process_estagio_comercial(getattr(processo, "estagio_comercial", None)) == "VENDA_FINALIZADA"
            and not contrato_ja_acionado
        )

        sla_analista_seconds = _compute_sla_seconds(processo, SLA_OWNER_ANALISTA, now)
        sla_corretor_seconds = _compute_sla_seconds(processo, SLA_OWNER_CORRETOR, now)
        sla_cca_seconds = _compute_sla_seconds(processo, SLA_OWNER_CCA, now)
        sla_analista_horas = sla_analista_seconds // 3600
        sla_corretor_horas = sla_corretor_seconds // 3600
        sla_cca_horas = sla_cca_seconds // 3600
        output.append(
            ProcessoOverviewOut(
                processo_id=processo.id,
                cliente_id=cliente.id,
                cliente_nome=cliente.nome,
                corretor=cliente.corretor,
                obra=cliente.obra,
                imobiliaria=getattr(cliente, 'imobiliaria', None),
                status_credito=processo.status_credito,
                status_geral=processo.status_geral,
                status_cca=status_cca_norm,
                status_agehab=status_agehab_norm,
                status_sinal=status_sinal_norm,
                valor_sinal=float(processo.valor_sinal) if getattr(processo, "valor_sinal", None) is not None else None,
                renda_bruta=float(processo.renda_bruta) if getattr(processo, "renda_bruta", None) is not None else None,
                renda_bruta_duplicada=bool(renda_duplicate),
                renda_bruta_duplicada_clientes=list(renda_duplicate.get("clientes") or []),
                renda_bruta_duplicada_tooltip=renda_duplicate.get("tooltip"),
                recolha_fgts=_process_recolha_fgts_status(getattr(processo, "recolha_fgts", None)),
                status_fiador=status_fiador_norm,
                estagio_comercial=estagio_comercial_norm,
                etapa_repasse=_process_etapa_repasse(processo.etapa_repasse),
                fila_atual=_fila_atual_from_processo(processo),
                cca_responsavel=processo.cca_responsavel,
                pendente_fiador=processo.pendente_fiador,
                pendente_sinal=processo.pendente_sinal,
                nao_contar_mes=_is_nao_contar_mes_active(processo, now),
                sla_credito_dias=sla_analista_horas // 24,
                sla_corretor_dias=sla_corretor_horas // 24,
                sla_cca_dias=sla_cca_horas // 24,
                sla_analista_horas=sla_analista_horas,
                sla_corretor_horas=sla_corretor_horas,
                sla_cca_horas=sla_cca_horas,
                sla_analista_seconds=sla_analista_seconds,
                sla_corretor_seconds=sla_corretor_seconds,
                sla_cca_seconds=sla_cca_seconds,
                sla_owner=_normalize_sla_owner(processo.sla_owner),
                sla_active_since=_as_utc(processo.sla_active_since),
                data_reserva_origem=getattr(cliente, "data_reserva_origem", None),
                data_cadastro_origem=getattr(cliente, "data_cadastro_origem", None),
                created_at=processo.created_at,
                observacao=processo.observacao,
                docs_total=docs_total,
                docs_recebidos=docs_recebidos,
                sem_documento_enviado=sem_documento_enviado,
                aviso_gerar_contrato_agehab=aviso_gerar_contrato_agehab,
                estagio_comercial_key=estagio_comercial_key,
                estagio_comercial_label=_process_view_label("geral", estagio_comercial_key),
                etapa_repasse_key=etapa_repasse_key,
                etapa_repasse_label=_process_view_label("repasse", etapa_repasse_key),
                repasse_fase_key=repasse_fase_key,
                repasse_fase_label=_process_view_label("repasse", repasse_fase_key),
                status_cca_key=status_cca_key,
                status_cca_label=_process_view_label("status_cca", status_cca_key),
                status_agehab_key=status_agehab_key,
                status_agehab_label=_process_view_label("status_agehab", status_agehab_key),
                status_sinal_key=status_sinal_key,
                status_sinal_label=_process_view_label("status_sinal", status_sinal_key),
                status_fiador_key=status_fiador_key,
                status_fiador_label=_process_view_label("status_fiador", status_fiador_key),
                docs_pendentes=docs_pendentes,
                documentos_resumo=documentos_resumo,
                observacao_resumo=observacao_resumo,
                status_pendencias=status_pendencias,
                status_pendencias_resumo=(
                    "Status OK: caixa, agehab, sinal e fiador alinhados"
                    if status_tudo_ok
                    else f"Pendencias: {'; '.join(status_pendencias)}"
                ),
                status_tudo_ok=status_tudo_ok,
            )
        )

    _set_cached_process_list(cache_key, output)
    return output


@app.get("/app/api/cca/analise", response_model=list[CcaAnaliseItemOut])
def app_list_cca_analise(
    session: dict[str, Any] = Depends(require_roles(ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    q: Optional[str] = Query(default=None),
    obra: Optional[str] = Query(default=None),
    status_cca: Optional[str] = Query(default=None),
    limit: int = Query(default=250, ge=1, le=1000),
):
    _ensure_process_archiving(db, _utcnow())
    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))

    query = (
        db.query(Processo, Cliente)
        .join(Cliente, Processo.cliente_id == Cliente.id)
        .filter(_processos_ativos_clause())
    )

    if role == ROLE_CCA:
        if not username:
            return []
        # CCA ve somente processos atribuidos ao proprio usuario.
        query = query.filter(func.lower(func.trim(func.coalesce(Processo.cca_responsavel, ""))) == username)

    termo = (q or "").strip().lower()
    obra_term = (obra or "").strip().lower()
    status_cca_norm = _process_caixa_status(status_cca, fallback="") if status_cca else ""

    if termo:
        like = f"%{termo}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(Cliente.nome, "")).like(like),
                func.lower(func.coalesce(Cliente.corretor, "")).like(like),
                func.lower(func.coalesce(Cliente.obra, "")).like(like),
                func.lower(func.coalesce(Processo.status_cca, "")).like(like),
            )
        )
    if obra_term:
        query = query.filter(func.lower(func.coalesce(Cliente.obra, "")).like(f"%{obra_term}%"))
    if status_cca_norm:
        query = query.filter(func.upper(func.coalesce(Processo.status_cca, "")) == status_cca_norm)

    rows = query.order_by(Processo.updated_at.desc(), Processo.created_at.desc()).limit(limit).all()
    processo_ids = [processo.id for processo, _ in rows]

    leads_by_processo: dict[uuid.UUID, LeadPreCadastro] = {}
    docs_stats: dict[uuid.UUID, dict[str, int]] = {}
    if processo_ids:
        lead_rows = (
            db.query(LeadPreCadastro)
            .filter(LeadPreCadastro.processo_id.in_(processo_ids))
            .order_by(LeadPreCadastro.updated_at.desc(), LeadPreCadastro.created_at.desc())
            .all()
        )
        for lead in lead_rows:
            pid = getattr(lead, "processo_id", None)
            if pid and pid not in leads_by_processo:
                leads_by_processo[pid] = lead

        docs_rows = (
            db.query(Documento.processo_id, Documento.status_doc)
            .filter(Documento.processo_id.in_(processo_ids))
            .all()
        )
        for processo_id, status_doc in docs_rows:
            stats = docs_stats.setdefault(processo_id, {"docs_total": 0, "docs_recebidos": 0})
            stats["docs_total"] += 1
            if _doc_is_done(status_doc):
                stats["docs_recebidos"] += 1

    output: list[CcaAnaliseItemOut] = []
    for processo, cliente in rows:
        lead_item = leads_by_processo.get(processo.id)
        if role == ROLE_CCA:
            # Para CCA, exibir apenas funil de entrada operacional do proprio CCA.
            estagio_lead = _lead_stage(getattr(lead_item, "estagio_lead", None), fallback="")
            if estagio_lead not in {"PRECADASTRO", "RESERVA"}:
                continue
        stats = docs_stats.get(processo.id, {"docs_total": 0, "docs_recebidos": 0})
        output.append(
            _build_cca_analise_item(
                db,
                processo,
                cliente,
                lead=lead_item,
                docs_total=int(stats.get("docs_total", 0)),
                docs_recebidos=int(stats.get("docs_recebidos", 0)),
            )
        )
    return output


@app.patch("/app/api/cca/analise/{processo_id}", response_model=CcaAnaliseItemOut)
def app_patch_cca_analise(
    processo_id: uuid.UUID,
    payload: CcaAnaliseUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    cliente = db.get(Cliente, processo.cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente nao encontrado")

    actor_role = _normalize_role(str(session.get("role", "")))
    actor_username = _normalize_username(str(session.get("username", "")))
    if actor_role == ROLE_CCA:
        responsavel_norm = _normalize_username(processo.cca_responsavel)
        if responsavel_norm != actor_username:
            raise HTTPException(status_code=403, detail="Sem permissao para este processo")

    changes = payload.model_dump(exclude_unset=True)
    if "status_cca" in changes:
        next_status_cca = _process_caixa_status(changes.get("status_cca"), fallback=processo.status_cca or "ANALISE_CREDITO")
        old_status_cca = processo.status_cca
        _validate_status_transition("status_cca", old_status_cca, next_status_cca)
        if old_status_cca != next_status_cca:
            processo.status_cca = next_status_cca
            _record_processo_event(
                db,
                processo_id=processo.id,
                actor_username=actor_username,
                actor_role=actor_role,
                event_type="STATUS_CHANGE",
                field_name="status_cca",
                old_value=old_status_cca,
                new_value=next_status_cca,
                details="decisao_cca",
            )

            # Regra de fluxo profissional: resposta da CCA sincroniza credito/geral.
            if next_status_cca in {"APROVADO", "DAR_QV"}:
                if processo.status_credito != "APROVADO":
                    _record_processo_event(
                        db,
                        processo_id=processo.id,
                        actor_username=actor_username,
                        actor_role=actor_role,
                        event_type="STATUS_CHANGE",
                        field_name="status_credito",
                        old_value=processo.status_credito,
                        new_value="APROVADO",
                        details="sincronizado_status_cca",
                    )
                    processo.status_credito = "APROVADO"
                if processo.status_geral in {"NOVO", "PENDENCIADO"}:
                    _record_processo_event(
                        db,
                        processo_id=processo.id,
                        actor_username=actor_username,
                        actor_role=actor_role,
                        event_type="STATUS_CHANGE",
                        field_name="status_geral",
                        old_value=processo.status_geral,
                        new_value="EM_ANDAMENTO",
                        details="sincronizado_status_cca",
                    )
                    processo.status_geral = "EM_ANDAMENTO"
            elif next_status_cca == "CONDICIONADO":
                if processo.status_credito != "PENDENCIADO":
                    _record_processo_event(
                        db,
                        processo_id=processo.id,
                        actor_username=actor_username,
                        actor_role=actor_role,
                        event_type="STATUS_CHANGE",
                        field_name="status_credito",
                        old_value=processo.status_credito,
                        new_value="PENDENCIADO",
                        details="sincronizado_status_cca",
                    )
                    processo.status_credito = "PENDENCIADO"
                if processo.status_geral != "PENDENCIADO":
                    _record_processo_event(
                        db,
                        processo_id=processo.id,
                        actor_username=actor_username,
                        actor_role=actor_role,
                        event_type="STATUS_CHANGE",
                        field_name="status_geral",
                        old_value=processo.status_geral,
                        new_value="PENDENCIADO",
                        details="sincronizado_status_cca",
                    )
                    processo.status_geral = "PENDENCIADO"
            elif next_status_cca in {"REPROVADO", "BLOQUEADO"}:
                if processo.status_credito != "REPROVADO":
                    _record_processo_event(
                        db,
                        processo_id=processo.id,
                        actor_username=actor_username,
                        actor_role=actor_role,
                        event_type="STATUS_CHANGE",
                        field_name="status_credito",
                        old_value=processo.status_credito,
                        new_value="REPROVADO",
                        details="sincronizado_status_cca",
                    )
                    processo.status_credito = "REPROVADO"
                if processo.status_geral != "REPROVADO":
                    _record_processo_event(
                        db,
                        processo_id=processo.id,
                        actor_username=actor_username,
                        actor_role=actor_role,
                        event_type="STATUS_CHANGE",
                        field_name="status_geral",
                        old_value=processo.status_geral,
                        new_value="REPROVADO",
                        details="sincronizado_status_cca",
                    )
                    processo.status_geral = "REPROVADO"

    if "recolha_fgts" in changes:
        old_fgts = getattr(processo, "recolha_fgts", None)
        processo.recolha_fgts = _process_recolha_fgts_status(changes.get("recolha_fgts"), fallback=old_fgts or "NAO_RECOLHIDO")
        if old_fgts != processo.recolha_fgts:
            _record_processo_event(
                db,
                processo_id=processo.id,
                actor_username=actor_username,
                actor_role=actor_role,
                event_type="PROCESSO_UPDATE",
                field_name="recolha_fgts",
                old_value=old_fgts,
                new_value=processo.recolha_fgts,
            )

    labels = {
        "renda_bruta": "Renda bruta",
        "renda_liquida": "Renda liquida",
        "valor_parcela": "Valor da parcela",
        "valor_imovel": "Valor do imovel",
        "valor_avaliacao": "Valor de avaliacao",
        "valor_financiamento": "Valor do financiamento",
        "valor_subsidio": "Valor do subsidio",
    }

    for field_name, label in labels.items():
        if field_name not in changes:
            continue
        old_value = getattr(processo, field_name, None)
        new_value = _coerce_optional_currency(changes.get(field_name), label)
        setattr(processo, field_name, new_value)
        if old_value != new_value:
            _record_processo_event(
                db,
                processo_id=processo.id,
                actor_username=actor_username,
                actor_role=actor_role,
                event_type="PROCESSO_UPDATE",
                field_name=field_name,
                old_value=old_value,
                new_value=new_value,
            )

    old_cheque = getattr(processo, "valor_cheque_moradia", None)
    auto_cheque = _resolve_cheque_moradia_valor(db, cliente.obra)
    processo.valor_cheque_moradia = auto_cheque
    if old_cheque != auto_cheque:
        _record_processo_event(
            db,
            processo_id=processo.id,
            actor_username=actor_username,
            actor_role=actor_role,
            event_type="PROCESSO_UPDATE",
            field_name="valor_cheque_moradia",
            old_value=old_cheque,
            new_value=auto_cheque,
            details=f"regra_automatica_empreendimento={cliente.obra or '-'}",
        )

    _refresh_sla_fixed_markers(processo, _utcnow())
    _apply_sla_rules(
        processo,
        has_enviado_docs=_process_has_enviado_docs(db, processo.id),
    )

    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="cca_analise",
        acao="CCA_ANALISE_FINANCEIRA_ATUALIZADA",
        entidade_tipo="processo",
        entidade_id=str(processo.id),
        details=(
            f"cliente={cliente.nome}; obra={cliente.obra or '-'}; "
            f"cheque_moradia={processo.valor_cheque_moradia or 0}; campos={','.join(sorted(changes.keys())) or 'auto'}"
        ),
    )
    db.commit()
    _invalidate_process_list_cache()

    lead = (
        db.query(LeadPreCadastro)
        .filter(LeadPreCadastro.processo_id == processo.id)
        .order_by(LeadPreCadastro.updated_at.desc(), LeadPreCadastro.created_at.desc())
        .first()
    )
    _sync_frankstein_events_for_processo(db, processo, lead=lead)
    docs_rows = (
        db.query(Documento.status_doc)
        .filter(Documento.processo_id == processo.id)
        .all()
    )
    docs_total = len(docs_rows)
    docs_recebidos = sum(1 for (status_doc,) in docs_rows if _doc_is_done(status_doc))
    db.refresh(processo)
    return _build_cca_analise_item(
        db,
        processo,
        cliente,
        lead=lead,
        docs_total=docs_total,
        docs_recebidos=docs_recebidos,
    )


@app.get("/app/api/processos/arquivados", response_model=ProcessoArquivadoListOut)
def app_list_processos_arquivados(
    _: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    q: Optional[str] = Query(default=None),
    obra: Optional[str] = Query(default=None),
    corretor: Optional[str] = Query(default=None),
    imobiliaria: Optional[str] = Query(default=None),
    ano: Optional[int] = Query(default=None),
    mes: Optional[int] = Query(default=None),
    limit: Optional[int] = Query(default=120, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
):
    _ensure_process_archiving(db, _utcnow())

    query = (
        db.query(Processo, Cliente)
        .join(Cliente, Processo.cliente_id == Cliente.id)
        .filter(Processo.arquivado.is_(True))
    )

    ano_periodo: Optional[int] = None
    mes_periodo: Optional[int] = None
    if ano is not None or mes is not None:
        ano_periodo, mes_periodo = _normalize_meta_period(ano, mes)
        periodo_inicio = date(ano_periodo, mes_periodo, 1)
        if mes_periodo == 12:
            proximo_mes = date(ano_periodo + 1, 1, 1)
        else:
            proximo_mes = date(ano_periodo, mes_periodo + 1, 1)
        periodo_inicio_dt = datetime.combine(periodo_inicio, time.min, tzinfo=timezone.utc)
        periodo_fim_exclusivo_dt = datetime.combine(proximo_mes, time.min, tzinfo=timezone.utc)

        query = query.filter(
            or_(
                and_(
                    Processo.arquivado_ref_ano == ano_periodo,
                    Processo.arquivado_ref_mes == mes_periodo,
                ),
                and_(
                    Processo.arquivado_ref_ano.is_(None),
                    Processo.arquivado_ref_mes.is_(None),
                    Processo.arquivado_em.is_not(None),
                    Processo.arquivado_em >= periodo_inicio_dt,
                    Processo.arquivado_em < periodo_fim_exclusivo_dt,
                ),
            )
        )

    term = (q or "").strip().lower()
    obra_term = (obra or "").strip().lower()
    corretor_term = (corretor or "").strip().lower()
    imobiliaria_term = (imobiliaria or "").strip().lower()

    if term:
        like = f"%{term}%"
        query = query.filter(
            or_(
                func.lower(func.coalesce(Cliente.nome, "")).like(like),
                func.lower(func.coalesce(Cliente.obra, "")).like(like),
                func.lower(func.coalesce(Cliente.corretor, "")).like(like),
                func.lower(func.coalesce(Cliente.imobiliaria, "")).like(like),
            )
        )
    if obra_term:
        query = query.filter(func.lower(func.coalesce(Cliente.obra, "")).like(f"%{obra_term}%"))
    if corretor_term:
        query = query.filter(func.lower(func.coalesce(Cliente.corretor, "")).like(f"%{corretor_term}%"))
    if imobiliaria_term:
        query = query.filter(func.lower(func.coalesce(Cliente.imobiliaria, "")).like(f"%{imobiliaria_term}%"))

    query = query.order_by(Processo.arquivado_em.desc().nullslast(), Processo.updated_at.desc()).offset(offset)
    if limit is not None:
        query = query.limit(limit)
    rows = query.all()

    itens = [
        ProcessoArquivadoOut(
            processo_id=processo.id,
            cliente_id=cliente.id,
            cliente_nome=cliente.nome,
            corretor=cliente.corretor,
            obra=cliente.obra,
            imobiliaria=getattr(cliente, "imobiliaria", None),
            status_geral=_process_geral_status(getattr(processo, "status_geral", None)),
            estagio_comercial=_process_estagio_comercial(getattr(processo, "estagio_comercial", None)),
            etapa_repasse=_process_etapa_repasse(getattr(processo, "etapa_repasse", None)),
            status_cca=processo.status_cca,
            status_agehab=processo.status_agehab,
            status_sinal=processo.status_sinal,
            status_fiador=processo.status_fiador,
            arquivado_em=_as_utc(getattr(processo, "arquivado_em", None)),
            arquivado_ref_ano=getattr(processo, "arquivado_ref_ano", None),
            arquivado_ref_mes=getattr(processo, "arquivado_ref_mes", None),
            data_reserva_origem=getattr(cliente, "data_reserva_origem", None),
            data_cadastro_origem=getattr(cliente, "data_cadastro_origem", None),
            created_at=processo.created_at,
            sla_analista_horas=(
                int(round((int(getattr(processo, "sla_analista_seconds", 0) or 0)) / 3600))
                if getattr(processo, "sla_analista_seconds", None) is not None
                else None
            ),
            sla_corretor_horas=(
                int(round((int(getattr(processo, "sla_corretor_seconds", 0) or 0)) / 3600))
                if getattr(processo, "sla_corretor_seconds", None) is not None
                else None
            ),
        )
        for processo, cliente in rows
    ]

    total_clientes_cadastrados = int(db.query(func.count(Cliente.id)).scalar() or 0)
    total_processos_ativos = int(db.query(func.count(Processo.id)).filter(_processos_ativos_clause()).scalar() or 0)
    total_processos_arquivados = int(
        db.query(func.count(Processo.id)).filter(Processo.arquivado.is_(True)).scalar() or 0
    )

    return ProcessoArquivadoListOut(
        total_clientes_cadastrados=total_clientes_cadastrados,
        total_processos_ativos=total_processos_ativos,
        total_processos_arquivados=total_processos_arquivados,
        itens=itens,
    )


@app.get("/app/api/gestor/dashboard")
def app_gestor_dashboard(
    _: dict[str, Any] = Depends(require_roles(ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    _ensure_process_archiving(db, _utcnow())
    rows = (
        db.query(Processo, Cliente)
        .join(Cliente, Processo.cliente_id == Cliente.id)
        .filter(_processos_ativos_clause())
        .all()
    )
    now = _utcnow()
    total_bruto = len(rows)
    total = 0

    total_comercial = 0
    total_credito = 0
    total_repasse = 0
    total_assinados = 0
    total_assinados_semana = 0
    provaveis_cair = 0
    perdas_mes = 0
    nao_contar_mes_total = 0
    chegadas_ultimos_7_dias = 0

    clientes_comercial: list[dict[str, Any]] = []
    clientes_repasse: list[dict[str, Any]] = []
    clientes_assinados: list[dict[str, Any]] = []
    clientes_assinados_semana: list[dict[str, Any]] = []
    clientes_prontos_repasse: list[dict[str, Any]] = []
    clientes_credito: list[dict[str, Any]] = []
    clientes_perdas_mes: list[dict[str, Any]] = []
    clientes_excluidos_mes: list[dict[str, Any]] = []
    clientes_chegadas_7d: list[dict[str, Any]] = []
    clientes_provaveis_cair: list[dict[str, Any]] = []
    clientes_estagios: list[dict[str, Any]] = []

    imob_map: dict[str, int] = {}
    imob_corretores: dict[str, dict[str, int]] = {}
    start_7d = now.date() - timedelta(days=6)
    semana_inicio = now.date() - timedelta(days=now.date().weekday())
    semana_fim_util = semana_inicio + timedelta(days=4)
    semana_fim_contagem = min(now.date(), semana_fim_util)
    dias_uteis_decorridos_semana = 0
    if semana_fim_contagem >= semana_inicio:
        dias_uteis_decorridos_semana = min(5, (semana_fim_contagem - semana_inicio).days + 1)
    dias_uteis_restantes_semana = max(0, 5 - dias_uteis_decorridos_semana)

    sla_comercial_sum = 0
    sla_comercial_count = 0
    sla_credito_sum = 0
    sla_credito_count = 0
    sla_cca_sum = 0
    sla_cca_count = 0
    processo_ids = [processo.id for processo, _ in rows]
    pendencias_docs_por_processo: dict[uuid.UUID, int] = {}
    docs_por_processo: dict[uuid.UUID, list[dict[str, str]]] = {}
    assinado_event_at_por_processo: dict[uuid.UUID, datetime] = {}
    perda_event_at_por_processo: dict[uuid.UUID, datetime] = {}
    if processo_ids:
        docs_rows = (
            db.query(
                Documento.processo_id,
                Documento.categoria,
                Documento.nome,
                Documento.status_credito,
                Documento.pendencia_info,
            )
            .filter(Documento.processo_id.in_(processo_ids))
            .all()
        )
        for processo_id, categoria, nome, status_credito, pendencia_info in docs_rows:
            status_doc = _status_token(status_credito) or "AGUARDANDO_ENVIO"
            if processo_id not in docs_por_processo:
                docs_por_processo[processo_id] = []
            docs_por_processo[processo_id].append(
                {
                    "categoria": str(categoria or "").strip(),
                    "nome": str(nome or "").strip(),
                    "status": status_doc,
                    "descricao": str(pendencia_info or "").strip(),
                }
            )
            if status_doc in {"PENDENCIADO", "REPROVADO"}:
                pendencias_docs_por_processo[processo_id] = pendencias_docs_por_processo.get(processo_id, 0) + 1

        eventos_status_rows = (
            db.query(
                ProcessoEvento.processo_id,
                ProcessoEvento.field_name,
                ProcessoEvento.new_value,
                ProcessoEvento.created_at,
            )
            .filter(
                ProcessoEvento.processo_id.in_(processo_ids),
                ProcessoEvento.field_name.in_(["status_cca", "status_geral"]),
            )
            .all()
        )
        for processo_id, field_name, new_value, created_at in eventos_status_rows:
            evento_time = _as_utc(created_at)
            if evento_time is None:
                continue
            token = _status_token(new_value)
            if field_name == "status_cca" and token in PROCESS_CCA_FINAL_STATUSES:
                prev = assinado_event_at_por_processo.get(processo_id)
                if prev is None or evento_time > prev:
                    assinado_event_at_por_processo[processo_id] = evento_time
            elif field_name == "status_geral" and token in {"CANCELADO", "DISTRATO"}:
                prev = perda_event_at_por_processo.get(processo_id)
                if prev is None or evento_time > prev:
                    perda_event_at_por_processo[processo_id] = evento_time

    for processo, cliente in rows:
        estagio = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
        etapa_repasse = _process_etapa_repasse(getattr(processo, "etapa_repasse", None))
        fila_atual = _fila_atual_from_processo(processo)
        status_geral = _status_token(getattr(processo, "status_geral", None))
        status_cca = _status_token(getattr(processo, "status_cca", None))
        status_agehab = _status_token(getattr(processo, "status_agehab", None))
        status_sinal = _status_token(getattr(processo, "status_sinal", None))
        status_fiador = _status_token(getattr(processo, "status_fiador", None))
        created_at_utc = _as_utc(processo.created_at)
        updated_at_utc = _as_utc(getattr(processo, "updated_at", None)) or created_at_utc
        assinado_event_at_utc = _as_utc(assinado_event_at_por_processo.get(processo.id))
        perda_event_at_utc = _as_utc(perda_event_at_por_processo.get(processo.id))
        created_date = created_at_utc.date() if created_at_utc else None
        data_referencia = _resolve_dashboard_reference_date(cliente, created_date, now.date())
        dias_em_aberto = (now.date() - data_referencia).days if data_referencia else None
        if dias_em_aberto is not None and (dias_em_aberto < 0 or dias_em_aberto > DASHBOARD_MAX_DIAS_EM_ABERTO):
            dias_em_aberto = None
        nao_contar_mes = _is_nao_contar_mes_active(processo, now)
        pendencias_docs = int(pendencias_docs_por_processo.get(processo.id, 0))
        tem_pendencia_status = _processo_has_pendencia(processo)
        tem_pendencia = tem_pendencia_status or pendencias_docs > 0
        documentos = docs_por_processo.get(processo.id, [])
        docs_total = len(documentos)
        docs_todos_aguardando = docs_total > 0 and all(
            doc.get("status") in {"AGUARDANDO_ENVIO", "ANALISE", "NAO_APLICA"} for doc in documentos
        )
        docs_com_pendencia = [doc for doc in documentos if doc.get("status") in {"PENDENCIADO", "REPROVADO"}]
        docs_tooltip_lines: list[str] = []
        if docs_total == 0 or docs_todos_aguardando:
            docs_tooltip = "Documentos: nao foi enviado doc."
        else:
            docs_para_listar = docs_com_pendencia
            if not docs_para_listar:
                docs_todos_aprovados = all(doc.get("status") in {"APROVADO", "NAO_APLICA"} for doc in documentos)
                docs_tooltip = (
                    "Documentos: todos aprovados."
                    if docs_todos_aprovados
                    else "Documentos: sem pendencias registradas."
                )
            else:
                for doc in docs_para_listar:
                    desc = str(doc.get("descricao") or "").strip()
                    desc_upper = desc.upper()
                    for prefix in (
                        "PENDENTE -",
                        "PENDENTE:",
                        "BLOQUEADO -",
                        "BLOQUEADO:",
                        "REPROVADO -",
                        "REPROVADO:",
                    ):
                        if desc_upper.startswith(prefix):
                            desc = desc[len(prefix):].strip(" -:")
                            break
                    nome_doc = str(doc.get("nome") or "-")
                    status_doc = str(doc.get("status") or "-")
                    if desc:
                        docs_tooltip_lines.append(f"- {nome_doc} [{status_doc}]: {desc}")
                    else:
                        docs_tooltip_lines.append(f"- {nome_doc} [{status_doc}]: sem descricao")
                docs_tooltip = "Documentos com pendencia:\n" + "\n".join(docs_tooltip_lines)
        sinal_ok = status_sinal in {"NAO_TEM", "PAGO"}
        fiador_ok = status_fiador in {"NAO_TEM", "FINALIZADO"}
        caixa_ok = _is_caixa_apta_para_assinatura(status_cca)
        pronto_para_repassar = (
            estagio == "VENDA_FINALIZADA"
            and status_agehab == "VALIDADO_AGEHAB"
            and sinal_ok
            and fiador_ok
            and caixa_ok
        )

        cliente_item = {
            "processo_id": str(processo.id),
            "cliente_nome": cliente.nome,
            "corretor": cliente.corretor,
            "obra": cliente.obra,
            "imobiliaria": getattr(cliente, "imobiliaria", None),
            "estagio_comercial": estagio,
            "etapa_repasse": etapa_repasse,
            "fila_atual": fila_atual,
            "status_geral": _process_geral_status(getattr(processo, "status_geral", None)),
            "status_credito": _credit_status(getattr(processo, "status_credito", None)),
            "status_cca": _process_caixa_status(getattr(processo, "status_cca", None)),
            "status_agehab": _process_agehab_status(getattr(processo, "status_agehab", None)),
            "status_sinal": _process_sinal_status(getattr(processo, "status_sinal", None)),
            "valor_sinal": float(processo.valor_sinal) if getattr(processo, "valor_sinal", None) is not None else None,
            "status_fiador": _process_fiador_status(getattr(processo, "status_fiador", None)),
            "recolha_fgts": _process_recolha_fgts_status(getattr(processo, "recolha_fgts", None)),
            "dias_em_aberto": dias_em_aberto,
            "data_cadastro_origem": data_referencia.isoformat() if data_referencia else None,
            "pendencias_documentos": pendencias_docs,
            "tem_pendencia_status": tem_pendencia_status,
            "tem_pendencia": tem_pendencia,
            "docs_total": docs_total,
            "docs_todos_aguardando": docs_todos_aguardando,
            "docs_pendentes_tooltip": docs_tooltip,
            "nao_contar_mes": nao_contar_mes,
            "pronto_para_repassar": pronto_para_repassar,
        }

        if nao_contar_mes:
            nao_contar_mes_total += 1
            clientes_excluidos_mes.append(cliente_item)
            continue

        total += 1
        clientes_estagios.append(cliente_item)
        if data_referencia and data_referencia >= start_7d:
            chegadas_ultimos_7_dias += 1
            clientes_chegadas_7d.append(cliente_item)

        if estagio == "EM_PROCESSO":
            total_comercial += 1
            clientes_comercial.append(cliente_item)
            sla_comercial_sum += _compute_sla_hours(processo, SLA_OWNER_CORRETOR, now)
            sla_comercial_count += 1
        else:
            total_credito += 1
            clientes_credito.append(cliente_item)
            sla_credito_sum += _compute_sla_hours(processo, SLA_OWNER_ANALISTA, now)
            sla_credito_count += 1
            sla_cca_sum += _compute_sla_hours(processo, SLA_OWNER_CCA, now)
            sla_cca_count += 1

        if estagio in ESTAGIOS_REPASSE_COMERCIAL:
            total_repasse += 1
            clientes_repasse.append(cliente_item)

        if status_cca in PROCESS_CCA_FINAL_STATUSES:
            total_assinados += 1
            clientes_assinados.append(cliente_item)
            assinado_data_ref = assinado_event_at_utc or updated_at_utc
            if assinado_data_ref is not None:
                dt_repasse = assinado_data_ref.date()
                if semana_inicio <= dt_repasse <= semana_fim_contagem and dt_repasse.weekday() < 5:
                    total_assinados_semana += 1
                    clientes_assinados_semana.append(cliente_item)

        if pronto_para_repassar:
            clientes_prontos_repasse.append(cliente_item)

        if estagio == "EM_PROCESSO" and dias_em_aberto is not None and dias_em_aberto > FALL_RISK_DAYS:
            provaveis_cair += 1
            clientes_provaveis_cair.append(cliente_item)

        if (
            status_geral in {"CANCELADO", "DISTRATO"}
            and (perda_event_at_utc or updated_at_utc) is not None
            and (perda_event_at_utc or updated_at_utc).year == now.year
            and (perda_event_at_utc or updated_at_utc).month == now.month
        ):
            perdas_mes += 1
            clientes_perdas_mes.append(cliente_item)

        imob_name = (getattr(cliente, "imobiliaria", None) or "Sem imobiliaria").strip() or "Sem imobiliaria"
        imob_map[imob_name] = imob_map.get(imob_name, 0) + 1
        corretor_name = (cliente.corretor or "Sem corretor").strip() or "Sem corretor"
        if imob_name not in imob_corretores:
            imob_corretores[imob_name] = {}
        imob_corretores[imob_name][corretor_name] = imob_corretores[imob_name].get(corretor_name, 0) + 1

    imobs = [
        {
            "nome": k,
            "total": v,
            "corretores": [
                {"nome": ck, "total": cv}
                for ck, cv in sorted(imob_corretores.get(k, {}).items(), key=lambda x: -x[1])
            ],
        }
        for k, v in sorted(imob_map.items(), key=lambda x: -x[1])
    ]

    hoje = now.date()
    dias_no_mes = calendar.monthrange(hoje.year, hoje.month)[1]
    dias_decorridos = max(1, hoje.day)
    meta_periodo_ano, meta_periodo_mes = hoje.year, hoje.month
    meta, meta_semanal, meta_source, meta_semanal_source = _resolve_gestor_meta_periodo(
        db,
        meta_periodo_ano,
        meta_periodo_mes,
    )
    real = total_assinados
    previsao = real + total
    dias_restantes = max(1, dias_no_mes - dias_decorridos)
    media_necessaria_dia = round(max(0, meta - real) / dias_restantes, 2)
    real_semanal = total_assinados_semana
    media_semana_dia_util = round(real_semanal / max(1, dias_uteis_decorridos_semana), 2)
    previsao_semanal = int(round(media_semana_dia_util * 5))
    media_necessaria_semana_dia_util = (
        round(max(0, meta_semanal - real_semanal) / dias_uteis_restantes_semana, 2)
        if dias_uteis_restantes_semana > 0
        else 0.0
    )

    sla_medio_comercial_horas = round(sla_comercial_sum / sla_comercial_count) if sla_comercial_count else 0
    sla_medio_credito_horas = round(sla_credito_sum / sla_credito_count) if sla_credito_count else 0
    sla_medio_cca_horas = round(sla_cca_sum / sla_cca_count) if sla_cca_count else 0
    media_chegadas_dia_7d = round(chegadas_ultimos_7_dias / 7, 2)
    projecao_chegadas_30_dias = int(round(media_chegadas_dia_7d * 30))

    clientes_por_fase: dict[str, list[dict[str, Any]]] = {
        "assinados": clientes_assinados,
        "conformidade_ok": clientes_prontos_repasse,
        "enviados_conformidade": clientes_comercial,
        "em_analise": clientes_credito,
        "com_pendencias": clientes_perdas_mes,
        "passiveis_cair": clientes_provaveis_cair,
        "nao_iniciado": clientes_excluidos_mes,
        "sla_comercial": clientes_comercial,
        "sla_credito": clientes_credito,
        "sla_cca": clientes_credito,
        "chegada": clientes_assinados,
        "media_necessaria": clientes_prontos_repasse,
        "comercial": clientes_comercial,
        "repasse": clientes_repasse,
        "realizados": clientes_assinados,
        "repasses_semana": clientes_assinados_semana,
        "perdas": clientes_perdas_mes,
    }

    return {
        "total": total,
        "total_bruto": total_bruto,
        "assinados": total_assinados,
        "conformidade_ok": len(clientes_prontos_repasse),
        "enviados_conformidade": total_comercial,
        "em_analise": total_credito,
        "com_pendencias": perdas_mes,
        "passiveis_cair": provaveis_cair,
        "nao_iniciado": nao_contar_mes_total,
        "processos_ativos": total,
        "sla_medio_comercial_horas": sla_medio_comercial_horas,
        "sla_medio_credito_horas": sla_medio_credito_horas,
        "sla_medio_cca_horas": sla_medio_cca_horas,
        "media_necessaria_dia": media_necessaria_dia,
        "chegadas_ultimos_7_dias": chegadas_ultimos_7_dias,
        "media_chegadas_dia_7d": media_chegadas_dia_7d,
        "projecao_chegadas_30_dias": projecao_chegadas_30_dias,
        "janela_chegadas_dias": 7,
        "janela_projecao_chegadas_dias": 30,
        "dias_estimativa_queda": FALL_RISK_DAYS,
        "clientes_por_fase": clientes_por_fase,
        "imobiliarias": imobs,
        "total_comercial": total_comercial,
        "total_repasse": total_repasse,
        "total_credito": total_credito,
        "provaveis_cair": provaveis_cair,
        "total_assinados": total_assinados,
        "meta": meta,
        "meta_semanal": meta_semanal,
        "meta_fonte": meta_source,
        "meta_semanal_fonte": meta_semanal_source,
        "meta_periodo_ano": meta_periodo_ano,
        "meta_periodo_mes": meta_periodo_mes,
        "real": real,
        "previsao": previsao,
        "real_semanal": real_semanal,
        "previsao_semanal": previsao_semanal,
        "media_necessaria_semana_dia_util": media_necessaria_semana_dia_util,
        "dias_uteis_decorridos_semana": dias_uteis_decorridos_semana,
        "dias_uteis_restantes_semana": dias_uteis_restantes_semana,
        "semana_inicio": semana_inicio.isoformat(),
        "semana_fim_util": semana_fim_util.isoformat(),
        "semana_fim_contagem": semana_fim_contagem.isoformat(),
        "perdas_mes": perdas_mes,
        "nao_contar_mes": nao_contar_mes_total,
        "clientes_estagios": clientes_estagios,
    }


@app.get("/app/api/gestor/meta", response_model=GestorMetaOut)
def app_get_gestor_meta(
    _: dict[str, Any] = Depends(require_roles(ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ANALISTA, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    ano, mes = _current_meta_period()
    meta, _, fonte_mensal, _ = _resolve_gestor_meta_periodo(db, ano, mes)
    return GestorMetaOut(meta=meta, fonte=fonte_mensal)


@app.put("/app/api/gestor/meta", response_model=GestorMetaOut)
def app_set_gestor_meta(
    payload: GestorMetaPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    ano, mes = _current_meta_period()
    monthly_period_key = _build_meta_period_runtime_key(META_MENSAL_RUNTIME_KEY, ano, mes)
    meta_value = max(0, int(payload.meta or 0))
    _set_runtime_meta(db, monthly_period_key, str(meta_value))
    _set_runtime_meta(db, META_MENSAL_RUNTIME_KEY, str(meta_value))
    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="analista_painel",
        acao="META_GESTOR_ATUALIZADA",
        entidade_tipo="configuracao",
        entidade_id=monthly_period_key,
        details=f"meta={meta_value};periodo={ano:04d}-{mes:02d}",
    )
    db.commit()
    meta, _, fonte, _ = _resolve_gestor_meta_periodo(db, ano, mes)
    return GestorMetaOut(meta=meta, fonte=fonte)


@app.get("/app/api/gestor/meta-periodo", response_model=GestorMetaPeriodoOut)
def app_get_gestor_meta_periodo(
    ano: Optional[int] = Query(default=None),
    mes: Optional[int] = Query(default=None),
    _: dict[str, Any] = Depends(require_roles(ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ANALISTA, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    ano_val, mes_val = _normalize_meta_period(ano, mes)
    meta_mensal, meta_semanal, fonte_mensal, fonte_semanal = _resolve_gestor_meta_periodo(db, ano_val, mes_val)
    return GestorMetaPeriodoOut(
        ano=ano_val,
        mes=mes_val,
        meta_mensal=meta_mensal,
        meta_semanal=meta_semanal,
        fonte_mensal=fonte_mensal,
        fonte_semanal=fonte_semanal,
    )


@app.put("/app/api/gestor/meta-periodo", response_model=GestorMetaPeriodoOut)
def app_set_gestor_meta_periodo(
    payload: GestorMetaPeriodoPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    ano_val, mes_val = _normalize_meta_period(payload.ano, payload.mes)
    meta_mensal = max(0, int(payload.meta_mensal or 0))
    meta_semanal = max(0, int(payload.meta_semanal or 0))

    monthly_period_key = _build_meta_period_runtime_key(META_MENSAL_RUNTIME_KEY, ano_val, mes_val)
    weekly_period_key = _build_meta_period_runtime_key(META_SEMANAL_RUNTIME_KEY, ano_val, mes_val)
    _set_runtime_meta(db, monthly_period_key, str(meta_mensal))
    _set_runtime_meta(db, weekly_period_key, str(meta_semanal))

    ano_atual, mes_atual = _current_meta_period()
    if ano_val == ano_atual and mes_val == mes_atual:
        _set_runtime_meta(db, META_MENSAL_RUNTIME_KEY, str(meta_mensal))
        _set_runtime_meta(db, META_SEMANAL_RUNTIME_KEY, str(meta_semanal))

    _record_system_log(
        db,
        actor_username=_normalize_username(str(session.get("username", ""))),
        actor_role=_normalize_role(str(session.get("role", ""))),
        tela="gestor_credito",
        acao="META_GESTOR_PERIODO_ATUALIZADA",
        entidade_tipo="configuracao",
        entidade_id=f"{ano_val:04d}-{mes_val:02d}",
        details=f"meta_mensal={meta_mensal};meta_semanal={meta_semanal}",
    )

    db.commit()
    meta_mensal_out, meta_semanal_out, fonte_mensal, fonte_semanal = _resolve_gestor_meta_periodo(db, ano_val, mes_val)
    return GestorMetaPeriodoOut(
        ano=ano_val,
        mes=mes_val,
        meta_mensal=meta_mensal_out,
        meta_semanal=meta_semanal_out,
        fonte_mensal=fonte_mensal,
        fonte_semanal=fonte_semanal,
    )


@app.get("/app/api/analista/planejamento", response_model=CreditoPlanejamentoDashboardOut)
def app_get_credito_planejamento_dashboard(
    dias: int = Query(default=10, ge=1, le=30),
    _: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    hoje = _utcnow().date()
    inicio_periodo = hoje - timedelta(days=2)
    fim_periodo = hoje + timedelta(days=max(1, dias))
    corte_recente = _utcnow() - timedelta(days=2)

    itens_db = (
        db.query(CreditoPlanejamentoItem)
        .filter(
            func.lower(func.coalesce(CreditoPlanejamentoItem.tipo, "")) != "anotacao",
            or_(
                and_(
                    func.lower(func.coalesce(CreditoPlanejamentoItem.tipo, "")) == "subtarefa",
                    or_(
                        CreditoPlanejamentoItem.data_referencia.is_(None),
                        CreditoPlanejamentoItem.data_referencia <= fim_periodo,
                    ),
                ),
                and_(
                    CreditoPlanejamentoItem.data_referencia.isnot(None),
                    CreditoPlanejamentoItem.data_referencia >= inicio_periodo,
                    CreditoPlanejamentoItem.data_referencia <= fim_periodo,
                ),
                and_(
                    CreditoPlanejamentoItem.urgente.is_(True),
                    func.lower(func.coalesce(CreditoPlanejamentoItem.status, "")) != "concluido",
                ),
                and_(
                    CreditoPlanejamentoItem.data_referencia.is_(None),
                    CreditoPlanejamentoItem.updated_at >= corte_recente,
                ),
            ),
        )
        .order_by(CreditoPlanejamentoItem.updated_at.desc())
        .limit(300)
        .all()
    )
    anotacoes_db = (
        db.query(CreditoPlanejamentoItem)
        .filter(func.lower(func.coalesce(CreditoPlanejamentoItem.tipo, "")) == "anotacao")
        .order_by(CreditoPlanejamentoItem.updated_at.desc())
        .limit(30)
        .all()
    )

    itens_out = [_credito_planejamento_item_out(item) for item in itens_db]
    itens_out.sort(
        key=lambda item: (
            0 if item.urgente else 1,
            item.data_referencia or date(9999, 12, 31),
            item.hora_inicio or "99:99",
            0 if item.status != "concluido" else 1,
            item.titulo.lower(),
        )
    )

    tarefas_dia = sorted(
        [item for item in itens_out if item.data_referencia == hoje and item.tipo in {"tarefa", "urgente"}],
        key=lambda item: (item.hora_inicio or "99:99", item.titulo.lower()),
    )
    agendamentos_dia = sorted(
        [item for item in itens_out if item.data_referencia == hoje and item.tipo == "agendamento"],
        key=lambda item: (item.hora_inicio or "99:99", item.titulo.lower()),
    )
    entregas_dia = sorted(
        [item for item in itens_out if item.data_referencia == hoje and item.tipo == "entrega"],
        key=lambda item: (0 if item.status != "concluido" else 1, item.hora_inicio or "99:99", item.titulo.lower()),
    )
    urgentes = sorted(
        [item for item in itens_out if item.urgente and item.status != "concluido"],
        key=lambda item: (
            0 if item.data_referencia == hoje else 1,
            item.data_referencia or date(9999, 12, 31),
            item.hora_inicio or "99:99",
            item.titulo.lower(),
        ),
    )[:18]

    evolucao_map: dict[str, dict[str, float]] = {}
    for item in itens_out:
        responsavel = _normalize_credito_planejamento_text(item.responsavel, max_len=120) or "Sem responsavel"
        bucket = evolucao_map.setdefault(
            responsavel,
            {"total": 0.0, "concluidas": 0.0, "pendentes": 0.0, "progresso_total": 0.0},
        )
        bucket["total"] += 1
        progresso = _credito_planejamento_progresso(item.progresso)
        if item.status == "concluido":
            bucket["concluidas"] += 1
            progresso = 100
        else:
            bucket["pendentes"] += 1
        bucket["progresso_total"] += progresso

    evolucao_time: list[CreditoPlanejamentoEvolucaoOut] = []
    for responsavel, bucket in evolucao_map.items():
        total = int(bucket["total"])
        concluidas = int(bucket["concluidas"])
        pendentes = int(bucket["pendentes"])
        progresso_medio = int(round((bucket["progresso_total"] / total), 0)) if total else 0
        taxa = round((concluidas / total) * 100, 1) if total else 0.0
        evolucao_time.append(
            CreditoPlanejamentoEvolucaoOut(
                responsavel=responsavel,
                total=total,
                concluidas=concluidas,
                pendentes=pendentes,
                progresso_medio=progresso_medio,
                taxa_conclusao=taxa,
            )
        )
    evolucao_time.sort(key=lambda row: (-row.taxa_conclusao, -row.total, row.responsavel.lower()))
    anotacoes_out = [_credito_planejamento_item_out(item) for item in anotacoes_db]
    pendentes_total = sum(1 for item in itens_out if item.status != "concluido")

    return CreditoPlanejamentoDashboardOut(
        referencia=hoje,
        pendentes_total=pendentes_total,
        tarefas_dia=tarefas_dia,
        agendamentos_dia=agendamentos_dia,
        entregas_dia=entregas_dia,
        urgentes=urgentes,
        evolucao_time=evolucao_time,
        anotacoes=anotacoes_out,
        itens=itens_out,
    )


@app.post("/app/api/frankstein/agenda/email-alertas/processar")
def app_process_frankstein_agenda_email_alerts(
    request: Request,
    token: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    supplied_token = (token or request.headers.get("x-siocred-alert-token") or "").strip()
    if SIOCRED_EMAIL_ALERT_TOKEN and supplied_token != SIOCRED_EMAIL_ALERT_TOKEN:
        raise HTTPException(status_code=403, detail="Token de alerta invalido.")
    return _processar_alertas_email_planejamento(db)


@app.post("/app/api/analista/planejamento/itens", response_model=CreditoPlanejamentoItemOut)
def app_create_credito_planejamento_item(
    payload: CreditoPlanejamentoItemCreate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))

    tipo = _credito_planejamento_tipo(payload.tipo, fallback="tarefa")
    titulo = _normalize_credito_planejamento_text(payload.titulo, max_len=180)
    if not titulo:
        raise HTTPException(status_code=422, detail="Titulo obrigatorio.")
    descricao = _normalize_credito_planejamento_text(payload.descricao, max_len=2200)
    responsavel = _normalize_credito_planejamento_text(payload.responsavel, max_len=120)
    status = _credito_planejamento_status(payload.status, fallback="pendente")
    progresso = _credito_planejamento_progresso(payload.progresso)
    hora_inicio = _credito_planejamento_hora(payload.hora_inicio, field_name="Hora inicio")
    hora_fim = _credito_planejamento_hora(payload.hora_fim, field_name="Hora fim")

    if hora_inicio and hora_fim and hora_fim < hora_inicio:
        raise HTTPException(status_code=422, detail="Hora fim nao pode ser menor que hora inicio.")
    urgente = bool(payload.urgente) or tipo == "urgente"
    if status == "concluido":
        progresso = 100
    if tipo == "anotacao":
        status = "pendente"
        progresso = 0

    item = CreditoPlanejamentoItem(
        tipo=tipo,
        titulo=titulo,
        descricao=descricao,
        responsavel=responsavel,
        data_referencia=payload.data_referencia,
        hora_inicio=hora_inicio,
        hora_fim=hora_fim,
        status=status,
        progresso=progresso,
        urgente=urgente,
        created_by_username=actor_username or None,
        updated_by_username=actor_username or None,
    )
    db.add(item)
    db.flush()
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_painel",
        acao="CREDITO_PLANEJAMENTO_ITEM_CRIADO",
        entidade_tipo="planejamento",
        entidade_id=str(item.id),
        details=(
            f"tipo={item.tipo}; titulo={item.titulo}; status={item.status}; "
            f"progresso={item.progresso}; urgente={item.urgente}"
        ),
    )
    db.commit()
    db.refresh(item)
    return _credito_planejamento_item_out(item)


@app.patch("/app/api/analista/planejamento/itens/{item_id}", response_model=CreditoPlanejamentoItemOut)
def app_update_credito_planejamento_item(
    item_id: uuid.UUID,
    payload: CreditoPlanejamentoItemUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    item = db.get(CreditoPlanejamentoItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item nao encontrado.")

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    changes = payload.model_dump(exclude_unset=True)
    if not changes:
        return _credito_planejamento_item_out(item)

    if "tipo" in changes:
        item.tipo = _credito_planejamento_tipo(changes.get("tipo"), fallback=_credito_planejamento_tipo(item.tipo, fallback="tarefa"))
    if "titulo" in changes:
        titulo = _normalize_credito_planejamento_text(changes.get("titulo"), max_len=180)
        if not titulo:
            raise HTTPException(status_code=422, detail="Titulo obrigatorio.")
        item.titulo = titulo
    if "descricao" in changes:
        item.descricao = _normalize_credito_planejamento_text(changes.get("descricao"), max_len=2200)
    if "responsavel" in changes:
        item.responsavel = _normalize_credito_planejamento_text(changes.get("responsavel"), max_len=120)
    if "data_referencia" in changes:
        item.data_referencia = changes.get("data_referencia")
    if "hora_inicio" in changes:
        item.hora_inicio = _credito_planejamento_hora(changes.get("hora_inicio"), field_name="Hora inicio")
    if "hora_fim" in changes:
        item.hora_fim = _credito_planejamento_hora(changes.get("hora_fim"), field_name="Hora fim")
    if item.hora_inicio and item.hora_fim and item.hora_fim < item.hora_inicio:
        raise HTTPException(status_code=422, detail="Hora fim nao pode ser menor que hora inicio.")
    if "status" in changes:
        item.status = _credito_planejamento_status(changes.get("status"), fallback=_credito_planejamento_status(item.status))
    if "progresso" in changes:
        item.progresso = _credito_planejamento_progresso(changes.get("progresso"))
    if "urgente" in changes:
        item.urgente = bool(changes.get("urgente"))

    if _credito_planejamento_tipo(item.tipo, fallback="tarefa") == "urgente":
        item.urgente = True
    if _credito_planejamento_status(item.status, fallback="pendente") == "concluido":
        item.progresso = 100
    if _credito_planejamento_tipo(item.tipo, fallback="tarefa") == "anotacao":
        item.status = "pendente"
        item.progresso = 0

    item.updated_by_username = actor_username or None
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_painel",
        acao="CREDITO_PLANEJAMENTO_ITEM_ATUALIZADO",
        entidade_tipo="planejamento",
        entidade_id=str(item.id),
        details=(
            f"tipo={item.tipo}; titulo={item.titulo}; status={item.status}; "
            f"progresso={item.progresso}; urgente={item.urgente}"
        ),
    )
    db.commit()
    db.refresh(item)
    return _credito_planejamento_item_out(item)


@app.delete("/app/api/analista/planejamento/itens/{item_id}")
def app_delete_credito_planejamento_item(
    item_id: uuid.UUID,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    item = db.get(CreditoPlanejamentoItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item nao encontrado.")

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_painel",
        acao="CREDITO_PLANEJAMENTO_ITEM_REMOVIDO",
        entidade_tipo="planejamento",
        entidade_id=str(item.id),
        details=f"tipo={item.tipo}; titulo={item.titulo}; status={item.status}; urgente={item.urgente}",
    )
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/app/api/analista/reuniao-comercial", response_model=AnalistaReuniaoComercialDashboardOut)
def app_get_analista_reuniao_comercial_dashboard(
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    limit: int = Query(default=400, ge=1, le=1000),
):
    _ensure_process_archiving(db, _utcnow())
    referencia = _utcnow().date()
    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))

    query = (
        db.query(Processo, Cliente)
        .join(Cliente, Processo.cliente_id == Cliente.id)
        .filter(
            _processos_ativos_clause(),
            func.upper(func.coalesce(Processo.estagio_comercial, "")).in_(tuple(ANALISTA_REUNIAO_ESTAGIOS)),
        )
    )
    if role == ROLE_CCA:
        if not username:
            return AnalistaReuniaoComercialDashboardOut(
                referencia=referencia,
                total_clientes=0,
                assinados=0,
                seguir=0,
                finalizar_hoje=0,
                entrega_hoje=0,
                risco_queda=0,
                solicitar_cancelamento=0,
                clientes_entrega_hoje=[],
                clientes=[],
            )
        query = query.filter(func.lower(func.trim(func.coalesce(Processo.cca_responsavel, ""))) == username)

    rows = query.order_by(Processo.created_at.asc()).limit(limit).all()
    processo_ids = [processo.id for processo, _ in rows]

    followup_map: dict[uuid.UUID, AnalistaReuniaoComercial] = {}
    compromissos_map: dict[uuid.UUID, list[AnalistaReuniaoCompromisso]] = {}
    if processo_ids:
        followups = (
            db.query(AnalistaReuniaoComercial)
            .filter(AnalistaReuniaoComercial.processo_id.in_(processo_ids))
            .all()
        )
        followup_map = {item.processo_id: item for item in followups}

        compromissos = (
            db.query(AnalistaReuniaoCompromisso)
            .filter(AnalistaReuniaoCompromisso.processo_id.in_(processo_ids))
            .order_by(AnalistaReuniaoCompromisso.data_prometida.desc(), AnalistaReuniaoCompromisso.created_at.desc())
            .all()
        )
        for compromisso in compromissos:
            compromissos_map.setdefault(compromisso.processo_id, []).append(compromisso)

    clientes_out: list[AnalistaReuniaoClienteOut] = []
    for processo, cliente in rows:
        cliente_out = _analista_reuniao_cliente_out(
            processo,
            cliente,
            followup_map.get(processo.id),
            compromissos_map.get(processo.id, []),
            referencia=referencia,
        )
        clientes_out.append(cliente_out)

    clientes_out.sort(key=lambda row: (-row.sla_dias, row.cliente_nome.lower()))

    assinados = sum(1 for row in clientes_out if row.status_followup == "assinado")
    seguir = sum(1 for row in clientes_out if row.status_followup == "seguir")
    finalizar_hoje = sum(1 for row in clientes_out if row.status_followup == "finalizar_hoje")
    entrega_hoje = sum(1 for row in clientes_out if row.entrega_hoje)
    risco_queda = sum(1 for row in clientes_out if row.probabilidade_queda)
    solicitar_cancelamento = sum(1 for row in clientes_out if row.solicitar_cancelamento)
    clientes_entrega_hoje = [row.cliente_nome for row in clientes_out if row.entrega_hoje]

    return AnalistaReuniaoComercialDashboardOut(
        referencia=referencia,
        total_clientes=len(clientes_out),
        assinados=assinados,
        seguir=seguir,
        finalizar_hoje=finalizar_hoje,
        entrega_hoje=entrega_hoje,
        risco_queda=risco_queda,
        solicitar_cancelamento=solicitar_cancelamento,
        clientes_entrega_hoje=clientes_entrega_hoje,
        clientes=clientes_out,
    )


@app.patch("/app/api/analista/reuniao-comercial/{processo_id}", response_model=AnalistaReuniaoClienteOut)
def app_update_analista_reuniao_comercial_item(
    processo_id: uuid.UUID,
    payload: AnalistaReuniaoComercialUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado.")

    cliente = db.get(Cliente, processo.cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente do processo nao encontrado.")

    estagio = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
    if estagio not in ANALISTA_REUNIAO_ESTAGIOS:
        raise HTTPException(status_code=422, detail="Processo fora das fases da reuniao comercial.")

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    changes = payload.model_dump(exclude_unset=True)
    ref_now = _utcnow()

    item = db.get(AnalistaReuniaoComercial, processo_id)
    if not item:
        item = AnalistaReuniaoComercial(
            processo_id=processo_id,
            conta_no_mes=not _is_nao_contar_mes_active(processo, ref_now),
        )
        db.add(item)
        db.flush()

    if "conta_no_mes" in changes:
        conta_no_mes = bool(changes.get("conta_no_mes"))
        item.conta_no_mes = conta_no_mes
        _set_nao_contar_mes_period(processo, not conta_no_mes, ref_now)
    if "data_prevista_entrega" in changes:
        item.data_prevista_entrega = changes.get("data_prevista_entrega")
    if "probabilidade_queda" in changes:
        item.probabilidade_queda = bool(changes.get("probabilidade_queda"))
    if "solicitar_cancelamento" in changes:
        item.solicitar_cancelamento = bool(changes.get("solicitar_cancelamento"))
    if item.solicitar_cancelamento:
        _set_nao_contar_mes_period(processo, True, ref_now)
    if "status_followup" in changes:
        item.status_followup = _analista_reuniao_followup_status(changes.get("status_followup"), fallback=item.status_followup or "seguir")
    if "observacao" in changes:
        item.observacao = _normalize_credito_planejamento_text(changes.get("observacao"), max_len=2500)
    if "justificativa_reincidencia" in changes:
        item.justificativa_reincidencia = _normalize_credito_planejamento_text(changes.get("justificativa_reincidencia"), max_len=2000)

    item.conta_no_mes = (not _is_nao_contar_mes_active(processo, ref_now)) and not item.solicitar_cancelamento
    item.updated_by_username = actor_username or None
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_reuniao_comercial",
        acao="REUNIAO_COMERCIAL_ITEM_ATUALIZADO",
        entidade_tipo="processo",
        entidade_id=str(processo_id),
        details=(
            f"conta_no_mes={item.conta_no_mes}; status_followup={item.status_followup}; "
            f"entrega_prevista={item.data_prevista_entrega or '-'}; risco_queda={item.probabilidade_queda}; "
            f"cancelamento={item.solicitar_cancelamento}"
        ),
    )

    db.commit()
    _invalidate_process_list_cache()

    compromissos = (
        db.query(AnalistaReuniaoCompromisso)
        .filter(AnalistaReuniaoCompromisso.processo_id == processo_id)
        .order_by(AnalistaReuniaoCompromisso.data_prometida.desc(), AnalistaReuniaoCompromisso.created_at.desc())
        .all()
    )
    return _analista_reuniao_cliente_out(processo, cliente, item, compromissos, referencia=ref_now.date())


@app.post("/app/api/analista/reuniao-comercial/{processo_id}/compromissos", response_model=AnalistaReuniaoCompromissoOut)
def app_create_analista_reuniao_compromisso(
    processo_id: uuid.UUID,
    payload: AnalistaReuniaoCompromissoCreate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado.")

    estagio = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
    if estagio not in ANALISTA_REUNIAO_ESTAGIOS:
        raise HTTPException(status_code=422, detail="Processo fora das fases da reuniao comercial.")

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    observacao = _normalize_credito_planejamento_text(payload.observacao, max_len=900)
    compromisso = AnalistaReuniaoCompromisso(
        processo_id=processo_id,
        data_prometida=payload.data_prometida,
        status="pendente",
        observacao=observacao,
        created_by_username=actor_username or None,
        updated_by_username=actor_username or None,
    )
    db.add(compromisso)
    db.flush()
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_reuniao_comercial",
        acao="REUNIAO_COMERCIAL_COMPROMISSO_CRIADO",
        entidade_tipo="processo",
        entidade_id=str(processo_id),
        details=f"data_prometida={payload.data_prometida.isoformat()}; status=pendente",
    )
    db.commit()
    db.refresh(compromisso)
    return _analista_reuniao_compromisso_out(compromisso)


@app.post("/app/api/analista/reuniao-comercial/compromissos/{compromisso_id}/nao-entregue")
def app_mark_analista_reuniao_compromisso_nao_entregue(
    compromisso_id: uuid.UUID,
    payload: AnalistaReuniaoCompromissoNaoEntreguePayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    compromisso = db.get(AnalistaReuniaoCompromisso, compromisso_id)
    if not compromisso:
        raise HTTPException(status_code=404, detail="Compromisso nao encontrado.")

    justificativa = _normalize_credito_planejamento_text(payload.justificativa, max_len=1400)
    if not justificativa:
        raise HTTPException(status_code=422, detail="Justificativa obrigatoria para nao entregue.")

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))

    compromisso.status = "nao_entregue"
    compromisso.justificativa = justificativa
    compromisso.updated_by_username = actor_username or None

    novo_compromisso_id: Optional[uuid.UUID] = None
    if payload.nova_data_prometida:
        novo = AnalistaReuniaoCompromisso(
            processo_id=compromisso.processo_id,
            data_prometida=payload.nova_data_prometida,
            status="pendente",
            observacao=_normalize_credito_planejamento_text(payload.nova_observacao, max_len=900),
            created_by_username=actor_username or None,
            updated_by_username=actor_username or None,
        )
        db.add(novo)
        db.flush()
        novo_compromisso_id = novo.id

    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_reuniao_comercial",
        acao="REUNIAO_COMERCIAL_COMPROMISSO_NAO_ENTREGUE",
        entidade_tipo="processo",
        entidade_id=str(compromisso.processo_id),
        details=(
            f"compromisso_id={compromisso.id}; justificativa={justificativa}; "
            f"nova_data={payload.nova_data_prometida or '-'}; novo_compromisso_id={novo_compromisso_id or '-'}"
        ),
    )
    db.commit()
    return {"ok": True, "novo_compromisso_id": str(novo_compromisso_id) if novo_compromisso_id else None}


@app.post("/app/api/analista/reuniao-comercial/compromissos/{compromisso_id}/entregue")
def app_mark_analista_reuniao_compromisso_entregue(
    compromisso_id: uuid.UUID,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    compromisso = db.get(AnalistaReuniaoCompromisso, compromisso_id)
    if not compromisso:
        raise HTTPException(status_code=404, detail="Compromisso nao encontrado.")

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    compromisso.status = "entregue"
    compromisso.updated_by_username = actor_username or None
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="analista_reuniao_comercial",
        acao="REUNIAO_COMERCIAL_COMPROMISSO_ENTREGUE",
        entidade_tipo="processo",
        entidade_id=str(compromisso.processo_id),
        details=f"compromisso_id={compromisso.id}",
    )
    db.commit()
    return {"ok": True}


@app.post("/app/api/processos/intake")
def app_create_intake(
    payload: ProcessoIntakeCreate,
    session: dict[str, Any] = Depends(require_roles(*PROCESSO_INTAKE_ROLES)),
    db: Session = Depends(get_db),
):
    username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    obra_nome = _resolve_empreendimento_nome(db, payload.obra)
    if payload.obra and not obra_nome:
        raise HTTPException(status_code=422, detail="Empreendimento invalido. Selecione um empreendimento cadastrado.")
    estagio = _process_estagio_comercial(payload.estagio_comercial, fallback="RESERVA")
    corretor_nome = (
        _normalize_corretor_nome_curto(username)
        if actor_role == ROLE_CORRETOR
        else (_normalize_corretor_nome_curto(payload.corretor) if payload.corretor else None)
    )

    cliente = Cliente(
        nome=payload.nome.strip(),
        corretor=corretor_nome,
        obra=obra_nome,
        imobiliaria=(payload.imobiliaria or "").strip() or None,
        data_reserva_origem=payload.data_reserva_origem,
        data_cadastro_origem=payload.data_cadastro_origem,
    )
    db.add(cliente)
    db.commit()
    db.refresh(cliente)

    processo = Processo(
        cliente_id=cliente.id,
        estagio_comercial=estagio,
        etapa_repasse="EM_REPASSE" if estagio in ESTAGIOS_REPASSE_COMERCIAL else None,
        valor_cheque_moradia=_resolve_cheque_moradia_valor(db, obra_nome),
        sla_comercial_inicio_at=_utc_start_of_day(payload.data_cadastro_origem) or _utcnow(),
    )
    _sync_estagio_repasse_rules(processo, _utcnow())
    _refresh_sla_fixed_markers(processo, _utcnow())
    _switch_sla_owner(processo, SLA_OWNER_ANALISTA, _utcnow())
    db.add(processo)
    db.commit()
    db.refresh(processo)
    _record_processo_event(
        db,
        processo_id=processo.id,
        actor_username=username,
        actor_role=actor_role,
        event_type="PROCESSO_CRIADO",
        details=f"Cliente: {cliente.nome}",
    )
    _record_system_log(
        db,
        actor_username=username,
        actor_role=actor_role,
        tela="checklist",
        acao="PROCESSO_CRIADO",
        entidade_tipo="processo",
        entidade_id=str(processo.id),
        details=(
            f"cliente={cliente.nome}; corretor={cliente.corretor or '-'}; obra={cliente.obra or '-'}; "
            f"imobiliaria={cliente.imobiliaria or '-'}; estagio={processo.estagio_comercial}"
        ),
    )
    db.commit()

    _ensure_default_documentos(db, processo.id)
    _invalidate_process_list_cache()
    return {"ok": True, "cliente_id": str(cliente.id), "processo_id": str(processo.id)}


@app.post("/app/api/processos/importar", response_model=ImportPlanilhaOut)
async def app_importar_processos_planilha(
    file: UploadFile = File(...),
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    filename = (file.filename or "").strip()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=422, detail="Arquivo vazio.")

    raw_rows = _load_import_rows(filename, content)
    provided_columns: set[str] = set()
    for row in raw_rows:
        provided_columns.update(row.keys())

    missing = sorted(IMPORT_REQUIRED_COLUMNS - provided_columns)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Colunas obrigatorias ausentes: {', '.join(missing)}",
        )

    actor_username = _normalize_username(str(session.get("username", "")))
    actor_role = _normalize_role(str(session.get("role", "")))
    existing_keys = {
        _normalize_cliente_key(name)
        for (name,) in db.query(Cliente.nome).all()
        if _normalize_cliente_key(name)
    }
    created_keys_in_batch: set[str] = set()

    resultados: list[ImportPlanilhaRowOut] = []
    importados = 0
    ignorados_existentes = 0
    invalidados = 0

    for row_number, row in enumerate(raw_rows, start=2):
        nome = " ".join(str(row.get("nome_cliente") or "").strip().split())
        stage_raw = row.get("estagio")
        estagio = _process_estagio_comercial(stage_raw, fallback="")
        reserva_raw = row.get("reserva")
        data_reserva_origem = _parse_import_date(reserva_raw)
        cadastro_raw = row.get("data_cadastro")
        data_cadastro_origem = _parse_import_date(cadastro_raw)

        if not nome:
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=None,
                    status="invalido",
                    motivo="nome_cliente vazio",
                )
            )
            continue

        if not estagio:
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="invalido",
                    motivo=f"estagio invalido: {stage_raw}",
                )
            )
            continue

        if not str(reserva_raw or "").strip():
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="invalido",
                    motivo="reserva obrigatoria",
                )
            )
            continue

        if data_reserva_origem is None:
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="invalido",
                    motivo="reserva invalida",
                )
            )
            continue

        if not str(cadastro_raw or "").strip():
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="invalido",
                    motivo="data_cadastro obrigatoria",
                )
            )
            continue

        if data_cadastro_origem is None:
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="invalido",
                    motivo="data_cadastro invalida",
                )
            )
            continue

        cliente_key = _normalize_cliente_key(nome)
        if not cliente_key:
            invalidados += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="invalido",
                    motivo="nome_cliente invalido",
                )
            )
            continue

        if cliente_key in existing_keys or cliente_key in created_keys_in_batch:
            ignorados_existentes += 1
            resultados.append(
                ImportPlanilhaRowOut(
                    linha=row_number,
                    nome_cliente=nome,
                    status="ignorado_existente",
                    motivo="cliente ja existe na base; importacao ignorada (sem substituicao)",
                )
            )
            continue

        empreendimento_raw = _normalize_empreendimento_nome(row.get("empreendimento"))
        empreendimento_resolvido = _resolve_empreendimento_nome(db, empreendimento_raw) if empreendimento_raw else None
        empreendimento = empreendimento_resolvido or empreendimento_raw or None

        cliente = Cliente(
            nome=nome,
            corretor=_normalize_corretor_nome_curto(row.get("corretor")) or None,
            obra=empreendimento,
            imobiliaria=" ".join(str(row.get("imobiliaria") or "").strip().split()) or None,
            data_reserva_origem=data_reserva_origem,
            data_cadastro_origem=data_cadastro_origem,
        )
        db.add(cliente)
        db.flush()

        processo = Processo(
            cliente_id=cliente.id,
            estagio_comercial=estagio,
            etapa_repasse="EM_REPASSE" if estagio in ESTAGIOS_REPASSE_COMERCIAL else None,
            valor_cheque_moradia=_resolve_cheque_moradia_valor(db, empreendimento),
            sla_comercial_inicio_at=_utc_start_of_day(data_cadastro_origem) or _utcnow(),
        )
        _sync_estagio_repasse_rules(processo, _utcnow())
        _refresh_sla_fixed_markers(processo, _utcnow())
        _switch_sla_owner(processo, SLA_OWNER_ANALISTA, _utcnow())
        db.add(processo)
        db.flush()

        _record_processo_event(
            db,
            processo_id=processo.id,
            actor_username=actor_username,
            actor_role=actor_role,
            event_type="PROCESSO_IMPORTADO_PLANILHA",
            details=f"cliente={cliente.nome}; estagio={processo.estagio_comercial}",
        )
        _record_system_log(
            db,
            actor_username=actor_username,
            actor_role=actor_role,
            tela="analista_painel",
            acao="IMPORTACAO_PLANILHA_ITEM",
            entidade_tipo="processo",
            entidade_id=str(processo.id),
            details=(
                f"linha={row_number}; cliente={cliente.nome}; estagio={processo.estagio_comercial}; "
                f"empreendimento={cliente.obra or '-'}"
            ),
        )
        _ensure_default_documentos(db, processo.id, autocommit=False)

        created_keys_in_batch.add(cliente_key)
        importados += 1
        resultados.append(
            ImportPlanilhaRowOut(
                linha=row_number,
                nome_cliente=nome,
                status="importado",
                motivo=None,
            )
        )

    db.commit()
    _invalidate_process_list_cache()

    return ImportPlanilhaOut(
        ok=True,
        total_linhas=len(raw_rows),
        importados=importados,
        ignorados_existentes=ignorados_existentes,
        invalidados=invalidados,
        resultados=resultados,
    )


@app.get("/app/api/processos/{processo_id}/full", response_model=ProcessoFullOut)
def app_get_processo_full(
    processo_id: uuid.UUID,
    session: dict[str, Any] = Depends(require_roles(*PROCESSO_FULL_ROLES)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    cliente = db.get(Cliente, processo.cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente nao encontrado")

    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))
    if role == ROLE_CORRETOR and _normalize_username(cliente.corretor) != username:
        raise HTTPException(status_code=403, detail="Sem permissao para acessar este processo")
    if role == ROLE_CCA and _normalize_username(processo.cca_responsavel) != username:
        raise HTTPException(status_code=403, detail="Sem permissao para acessar este processo")

    _ensure_default_documentos(db, processo.id)
    documentos = (
        db.query(Documento)
        .filter(Documento.processo_id == processo.id)
        .order_by(Documento.categoria.asc(), Documento.nome.asc())
        .all()
    )

    now = _utcnow()
    processo_out = ProcessoOut.model_validate(processo)
    processo_out.sla_corretor_seconds = _compute_sla_seconds(processo, SLA_OWNER_CORRETOR, now)
    processo_out.sla_analista_seconds = _compute_sla_seconds(processo, SLA_OWNER_ANALISTA, now)
    processo_out.sla_cca_seconds = _compute_sla_seconds(processo, SLA_OWNER_CCA, now)
    processo_out.sla_corretor_horas = processo_out.sla_corretor_seconds // 3600
    processo_out.sla_analista_horas = processo_out.sla_analista_seconds // 3600
    processo_out.sla_cca_horas = processo_out.sla_cca_seconds // 3600
    processo_out.sla_corretor_dias = processo_out.sla_corretor_horas // 24
    processo_out.sla_credito_dias = processo_out.sla_analista_horas // 24
    processo_out.sla_cca_dias = processo_out.sla_cca_horas // 24
    processo_out.sla_owner = _normalize_sla_owner(processo.sla_owner)
    processo_out.sla_active_since = _as_utc(processo.sla_active_since)
    processo_out.status_cca = _process_caixa_status(processo.status_cca)
    processo_out.status_agehab = _process_agehab_status(processo.status_agehab)
    processo_out.recolha_fgts = _process_recolha_fgts_status(getattr(processo, "recolha_fgts", None))
    processo_out.estagio_comercial = _process_estagio_comercial(processo.estagio_comercial)
    processo_out.etapa_repasse = _process_etapa_repasse(processo.etapa_repasse)
    processo_out.fila_atual = _fila_atual_from_processo(processo)
    processo_out.nao_contar_mes = _is_nao_contar_mes_active(processo, now)

    return ProcessoFullOut(
        processo=processo_out,
        cliente=ClienteOut.model_validate(cliente),
        documentos=[DocumentoOut.model_validate(doc) for doc in documentos],
    )


@app.get("/app/api/processos/{processo_id}/eventos", response_model=list[ProcessoEventoOut])
def app_list_processo_eventos(
    processo_id: uuid.UUID,
    session: dict[str, Any] = Depends(require_roles(ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
    limit: int = Query(default=200, ge=1, le=1000),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    cliente = db.get(Cliente, processo.cliente_id)
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente nao encontrado")

    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))
    if role == ROLE_CORRETOR and _normalize_username(cliente.corretor) != username:
        raise HTTPException(status_code=403, detail="Sem permissao para acessar este processo")
    if role == ROLE_CCA and _normalize_username(processo.cca_responsavel) != username:
        raise HTTPException(status_code=403, detail="Sem permissao para acessar este processo")

    return (
        db.query(ProcessoEvento)
        .filter(ProcessoEvento.processo_id == processo_id)
        .order_by(ProcessoEvento.created_at.desc())
        .limit(limit)
        .all()
    )


@app.get("/app/api/metricas/processos", response_model=ProcessoMetricasOut)
def app_get_metricas_processos(
    session: dict[str, Any] = Depends(require_roles(ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    _ensure_process_archiving(db, _utcnow())
    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))
    processos = _query_processos_by_scope(db, role, username).order_by(Processo.created_at.desc()).all()

    total = len(processos)
    funnel = {
        "novo": 0,
        "em_andamento": 0,
        "pendenciado": 0,
        "aprovado": 0,
        "reprovado": 0,
        "distrato": 0,
        "cancelado": 0,
    }
    sla = {
        "analista_alerta_24h": 0,
        "analista_critico_48h": 0,
        "corretor_alerta_24h": 0,
        "corretor_critico_48h": 0,
        "cca_alerta_24h": 0,
        "cca_critico_48h": 0,
    }

    now = _utcnow()
    process_ids: list[uuid.UUID] = []
    pendencias_atuais: set[uuid.UUID] = set()
    for processo in processos:
        process_ids.append(processo.id)
        geral = _status_token(processo.status_geral).lower()
        if geral in funnel:
            funnel[geral] += 1

        sla_analista = _compute_sla_hours(processo, SLA_OWNER_ANALISTA, now)
        sla_corretor = _compute_sla_hours(processo, SLA_OWNER_CORRETOR, now)
        sla_cca = _compute_sla_hours(processo, SLA_OWNER_CCA, now)

        if sla_analista >= 24:
            sla["analista_alerta_24h"] += 1
        if sla_analista >= 48:
            sla["analista_critico_48h"] += 1
        if sla_corretor >= 24:
            sla["corretor_alerta_24h"] += 1
        if sla_corretor >= 48:
            sla["corretor_critico_48h"] += 1
        if sla_cca >= 24:
            sla["cca_alerta_24h"] += 1
        if sla_cca >= 48:
            sla["cca_critico_48h"] += 1

        if _processo_has_pendencia(processo):
            pendencias_atuais.add(processo.id)

    pendencias_historicas: set[uuid.UUID] = set()
    total_eventos_pendencia = 0
    if process_ids:
        pendencia_rows = (
            db.query(ProcessoEvento.processo_id, func.count(ProcessoEvento.id))
            .filter(ProcessoEvento.processo_id.in_(process_ids), ProcessoEvento.event_type == "PENDENCIA")
            .group_by(ProcessoEvento.processo_id)
            .all()
        )
        for processo_id, count in pendencia_rows:
            pendencias_historicas.add(processo_id)
            total_eventos_pendencia += int(count or 0)

    processos_com_pendencia = len(pendencias_atuais.union(pendencias_historicas))
    first_pass_yield = round(((total - processos_com_pendencia) / total) * 100, 2) if total > 0 else 0.0
    media_retrabalho = round(total_eventos_pendencia / total, 2) if total > 0 else 0.0

    return ProcessoMetricasOut(
        total_processos=total,
        funnel=ProcessoMetricasFunnelOut(**funnel),
        sla=ProcessoMetricasSlaOut(**sla),
        qualidade=ProcessoMetricasQualidadeOut(
            processos_com_pendencia=processos_com_pendencia,
            first_pass_yield_percent=first_pass_yield,
            media_retrabalho_por_processo=media_retrabalho,
        ),
        updated_at=now,
    )


@app.patch("/app/api/processos/{processo_id}", response_model=ProcessoOut)
def app_patch_processo(
    processo_id: uuid.UUID,
    payload: ProcessoUpdate,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    actor_role = _normalize_role(str(session.get("role", "")))
    actor_username = _normalize_username(str(session.get("username", "")))
    changes = payload.model_dump(exclude_unset=True)
    # Campos derivados de status nao devem ser alterados diretamente via API.
    changes.pop("pendente_fiador", None)
    changes.pop("pendente_sinal", None)
    sla_trigger: Optional[str] = None
    old_sla_owner = _normalize_sla_owner(processo.sla_owner)
    estagio_changed = False
    assinatura_caixa_requested = False

    for field, value in changes.items():
        if field == "status_credito":
            status_credito = _process_credit_status(value)
            _validate_status_transition(field, processo.status_credito, status_credito)
            old_value = processo.status_credito
            processo.status_credito = status_credito
            if status_credito == "PENDENCIADO":
                sla_trigger = "analista_pendenciou"
            if old_value != status_credito:
                event_type = "RETORNO_PENDENCIA" if _is_pendencia_status(field, old_value) and not _is_pendencia_status(field, status_credito) else ("PENDENCIA" if _is_pendencia_status(field, status_credito) else "STATUS_CHANGE")
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type=event_type,
                    field_name=field,
                    old_value=old_value,
                    new_value=status_credito,
                )
        elif field == "status_geral":
            status_geral = _process_geral_status(value)
            _validate_status_transition(field, processo.status_geral, status_geral)
            old_value = processo.status_geral
            processo.status_geral = status_geral
            if status_geral == "PENDENCIADO":
                sla_trigger = "analista_pendenciou"
            if old_value != status_geral:
                event_type = "RETORNO_PENDENCIA" if _is_pendencia_status(field, old_value) and not _is_pendencia_status(field, status_geral) else ("PENDENCIA" if _is_pendencia_status(field, status_geral) else "STATUS_CHANGE")
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type=event_type,
                    field_name=field,
                    old_value=old_value,
                    new_value=status_geral,
                )
        elif field == "status_cca":
            status_cca = _process_caixa_status(value)
            _validate_status_transition(field, processo.status_cca, status_cca)
            old_value = processo.status_cca
            processo.status_cca = status_cca
            if status_cca in {"APROVADO", "DAR_QV"}:
                processo.status_credito = "APROVADO"
                if processo.status_geral in {"NOVO", "PENDENCIADO"}:
                    processo.status_geral = "EM_ANDAMENTO"
            elif status_cca == "CONDICIONADO":
                processo.status_credito = "PENDENCIADO"
                processo.status_geral = "PENDENCIADO"
                if old_sla_owner == SLA_OWNER_CCA:
                    sla_trigger = "cca_pendenciou"
            elif status_cca in {"REPROVADO", "BLOQUEADO"}:
                processo.status_credito = "REPROVADO"
                processo.status_geral = "REPROVADO"
            if status_cca == "ASSINATURA_CAIXA":
                assinatura_caixa_requested = True
            if old_sla_owner == SLA_OWNER_CCA and status_cca == "PENDENTE_CCA":
                sla_trigger = "cca_pendenciou"
            if old_value != status_cca:
                event_type = "RETORNO_PENDENCIA" if _is_pendencia_status(field, old_value) and not _is_pendencia_status(field, status_cca) else ("PENDENCIA" if _is_pendencia_status(field, status_cca) else "STATUS_CHANGE")
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type=event_type,
                    field_name=field,
                    old_value=old_value,
                    new_value=status_cca,
                )
        elif field == "status_agehab":
            status_agehab = _process_agehab_status(value)
            _validate_status_transition(field, processo.status_agehab, status_agehab)
            old_value = processo.status_agehab
            processo.status_agehab = status_agehab
            if old_sla_owner == SLA_OWNER_CCA and status_agehab == "PENDENTE_AGEHAB":
                sla_trigger = "cca_pendenciou"
            if old_value != status_agehab:
                event_type = "RETORNO_PENDENCIA" if _is_pendencia_status(field, old_value) and not _is_pendencia_status(field, status_agehab) else ("PENDENCIA" if _is_pendencia_status(field, status_agehab) else "STATUS_CHANGE")
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type=event_type,
                    field_name=field,
                    old_value=old_value,
                    new_value=status_agehab,
                )
        elif field == "status_sinal":
            status_sinal = _process_sinal_status(value)
            _validate_status_transition(field, processo.status_sinal, status_sinal)
            old_value = processo.status_sinal
            processo.status_sinal = status_sinal
            processo.pendente_sinal = status_sinal == "PENDENTE"
            if status_sinal == "NAO_TEM" and getattr(processo, "valor_sinal", None) is not None:
                old_sinal_valor = processo.valor_sinal
                processo.valor_sinal = None
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="PROCESSO_UPDATE",
                    field_name="valor_sinal",
                    old_value=old_sinal_valor,
                    new_value=None,
                    details="reset_automatico_status_sinal_nao_tem",
                )
            if old_value != status_sinal:
                event_type = "RETORNO_PENDENCIA" if _is_pendencia_status(field, old_value) and not _is_pendencia_status(field, status_sinal) else ("PENDENCIA" if _is_pendencia_status(field, status_sinal) else "STATUS_CHANGE")
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type=event_type,
                    field_name=field,
                    old_value=old_value,
                    new_value=status_sinal,
                )
        elif field == "valor_sinal":
            if value is None or str(value).strip() == "":
                new_valor_sinal = None
            else:
                try:
                    new_valor_sinal = round(float(value), 2)
                except (TypeError, ValueError):
                    raise HTTPException(status_code=422, detail="Valor de sinal invalido.")
                if new_valor_sinal < 0:
                    raise HTTPException(status_code=422, detail="Valor de sinal nao pode ser negativo.")
            old_value = getattr(processo, "valor_sinal", None)
            processo.valor_sinal = new_valor_sinal
            if old_value != new_valor_sinal:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="PROCESSO_UPDATE",
                    field_name=field,
                    old_value=old_value,
                    new_value=new_valor_sinal,
                )
        elif field == "status_fiador":
            status_fiador = _process_fiador_status(value)
            _validate_status_transition(field, processo.status_fiador, status_fiador)
            old_value = processo.status_fiador
            processo.status_fiador = status_fiador
            processo.pendente_fiador = status_fiador == "PENDENTE"
            if old_value != status_fiador:
                event_type = "RETORNO_PENDENCIA" if _is_pendencia_status(field, old_value) and not _is_pendencia_status(field, status_fiador) else ("PENDENCIA" if _is_pendencia_status(field, status_fiador) else "STATUS_CHANGE")
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type=event_type,
                    field_name=field,
                    old_value=old_value,
                    new_value=status_fiador,
                )
        elif field == "recolha_fgts":
            old_value = getattr(processo, "recolha_fgts", None)
            processo.recolha_fgts = _process_recolha_fgts_status(value, fallback=old_value or "NAO_RECOLHIDO")
            if old_value != processo.recolha_fgts:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="PROCESSO_UPDATE",
                    field_name=field,
                    old_value=old_value,
                    new_value=processo.recolha_fgts,
                )
        elif field == "nao_contar_mes":
            old_value = _is_nao_contar_mes_active(processo, _utcnow())
            _set_nao_contar_mes_period(processo, bool(value), _utcnow())
            new_value = _is_nao_contar_mes_active(processo, _utcnow())
            details = None
            if bool(value):
                details = (
                    f"periodo_referencia="
                    f"{int(processo.nao_contar_mes_ref_ano or 0):04d}-{int(processo.nao_contar_mes_ref_mes or 0):02d}"
                )
            if old_value != new_value:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="PROCESSO_UPDATE",
                    field_name=field,
                    old_value=old_value,
                    new_value=new_value,
                    details=details,
                )
        elif field == "estagio_comercial":
            next_stage = _process_estagio_comercial(value, fallback="")
            if not next_stage:
                raise HTTPException(status_code=422, detail="Estagio comercial invalido.")
            old_value = _process_estagio_comercial(processo.estagio_comercial)
            _validate_estagio_comercial_transition(old_value, next_stage)
            processo.estagio_comercial = next_stage
            estagio_changed = old_value != next_stage
            if old_value != next_stage:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="ESTAGIO_COMERCIAL_UPDATE",
                    field_name=field,
                    old_value=old_value,
                    new_value=next_stage,
                )
            etapa_before = _process_etapa_repasse(processo.etapa_repasse)
            entered_repasse = old_value not in ESTAGIOS_REPASSE_COMERCIAL and next_stage in ESTAGIOS_REPASSE_COMERCIAL
            if entered_repasse:
                processo.etapa_repasse = "EM_REPASSE"
            _sync_estagio_repasse_rules(processo, _utcnow())
            etapa_after = _process_etapa_repasse(processo.etapa_repasse)
            if etapa_before != etapa_after:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="ETAPA_REPASSE_AUTO",
                    field_name="etapa_repasse",
                    old_value=etapa_before,
                    new_value=etapa_after,
                    details="entrada_automatica_repasse",
                )
        elif field == "etapa_repasse":
            next_step = _process_etapa_repasse(value, fallback=None)
            if value is not None and str(value).strip() and not next_step:
                raise HTTPException(status_code=422, detail="Etapa de repasse invalida.")
            current_stage = _process_estagio_comercial(getattr(processo, "estagio_comercial", None))
            if next_step in {"EM_REPASSE", "INICIO_REPASSE"} and current_stage not in ESTAGIOS_REPASSE_COMERCIAL:
                raise HTTPException(
                    status_code=422,
                    detail="Etapa de repasse so pode ser usada a partir de ASSINATURA_DIRETORIA.",
                )
            if next_step == "ASSINATURA_AUTORIZADA" and not _can_set_assinatura_autorizada(processo):
                raise HTTPException(
                    status_code=422,
                    detail=(
                        "Nao pode avancar para ASSINATURA_AUTORIZADA. "
                        "Exige estagio VENDA_FINALIZADA, sinal/fiador regular, Agehab validada e Caixa aprovado."
                    ),
                )
            old_value = _process_etapa_repasse(processo.etapa_repasse)
            processo.etapa_repasse = next_step
            if old_value != next_step:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="ETAPA_REPASSE_UPDATE",
                    field_name=field,
                    old_value=old_value,
                    new_value=next_step,
                )
        elif field == "cca_responsavel":
            old_value = processo.cca_responsavel
            cca_username = _normalize_username(value)
            if not cca_username:
                processo.cca_responsavel = None
            else:
                cca_user = _get_user_by_username(db, cca_username)
                if not cca_user or _normalize_role(cca_user.role) != ROLE_CCA or not cca_user.is_active:
                    raise HTTPException(status_code=422, detail="CCA invalido")
                processo.cca_responsavel = cca_user.username
            if old_value != processo.cca_responsavel:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="ATRIBUICAO_CCA",
                    field_name=field,
                    old_value=old_value,
                    new_value=processo.cca_responsavel,
                )
        elif field in {"sla_credito_dias", "sla_corretor_dias", "sla_cca_dias"}:
            # SLA passa a ser calculado automaticamente pelo backend.
            continue
        else:
            old_value = getattr(processo, field, None)
            setattr(processo, field, value)
            new_value = getattr(processo, field, None)
            if old_value != new_value:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=actor_username,
                    actor_role=actor_role,
                    event_type="PROCESSO_UPDATE" if field != "observacao" else "OBSERVACAO_UPDATE",
                    field_name=field,
                    old_value=old_value,
                    new_value=new_value,
                )

    if estagio_changed:
        _sync_estagio_repasse_rules(processo, _utcnow())

    if assinatura_caixa_requested:
        if not _can_set_assinatura_autorizada(processo):
            raise HTTPException(
                status_code=422,
                detail=(
                    "Nao pode avancar para ASSINATURA_CAIXA. "
                    "Exige VENDA_FINALIZADA, sinal/fiador regular, Agehab validada e Caixa aprovado."
                ),
            )
        old_etapa = _process_etapa_repasse(processo.etapa_repasse)
        if old_etapa != "ASSINATURA_AUTORIZADA":
            processo.etapa_repasse = "ASSINATURA_AUTORIZADA"
            _record_processo_event(
                db,
                processo_id=processo.id,
                actor_username=actor_username,
                actor_role=actor_role,
                event_type="ETAPA_REPASSE_AUTO",
                field_name="etapa_repasse",
                old_value=old_etapa,
                new_value="ASSINATURA_AUTORIZADA",
                details="assinatura_caixa_confirmada",
            )
        if _status_token(processo.status_geral) not in {"CANCELADO", "DISTRATO"}:
            processo.status_geral = "APROVADO"
        processo.status_credito = "APROVADO"

    if changes:
        _ensure_credito_sla_start(processo, actor_role, _utcnow())
        _refresh_sla_fixed_markers(processo, _utcnow())

    _apply_sla_rules(
        processo,
        trigger=sla_trigger,
        has_enviado_docs=_process_has_enviado_docs(db, processo.id),
    )
    ref_now = _utcnow()
    status_geral_token = _status_token(getattr(processo, "status_geral", None))
    if status_geral_token in PROCESS_GERAL_ARQUIVO_IMEDIATO:
        old_nao_contar_mes = _is_nao_contar_mes_active(processo, ref_now)
        _set_nao_contar_mes_period(processo, True, ref_now)
        if not old_nao_contar_mes:
            _record_processo_event(
                db,
                processo_id=processo.id,
                actor_username=actor_username,
                actor_role=actor_role,
                event_type="PROCESSO_UPDATE",
                field_name="nao_contar_mes",
                old_value=old_nao_contar_mes,
                new_value=True,
                details=f"auto_arquivado_status_geral={status_geral_token.lower()}",
            )
        _set_processo_archived(
            processo,
            archived_at=ref_now,
            reference_date=ref_now.date(),
        )
    if "nao_contar_mes" in changes or status_geral_token in PROCESS_GERAL_ARQUIVO_IMEDIATO:
        _sync_reuniao_conta_no_mes_with_processo(db, processo, ref_now)
    new_sla_owner = _normalize_sla_owner(processo.sla_owner)
    if old_sla_owner != new_sla_owner:
        _record_processo_event(
            db,
            processo_id=processo.id,
            actor_username=actor_username,
            actor_role=actor_role,
            event_type="SLA_OWNER_CHANGE",
            field_name="sla_owner",
            old_value=old_sla_owner,
            new_value=new_sla_owner,
            details=(sla_trigger or "regra_automatica"),
        )

    if changes:
        ordered_fields = ",".join(sorted(changes.keys()))
        _record_system_log(
            db,
            actor_username=actor_username,
            actor_role=actor_role,
            tela="analista",
            acao="PROCESSO_ATUALIZADO",
            entidade_tipo="processo",
            entidade_id=str(processo.id),
            details=(
                f"campos={ordered_fields}; "
                f"status_geral={processo.status_geral}; status_credito={processo.status_credito}; "
                f"status_cca={processo.status_cca}; status_agehab={processo.status_agehab}"
            ),
        )

    db.commit()
    db.refresh(processo)
    _sync_frankstein_events_for_processo(db, processo)
    _invalidate_process_list_cache()

    now = _utcnow()
    processo_out = ProcessoOut.model_validate(processo)
    processo_out.sla_corretor_seconds = _compute_sla_seconds(processo, SLA_OWNER_CORRETOR, now)
    processo_out.sla_analista_seconds = _compute_sla_seconds(processo, SLA_OWNER_ANALISTA, now)
    processo_out.sla_cca_seconds = _compute_sla_seconds(processo, SLA_OWNER_CCA, now)
    processo_out.sla_corretor_horas = processo_out.sla_corretor_seconds // 3600
    processo_out.sla_analista_horas = processo_out.sla_analista_seconds // 3600
    processo_out.sla_cca_horas = processo_out.sla_cca_seconds // 3600
    processo_out.sla_corretor_dias = processo_out.sla_corretor_horas // 24
    processo_out.sla_credito_dias = processo_out.sla_analista_horas // 24
    processo_out.sla_cca_dias = processo_out.sla_cca_horas // 24
    processo_out.sla_owner = _normalize_sla_owner(processo.sla_owner)
    processo_out.sla_active_since = _as_utc(processo.sla_active_since)
    processo_out.status_cca = _process_caixa_status(processo.status_cca)
    processo_out.status_agehab = _process_agehab_status(processo.status_agehab)
    processo_out.recolha_fgts = _process_recolha_fgts_status(getattr(processo, "recolha_fgts", None))
    processo_out.estagio_comercial = _process_estagio_comercial(processo.estagio_comercial)
    processo_out.etapa_repasse = _process_etapa_repasse(processo.etapa_repasse)
    processo_out.fila_atual = _fila_atual_from_processo(processo)
    processo_out.nao_contar_mes = _is_nao_contar_mes_active(processo, now)
    return processo_out


@app.put("/app/api/processos/{processo_id}/documentos", response_model=list[DocumentoOut])
def app_bulk_upsert_documentos(
    processo_id: uuid.UUID,
    payload: DocumentoBulkUpsert,
    session: dict[str, Any] = Depends(require_roles(*PROCESSO_DOCUMENTOS_UPSERT_ROLES)),
    db: Session = Depends(get_db),
):
    processo = db.get(Processo, processo_id)
    if not processo:
        raise HTTPException(status_code=404, detail="Processo nao encontrado")

    role = _normalize_role(str(session.get("role", "")))
    username = _normalize_username(str(session.get("username", "")))
    if role == ROLE_CORRETOR:
        cliente = db.get(Cliente, processo.cliente_id)
        if not cliente:
            raise HTTPException(status_code=404, detail="Cliente nao encontrado")
        if _normalize_username(cliente.corretor) != username:
            raise HTTPException(status_code=403, detail="Sem permissao para atualizar este processo")
    if role == ROLE_CCA:
        if _normalize_username(processo.cca_responsavel) != username:
            raise HTTPException(status_code=403, detail="Sem permissao para atualizar este processo")
        raise HTTPException(status_code=403, detail="Perfil CCA possui acesso somente leitura nesta tela")

    old_status_credito = processo.status_credito
    old_status_geral = processo.status_geral
    old_sla_owner = _normalize_sla_owner(processo.sla_owner)

    if not payload.documentos:
        raise HTTPException(status_code=422, detail="Lista de documentos vazia")

    dedup_map: dict[tuple[str, str], DocumentoBulkItem] = {}
    for item in payload.documentos:
        categoria = item.categoria.strip()
        nome = item.nome.strip()
        if not categoria or not nome:
            continue
        dedup_map[(categoria, nome)] = item

    if not dedup_map:
        raise HTTPException(status_code=422, detail="Nenhum documento valido para salvar")

    can_update_credit = role in {ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO}
    existing_docs = (
        db.query(Documento)
        .filter(Documento.processo_id == processo_id)
        .all()
    )
    docs_by_key: dict[tuple[str, str], Documento] = {
        (doc.categoria, doc.nome): doc for doc in existing_docs
    }

    for (categoria, nome), item in dedup_map.items():
        documento = docs_by_key.get((categoria, nome))

        status_doc = _doc_status(item.status_doc) if item.status_doc is not None else None
        status_credito = (
            _credit_status(item.status_credito)
            if can_update_credit and item.status_credito is not None
            else None
        )
        pendencia_info = _normalize_pendencia_info(item.pendencia_info)

        if documento:
            old_status_doc = documento.status_doc
            old_status_credito_doc = documento.status_credito
            old_pendencia_info = _normalize_pendencia_info(documento.pendencia_info)
            current_credit = _credit_status(documento.status_credito, fallback="AGUARDANDO_ENVIO")
            if status_doc is not None:
                if not (role == ROLE_CORRETOR and _doc_is_done(documento.status_doc)):
                    documento.status_doc = status_doc
                    if role == ROLE_CORRETOR:
                        if _doc_is_done(status_doc) and current_credit == "AGUARDANDO_ENVIO":
                            documento.status_credito = "ANALISE"
                        elif status_doc == "PENDENTE" and current_credit in {"AGUARDANDO_ENVIO", "ANALISE"}:
                            documento.status_credito = "AGUARDANDO_ENVIO"
            if status_credito is not None:
                documento.status_credito = status_credito
            elif documento.status_doc == "NAO_APLICA":
                documento.status_credito = "NAO_APLICA"
            elif _doc_is_done(documento.status_doc) and _credit_status(documento.status_credito, fallback="AGUARDANDO_ENVIO") == "AGUARDANDO_ENVIO":
                documento.status_credito = "ANALISE"
            next_credit = _credit_status(documento.status_credito, fallback="AGUARDANDO_ENVIO")
            became_pending_doc = old_status_doc != "PENDENTE" and documento.status_doc == "PENDENTE"
            if not _doc_is_done(documento.status_doc) and next_credit != "PENDENCIADO":
                documento.status_credito = "AGUARDANDO_ENVIO"
                next_credit = "AGUARDANDO_ENVIO"
            if (
                documento.status_doc == "PENDENTE"
                and next_credit == "PENDENCIADO"
                and not became_pending_doc
                and not pendencia_info
                and not old_pendencia_info
            ):
                documento.status_credito = "AGUARDANDO_ENVIO"
                next_credit = "AGUARDANDO_ENVIO"
            if next_credit in {"PENDENCIADO", "REPROVADO"}:
                if pendencia_info:
                    documento.pendencia_info = pendencia_info
                elif role == ROLE_ANALISTA and (
                    next_credit == "REPROVADO"
                    or became_pending_doc
                    or not old_pendencia_info
                ):
                    raise HTTPException(
                        status_code=422,
                        detail=(f"Documento '{nome}' foi {next_credit.lower()}. Informe pendencia_info."),
                    )
                elif old_pendencia_info:
                    documento.pendencia_info = old_pendencia_info
            else:
                documento.pendencia_info = None

            normalized_pendencia_info = _normalize_pendencia_info(documento.pendencia_info)
            documento.pendencia_info = normalized_pendencia_info
            if old_status_doc != documento.status_doc:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=username,
                    actor_role=role,
                    event_type="DOCUMENTO_STATUS_DOC",
                    field_name=f"{categoria}:{nome}:status_doc",
                    old_value=old_status_doc,
                    new_value=documento.status_doc,
                )
            if old_status_credito_doc != documento.status_credito:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=username,
                    actor_role=role,
                    event_type="DOCUMENTO_STATUS_CREDITO",
                    field_name=f"{categoria}:{nome}:status_credito",
                    old_value=old_status_credito_doc,
                    new_value=documento.status_credito,
                )
            if old_pendencia_info != documento.pendencia_info:
                _record_processo_event(
                    db,
                    processo_id=processo.id,
                    actor_username=username,
                    actor_role=role,
                    event_type="DOCUMENTO_PENDENCIA_INFO",
                    field_name=f"{categoria}:{nome}:pendencia_info",
                    old_value=old_pendencia_info,
                    new_value=documento.pendencia_info,
                )
        else:
            if can_update_credit:
                if (status_doc or "").upper() == "NAO_APLICA":
                    credito_default = status_credito or "NAO_APLICA"
                else:
                    credito_default = status_credito or ("ANALISE" if _doc_is_done(status_doc) else "AGUARDANDO_ENVIO")
            else:
                credito_default = "NAO_APLICA" if (status_doc or "").upper() == "NAO_APLICA" else ("ANALISE" if _doc_is_done(status_doc) else "AGUARDANDO_ENVIO")
            credito_default_norm = _credit_status(credito_default, fallback="AGUARDANDO_ENVIO")
            status_doc_norm = status_doc or "PENDENTE"
            if not _doc_is_done(status_doc_norm):
                if credito_default_norm != "PENDENCIADO":
                    credito_default_norm = "AGUARDANDO_ENVIO"
                elif not pendencia_info:
                    raise HTTPException(
                        status_code=422,
                        detail=f"Documento '{nome}' foi pendenciado. Informe pendencia_info.",
                    )
            if credito_default_norm in {"PENDENCIADO", "REPROVADO"} and not pendencia_info:
                raise HTTPException(
                    status_code=422,
                    detail=f"Documento '{nome}' foi {credito_default_norm.lower()}. Informe pendencia_info.",
                )
            novo_documento = Documento(
                processo_id=processo_id,
                categoria=categoria,
                nome=nome,
                status_doc=status_doc_norm,
                status_credito=credito_default_norm,
                pendencia_info=pendencia_info if credito_default_norm in {"PENDENCIADO", "REPROVADO"} else None,
            )
            docs_by_key[(categoria, nome)] = novo_documento
            db.add(novo_documento)
            _record_processo_event(
                db,
                processo_id=processo.id,
                actor_username=username,
                actor_role=role,
                event_type="DOCUMENTO_CRIADO",
                field_name=f"{categoria}:{nome}",
                new_value=f"{novo_documento.status_doc}|{novo_documento.status_credito}",
            )

    db.flush()

    documentos = (
        db.query(Documento)
        .filter(Documento.processo_id == processo_id)
        .order_by(Documento.categoria.asc(), Documento.nome.asc())
        .all()
    )
    has_enviado_docs = any(_doc_is_done(doc.status_doc) for doc in documentos)
    enviados_count = sum(1 for doc in documentos if _doc_is_done(doc.status_doc))

    sla_trigger: Optional[str] = None
    if role == ROLE_CORRETOR and has_enviado_docs:
        if _status_token(processo.status_credito) == "PENDENCIADO":
            processo.status_credito = "EM_ANALISE"
        if _status_token(processo.status_geral) in {"NOVO", "PENDENCIADO"}:
            processo.status_geral = "EM_ANDAMENTO"
        sla_trigger = "corretor_enviou"

    if old_status_credito != processo.status_credito:
        _record_processo_event(
            db,
            processo_id=processo.id,
            actor_username=username,
            actor_role=role,
            event_type="STATUS_CHANGE",
            field_name="status_credito",
            old_value=old_status_credito,
            new_value=processo.status_credito,
            details="ajuste_automatico_documentos",
        )
    if old_status_geral != processo.status_geral:
        event_type = "RETORNO_PENDENCIA" if _is_pendencia_status("status_geral", old_status_geral) and not _is_pendencia_status("status_geral", processo.status_geral) else "STATUS_CHANGE"
        _record_processo_event(
            db,
            processo_id=processo.id,
            actor_username=username,
            actor_role=role,
            event_type=event_type,
            field_name="status_geral",
            old_value=old_status_geral,
            new_value=processo.status_geral,
            details="ajuste_automatico_documentos",
        )

    if dedup_map:
        _ensure_credito_sla_start(processo, role, _utcnow())
        _refresh_sla_fixed_markers(processo, _utcnow())

    _apply_sla_rules(processo, trigger=sla_trigger, has_enviado_docs=has_enviado_docs)
    new_sla_owner = _normalize_sla_owner(processo.sla_owner)
    if old_sla_owner != new_sla_owner:
        _record_processo_event(
            db,
            processo_id=processo.id,
            actor_username=username,
            actor_role=role,
            event_type="SLA_OWNER_CHANGE",
            field_name="sla_owner",
            old_value=old_sla_owner,
            new_value=new_sla_owner,
            details=(sla_trigger or "regra_automatica"),
        )

    tela_origem = "checklist" if role == ROLE_CORRETOR else ("analista" if role == ROLE_ANALISTA else "admin")
    _record_system_log(
        db,
        actor_username=username,
        actor_role=role,
        tela=tela_origem,
        acao="DOCUMENTOS_SALVOS",
        entidade_tipo="processo",
        entidade_id=str(processo.id),
        details=(
            f"documentos_recebidos={len(dedup_map)}; "
            f"documentos_enviados={enviados_count}; "
            f"status_geral={processo.status_geral}; status_credito={processo.status_credito}"
        ),
    )

    db.commit()
    _sync_frankstein_events_for_processo(db, processo)
    _invalidate_process_list_cache()
    documentos = (
        db.query(Documento)
        .filter(Documento.processo_id == processo_id)
        .order_by(Documento.categoria.asc(), Documento.nome.asc())
        .all()
    )
    return documentos


@app.post("/app/api/tabela-precos/upload", response_model=TabelaPrecoUploadResponse)
async def upload_tabela_precos(
    _: dict[str, Any] = Depends(require_roles(ROLE_ADMIN)),
    file: UploadFile = File(...),
):
    rows = await processar_tabela_upload(file)
    salvar_tabela_precos(rows)
    return TabelaPrecoUploadResponse(linhas=len(rows), filename=file.filename or "")


@app.get("/app/api/tabela-precos", response_model=list[TabelaPrecoItem])
async def listar_tabela_precos(
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
):
    return carregar_tabela_precos()


@app.post("/app/api/simulacao")
async def simular_proposta(
    payload: SimulacaoInput,
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
):
    resultado = engine_calculo_imobiliario(payload)
    if isinstance(resultado, dict) and resultado.get("erro_politica"):
        raise HTTPException(status_code=422, detail=resultado.get("mensagem", "Valor abaixo da polÃ­tica comercial"))
    return resultado


@app.post("/app/api/recomendacao")
async def recomendar_proposta(
    payload: SimulacaoInput,
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
):
    resultado = engine_calculo_imobiliario(payload)
    if isinstance(resultado, dict) and resultado.get("erro_politica"):
        raise HTTPException(status_code=422, detail=resultado.get("mensagem", "Valor abaixo da polÃ­tica comercial"))

    preco_sugerido = resultado["apresentacao_cliente"]["valor_imovel"]
    status_ia = resultado["leitura_executiva_corretor"]["status_ia"]
    risco = resultado["leitura_executiva_corretor"]["risco_exposicao"]
    confianca = 0.62  # placeholder heurÃ­stico

    return {
        "preco_sugerido": preco_sugerido,
        "status_ia": status_ia,
        "risco_exposicao": risco,
        "confianca": confianca,
        "motivo": resultado["leitura_executiva_corretor"].get("motivo_auditoria", ""),
    }


@app.post("/app/api/recomendacao/feedback")
async def feedback_recomendacao(
    payload: dict,
    session: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    feedback = IaFeedback(
        aceitou=bool(payload.get("aceitou")),
        preco_sugerido=float(payload.get("preco_sugerido") or 0.0) if payload.get("preco_sugerido") is not None else None,
        contexto_json=json.dumps(payload.get("contexto"), ensure_ascii=False) if payload.get("contexto") is not None else None,
        origem=_normalize_role(str(session.get("role", ""))),
    )
    db.add(feedback)
    db.commit()
    return {"ok": True}


def _safe_load_json_list(path: Path) -> list[dict[str, Any]]:
    if path == FRANKSTEIN_EVENTS_PATH and (not path.exists()) and FRANKSTEIN_EVENTS_LEGACY_PATH.exists():
        path = FRANKSTEIN_EVENTS_LEGACY_PATH
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


def _persist_json_list(path: Path, items: list[dict[str, Any]], *, max_len: int = 5000) -> None:
    trimmed = items[-max_len:]
    path.parent.mkdir(exist_ok=True, parents=True)
    path.write_text(json.dumps(trimmed, ensure_ascii=False, indent=2), encoding="utf-8")


def _clean_optional_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_optional_uuid(value: Any) -> Optional[uuid.UUID]:
    text = _clean_optional_text(value)
    if not text:
        return None
    try:
        return uuid.UUID(text)
    except (TypeError, ValueError, AttributeError):
        return None


def _clean_optional_float(value: Any, *, digits: int = 6) -> Optional[float]:
    if value is None:
        return None
    try:
        return round(float(value), digits)
    except (TypeError, ValueError):
        return None


def _clean_optional_bool(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "sim"}:
        return True
    if text in {"0", "false", "no", "nao", "nÃ£o"}:
        return False
    return None


def _load_json_file(path: Path) -> Any:
    if not path.exists():
        return None
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        logger.exception("Falha ao carregar arquivo JSON: %s", path)
        return None
    return loaded


def _load_sklearn_frankstein_model_bundle() -> Optional[dict[str, Any]]:
    model_path = FRANKSTEIN_CURRENT_SKLEARN_MODEL_PATH
    if not model_path.exists() and FRANKSTEIN_CURRENT_SKLEARN_MODEL_LEGACY_PATH.exists():
        model_path = FRANKSTEIN_CURRENT_SKLEARN_MODEL_LEGACY_PATH
    if not model_path.exists():
        return None
    try:
        import joblib
    except Exception:
        logger.exception("Joblib indisponivel para carregar modelo %s atual.", AI_ASSISTANT_DISPLAY_NAME)
        return None

    try:
        bundle: dict[str, Any] = {
            "_kind": "sklearn_pipeline",
            "model": joblib.load(model_path),
            "feature_columns": [],
            "metrics": {},
            "model_info": {},
        }
        feature_columns = _load_json_file(FRANKSTEIN_CURRENT_FEATURE_COLUMNS_PATH)
        bundle["feature_columns"] = feature_columns if isinstance(feature_columns, list) else []
        bundle["metrics"] = _load_json_file(FRANKSTEIN_CURRENT_METRICS_PATH) or {}
        bundle["model_info"] = _load_json_file(FRANKSTEIN_CURRENT_MODEL_INFO_PATH) or {}
        return bundle
    except Exception:
        logger.exception("Falha ao carregar bundle sklearn atual do %s.", AI_ASSISTANT_DISPLAY_NAME)
        return None


def _load_frankstein_model() -> Optional[dict[str, Any]]:
    sklearn_bundle = _load_sklearn_frankstein_model_bundle()
    if sklearn_bundle:
        return sklearn_bundle
    model_path = FRANKSTEIN_MODEL_PATH if FRANKSTEIN_MODEL_PATH.exists() else FRANKSTEIN_MODEL_LEGACY_PATH
    if not model_path.exists():
        return None
    try:
        loaded = json.loads(model_path.read_text(encoding="utf-8"))
        if isinstance(loaded, dict):
            loaded["_kind"] = "legacy_json"
            return loaded
        return None
    except Exception:
        logger.exception("Falha ao carregar modelo frankstein; seguindo sem modelo.")
        return None


def _sigmoid(x: float) -> float:
    try:
        return 1 / (1 + math.exp(-x))
    except OverflowError:
        return 0.0 if x < 0 else 1.0


def _log_frankstein_event(event: dict[str, Any]) -> None:
    if not _log_frankstein_event_db(event) and not FRANKSTEIN_DB_URL:
        events = _safe_load_json_list(FRANKSTEIN_EVENTS_PATH)
        events.append(event)
        _persist_json_list(FRANKSTEIN_EVENTS_PATH, events)


class FranksteinSugestaoOut(BaseModel):
    preco_heuristico: float
    preco_frankstein: float
    prob_aceite: float
    status_ia: str
    risco_exposicao: str
    confianca: float
    motivos: list[str] = Field(default_factory=list)
    sugestoes: list[str] = Field(default_factory=list)

    model_config = ConfigDict(arbitrary_types_allowed=True)


class FranksteinRecomendacaoOut(BaseModel):
    event_id: Optional[str] = None
    modelo_versao: Optional[str] = None
    heuristica: dict[str, Any]
    frankstein: FranksteinSugestaoOut

    model_config = ConfigDict(arbitrary_types_allowed=True)


class FranksteinEventResultadoUpdate(BaseModel):
    teve_pendencia_cca: Optional[bool] = None
    teve_pendencia_agehab: Optional[bool] = None
    foi_aprovado: Optional[bool] = None
    foi_reprovado: Optional[bool] = None
    virou_condicionado: Optional[bool] = None
    assinou_caixa: Optional[bool] = None
    finalizou: Optional[bool] = None
    tempo_ate_assinatura_horas: Optional[float] = Field(None, ge=0)
    tempo_total_processo_horas: Optional[float] = Field(None, ge=0)
    retorno_cca: Optional[str] = None
    resultado_real: Optional[str] = None


def _safe_ratio_percent(value: Any) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip().replace("%", "").replace(",", ".")
    if not text:
        return None
    try:
        return round(float(text) / 100.0, 6)
    except (TypeError, ValueError):
        return None


def _frankstein_model_version(model: Optional[dict[str, Any]]) -> Optional[str]:
    if not model:
        return None
    model_info = model.get("model_info") if isinstance(model.get("model_info"), dict) else {}
    return (
        _clean_optional_text(model_info.get("version"))
        or _clean_optional_text(model.get("version"))
        or _clean_optional_text(model_info.get("model_name"))
    )


def _frankstein_features(payload: SimulacaoInput, heuristica: dict[str, Any]) -> list[float]:
    leitura = heuristica.get("leitura_executiva_corretor", {})
    expo = _safe_ratio_percent(leitura.get("risco_exposicao")) or 0.0
    is_pos = _safe_ratio_percent(leitura.get("is_pos_chaves")) or 0.0
    return [
        float(payload.renda_bruta),
        float(payload.valor_tabela),
        float(payload.sobrepreco_vila),
        float(payload.valor_obtido),
        float(payload.parcela_caixa),
        float(payload.preco_digitado_corretor or 0.0),
        expo,
        is_pos,
    ]


def _frankstein_feature_payload(payload: SimulacaoInput, heuristica: dict[str, Any]) -> dict[str, Any]:
    apresentacao = heuristica.get("apresentacao_cliente", {})
    leitura = heuristica.get("leitura_executiva_corretor", {})
    preco_base_politica = round(max(float(payload.valor_tabela), float(payload.valor_obtido)) + float(payload.sobrepreco_vila), 2)
    preco_final = _clean_optional_float(apresentacao.get("valor_imovel"), digits=2)
    entrada_liquida = _clean_optional_float(apresentacao.get("entrada_facilitada"), digits=2)
    valor_parcela_entrada = _clean_optional_float(apresentacao.get("valor_parcela"), digits=2)
    exposicao_risco = round(
        max(0.0, (float(payload.valor_tabela) - float(payload.valor_obtido)) / max(float(payload.valor_tabela), 1.0)),
        6,
    )
    is_pos_chaves = round(
        (float(payload.parcela_caixa) + float(valor_parcela_entrada or 0.0)) / max(float(payload.renda_bruta), 1.0),
        6,
    )
    return {
        "perfil": _clean_optional_text(getattr(payload, "perfil", None)),
        "empreendimento": _clean_optional_text(getattr(payload, "empreendimento", None)),
        "renda_bruta": _clean_optional_float(payload.renda_bruta, digits=2),
        "valor_tabela": _clean_optional_float(payload.valor_tabela, digits=2),
        "sobrepreco_vila": _clean_optional_float(payload.sobrepreco_vila, digits=2),
        "valor_obtido": _clean_optional_float(payload.valor_obtido, digits=2),
        "parcela_caixa": _clean_optional_float(payload.parcela_caixa, digits=2),
        "preco_digitado_corretor": _clean_optional_float(payload.preco_digitado_corretor or 0.0, digits=2),
        "preco_base_politica": preco_base_politica,
        "preco_final": preco_final,
        "entrada_liquida": entrada_liquida,
        "valor_parcela_entrada": valor_parcela_entrada,
        "exposicao_risco": exposicao_risco,
        "is_pos_chaves": is_pos_chaves,
        "garantidores_necessarios": leitura.get("garantidores_necessarios"),
        "status_ia_heuristica": _clean_optional_text(leitura.get("status_ia")),
        "alerta_preco": _clean_optional_text(leitura.get("alerta_preco")),
        "bloqueio_critico": bool(leitura.get("bloqueio_critico", False)),
    }


def _frankstein_operacional_features(payload: FranksteinAnaliseInput, resposta: FranksteinRespostaOperacional) -> dict[str, Any]:
    faltante = 0.0
    cobertura_total = float(payload.garantido) + float(payload.cheque_moradia)
    if cobertura_total < float(payload.valor_venda):
        faltante = round(float(payload.valor_venda) - cobertura_total, 2)
    return {
        "perfil": _clean_optional_text(payload.perfil),
        "empreendimento": _clean_optional_text(getattr(payload, "empreendimento", None)),
        "valor_venda": _clean_optional_float(payload.valor_venda, digits=2),
        "garantido": _clean_optional_float(payload.garantido, digits=2),
        "cheque_moradia": _clean_optional_float(payload.cheque_moradia, digits=2),
        "faltante": _clean_optional_float(faltante, digits=2),
        "qtd_problemas_documentais": len(resposta.frankstein.campos_com_problema),
        "score_risco_regra": _clean_optional_float(resposta.frankstein.score.valor, digits=4),
        "status_geral_regra": _clean_optional_text(resposta.frankstein.status_geral),
        "decisao_recomendada_regra": _clean_optional_text(resposta.frankstein.decisao_recomendada.codigo),
        "doc_rg_cpf_ok": bool(payload.documentos.rg_cpf),
        "doc_comprovante_residencia_ok": bool(payload.documentos.comprovante_residencia),
        "doc_comprovante_renda_ok": bool(payload.documentos.comprovante_renda),
        "doc_fgts_validado": bool(payload.documentos.fgts_validado),
        "renda_informada": _clean_optional_float(payload.renda_informada, digits=2),
        "score_classificacao_regra": _clean_optional_text(resposta.frankstein.score.classificacao),
        "resumo_regra": _clean_optional_text(resposta.frankstein.resumo),
    }


def _find_latest_lead_for_processo(db: Session, processo_id: uuid.UUID) -> Optional[LeadPreCadastro]:
    return (
        db.query(LeadPreCadastro)
        .filter(LeadPreCadastro.processo_id == processo_id)
        .order_by(LeadPreCadastro.updated_at.desc(), LeadPreCadastro.created_at.desc())
        .first()
    )


def _resolve_frankstein_process_context(
    db: Session,
    *,
    processo_id: Any = None,
    cliente_id: Any = None,
    reserva_id: Any = None,
    lead_id: Any = None,
) -> tuple[Optional[Processo], Optional[LeadPreCadastro]]:
    processo: Optional[Processo] = None
    lead: Optional[LeadPreCadastro] = None

    processo_uuid = _clean_optional_uuid(processo_id)
    if processo_uuid is not None:
        processo = db.get(Processo, processo_uuid)

    lead_uuid = _clean_optional_uuid(lead_id) or _clean_optional_uuid(reserva_id)
    if lead_uuid is not None:
        lead = db.get(LeadPreCadastro, lead_uuid)
        if lead is not None and getattr(lead, "processo_id", None) and processo is None:
            processo = db.get(Processo, lead.processo_id)

    cliente_uuid = _clean_optional_uuid(cliente_id)
    if processo is None and cliente_uuid is not None:
        processo = (
            db.query(Processo)
            .filter(Processo.cliente_id == cliente_uuid)
            .order_by(Processo.arquivado.asc(), Processo.updated_at.desc(), Processo.created_at.desc())
            .first()
        )

    if processo is not None and lead is None:
        lead = _find_latest_lead_for_processo(db, processo.id)

    return processo, lead


def _attach_frankstein_event_context(
    event: dict[str, Any],
    *,
    processo: Optional[Processo],
    lead: Optional[LeadPreCadastro],
) -> dict[str, Any]:
    if processo is not None:
        event["processo_id"] = str(processo.id)
        event["cliente_id"] = event.get("cliente_id") or str(processo.cliente_id)
    if lead is not None:
        lead_id = str(lead.id)
        event["lead_id"] = lead_id
        event["reserva_id"] = event.get("reserva_id") or lead_id
    return event


def _frankstein_event_timestamp(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return _as_utc(value)
    text = _clean_optional_text(value)
    if not text:
        return None
    try:
        return _as_utc(datetime.fromisoformat(text))
    except ValueError:
        return None


def _remember_processo_frankstein_event(processo: Processo, event: dict[str, Any]) -> None:
    event_id = _clean_optional_text(event.get("event_id"))
    event_at = _frankstein_event_timestamp(event.get("timestamp"))
    if event_id:
        processo.frankstein_last_event_id = event_id
    if event_at is not None:
        processo.frankstein_last_event_at = event_at


def _cliente_has_single_processo(db: Session, cliente_id: uuid.UUID) -> bool:
    total = int(db.query(func.count(Processo.id)).filter(Processo.cliente_id == cliente_id).scalar() or 0)
    return total == 1


def _link_frankstein_events_to_processo(
    db: Session,
    processo: Processo,
    *,
    lead: Optional[LeadPreCadastro] = None,
) -> int:
    if processo is None:
        return 0

    resolved_lead = lead or _find_latest_lead_for_processo(db, processo.id)
    processo_id_text = str(processo.id)
    cliente_id_text = str(processo.cliente_id)
    lead_id_text = str(resolved_lead.id) if resolved_lead is not None else None
    allow_cliente_match = _cliente_has_single_processo(db, processo.cliente_id)
    updated = 0

    conn = _frankstein_db_conn()
    if conn is not None:
        try:
            with conn, conn.cursor() as cur:
                if lead_id_text:
                    cur.execute(
                        """
                        update frankstein_events
                           set processo_id = %s,
                               lead_id = coalesce(lead_id, %s),
                               updated_at = now()
                         where (processo_id is null or trim(coalesce(processo_id, '')) = '')
                           and (
                               trim(coalesce(lead_id, '')) = %s
                               or trim(coalesce(reserva_id, '')) = %s
                           )
                        """,
                        (processo_id_text, lead_id_text, lead_id_text, lead_id_text),
                    )
                    updated += cur.rowcount or 0
                if allow_cliente_match:
                    cur.execute(
                        """
                        update frankstein_events
                           set processo_id = %s,
                               lead_id = coalesce(lead_id, %s),
                               updated_at = now()
                         where (processo_id is null or trim(coalesce(processo_id, '')) = '')
                           and trim(coalesce(cliente_id, '')) = %s
                        """,
                        (processo_id_text, lead_id_text, cliente_id_text),
                    )
                    updated += cur.rowcount or 0
                if lead_id_text:
                    cur.execute(
                        """
                        update frankstein_events
                           set lead_id = %s,
                               updated_at = now()
                         where trim(coalesce(processo_id, '')) = %s
                           and (lead_id is null or trim(coalesce(lead_id, '')) = '')
                        """,
                        (lead_id_text, processo_id_text),
                    )
        except Exception:
            logger.exception("Falha ao vincular eventos %s ao processo %s no banco; fallback arquivo.", AI_ASSISTANT_DISPLAY_NAME, processo_id_text)
        finally:
            conn.close()

    events = _safe_load_json_list(FRANKSTEIN_EVENTS_PATH)
    if not events:
        return updated

    file_updated = False
    for event in events:
        event_processo_id = _clean_optional_text(event.get("processo_id"))
        event_cliente_id = _clean_optional_text(event.get("cliente_id"))
        event_lead_id = _clean_optional_text(event.get("lead_id"))
        event_reserva_id = _clean_optional_text(event.get("reserva_id"))
        matches_processo = event_processo_id == processo_id_text
        matches_lead = bool(lead_id_text) and (event_lead_id == lead_id_text or event_reserva_id == lead_id_text)
        matches_cliente = allow_cliente_match and event_cliente_id == cliente_id_text
        if matches_processo or matches_lead or matches_cliente:
            if event_processo_id != processo_id_text:
                event["processo_id"] = processo_id_text
                updated += 1
                file_updated = True
            if lead_id_text and event_lead_id != lead_id_text:
                event["lead_id"] = lead_id_text
                file_updated = True

    if file_updated:
        _persist_json_list(FRANKSTEIN_EVENTS_PATH, events)
    return updated


def _load_frankstein_status_history(
    db: Session,
    processo_id: uuid.UUID,
) -> list[tuple[str, str, Optional[datetime]]]:
    rows = (
        db.query(ProcessoEvento.field_name, ProcessoEvento.new_value, ProcessoEvento.created_at)
        .filter(
            ProcessoEvento.processo_id == processo_id,
            ProcessoEvento.field_name.in_(["status_credito", "status_geral", "status_cca", "status_agehab"]),
        )
        .order_by(ProcessoEvento.created_at.asc())
        .all()
    )
    history: list[tuple[str, str, Optional[datetime]]] = []
    for field_name, new_value, created_at in rows:
        history.append(((field_name or "").strip().lower(), _status_token(new_value), _as_utc(created_at)))
    return history


def _first_frankstein_status_change_at(
    history: list[tuple[str, str, Optional[datetime]]],
    *,
    field_names: set[str],
    new_values: set[str],
) -> Optional[datetime]:
    for field_name, new_value, created_at in history:
        if field_name in field_names and new_value in new_values and created_at is not None:
            return created_at
    return None


def _frankstein_resultado_real_from_statuses(
    *,
    status_credito: str,
    status_geral: str,
    status_cca: str,
    status_agehab: str,
) -> Optional[str]:
    if status_cca == "FINALIZADO":
        return "FINALIZADO"
    if status_cca == "ASSINATURA_CAIXA":
        return "ASSINATURA_CAIXA"
    if status_geral in {"DISTRATO", "CANCELADO"}:
        return status_geral
    if status_cca in {"REPROVADO", "BLOQUEADO"} or status_credito == "REPROVADO" or status_geral == "REPROVADO":
        return "REPROVADO"
    if status_cca in {"APROVADO", "DAR_QV"} or status_credito == "APROVADO":
        return "APROVADO"
    if status_cca == "CONDICIONADO":
        return "CONDICIONADO"
    if status_cca == "PENDENTE_CCA":
        return "PENDENCIA_CCA"
    if status_agehab == "PENDENTE_AGEHAB":
        return "PENDENCIA_AGEHAB"
    if status_geral == "PENDENCIADO" or status_credito == "PENDENCIADO":
        return "PENDENCIADO"
    return status_cca or status_credito or status_geral or status_agehab or None


def _frankstein_hours_between(start_at: Optional[datetime], end_at: Optional[datetime]) -> Optional[float]:
    start_utc = _as_utc(start_at)
    end_utc = _as_utc(end_at)
    if start_utc is None or end_utc is None or end_utc < start_utc:
        return None
    return round((end_utc - start_utc).total_seconds() / 3600.0, 2)


def _build_frankstein_process_outcome(
    db: Session,
    processo: Processo,
) -> tuple[dict[str, Any], Optional[datetime], Optional[datetime]]:
    history = _load_frankstein_status_history(db, processo.id)
    status_credito = _status_token(getattr(processo, "status_credito", None))
    status_geral = _status_token(getattr(processo, "status_geral", None))
    status_cca = _status_token(getattr(processo, "status_cca", None))
    status_agehab = _status_token(getattr(processo, "status_agehab", None))

    teve_pendencia_cca = status_cca in {"PENDENTE_CCA", "CONDICIONADO"} or any(
        field_name == "status_cca" and new_value in {"PENDENTE_CCA", "CONDICIONADO"}
        for field_name, new_value, _ in history
    )
    teve_pendencia_agehab = status_agehab == "PENDENTE_AGEHAB" or any(
        field_name == "status_agehab" and new_value == "PENDENTE_AGEHAB"
        for field_name, new_value, _ in history
    )
    virou_condicionado = status_cca == "CONDICIONADO" or any(
        field_name == "status_cca" and new_value == "CONDICIONADO"
        for field_name, new_value, _ in history
    )
    assinatura_at = _first_frankstein_status_change_at(
        history,
        field_names={"status_cca"},
        new_values={"ASSINATURA_CAIXA", "FINALIZADO"},
    )

    finalizacao_at: Optional[datetime] = None
    if status_cca == "FINALIZADO":
        finalizacao_at = _first_frankstein_status_change_at(
            history,
            field_names={"status_cca"},
            new_values={"FINALIZADO"},
        )
    elif status_geral in {"REPROVADO", "DISTRATO", "CANCELADO"}:
        finalizacao_at = _first_frankstein_status_change_at(
            history,
            field_names={"status_geral"},
            new_values={status_geral},
        )
    elif status_credito == "REPROVADO" or status_cca in {"REPROVADO", "BLOQUEADO"}:
        finalizacao_at = _first_frankstein_status_change_at(
            history,
            field_names={"status_credito", "status_cca"},
            new_values={"REPROVADO", "BLOQUEADO"},
        )

    assinou_caixa = status_cca in {"ASSINATURA_CAIXA", "FINALIZADO"} or assinatura_at is not None
    foi_aprovado = status_credito == "APROVADO" or status_cca in {"APROVADO", "DAR_QV", "ASSINATURA_CAIXA", "FINALIZADO"}
    foi_reprovado = status_credito == "REPROVADO" or status_geral == "REPROVADO" or status_cca in {"REPROVADO", "BLOQUEADO"}
    finalizou = (
        status_cca == "FINALIZADO"
        or status_geral in {"REPROVADO", "DISTRATO", "CANCELADO"}
        or finalizacao_at is not None
    )

    outcome = {
        "teve_pendencia_cca": teve_pendencia_cca,
        "teve_pendencia_agehab": teve_pendencia_agehab,
        "foi_aprovado": foi_aprovado,
        "foi_reprovado": foi_reprovado,
        "virou_condicionado": virou_condicionado,
        "assinou_caixa": assinou_caixa,
        "finalizou": finalizou,
        "retorno_cca": status_cca or None,
        "resultado_real": _frankstein_resultado_real_from_statuses(
            status_credito=status_credito,
            status_geral=status_geral,
            status_cca=status_cca,
            status_agehab=status_agehab,
        ),
    }
    return outcome, assinatura_at, finalizacao_at


def _sync_frankstein_events_for_processo(
    db: Session,
    processo: Processo,
    *,
    lead: Optional[LeadPreCadastro] = None,
) -> int:
    if processo is None:
        return 0

    try:
        resolved_lead = lead or _find_latest_lead_for_processo(db, processo.id)
        _link_frankstein_events_to_processo(db, processo, lead=resolved_lead)
        outcome, assinatura_at, finalizacao_at = _build_frankstein_process_outcome(db, processo)
        processo_id_text = str(processo.id)

        conn = _frankstein_db_conn()
        if conn is not None:
            try:
                updated = 0
                with conn, conn.cursor() as cur:
                    cur.execute(
                        'select event_id, "timestamp" from frankstein_events where trim(coalesce(processo_id, \'\')) = %s',
                        (processo_id_text,),
                    )
                    for event_id, event_timestamp in cur.fetchall():
                        event_at = _frankstein_event_timestamp(event_timestamp)
                        changes = dict(outcome)
                        changes["tempo_ate_assinatura_horas"] = _frankstein_hours_between(event_at, assinatura_at)
                        changes["tempo_total_processo_horas"] = _frankstein_hours_between(event_at, finalizacao_at)
                        assignments = ", ".join(f"{field} = %s" for field in changes.keys())
                        params = [changes[field] for field in changes.keys()]
                        params.extend([datetime.now(timezone.utc), event_id])
                        cur.execute(
                            f"""
                            update frankstein_events
                               set {assignments},
                                   updated_at = %s
                             where event_id = %s::uuid
                            """,
                            params,
                        )
                        updated += cur.rowcount or 0
                if updated > 0:
                    return updated
            except Exception:
                logger.exception("Falha ao sincronizar eventos %s do processo %s no banco; fallback arquivo.", AI_ASSISTANT_DISPLAY_NAME, processo_id_text)
            finally:
                conn.close()

        events = _safe_load_json_list(FRANKSTEIN_EVENTS_PATH)
        if not events:
            return 0

        updated = 0
        for event in events:
            if _clean_optional_text(event.get("processo_id")) != processo_id_text:
                continue
            event_at = _frankstein_event_timestamp(event.get("timestamp"))
            event.update(outcome)
            event["tempo_ate_assinatura_horas"] = _frankstein_hours_between(event_at, assinatura_at)
            event["tempo_total_processo_horas"] = _frankstein_hours_between(event_at, finalizacao_at)
            updated += 1

        if updated > 0:
            _persist_json_list(FRANKSTEIN_EVENTS_PATH, events)
        return updated
    except Exception:
        logger.exception("Falha ao sincronizar resultados %s para o processo %s.", AI_ASSISTANT_DISPLAY_NAME, getattr(processo, "id", None))
        return 0


def _frankstein_predict_prob(model: dict[str, Any], feats: list[float], feature_payload: Optional[dict[str, Any]] = None) -> float:
    kind = _clean_optional_text(model.get("_kind")) or "legacy_json"
    if kind == "sklearn_pipeline":
        try:
            import pandas as pd
        except Exception:
            logger.exception("Pandas indisponivel para inferencia do modelo sklearn do %s.", AI_ASSISTANT_DISPLAY_NAME)
            return 0.62

        predictor = model.get("model")
        feature_columns = model.get("feature_columns") or []
        if predictor is None or not feature_columns:
            return 0.62
        row = {column: (feature_payload or {}).get(column) for column in feature_columns}
        df = pd.DataFrame([row], columns=feature_columns)
        try:
            return max(0.0, min(1.0, float(predictor.predict_proba(df)[0][1])))
        except Exception:
            logger.exception("Falha na inferencia do modelo sklearn do %s.", AI_ASSISTANT_DISPLAY_NAME)
            return 0.62

    means = model.get("means", [])
    stds = model.get("stds", [])
    weights = model.get("weights", [])
    bias = float(model.get("bias", 0.0))
    if not weights or len(weights) != len(feats):
        return 0.62
    total = bias
    for x, m, s, w in zip(feats, means, stds, weights):
        s = s if s and s != 0 else 1.0
        total += ((x - m) / s) * w
    return max(0.0, min(1.0, _sigmoid(total)))


def _build_frankstein_recomendacao_event(
    payload: SimulacaoInput,
    heuristica: dict[str, Any],
    frankstein_payload: FranksteinSugestaoOut,
    request: Request,
    *,
    model_version: Optional[str] = None,
) -> dict[str, Any]:
    feature_payload = _frankstein_feature_payload(payload, heuristica)
    event_timestamp = datetime.now().astimezone().isoformat()
    input_payload = payload.model_dump()
    frankstein_dict = frankstein_payload.model_dump()

    return {
        "event_id": str(uuid.uuid4()),
        "timestamp": event_timestamp,
        "processo_id": _clean_optional_text(getattr(payload, "processo_id", None)),
        "lead_id": _clean_optional_text(getattr(payload, "lead_id", None)),
        "cliente_id": _clean_optional_text(getattr(payload, "cliente_id", None)),
        "reserva_id": _clean_optional_text(getattr(payload, "reserva_id", None)),
        "corretor_id": _clean_optional_text(getattr(payload, "corretor_id", None)),
        "empreendimento": feature_payload.get("empreendimento"),
        "perfil": feature_payload.get("perfil"),
        "renda_bruta": feature_payload.get("renda_bruta"),
        "valor_tabela": feature_payload.get("valor_tabela"),
        "sobrepreco_vila": feature_payload.get("sobrepreco_vila"),
        "valor_obtido": feature_payload.get("valor_obtido"),
        "parcela_caixa": feature_payload.get("parcela_caixa"),
        "preco_digitado_corretor": feature_payload.get("preco_digitado_corretor"),
        "preco_base_politica": feature_payload.get("preco_base_politica"),
        "preco_final": feature_payload.get("preco_final"),
        "entrada_liquida": feature_payload.get("entrada_liquida"),
        "valor_parcela_entrada": feature_payload.get("valor_parcela_entrada"),
        "is_pos_chaves": feature_payload.get("is_pos_chaves"),
        "status_ia_heuristica": feature_payload.get("status_ia_heuristica"),
        "bloqueio_critico": feature_payload.get("bloqueio_critico"),
        "motivo_auditoria": _clean_optional_text(heuristica.get("leitura_executiva_corretor", {}).get("motivo_auditoria")),
        "garantidores_necessarios": heuristica.get("leitura_executiva_corretor", {}).get("garantidores_necessarios"),
        "exposicao_risco": feature_payload.get("exposicao_risco"),
        "alerta_preco": feature_payload.get("alerta_preco"),
        "valor_venda": None,
        "garantido": None,
        "cheque_moradia": None,
        "faltante": None,
        "qtd_problemas_documentais": None,
        "score_risco_regra": None,
        "status_geral_regra": None,
        "decisao_recomendada_regra": None,
        "probabilidade_modelo": _clean_optional_float(frankstein_payload.prob_aceite, digits=6),
        "preco_frankstein": _clean_optional_float(frankstein_payload.preco_frankstein, digits=2),
        "confianca_modelo": _clean_optional_float(frankstein_payload.confianca, digits=6),
        "modelo_versao": _clean_optional_text(model_version),
        "teve_pendencia_cca": None,
        "teve_pendencia_agehab": None,
        "foi_aprovado": None,
        "foi_reprovado": None,
        "virou_condicionado": None,
        "assinou_caixa": None,
        "finalizou": None,
        "tempo_ate_assinatura_horas": None,
        "tempo_total_processo_horas": None,
        "retorno_cca": None,
        "resultado_real": None,
        "input_json": input_payload,
        "heuristica_json": heuristica,
        "frankstein_json": frankstein_dict,
        "features_json": feature_payload,
        "input": input_payload,
        "heuristica": heuristica,
        "frankstein": frankstein_dict,
        "features": feature_payload,
        "origem": "app_frankstein_recomendacao",
        "user_agent": request.headers.get("user-agent"),
        "client_host": request.client.host if request.client else None,
    }


def _build_frankstein_operacional_event(
    payload: FranksteinAnaliseInput,
    resposta: FranksteinRespostaOperacional,
    request: Request,
) -> dict[str, Any]:
    feature_payload = _frankstein_operacional_features(payload, resposta)
    event_timestamp = datetime.now().astimezone().isoformat()
    input_payload = payload.model_dump()
    frankstein_payload = resposta.model_dump()

    return {
        "event_id": str(uuid.uuid4()),
        "timestamp": event_timestamp,
        "processo_id": _clean_optional_text(getattr(payload, "processo_id", None)),
        "lead_id": _clean_optional_text(getattr(payload, "lead_id", None)),
        "cliente_id": _clean_optional_text(getattr(payload, "cliente_id", None)),
        "reserva_id": _clean_optional_text(getattr(payload, "reserva_id", None)),
        "corretor_id": _clean_optional_text(getattr(payload, "corretor_id", None)),
        "empreendimento": feature_payload.get("empreendimento"),
        "perfil": feature_payload.get("perfil"),
        "renda_bruta": feature_payload.get("renda_informada"),
        "valor_tabela": None,
        "sobrepreco_vila": None,
        "valor_obtido": None,
        "parcela_caixa": None,
        "preco_digitado_corretor": None,
        "preco_base_politica": None,
        "preco_final": None,
        "entrada_liquida": None,
        "valor_parcela_entrada": None,
        "is_pos_chaves": None,
        "status_ia_heuristica": None,
        "bloqueio_critico": None,
        "motivo_auditoria": None,
        "garantidores_necessarios": None,
        "exposicao_risco": None,
        "alerta_preco": None,
        "valor_venda": feature_payload.get("valor_venda"),
        "garantido": feature_payload.get("garantido"),
        "cheque_moradia": feature_payload.get("cheque_moradia"),
        "faltante": feature_payload.get("faltante"),
        "qtd_problemas_documentais": feature_payload.get("qtd_problemas_documentais"),
        "score_risco_regra": feature_payload.get("score_risco_regra"),
        "status_geral_regra": feature_payload.get("status_geral_regra"),
        "decisao_recomendada_regra": feature_payload.get("decisao_recomendada_regra"),
        "probabilidade_modelo": None,
        "preco_frankstein": None,
        "confianca_modelo": _clean_optional_float(resposta.frankstein.metricas_operacionais.confianca_modelo, digits=6),
        "modelo_versao": _clean_optional_text(getattr(resposta, "modelo_versao", None) or resposta.frankstein.auditoria.versao),
        "teve_pendencia_cca": None,
        "teve_pendencia_agehab": None,
        "foi_aprovado": None,
        "foi_reprovado": None,
        "virou_condicionado": None,
        "assinou_caixa": None,
        "finalizou": None,
        "tempo_ate_assinatura_horas": None,
        "tempo_total_processo_horas": None,
        "retorno_cca": None,
        "resultado_real": None,
        "input_json": input_payload,
        "heuristica_json": None,
        "frankstein_json": frankstein_payload,
        "features_json": feature_payload,
        "input": input_payload,
        "heuristica": None,
        "frankstein": frankstein_payload,
        "features": feature_payload,
        "origem": "app_frankstein_operacional",
        "user_agent": request.headers.get("user-agent"),
        "client_host": request.client.host if request.client else None,
    }


def _serialize_jsonb_payload(value: Any) -> Optional[str]:
    if value is None:
        return None
    return json.dumps(value, ensure_ascii=False)


def _frankstein_db_conn():
    if not FRANKSTEIN_DB_URL:
        return None
    try:
        return psycopg.connect(FRANKSTEIN_DB_URL, connect_timeout=10)
    except Exception:
        logger.exception("Falha ao conectar FRANKSTEIN_DB_URL; fallback para arquivo.")
        return None


def _ensure_frankstein_tables():
    runtime_ensure_frankstein_tables(conn_factory=_frankstein_db_conn, logger=logger)


def _log_frankstein_event_db(event: dict[str, Any]) -> bool:
    conn = _frankstein_db_conn()
    if conn is None:
        return False
    try:
        with conn, conn.cursor() as cur:
            cur.execute(
                """
                insert into frankstein_events(
                    event_id, "timestamp", processo_id, lead_id, cliente_id, reserva_id, corretor_id, empreendimento, perfil,
                    renda_bruta, valor_tabela, sobrepreco_vila, valor_obtido, parcela_caixa, preco_digitado_corretor,
                    preco_base_politica, preco_final, entrada_liquida, valor_parcela_entrada, is_pos_chaves,
                    status_ia_heuristica, bloqueio_critico, motivo_auditoria, garantidores_necessarios,
                    exposicao_risco, alerta_preco, valor_venda, garantido, cheque_moradia, faltante,
                    qtd_problemas_documentais, score_risco_regra, status_geral_regra, decisao_recomendada_regra,
                    probabilidade_modelo, preco_frankstein, confianca_modelo, modelo_versao,
                    teve_pendencia_cca, teve_pendencia_agehab, foi_aprovado, foi_reprovado, virou_condicionado,
                    assinou_caixa, finalizou, tempo_ate_assinatura_horas, tempo_total_processo_horas,
                    retorno_cca, resultado_real, input_json, heuristica_json, frankstein_json, features_json, origem,
                    updated_at
                )
                values (
                    %s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s,%s,%s,
                    %s,%s,%s,%s,
                    %s,%s,%s::jsonb,%s::jsonb,%s::jsonb,%s::jsonb,%s,now()
                )
                """,
                (
                    event.get("event_id"),
                    event.get("timestamp"),
                    event.get("processo_id"),
                    event.get("lead_id"),
                    event.get("cliente_id"),
                    event.get("reserva_id"),
                    event.get("corretor_id"),
                    event.get("empreendimento"),
                    event.get("perfil"),
                    event.get("renda_bruta"),
                    event.get("valor_tabela"),
                    event.get("sobrepreco_vila"),
                    event.get("valor_obtido"),
                    event.get("parcela_caixa"),
                    event.get("preco_digitado_corretor"),
                    event.get("preco_base_politica"),
                    event.get("preco_final"),
                    event.get("entrada_liquida"),
                    event.get("valor_parcela_entrada"),
                    event.get("is_pos_chaves"),
                    event.get("status_ia_heuristica"),
                    event.get("bloqueio_critico"),
                    event.get("motivo_auditoria"),
                    event.get("garantidores_necessarios"),
                    event.get("exposicao_risco"),
                    event.get("alerta_preco"),
                    event.get("valor_venda"),
                    event.get("garantido"),
                    event.get("cheque_moradia"),
                    event.get("faltante"),
                    event.get("qtd_problemas_documentais"),
                    event.get("score_risco_regra"),
                    event.get("status_geral_regra"),
                    event.get("decisao_recomendada_regra"),
                    event.get("probabilidade_modelo"),
                    event.get("preco_frankstein"),
                    event.get("confianca_modelo"),
                    event.get("modelo_versao"),
                    event.get("teve_pendencia_cca"),
                    event.get("teve_pendencia_agehab"),
                    event.get("foi_aprovado"),
                    event.get("foi_reprovado"),
                    event.get("virou_condicionado"),
                    event.get("assinou_caixa"),
                    event.get("finalizou"),
                    event.get("tempo_ate_assinatura_horas"),
                    event.get("tempo_total_processo_horas"),
                    event.get("retorno_cca"),
                    event.get("resultado_real"),
                    _serialize_jsonb_payload(event.get("input_json") or event.get("input")),
                    _serialize_jsonb_payload(event.get("heuristica_json") or event.get("heuristica")),
                    _serialize_jsonb_payload(event.get("frankstein_json") or event.get("frankstein")),
                    _serialize_jsonb_payload(event.get("features_json") or event.get("features")),
                    event.get("origem") or "app",
                ),
            )
        return True
    except Exception:
        logger.exception("Falha ao gravar evento frankstein no banco; fallback arquivo.")
        return False
    finally:
        conn.close()


def _update_frankstein_event_store(event_id: str, changes: dict[str, Any]) -> bool:
    normalized_event_id = _clean_optional_text(event_id)
    if not normalized_event_id:
        return False

    allowed_fields = {
        "teve_pendencia_cca",
        "teve_pendencia_agehab",
        "foi_aprovado",
        "foi_reprovado",
        "virou_condicionado",
        "assinou_caixa",
        "finalizou",
        "tempo_ate_assinatura_horas",
        "tempo_total_processo_horas",
        "retorno_cca",
        "resultado_real",
    }
    updates = {key: value for key, value in changes.items() if key in allowed_fields and value is not None}
    if not updates:
        return False

    conn = _frankstein_db_conn()
    if conn is not None:
        try:
            assignments = ", ".join(f"{field} = %s" for field in updates.keys())
            params = [updates[field] for field in updates.keys()]
            params.extend([datetime.now(timezone.utc), normalized_event_id])
            with conn, conn.cursor() as cur:
                cur.execute(
                    f"""
                    update frankstein_events
                    set {assignments}, updated_at = %s
                    where event_id = %s::uuid
                    """,
                    params,
                )
                if cur.rowcount > 0:
                    return True
        except Exception:
            logger.exception("Falha ao atualizar frankstein_event %s no banco; fallback arquivo.", normalized_event_id)
        finally:
            conn.close()

    events = _safe_load_json_list(FRANKSTEIN_EVENTS_PATH)
    updated = False
    for event in events:
        if _clean_optional_text(event.get("event_id")) != normalized_event_id:
            continue
        event.update(updates)
        updated = True
        break
    if updated:
        _persist_json_list(FRANKSTEIN_EVENTS_PATH, events)
    return updated


def _load_latest_frankstein_model_from_db() -> Optional[dict[str, Any]]:
    conn = _frankstein_db_conn()
    if conn is None:
        return None
    try:
        with conn, conn.cursor() as cur:
            cur.execute("select artifact from frankstein_models order by created_at desc limit 1;")
            row = cur.fetchone()
            if row and row[0]:
                artifact = row[0]
                if isinstance(artifact, dict):
                    artifact["_kind"] = "legacy_json_db"
                    return artifact
                return None
    except Exception:
        logger.exception("Falha ao carregar modelo frankstein do banco.")
    finally:
        conn.close()
    return None


@app.post("/app/api/frankstein/recomendacao", response_model=FranksteinRecomendacaoOut)
async def recomendar_proposta_frankstein(
    payload: SimulacaoInput,
    request: Request,
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    """
    Camada \"frankstein\":
    - MantÃ©m cÃ¡lculo determinÃ­stico existente.
    - Acrescenta sugestÃ£o baseada em regras suaves ou modelo leve, e registra evento para futuro treinamento.
    """
    resultado = engine_calculo_imobiliario(payload)
    heur_valor = float(resultado["apresentacao_cliente"]["valor_imovel"])
    status_ia = resultado["leitura_executiva_corretor"]["status_ia"]
    risco = resultado["leitura_executiva_corretor"]["risco_exposicao"]
    bloqueio = bool(resultado["leitura_executiva_corretor"].get("bloqueio_critico", False))

    preco_frankstein = heur_valor
    sugestoes: list[str] = []
    motivos: list[str] = []
    prob_aceite = 0.62

    model = _load_frankstein_model() or _load_latest_frankstein_model_from_db()
    feats = _frankstein_features(payload, resultado)
    feature_payload = _frankstein_feature_payload(payload, resultado)
    if model:
        prob_aceite = _frankstein_predict_prob(model, feats, feature_payload)
        motivos.append("Probabilidade estimada via modelo frankstein.")
    if bloqueio:
        preco_frankstein = round(heur_valor * 0.97, 2)
        sugestoes.append("Reduzir preÃ§o em ~3% para aliviar IS pÃ³s-chaves.")
        motivos.append("Bloqueio crÃ­tico pela regra de IS >= 40%.")
        prob_aceite = min(prob_aceite, 0.4)
    else:
        motivos.append("Perfil dentro dos limites heurÃ­sticos atuais.")
        if risco:
            motivos.append(f"Risco de exposiÃ§Ã£o: {risco}.")

    # Ajuste suave de preÃ§o baseado na probabilidade.
    preco_frankstein = round(heur_valor * (0.97 + 0.06 * prob_aceite), 2)
    confianca = prob_aceite

    frankstein_payload = FranksteinSugestaoOut(
        preco_heuristico=heur_valor,
        preco_frankstein=preco_frankstein,
        prob_aceite=prob_aceite,
        status_ia=status_ia,
        risco_exposicao=risco,
        confianca=confianca,
        motivos=motivos,
        sugestoes=sugestoes,
    )

    event = _build_frankstein_recomendacao_event(
        payload,
        resultado,
        frankstein_payload,
        request,
        model_version=_frankstein_model_version(model) or "fallback-0.62",
    )
    processo, lead = _resolve_frankstein_process_context(
        db,
        processo_id=getattr(payload, "processo_id", None),
        cliente_id=getattr(payload, "cliente_id", None),
        reserva_id=getattr(payload, "reserva_id", None),
        lead_id=getattr(payload, "lead_id", None),
    )
    event = _attach_frankstein_event_context(event, processo=processo, lead=lead)
    _log_frankstein_event(event)
    if processo is not None:
        _remember_processo_frankstein_event(processo, event)
        try:
            db.commit()
        except SQLAlchemyError:
            db.rollback()
            logger.exception("Falha ao persistir rastreio do ultimo evento %s para processo %s.", AI_ASSISTANT_DISPLAY_NAME, processo.id)
        else:
            _sync_frankstein_events_for_processo(db, processo, lead=lead)

    return FranksteinRecomendacaoOut(
        event_id=event["event_id"],
        modelo_versao=event.get("modelo_versao"),
        heuristica=resultado,
        frankstein=frankstein_payload,
    )


@app.post("/app/api/frankstein/analisar", response_model=FranksteinRespostaOperacional)
@app.post("/frankstein/analisar", response_model=FranksteinRespostaOperacional)
def frankstein_analisar_operacional(
    payload: FranksteinAnaliseInput,
    request: Request,
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
) -> FranksteinRespostaOperacional:
    resposta = analisar_operacao_frankstein(payload)
    event = _build_frankstein_operacional_event(payload, resposta, request)
    processo, lead = _resolve_frankstein_process_context(
        db,
        processo_id=getattr(payload, "processo_id", None),
        cliente_id=getattr(payload, "cliente_id", None),
        reserva_id=getattr(payload, "reserva_id", None),
        lead_id=getattr(payload, "lead_id", None),
    )
    event = _attach_frankstein_event_context(event, processo=processo, lead=lead)
    _log_frankstein_event(event)
    if processo is not None:
        _remember_processo_frankstein_event(processo, event)
        try:
            db.commit()
        except SQLAlchemyError:
            db.rollback()
            logger.exception("Falha ao persistir rastreio do ultimo evento %s para processo %s.", AI_ASSISTANT_DISPLAY_NAME, processo.id)
        else:
            _sync_frankstein_events_for_processo(db, processo, lead=lead)
    resposta.event_id = event["event_id"]
    resposta.modelo_versao = event.get("modelo_versao")
    return resposta


@app.patch("/app/api/frankstein/events/{event_id}")
def atualizar_resultado_frankstein_event(
    event_id: str,
    payload: FranksteinEventResultadoUpdate,
    _: dict[str, Any] = Depends(require_roles(ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
):
    changes = payload.model_dump(exclude_none=True)
    updated = _update_frankstein_event_store(event_id, changes)
    if not updated:
        raise HTTPException(status_code=404, detail=f"Evento {AI_ASSISTANT_DISPLAY_NAME} nao encontrado")
    return {"ok": True, "event_id": event_id, "updated_fields": sorted(changes.keys())}


@app.post("/app/api/frankstein/analista-aprendizado", response_model=FranksteinAnalistaAprendizadoOut)
def registrar_aprendizado_analista_frankstein(
    payload: FranksteinAnalistaAprendizadoPayload,
    session: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
) -> FranksteinAnalistaAprendizadoOut:
    actor_username = str(session.get("username") or "analista")
    actor_role = _normalize_role(str(session.get("role") or ROLE_ANALISTA))
    details_payload = {
        "processo_id": str(payload.processo_id) if payload.processo_id else None,
        "contexto": payload.contexto,
        "decisao_analista": payload.decisao_analista,
        "regras_frankstein": payload.regras_frankstein[:40],
    }
    _record_system_log(
        db,
        actor_username=actor_username,
        actor_role=actor_role,
        tela="frankstein",
        acao="ANALISTA_APRENDIZADO",
        entidade_tipo="processo",
        entidade_id=str(payload.processo_id) if payload.processo_id else None,
        details=json.dumps(details_payload, ensure_ascii=False, default=str),
    )
    db.commit()
    padroes = _frankstein_learning_suggestions(db, payload.contexto)
    total_interacoes = int(
        db.query(func.count(SistemaLog.id))
        .filter(SistemaLog.tela == "frankstein")
        .filter(SistemaLog.acao == "ANALISTA_APRENDIZADO")
        .scalar()
        or 0
    )
    return FranksteinAnalistaAprendizadoOut(ok=True, total_interacoes=total_interacoes, padroes=padroes)


@app.post("/app/api/frankstein/analista-aprendizado/sugestoes", response_model=FranksteinAnalistaAprendizadoOut)
def sugestoes_aprendizado_analista_frankstein(
    payload: FranksteinAnalistaAprendizadoPayload,
    _: dict[str, Any] = Depends(require_roles(ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
) -> FranksteinAnalistaAprendizadoOut:
    total_interacoes = int(
        db.query(func.count(SistemaLog.id))
        .filter(SistemaLog.tela == "frankstein")
        .filter(SistemaLog.acao == "ANALISTA_APRENDIZADO")
        .scalar()
        or 0
    )
    return FranksteinAnalistaAprendizadoOut(
        ok=True,
        total_interacoes=total_interacoes,
        padroes=_frankstein_learning_suggestions(db, payload.contexto),
    )


@app.post("/app/api/analises")
async def criar_analise(
    payload: AnaliseCreate,
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    registro = AnaliseRegistroDB(
        empreendimento=payload.empreendimento,
        unidade=payload.unidade,
        preco_imovel=payload.preco_imovel,
        valor_obtido=payload.valor_obtido,
        prosoluto_calculado=payload.prosoluto_calculado,
        prosoluto_liquido=payload.prosoluto_liquido,
        sinal=payload.sinal,
        sinal_produto=payload.sinal_produto,
        financiamento=payload.financiamento,
        subsidio=payload.subsidio,
        cheque_moradia=payload.cheque_moradia,
        renda_bruta=payload.renda_bruta,
        perc_construcao=payload.perc_construcao,
        is_agora=payload.is_agora,
        is_pos_chaves=payload.is_pos_chaves,
        tabela_referencia=json.dumps(payload.tabela_referencia, ensure_ascii=False),
        data_referencia=payload.data_referencia,
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)
    return {"id": registro.id, "created_at": registro.created_at}


@app.get("/app/api/analises")
async def listar_analises(
    limit: int = 50,
    _: dict[str, Any] = Depends(require_roles(ROLE_CORRETOR, ROLE_CCA, ROLE_ANALISTA, ROLE_GESTOR, ROLE_GESTOR_CREDITO, ROLE_ADMIN)),
    db: Session = Depends(get_db),
):
    rows = (
        db.query(AnaliseRegistroDB)
        .order_by(AnaliseRegistroDB.created_at.desc())
        .limit(min(max(limit, 1), 200))
        .all()
    )
    return [
        {
            "id": row.id,
            "created_at": row.created_at,
            "empreendimento": row.empreendimento,
            "unidade": row.unidade,
            "preco_imovel": row.preco_imovel,
            "valor_obtido": row.valor_obtido,
            "prosoluto_calculado": row.prosoluto_calculado,
            "prosoluto_liquido": row.prosoluto_liquido,
            "sinal": row.sinal,
            "sinal_produto": row.sinal_produto,
            "financiamento": row.financiamento,
            "subsidio": row.subsidio,
            "cheque_moradia": row.cheque_moradia,
            "renda_bruta": row.renda_bruta,
            "perc_construcao": row.perc_construcao,
            "is_agora": row.is_agora,
            "is_pos_chaves": row.is_pos_chaves,
            "tabela_referencia": json.loads(row.tabela_referencia or "[]"),
            "data_referencia": row.data_referencia,
        }
        for row in rows
    ]

